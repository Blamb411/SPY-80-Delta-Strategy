#!/usr/bin/env python3
"""
Alpha Picks Options P&L Model — ThetaData Historical Validation
================================================================
Replaces Black-Scholes synthetic pricing with real historical options data
from ThetaData. For each Alpha Pick, fetches actual bid/ask quotes, IV,
and Greeks to model realistic P&L under three execution scenarios:

  1. Mid-price: (bid + ask) / 2
  2. Natural (worst fill): buy at ask, sell at bid
  3. 25% slippage: mid + 25% of spread to buy, mid - 25% of spread to sell

Falls back to B-S synthetic pricing when ThetaData has no data for a ticker.

Requires:
  - Theta Terminal v3 running locally
  - price_cache.db (stock prices, already populated)
  - ProQuant History 1_29_2026.xlsx (Alpha Picks list)

Usage:
    python alpha_picks_options_thetadata.py
"""

import os
import sys
import math
import sqlite3
import logging
from datetime import datetime, date, timedelta
from collections import defaultdict

import openpyxl

# Add parent dir so we can import backtest modules
_this_dir = os.path.dirname(os.path.abspath(__file__))
_project_dir = os.path.dirname(_this_dir)
sys.path.insert(0, _project_dir)

from backtest.thetadata_client import ThetaDataClient
from backtest.black_scholes import black_scholes_price, find_strike_for_delta

log = logging.getLogger("alpha_thetadata")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
XLSX_PATH = os.path.join(_this_dir, "ProQuant History 1_29_2026.xlsx")
STOCK_DB_PATH = os.path.join(_this_dir, "price_cache.db")

DELTAS = [0.80, 0.50]
DTE_TARGET = 60           # target DTE at entry
DTE_MIN = 45              # minimum acceptable DTE
DTE_MAX = 90              # maximum acceptable DTE
RATE = 0.05               # risk-free rate for B-S fallback

# IV estimation for B-S strike targeting (used to find initial strike guess)
IV_LOOKBACK = 30
IV_PREMIUM = 1.2
IV_FLOOR = 0.15
IV_CAP = 1.00

# Windows for fixed-window analysis
WINDOWS = [1, 10, 30, 60]

# Rule-based exit parameters
STOP_LOSSES = [None, -0.30, -0.50, -0.70]
PROFIT_TARGETS = [None, 0.50, 1.00, 2.00]
MAX_HOLDS = [10, 30, 60]
MAX_SIM_DAYS = 60

# Execution scenarios
SCENARIOS = ["mid", "natural", "slippage25"]


# ---------------------------------------------------------------------------
# Data loading (same as synthetic model)
# ---------------------------------------------------------------------------

def load_alpha_picks():
    """Load Alpha Picks from Excel."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["AlphaPicks"]
    picks = []
    seen = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] is None or row[2] is None:
            continue
        symbol = str(row[1]).strip()
        pick_date = row[2]
        if isinstance(pick_date, datetime):
            date_str = pick_date.strftime("%Y-%m-%d")
        else:
            date_str = str(pick_date)[:10]
        buy_price = float(row[3]) if row[3] else None
        key = (symbol, date_str)
        if key in seen:
            continue
        seen.add(key)
        picks.append({
            "symbol": symbol,
            "pick_date": date_str,
            "buy_price": buy_price,
        })
    return picks


def load_prices_from_cache():
    """Load all stock prices from SQLite cache."""
    prices = defaultdict(dict)
    if not os.path.exists(STOCK_DB_PATH):
        return prices
    conn = sqlite3.connect(STOCK_DB_PATH)
    c = conn.cursor()
    c.execute("SELECT symbol, date, close FROM daily_prices ORDER BY symbol, date")
    for symbol, dt, close in c.fetchall():
        prices[symbol][dt] = close
    conn.close()
    return prices


def get_trading_dates(prices):
    """Get sorted list of all trading dates from SPY."""
    if "SPY" in prices:
        return sorted(prices["SPY"].keys())
    all_dates = set()
    for sym_prices in prices.values():
        all_dates.update(sym_prices.keys())
    return sorted(all_dates)


def find_trading_date(trading_dates, target, direction="on_or_before"):
    """Find the nearest trading date to target."""
    if target in trading_dates:
        return target
    for i, d in enumerate(trading_dates):
        if d > target:
            if direction == "on_or_before":
                return trading_dates[i - 1] if i > 0 else None
            else:
                return d
    if direction == "on_or_before":
        return trading_dates[-1] if trading_dates else None
    return None


def offset_trading_days(trading_dates, base_date, offset):
    """Move N trading days from base_date."""
    if base_date not in trading_dates:
        base_date = find_trading_date(trading_dates, base_date, "on_or_after")
        if not base_date:
            return None
    idx = trading_dates.index(base_date)
    target_idx = idx + offset
    if 0 <= target_idx < len(trading_dates):
        return trading_dates[target_idx]
    return None


# ---------------------------------------------------------------------------
# IV estimation (for B-S strike targeting only)
# ---------------------------------------------------------------------------

def estimate_iv(prices_dict, symbol, date_str, trading_dates, lookback=IV_LOOKBACK):
    """Estimate IV from trailing realized vol (used to find initial strike guess)."""
    if symbol not in prices_dict:
        return None
    sym_prices = prices_dict[symbol]
    if date_str not in trading_dates:
        date_str = find_trading_date(trading_dates, date_str, "on_or_before")
        if not date_str:
            return None
    idx = trading_dates.index(date_str)
    start_idx = idx - lookback
    if start_idx < 0:
        return None
    closes = []
    for i in range(start_idx, idx + 1):
        d = trading_dates[i]
        if d in sym_prices:
            closes.append(sym_prices[d])
    if len(closes) < lookback // 2:
        return None
    log_returns = []
    for i in range(1, len(closes)):
        if closes[i] > 0 and closes[i - 1] > 0:
            log_returns.append(math.log(closes[i] / closes[i - 1]))
    if len(log_returns) < 5:
        return None
    mean_ret = sum(log_returns) / len(log_returns)
    variance = sum((r - mean_ret) ** 2 for r in log_returns) / (len(log_returns) - 1)
    realized_vol = math.sqrt(variance * 252)
    iv = realized_vol * IV_PREMIUM
    return max(IV_FLOOR, min(IV_CAP, iv))


# ---------------------------------------------------------------------------
# Execution price helpers
# ---------------------------------------------------------------------------

def entry_price_for_scenario(bid, ask, scenario):
    """Compute the price to BUY an option under each scenario."""
    if bid is None or ask is None or bid <= 0 or ask <= 0:
        return None
    mid = (bid + ask) / 2.0
    spread = ask - bid
    if scenario == "mid":
        return mid
    elif scenario == "natural":
        return ask  # buy at ask (worst fill)
    elif scenario == "slippage25":
        return mid + 0.25 * spread
    return mid


def exit_price_for_scenario(bid, ask, scenario):
    """Compute the price to SELL an option under each scenario."""
    if bid is None or ask is None or bid <= 0 or ask <= 0:
        return None
    mid = (bid + ask) / 2.0
    spread = ask - bid
    if scenario == "mid":
        return mid
    elif scenario == "natural":
        return bid  # sell at bid (worst fill)
    elif scenario == "slippage25":
        return mid - 0.25 * spread
    return mid


# ---------------------------------------------------------------------------
# Core: model one pick using ThetaData
# ---------------------------------------------------------------------------

def model_pick_thetadata(pick, prices, trading_dates, client, entry_offset=0):
    """
    Model call option P&L for one Alpha Pick using real ThetaData data.

    Args:
        pick: dict with symbol, pick_date, buy_price
        prices: stock price cache
        trading_dates: sorted date list
        client: ThetaDataClient instance
        entry_offset: -1 = T-1 close, 0 = T0 close

    Returns dict with results per delta and scenario, or None.
    """
    sym = pick["symbol"]
    pick_date = pick["pick_date"]

    if sym not in prices or not prices[sym]:
        return None

    # T-1: trading day before announcement
    t_minus1 = find_trading_date(trading_dates, pick_date, "on_or_before")
    if t_minus1 == pick_date:
        idx = trading_dates.index(t_minus1)
        t_minus1 = trading_dates[idx - 1] if idx > 0 else None
    if not t_minus1:
        return None

    # T0: announcement day
    t0 = find_trading_date(trading_dates, pick_date, "on_or_after")
    if not t0:
        return None

    # Entry date
    entry_date = t0 if entry_offset == 0 else t_minus1

    spot_entry = prices[sym].get(entry_date)
    if not spot_entry:
        return None

    spy_entry = prices.get("SPY", {}).get(entry_date)

    # Estimate IV for B-S strike targeting
    est_iv = estimate_iv(prices, sym, entry_date, trading_dates)
    if est_iv is None:
        est_iv = 0.35  # fallback for strike estimation

    result = {
        "symbol": sym,
        "pick_date": pick_date,
        "t_minus1": t_minus1,
        "t0": t0,
        "entry_date": entry_date,
        "entry_offset": entry_offset,
        "spot_entry": spot_entry,
        "spy_entry": spy_entry,
        "data_source": "thetadata",  # may be changed to "synthetic" below
    }

    # Find nearest expiration ~60 DTE
    expiration = client.find_nearest_expiration(
        sym, entry_date, target_dte=DTE_TARGET, dte_min=DTE_MIN, dte_max=DTE_MAX
    )
    if not expiration:
        result["data_source"] = "no_expiration"
        return None

    result["expiration"] = expiration
    entry_dt = datetime.strptime(entry_date, "%Y-%m-%d").date()
    exp_dt = datetime.strptime(expiration, "%Y-%m-%d").date()
    actual_dte = (exp_dt - entry_dt).days
    result["actual_dte"] = actual_dte

    for delta in DELTAS:
        dk = f"d{int(delta * 100)}"

        # Use B-S to estimate the strike at target delta
        t_years = actual_dte / 365.0
        bs_strike = find_strike_for_delta(spot_entry, t_years, RATE, est_iv, delta, 'C')
        if bs_strike is None:
            result[f"{dk}_strike"] = None
            continue

        # Snap to nearest real strike
        real_strike = client.snap_strike(sym, expiration, bs_strike)
        if real_strike is None:
            result[f"{dk}_strike"] = None
            continue

        result[f"{dk}_strike"] = real_strike
        result[f"{dk}_bs_strike"] = bs_strike

        # Prefetch all EOD data for this option from entry through expiration
        eod_data = client.prefetch_option_life(
            sym, expiration, real_strike, "C", entry_date
        )
        if not eod_data:
            result[f"{dk}_strike"] = None
            log.debug("No EOD data for %s %s C%.2f from %s",
                      sym, expiration, real_strike, entry_date)
            continue

        # Also fetch Greeks for the entry date
        greeks_data = client.get_option_greeks(
            sym, expiration, real_strike, "C", entry_date, expiration
        )

        # Build lookup dicts for EOD and Greeks by date
        eod_by_date = {row["bar_date"]: row for row in eod_data}
        greeks_by_date = {row["greeks_date"]: row for row in greeks_data}

        # Entry data
        entry_eod = eod_by_date.get(entry_date)
        entry_greeks = greeks_by_date.get(entry_date)

        if entry_eod is None:
            # Try nearest available date
            available_dates = sorted(eod_by_date.keys())
            nearest = None
            for d in available_dates:
                if d >= entry_date:
                    nearest = d
                    break
            if nearest is None and available_dates:
                nearest = available_dates[-1]
            if nearest:
                entry_eod = eod_by_date[nearest]
                entry_greeks = greeks_by_date.get(nearest)
                result[f"{dk}_actual_entry_date"] = nearest
            else:
                result[f"{dk}_strike"] = None
                continue

        # Record entry Greeks
        if entry_greeks:
            result[f"{dk}_actual_iv"] = entry_greeks.get("iv")
            result[f"{dk}_actual_delta"] = entry_greeks.get("delta")
            result[f"{dk}_theta"] = entry_greeks.get("theta")

        entry_bid = entry_eod.get("bid", 0)
        entry_ask = entry_eod.get("ask", 0)
        result[f"{dk}_entry_bid"] = entry_bid
        result[f"{dk}_entry_ask"] = entry_ask

        # Check for valid bid/ask
        if entry_bid <= 0 or entry_ask <= 0:
            # Fall back to close price if bid/ask are zero
            close_price = entry_eod.get("close", 0)
            if close_price > 0:
                entry_bid = close_price * 0.97  # estimate 3% spread
                entry_ask = close_price * 1.03
                result[f"{dk}_entry_bid"] = entry_bid
                result[f"{dk}_entry_ask"] = entry_ask
                result[f"{dk}_bid_ask_estimated"] = True
            else:
                result[f"{dk}_strike"] = None
                continue

        # Compute entry prices for each scenario
        for scenario in SCENARIOS:
            ep = entry_price_for_scenario(entry_bid, entry_ask, scenario)
            result[f"{dk}_{scenario}_entry"] = ep

        # Fixed-window exits
        for w in WINDOWS:
            t_exit = offset_trading_days(trading_dates, t0, w)
            if not t_exit:
                for scenario in SCENARIOS:
                    result[f"{dk}_{scenario}_pnl_{w}d"] = None
                    result[f"{dk}_{scenario}_ret_{w}d"] = None
                result[f"{dk}_stock_ret_{w}d"] = None
                result[f"{dk}_spy_ret_{w}d"] = None
                continue

            spot_exit = prices[sym].get(t_exit)
            exit_eod = eod_by_date.get(t_exit)

            # Stock return
            if spot_exit:
                result[f"{dk}_stock_ret_{w}d"] = (spot_exit / spot_entry - 1)
            else:
                result[f"{dk}_stock_ret_{w}d"] = None

            # SPY return
            spy_exit = prices.get("SPY", {}).get(t_exit)
            if spy_entry and spy_exit:
                result[f"{dk}_spy_ret_{w}d"] = (spy_exit / spy_entry - 1)
            else:
                result[f"{dk}_spy_ret_{w}d"] = None

            if exit_eod is None:
                for scenario in SCENARIOS:
                    result[f"{dk}_{scenario}_pnl_{w}d"] = None
                    result[f"{dk}_{scenario}_ret_{w}d"] = None
                continue

            exit_bid = exit_eod.get("bid", 0)
            exit_ask = exit_eod.get("ask", 0)

            # Fall back to close if bid/ask zero
            if exit_bid <= 0 or exit_ask <= 0:
                close_price = exit_eod.get("close", 0)
                if close_price > 0:
                    exit_bid = close_price * 0.97
                    exit_ask = close_price * 1.03
                else:
                    for scenario in SCENARIOS:
                        result[f"{dk}_{scenario}_pnl_{w}d"] = None
                        result[f"{dk}_{scenario}_ret_{w}d"] = None
                    continue

            for scenario in SCENARIOS:
                ep = result.get(f"{dk}_{scenario}_entry")
                if ep is None or ep <= 0:
                    result[f"{dk}_{scenario}_pnl_{w}d"] = None
                    result[f"{dk}_{scenario}_ret_{w}d"] = None
                    continue
                xp = exit_price_for_scenario(exit_bid, exit_ask, scenario)
                if xp is None:
                    result[f"{dk}_{scenario}_pnl_{w}d"] = None
                    result[f"{dk}_{scenario}_ret_{w}d"] = None
                    continue
                pnl = (xp - ep) * 100  # per contract
                ret = (xp / ep - 1) if ep > 0 else None
                result[f"{dk}_{scenario}_pnl_{w}d"] = pnl
                result[f"{dk}_{scenario}_ret_{w}d"] = ret

        # Daily series for rule-based exits
        daily_series = {}
        for scenario in SCENARIOS:
            daily_series[scenario] = []

        # Build daily data from T0+1 through max simulation days
        for day in range(1, MAX_SIM_DAYS + 1):
            t_date = offset_trading_days(trading_dates, t0, day)
            if not t_date:
                break

            spot = prices[sym].get(t_date)
            day_eod = eod_by_date.get(t_date)

            if spot is None or day_eod is None:
                continue

            day_bid = day_eod.get("bid", 0)
            day_ask = day_eod.get("ask", 0)

            # Fall back to close
            if day_bid <= 0 or day_ask <= 0:
                close_price = day_eod.get("close", 0)
                if close_price > 0:
                    day_bid = close_price * 0.97
                    day_ask = close_price * 1.03
                else:
                    continue

            stock_return = (spot / spot_entry - 1) if spot_entry > 0 else 0

            for scenario in SCENARIOS:
                ep = result.get(f"{dk}_{scenario}_entry")
                if ep is None or ep <= 0:
                    continue
                xp = exit_price_for_scenario(day_bid, day_ask, scenario)
                if xp is None:
                    continue
                opt_return = (xp / ep - 1) if ep > 0 else 0
                daily_series[scenario].append({
                    "day": day,
                    "date": t_date,
                    "spot": spot,
                    "opt_price": xp,
                    "opt_return": opt_return,
                    "stock_return": stock_return,
                })

        for scenario in SCENARIOS:
            result[f"{dk}_{scenario}_daily"] = daily_series[scenario] if daily_series[scenario] else None

    return result


# ---------------------------------------------------------------------------
# Rule-based exits (same logic as synthetic model)
# ---------------------------------------------------------------------------

def apply_rules(daily_series, stop_loss, profit_target, max_hold):
    """Simulate a trade with stop-loss, profit-target, and max-hold rules."""
    if not daily_series:
        return None
    last_valid = None
    for entry in daily_series:
        day = entry["day"]
        ret = entry["opt_return"]
        if day > max_hold:
            break
        last_valid = entry
        if stop_loss is not None and ret <= stop_loss:
            return {
                "exit_day": day, "exit_date": entry["date"],
                "exit_reason": "stop_loss", "return_pct": ret,
                "exit_price": entry["opt_price"], "spot_exit": entry["spot"],
                "stock_return": entry["stock_return"],
            }
        if profit_target is not None and ret >= profit_target:
            return {
                "exit_day": day, "exit_date": entry["date"],
                "exit_reason": "profit_target", "return_pct": ret,
                "exit_price": entry["opt_price"], "spot_exit": entry["spot"],
                "stock_return": entry["stock_return"],
            }
    if last_valid:
        return {
            "exit_day": last_valid["day"], "exit_date": last_valid["date"],
            "exit_reason": "max_hold", "return_pct": last_valid["opt_return"],
            "exit_price": last_valid["opt_price"], "spot_exit": last_valid["spot"],
            "stock_return": last_valid["stock_return"],
        }
    return None


def compute_stats(values):
    """Compute mean, median, win rate, min, max."""
    if not values:
        return None
    n = len(values)
    mean = sum(values) / n
    sorted_v = sorted(values)
    median = sorted_v[n // 2]
    win_rate = sum(1 for v in values if v > 0) / n
    var = sum((r - mean) ** 2 for r in values) / n if n > 1 else 0
    std = math.sqrt(var) if var > 0 else 0.001
    return {
        "n": n, "mean": mean, "median": median,
        "win_rate": win_rate, "min": min(values), "max": max(values),
        "std": std, "sharpe_like": mean / std if std > 0 else 0,
    }


def run_rule_analysis(results, scenario="mid"):
    """Run all rule combinations for a given scenario."""
    all_stats = {}
    for dk in ["d80", "d50"]:
        rule_results = []
        for mh in MAX_HOLDS:
            for sl in STOP_LOSSES:
                for pt in PROFIT_TARGETS:
                    trade_returns = []
                    trade_days = []
                    exit_reasons = {"stop_loss": 0, "profit_target": 0, "max_hold": 0}
                    n_trades = 0
                    for r in results:
                        daily = r.get(f"{dk}_{scenario}_daily")
                        if daily is None:
                            continue
                        exit_info = apply_rules(daily, sl, pt, mh)
                        if exit_info is None:
                            continue
                        n_trades += 1
                        trade_returns.append(exit_info["return_pct"])
                        trade_days.append(exit_info["exit_day"])
                        reason = exit_info["exit_reason"]
                        if reason in exit_reasons:
                            exit_reasons[reason] += 1
                    if not trade_returns:
                        continue
                    s = compute_stats(trade_returns)
                    s["avg_days"] = sum(trade_days) / len(trade_days)
                    s["exit_reasons"] = exit_reasons
                    s["pct_sl"] = exit_reasons["stop_loss"] / n_trades if n_trades else 0
                    s["pct_pt"] = exit_reasons["profit_target"] / n_trades if n_trades else 0
                    s["pct_mh"] = exit_reasons["max_hold"] / n_trades if n_trades else 0
                    rule_results.append(((sl, pt, mh), s))
        all_stats[dk] = rule_results
    return all_stats


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def fmt_sl(sl):
    return "None" if sl is None else f"{sl:+.0%}"

def fmt_pt(pt):
    return "None" if pt is None else f"+{pt:.0%}"


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def print_data_coverage(results, skipped_symbols):
    """Report on data coverage."""
    print()
    print("=" * 120)
    print("DATA COVERAGE")
    print("=" * 120)
    print(f"  Picks with ThetaData options data: {len(results)}")
    print(f"  Picks skipped (no data):           {len(skipped_symbols)}")
    if skipped_symbols:
        print(f"  Skipped symbols: {', '.join(sorted(skipped_symbols))}")
    print()

    # Check how many have valid d80/d50 data per scenario
    for dk, label in [("d80", "80-Delta"), ("d50", "50-Delta")]:
        for scenario in SCENARIOS:
            count = sum(1 for r in results
                        if r.get(f"{dk}_{scenario}_entry") is not None
                        and r.get(f"{dk}_{scenario}_entry", 0) > 0)
            print(f"  {label} {scenario:>10}: {count} picks with valid entry pricing")

    # IV and delta stats
    print()
    for dk, label in [("d80", "80-Delta"), ("d50", "50-Delta")]:
        ivs = [r[f"{dk}_actual_iv"] for r in results
               if r.get(f"{dk}_actual_iv") is not None and r[f"{dk}_actual_iv"] > 0]
        deltas = [r[f"{dk}_actual_delta"] for r in results
                  if r.get(f"{dk}_actual_delta") is not None]
        if ivs:
            print(f"  {label} actual IV:    mean={sum(ivs)/len(ivs):.2%}"
                  f"  median={sorted(ivs)[len(ivs)//2]:.2%}"
                  f"  range=[{min(ivs):.2%}, {max(ivs):.2%}]  N={len(ivs)}")
        if deltas:
            print(f"  {label} actual delta: mean={sum(deltas)/len(deltas):.3f}"
                  f"  median={sorted(deltas)[len(deltas)//2]:.3f}"
                  f"  range=[{min(deltas):.3f}, {max(deltas):.3f}]  N={len(deltas)}")

    # Bid-ask spread stats
    print()
    for dk, label in [("d80", "80-Delta"), ("d50", "50-Delta")]:
        spreads = []
        for r in results:
            bid = r.get(f"{dk}_entry_bid", 0)
            ask = r.get(f"{dk}_entry_ask", 0)
            if bid > 0 and ask > 0:
                mid = (bid + ask) / 2
                pct_spread = (ask - bid) / mid if mid > 0 else 0
                spreads.append(pct_spread)
        if spreads:
            print(f"  {label} bid-ask spread: mean={sum(spreads)/len(spreads):.2%}"
                  f"  median={sorted(spreads)[len(spreads)//2]:.2%}"
                  f"  N={len(spreads)}")


def print_fixed_window_summary(results, scenario="mid"):
    """Print summary stats for fixed-window analysis."""
    print()
    print("=" * 140)
    scenario_label = {"mid": "Mid-Price", "natural": "Natural (Worst Fill)",
                      "slippage25": "25% Slippage"}
    print(f"FIXED-WINDOW SUMMARY — {scenario_label.get(scenario, scenario)}")
    print("=" * 140)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        print(f"\n--- {label} Calls ---")
        print(f"{'Window':<12} | {'N':>4} | {'Mean':>8} | {'Median':>8}"
              f" | {'Win%':>6} | {'Min':>8} | {'Max':>8}"
              f" | {'Stock':>8} | {'SPY':>8} | {'Alpha':>8}")
        print("-" * 120)

        for w in WINDOWS:
            rets = [r[f"{dk}_{scenario}_ret_{w}d"] for r in results
                    if r.get(f"{dk}_{scenario}_ret_{w}d") is not None]
            stock_rets = [r[f"{dk}_stock_ret_{w}d"] for r in results
                         if r.get(f"{dk}_stock_ret_{w}d") is not None
                         and r.get(f"{dk}_{scenario}_ret_{w}d") is not None]
            spy_rets = [r[f"{dk}_spy_ret_{w}d"] for r in results
                        if r.get(f"{dk}_spy_ret_{w}d") is not None
                        and r.get(f"{dk}_{scenario}_ret_{w}d") is not None]

            s = compute_stats(rets)
            if not s:
                continue

            stock_mean = sum(stock_rets) / len(stock_rets) if stock_rets else 0
            spy_mean = sum(spy_rets) / len(spy_rets) if spy_rets else 0
            alpha = stock_mean - spy_mean

            print(f"T+{w:<9} | {s['n']:>4} | {s['mean']:>+7.2%} | {s['median']:>+7.2%}"
                  f" | {s['win_rate']:>5.1%} | {s['min']:>+7.1%} | {s['max']:>+7.1%}"
                  f" | {stock_mean:>+7.2%} | {spy_mean:>+7.2%} | {alpha:>+7.2%}")


def print_scenario_comparison(results):
    """Compare the three execution scenarios side by side."""
    print()
    print("=" * 150)
    print("EXECUTION SCENARIO COMPARISON — Fixed-Window Returns")
    print("=" * 150)
    print("  mid       = buy/sell at midpoint of bid-ask")
    print("  natural   = buy at ask, sell at bid (worst realistic fill)")
    print("  slippage  = mid +/- 25% of spread")
    print()

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        print(f"\n  --- {label} ---")
        print(f"  {'Window':<10}"
              f" | {'--- Mid-Price ---':>30}"
              f" | {'--- Natural ---':>30}"
              f" | {'--- 25% Slippage ---':>30}")
        print(f"  {'':10}"
              f" | {'N':>4} {'Mean':>8} {'Med':>8} {'Win%':>6}"
              f" | {'N':>4} {'Mean':>8} {'Med':>8} {'Win%':>6}"
              f" | {'N':>4} {'Mean':>8} {'Med':>8} {'Win%':>6}")
        print(f"  {'-' * 105}")

        for w in WINDOWS:
            parts = [f"  T+{w:<8}"]
            for scenario in SCENARIOS:
                rets = [r[f"{dk}_{scenario}_ret_{w}d"] for r in results
                        if r.get(f"{dk}_{scenario}_ret_{w}d") is not None]
                s = compute_stats(rets)
                if s:
                    parts.append(
                        f" | {s['n']:>4} {s['mean']:>+7.2%} {s['median']:>+7.2%} {s['win_rate']:>5.1%}"
                    )
                else:
                    parts.append(f" |  --- no data ---            ")
            print("".join(parts))


def print_rule_tables(all_stats, scenario="mid"):
    """Print rule-based exit analysis for a given scenario."""
    scenario_label = {"mid": "Mid-Price", "natural": "Natural (Worst Fill)",
                      "slippage25": "25% Slippage"}
    print()
    print("=" * 160)
    print(f"RULE-BASED EXIT ANALYSIS — {scenario_label.get(scenario, scenario)}")
    print("=" * 160)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        rules = all_stats.get(dk, [])
        if not rules:
            print(f"\n  {label}: No data.")
            continue

        rules_sorted = sorted(rules, key=lambda x: x[1]["median"], reverse=True)

        print(f"\n  {label} — Top 20 Rule Combinations (by Median Return)")
        print(f"  {'Stop-Loss':>10} | {'Profit-Tgt':>10} | {'MaxHold':>7}"
              f" | {'N':>4} | {'Mean':>8} | {'Median':>8} | {'Win%':>6}"
              f" | {'Min':>8} | {'Max':>8} | {'AvgDays':>7}"
              f" | {'%SL':>5} | {'%PT':>5} | {'%MH':>5}"
              f" | {'Sharpe':>6}")
        print(f"  {'-' * 140}")

        for (sl, pt, mh), s in rules_sorted[:20]:
            print(f"  {fmt_sl(sl):>10} | {fmt_pt(pt):>10} | {mh:>5}d"
                  f" | {s['n']:>4} | {s['mean']:>+7.2%} | {s['median']:>+7.2%} | {s['win_rate']:>5.1%}"
                  f" | {s['min']:>+7.1%} | {s['max']:>+7.1%} | {s['avg_days']:>6.1f}"
                  f" | {s['pct_sl']:>4.0%} | {s['pct_pt']:>4.0%} | {s['pct_mh']:>4.0%}"
                  f" | {s['sharpe_like']:>5.2f}")


def print_rule_highlights(all_stats, scenario="mid"):
    """Print the best rule sets by various metrics."""
    scenario_label = {"mid": "Mid-Price", "natural": "Natural (Worst Fill)",
                      "slippage25": "25% Slippage"}
    print()
    print("=" * 150)
    print(f"BEST RULE SETS — {scenario_label.get(scenario, scenario)}")
    print("=" * 150)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        rules = all_stats.get(dk, [])
        if not rules:
            continue

        print(f"\n--- {label} ---")

        by_median = sorted(rules, key=lambda x: x[1]["median"], reverse=True)
        by_winrate = sorted(rules, key=lambda x: x[1]["win_rate"], reverse=True)
        by_sharpe = sorted(rules, key=lambda x: x[1]["sharpe_like"], reverse=True)
        by_mean = sorted(rules, key=lambda x: x[1]["mean"], reverse=True)

        categories = [
            ("Highest Median Return", by_median),
            ("Highest Win Rate", by_winrate),
            ("Best Risk-Adjusted (Sharpe)", by_sharpe),
            ("Highest Mean Return", by_mean),
        ]

        for cat_label, ranked in categories:
            (sl, pt, mh), s = ranked[0]
            print(f"\n  {cat_label}:")
            print(f"    Rule: SL={fmt_sl(sl)}, PT={fmt_pt(pt)}, MaxHold={mh}d")
            print(f"    N={s['n']}  Mean={s['mean']:+.2%}  Median={s['median']:+.2%}"
                  f"  Win%={s['win_rate']:.1%}  AvgDays={s['avg_days']:.1f}"
                  f"  Sharpe={s['sharpe_like']:.2f}")
            print(f"    Exits: {s['pct_sl']:.0%} stop-loss, {s['pct_pt']:.0%} profit-target,"
                  f" {s['pct_mh']:.0%} max-hold")


def print_rule_grid(all_stats, scenario="mid"):
    """Print compact grid: SL vs PT for each max-hold and delta."""
    scenario_label = {"mid": "Mid-Price", "natural": "Natural (Worst Fill)",
                      "slippage25": "25% Slippage"}
    print()
    print("=" * 150)
    print(f"RULE GRID — Median Return by SL x PT — {scenario_label.get(scenario, scenario)}")
    print("=" * 150)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        rules = all_stats.get(dk, [])
        if not rules:
            continue
        lookup = {}
        for (sl, pt, mh), s in rules:
            lookup[(sl, pt, mh)] = s

        for mh in MAX_HOLDS:
            print(f"\n  {label} — Max Hold: {mh} days")
            sl_pt_label = "SL \\ PT"
            pt_header = f"  {sl_pt_label:<10}"
            for pt in PROFIT_TARGETS:
                pt_header += f" | {fmt_pt(pt):>10}"
            print(pt_header)
            print(f"  {'-' * (12 + 13 * len(PROFIT_TARGETS))}")

            for sl in STOP_LOSSES:
                row = f"  {fmt_sl(sl):<10}"
                for pt in PROFIT_TARGETS:
                    s = lookup.get((sl, pt, mh))
                    if s:
                        med = s["median"]
                        wr = s["win_rate"]
                        row += f" | {med:>+5.0%} {wr:>4.0%}w"
                    else:
                        row += f" |        ---"
                print(row)


def print_per_pick_detail(results, delta_key="d80", scenario="mid"):
    """Per-pick detail table."""
    print()
    print("=" * 180)
    label = "80-Delta ITM" if delta_key == "d80" else "50-Delta ATM"
    scenario_label = {"mid": "Mid-Price", "natural": "Natural", "slippage25": "25% Slippage"}
    print(f"PER-PICK DETAIL — {label} Calls — {scenario_label.get(scenario, scenario)}")
    print("=" * 180)

    header = (f"{'Symbol':<8} | {'Pick Date':<12} | {'Spot':>8} | {'IV':>5} | {'Delta':>6}"
              f" | {'Strike':>8} | {'Bid':>7} | {'Ask':>7} | {'Entry$':>7}")
    for w in WINDOWS:
        header += f" | {'T+' + str(w) + ' Ret':>9}"
    print(header)
    print("-" * 180)

    for r in sorted(results, key=lambda x: x["pick_date"]):
        strike = r.get(f"{delta_key}_strike")
        entry = r.get(f"{delta_key}_{scenario}_entry")
        if strike is None or entry is None:
            continue

        iv = r.get(f"{delta_key}_actual_iv")
        actual_delta = r.get(f"{delta_key}_actual_delta")
        bid = r.get(f"{delta_key}_entry_bid", 0)
        ask = r.get(f"{delta_key}_entry_ask", 0)

        iv_str = f"{iv:>4.0%}" if iv else " ---"
        delta_str = f"{actual_delta:>5.3f}" if actual_delta else "  ---"

        line = (f"{r['symbol']:<8} | {r['pick_date']:<12}"
                f" | {r['spot_entry']:>8.2f} | {iv_str}"
                f" | {delta_str}"
                f" | {strike:>8.2f} | {bid:>7.2f} | {ask:>7.2f} | {entry:>7.2f}")

        for w in WINDOWS:
            ret = r.get(f"{delta_key}_{scenario}_ret_{w}d")
            if ret is not None:
                line += f" | {ret:>+8.1%}"
            else:
                line += f" |      ---"
        print(line)


def print_capital_analysis(results, scenario="mid"):
    """Capital analysis for 1 contract per pick."""
    print()
    print("=" * 120)
    scenario_label = {"mid": "Mid-Price", "natural": "Natural", "slippage25": "25% Slippage"}
    print(f"CAPITAL ANALYSIS — 1 Contract Per Pick — {scenario_label.get(scenario, scenario)}")
    print("=" * 120)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        print(f"\n--- {label} ---")

        entries = [(r[f"{dk}_{scenario}_entry"], r) for r in results
                   if r.get(f"{dk}_{scenario}_entry") is not None
                   and r[f"{dk}_{scenario}_entry"] > 0]
        if not entries:
            print("  No valid entries.")
            continue

        total_capital = sum(e * 100 for e, _ in entries)
        n_picks = len(entries)
        avg_cost = total_capital / n_picks

        print(f"  Picks with valid pricing: {n_picks}")
        print(f"  Total capital (1 contract each): ${total_capital:,.0f}")
        print(f"  Avg cost per contract: ${avg_cost:,.0f}")
        print()

        print(f"  {'Window':<12} | {'Total P&L':>12} | {'Total Ret':>10} | {'Avg P&L':>10}")
        print(f"  {'-'*55}")

        for w in WINDOWS:
            pnls = [r[f"{dk}_{scenario}_pnl_{w}d"] for r in results
                    if r.get(f"{dk}_{scenario}_pnl_{w}d") is not None]
            if not pnls:
                continue
            total_pnl = sum(pnls)
            total_ret = total_pnl / total_capital if total_capital > 0 else 0
            avg_pnl = total_pnl / len(pnls)
            print(f"  T+{w:<9} | ${total_pnl:>+11,.0f} | {total_ret:>+9.2%} | ${avg_pnl:>+9,.0f}")


def print_best_worst(results, delta_key="d80", window=30, scenario="mid"):
    """Top 5 and bottom 5 picks by options return."""
    print()
    print("=" * 120)
    label = "80-Delta ITM" if delta_key == "d80" else "50-Delta ATM"
    scenario_label = {"mid": "Mid-Price", "natural": "Natural", "slippage25": "25% Slippage"}
    print(f"TOP 5 AND BOTTOM 5 — {label} at T+{window} — {scenario_label.get(scenario, scenario)}")
    print("=" * 120)

    ret_key = f"{delta_key}_{scenario}_ret_{window}d"
    valid = [r for r in results if r.get(ret_key) is not None]
    if not valid:
        print("  No data for this window.")
        return

    ranked = sorted(valid, key=lambda x: x[ret_key], reverse=True)

    header = (f"{'Symbol':<8} | {'Pick Date':<12} | {'Spot':>8} | {'IV':>5}"
              f" | {'Entry$':>7} | {'Opt Ret':>9} | {'Stock Ret':>9}")
    for section_label, subset in [("TOP 5", ranked[:5]), ("BOTTOM 5", ranked[-5:])]:
        print(f"\n{section_label}:")
        print(header)
        print("-" * 80)
        for r in subset:
            entry = r.get(f"{delta_key}_{scenario}_entry", 0)
            ret = r.get(ret_key, 0)
            stock_ret = r.get(f"{delta_key}_stock_ret_{window}d", 0) or 0
            iv = r.get(f"{delta_key}_actual_iv")
            iv_str = f"{iv:>4.0%}" if iv else " ---"
            print(f"{r['symbol']:<8} | {r['pick_date']:<12}"
                  f" | {r['spot_entry']:>8.2f} | {iv_str}"
                  f" | {entry:>7.2f} | {ret:>+8.1%} | {stock_ret:>+8.1%}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        datefmt="%H:%M:%S",
    )

    print("=" * 80)
    print("ALPHA PICKS OPTIONS P&L MODEL — ThetaData Historical Validation")
    print("=" * 80)
    print(f"  Target DTE: {DTE_TARGET} days (range {DTE_MIN}-{DTE_MAX})")
    print(f"  Deltas: {DELTAS}")
    print(f"  Hold windows: {WINDOWS} trading days")
    print(f"  Execution scenarios: {SCENARIOS}")
    print(f"  Risk-free rate: {RATE:.0%}")
    print()

    # Load stock data
    print("Loading Alpha Picks...")
    picks = load_alpha_picks()
    print(f"  {len(picks)} picks loaded")

    print("Loading stock price cache...")
    prices = load_prices_from_cache()
    print(f"  {len(prices)} symbols in cache")

    trading_dates = get_trading_dates(prices)
    print(f"  Trading dates: {trading_dates[0]} to {trading_dates[-1]}")

    # Connect to ThetaData
    print("\nConnecting to ThetaData Terminal...")
    client = ThetaDataClient()
    connected = client.connect()
    if not connected:
        print("\nFATAL: Cannot proceed without ThetaData Terminal.")
        print("Please start Theta Terminal and try again.")
        sys.exit(1)
    print("  Connected.")

    # =================================================================
    # Model each pick — T0 entry (actionable on public info)
    # =================================================================
    print(f"\nModelling options P&L (T0 entry, ThetaData)...")
    print(f"  Processing {len(picks)} picks...")

    results = []
    skipped_symbols = []
    for i, pick in enumerate(picks):
        sym = pick["symbol"]
        print(f"  [{i+1:3d}/{len(picks)}] {sym:<8} {pick['pick_date']}...", end="", flush=True)

        r = model_pick_thetadata(pick, prices, trading_dates, client, entry_offset=0)
        if r:
            # Check if we got at least some valid data
            has_data = False
            for dk in ["d80", "d50"]:
                if r.get(f"{dk}_strike") is not None:
                    has_data = True
                    break
            if has_data:
                results.append(r)
                print(f" OK (exp={r.get('expiration', '?')}, dte={r.get('actual_dte', '?')})")
            else:
                skipped_symbols.append(sym)
                print(" SKIP (no valid strikes)")
        else:
            skipped_symbols.append(sym)
            print(" SKIP (no data)")

    print(f"\n  Results: {len(results)} picks with data, {len(skipped_symbols)} skipped")

    if not results:
        print("\nNo picks had valid ThetaData options data. Exiting.")
        client.close()
        sys.exit(1)

    # =================================================================
    # PART 1: Data Coverage Report
    # =================================================================
    print_data_coverage(results, skipped_symbols)

    # =================================================================
    # PART 2: Fixed-Window Analysis — All Scenarios
    # =================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 2: FIXED-WINDOW ANALYSIS — T0 ENTRY")
    print("#" * 160)

    for scenario in SCENARIOS:
        print_fixed_window_summary(results, scenario)

    print_scenario_comparison(results)

    # =================================================================
    # PART 3: Per-Pick Detail (Mid-Price)
    # =================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 3: PER-PICK DETAIL (Mid-Price)")
    print("#" * 160)

    print_per_pick_detail(results, "d80", "mid")
    print_per_pick_detail(results, "d50", "mid")

    # =================================================================
    # PART 4: Capital Analysis
    # =================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 4: CAPITAL ANALYSIS")
    print("#" * 160)

    for scenario in SCENARIOS:
        print_capital_analysis(results, scenario)

    # =================================================================
    # PART 5: Best/Worst Picks
    # =================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 5: BEST AND WORST PICKS")
    print("#" * 160)

    print_best_worst(results, "d80", 30, "mid")
    print_best_worst(results, "d50", 30, "mid")

    # =================================================================
    # PART 6: Rule-Based Exit Analysis — All Scenarios
    # =================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 6: RULE-BASED EXIT ANALYSIS")
    print("#" * 160)

    for scenario in SCENARIOS:
        stats = run_rule_analysis(results, scenario)
        print_rule_grid(stats, scenario)
        print_rule_highlights(stats, scenario)
        print_rule_tables(stats, scenario)

    # =================================================================
    # Closing notes
    # =================================================================
    print()
    print("=" * 80)
    print("NOTES:")
    print("  - Option prices from ThetaData historical EOD data (real market quotes)")
    print("  - Entry strike found via B-S delta targeting, snapped to real strike grid")
    print("  - Actual IV and delta from ThetaData Greeks endpoint")
    print("  - Three execution scenarios: mid, natural (worst fill), 25% slippage")
    print("  - Where bid/ask = 0, close price with 3% estimated spread used as fallback")
    print("  - T0 entry: buy at close on announcement day (public info)")
    print("  - Forward windows (T+1, T+10, etc.) measured from T0")
    print("  - Daily rule-based exits use EOD bid/ask for realistic exit pricing")
    print("=" * 80)

    client.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
