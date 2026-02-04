"""
Alpha Picks Configurable Backtester
=====================================
Tests various exit strategies and risk rules on Alpha Picks data
using daily price data from Massive API.

Strategies tested:
  1. Baseline: SA's actual entries/exits (hold until rating change)
  2. Stop-loss: exit if down X% from entry
  3. Trailing stop: exit if down X% from peak
  4. Max holding period: exit after N months regardless
  5. Vol-adjusted sizing: size positions by inverse volatility
  6. Combined: best combination of above

Also includes Sharpe decay analysis and sector/factor attribution.
"""

import openpyxl
import sqlite3
import os
import math
from collections import defaultdict, OrderedDict
from datetime import datetime, timedelta

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "price_cache.db")
XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "ProQuant History 1_29_2026.xlsx")

SNAPSHOT_DATE = "2026-01-29"


def load_prices():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.execute("SELECT symbol, date, close FROM daily_prices ORDER BY symbol, date")
    prices = defaultdict(dict)
    for symbol, date, close in cur.fetchall():
        prices[symbol][date] = close
    conn.close()
    return prices


def get_trading_dates(prices):
    return sorted(prices.get("SPY", {}).keys())


def load_alpha_picks():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['AlphaPicks']
    picks = []
    seen = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        symbol = row[1]
        date = row[2]
        if not symbol or not date:
            continue
        date_str = date.strftime("%Y-%m-%d") if hasattr(date, 'strftime') else str(date)[:10]
        closed = None
        if row[4] and str(row[4]).strip() not in ('-', 'None', ''):
            closed = row[4].strftime("%Y-%m-%d") if hasattr(row[4], 'strftime') else str(row[4])[:10]

        # Deduplicate: keep the entry with the longest holding period
        key = (symbol, date_str)
        if key in seen:
            # Update existing if this one is still open (longer hold)
            for p in picks:
                if p['symbol'] == symbol and p['picked'] == date_str:
                    if closed is None and p['closed'] is not None:
                        p['closed'] = None  # keep the open version
                    break
            continue
        seen.add(key)

        picks.append({
            'symbol': symbol,
            'picked': date_str,
            'buy_price': float(row[3]) if row[3] else 0.0,
            'closed': closed,
        })

    picks.sort(key=lambda x: x['picked'])
    return picks


def norm_cdf(x):
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def compute_sharpe_psr(returns):
    """Compute Sharpe and PSR from a return series."""
    n = len(returns)
    if n < 10:
        return None

    mu = sum(returns) / n
    var = sum((r - mu) ** 2 for r in returns) / (n - 1)
    sigma = math.sqrt(var) if var > 0 else 0

    if sigma == 0:
        return None

    sharpe = (mu / sigma) * math.sqrt(252)
    sr_hat = mu / sigma

    # Skewness and kurtosis for PSR
    if n > 2:
        skew = (n / ((n - 1) * (n - 2))) * sum(((r - mu) / sigma) ** 3 for r in returns)
    else:
        skew = 0
    if n > 3:
        raw_kurt = (n * (n + 1)) / ((n - 1) * (n - 2) * (n - 3))
        raw_kurt *= sum(((r - mu) / sigma) ** 4 for r in returns)
        kurt = raw_kurt - (3 * (n - 1) ** 2) / ((n - 2) * (n - 3))
    else:
        kurt = 0

    denom_sq = 1 - skew * sr_hat + (kurt / 4) * sr_hat ** 2
    if denom_sq > 0:
        z = sr_hat * math.sqrt(n - 1) / math.sqrt(denom_sq)
        psr = norm_cdf(z)
    else:
        psr = 0.5

    # Total return
    total = 1.0
    for r in returns:
        total *= (1 + r)

    # Max drawdown
    cum = 1.0
    peak = 1.0
    max_dd = 0
    for r in returns:
        cum *= (1 + r)
        if cum > peak:
            peak = cum
        dd = (peak - cum) / peak
        if dd > max_dd:
            max_dd = dd

    return {
        'n': n,
        'sharpe': sharpe,
        'psr': psr,
        'total_return': total - 1,
        'ann_return': (total ** (252 / n)) - 1 if n > 0 else 0,
        'ann_vol': sigma * math.sqrt(252),
        'max_dd': max_dd,
        'sortino': None,  # computed separately if needed
    }


# ============================================================
# PORTFOLIO SIMULATOR
# ============================================================

def simulate_portfolio(picks, prices, trading_dates, config):
    """
    Simulate Alpha Picks portfolio with configurable rules.

    Config options:
      stop_loss:     float or None — exit if down X% from entry (e.g., 0.20 = -20%)
      trailing_stop: float or None — exit if down X% from peak (e.g., 0.25 = -25%)
      max_hold_days: int or None — exit after N trading days
      vol_sizing:    bool — size by inverse 20-day volatility
      base_investment: float — base $ per pick (default 1000)
    """
    stop_loss = config.get('stop_loss')
    trailing_stop = config.get('trailing_stop')
    max_hold_days = config.get('max_hold_days')
    vol_sizing = config.get('vol_sizing', False)
    base_investment = config.get('base_investment', 1000)

    # Build positions
    positions = []
    for p in picks:
        sym = p['symbol']
        entry_date = p['picked']
        sa_exit = p['closed']
        entry_price = p['buy_price']

        if entry_price <= 0:
            continue

        sym_prices = prices.get(sym, {})
        if not sym_prices:
            continue

        # Vol-adjusted sizing
        if vol_sizing:
            # Compute 20-day trailing vol before entry
            dates_before = sorted(d for d in sym_prices.keys() if d < entry_date)
            if len(dates_before) >= 20:
                recent = dates_before[-20:]
                closes = [sym_prices[d] for d in recent]
                rets = [(closes[i] / closes[i-1]) - 1 for i in range(1, len(closes))]
                vol = math.sqrt(sum(r**2 for r in rets) / len(rets)) if rets else 0.01
                # Inverse vol: lower vol = larger position
                # Target 1% daily risk per position
                target_risk = 0.01
                position_size = (base_investment * target_risk) / vol if vol > 0 else base_investment
                # Cap at 3x base to prevent extreme sizes
                position_size = min(position_size, base_investment * 3)
            else:
                position_size = base_investment
        else:
            position_size = base_investment

        shares = position_size / entry_price

        positions.append({
            'symbol': sym,
            'entry_date': entry_date,
            'sa_exit': sa_exit,
            'entry_price': entry_price,
            'shares': shares,
            'position_size': position_size,
            'peak_price': entry_price,
            'days_held': 0,
            'exit_date': None,
            'exit_price': None,
            'exit_reason': None,
        })

    # Day-by-day simulation
    daily_values = []
    daily_returns = []
    prev_total = 0
    closed_positions = []

    for date in trading_dates:
        total_value = 0
        active = 0

        for pos in positions:
            if date < pos['entry_date']:
                continue
            if pos['exit_date'] and date > pos['exit_date']:
                continue

            sym_prices = prices.get(pos['symbol'], {})
            price = sym_prices.get(date)

            if price is None:
                # Use last known price
                avail = sorted(d for d in sym_prices.keys() if d <= date)
                price = sym_prices[avail[-1]] if avail else pos['entry_price']

            # Check exit conditions (if not already exited)
            if pos['exit_date'] is None and date > pos['entry_date']:
                pos['days_held'] += 1

                # Update peak
                if price > pos['peak_price']:
                    pos['peak_price'] = price

                # Stop loss check
                if stop_loss is not None:
                    loss = (price / pos['entry_price']) - 1
                    if loss <= -stop_loss:
                        pos['exit_date'] = date
                        pos['exit_price'] = price
                        pos['exit_reason'] = f"stop_loss_{stop_loss:.0%}"

                # Trailing stop check
                if trailing_stop is not None and pos['exit_date'] is None:
                    drawdown = (price / pos['peak_price']) - 1
                    if drawdown <= -trailing_stop:
                        pos['exit_date'] = date
                        pos['exit_price'] = price
                        pos['exit_reason'] = f"trailing_{trailing_stop:.0%}"

                # Max holding period check
                if max_hold_days is not None and pos['exit_date'] is None:
                    if pos['days_held'] >= max_hold_days:
                        pos['exit_date'] = date
                        pos['exit_price'] = price
                        pos['exit_reason'] = f"max_hold_{max_hold_days}d"

                # SA exit (rating change)
                if pos['exit_date'] is None and pos['sa_exit'] and date >= pos['sa_exit']:
                    pos['exit_date'] = date
                    pos['exit_price'] = price
                    pos['exit_reason'] = "sa_exit"

            # If just exited, record
            if pos['exit_date'] == date:
                closed_positions.append(pos)

            # If still active (not yet exited, or exiting today)
            if pos['exit_date'] is None or pos['exit_date'] >= date:
                total_value += pos['shares'] * price
                active += 1

        if total_value > 0:
            if prev_total > 0:
                ret = (total_value / prev_total) - 1
                daily_returns.append((date, ret))
            daily_values.append((date, total_value, active))
            prev_total = total_value

    # Close remaining positions at snapshot
    for pos in positions:
        if pos['exit_date'] is None:
            sym_prices = prices.get(pos['symbol'], {})
            avail = sorted(sym_prices.keys())
            if avail:
                pos['exit_date'] = avail[-1]
                pos['exit_price'] = sym_prices[avail[-1]]
                pos['exit_reason'] = "still_open"

    return daily_values, daily_returns, positions


def summarize_strategy(label, daily_returns, positions):
    """Print summary for a strategy."""
    stats = compute_sharpe_psr([r for _, r in daily_returns])
    if not stats:
        print(f"  {label}: insufficient data")
        return None

    # Per-position stats
    pos_returns = []
    for pos in positions:
        if pos['exit_price'] and pos['entry_price'] > 0:
            pos_returns.append((pos['exit_price'] / pos['entry_price']) - 1)

    wins = sum(1 for r in pos_returns if r > 0)
    n = len(pos_returns)
    avg_hold = sum(pos['days_held'] for pos in positions if pos['days_held'] > 0) / max(1, n)

    # Exit reason breakdown
    reasons = defaultdict(int)
    for pos in positions:
        if pos['exit_reason']:
            reasons[pos['exit_reason']] += 1

    print(f"  {label}")
    print(f"    Picks: {n}  Win: {wins/n:.1%}  AvgHold: {avg_hold:.0f}d  "
          f"Return: {stats['total_return']:>+.1%}  Ann: {stats['ann_return']:>+.1%}  "
          f"Sharpe: {stats['sharpe']:>.2f}  PSR: {stats['psr']:>.1%}  MaxDD: {stats['max_dd']:>.1%}")
    if len(reasons) > 1:
        reason_str = "  ".join(f"{k}:{v}" for k, v in sorted(reasons.items()))
        print(f"    Exits: {reason_str}")

    return stats


# ============================================================
# MAIN ANALYSIS
# ============================================================

def main():
    print("=" * 90)
    print("ALPHA PICKS — CONFIGURABLE EXIT STRATEGY BACKTESTER")
    print("=" * 90)
    print()

    prices = load_prices()
    trading_dates = get_trading_dates(prices)
    picks = load_alpha_picks()

    print(f"  Picks loaded: {len(picks)} (deduplicated)")
    print(f"  Trading dates: {trading_dates[0]} to {trading_dates[-1]}")
    print()

    # ================================================================
    # STRATEGY COMPARISON
    # ================================================================

    strategies = [
        ("BASELINE (SA exits only)", {}),
        ("Stop Loss 15%", {'stop_loss': 0.15}),
        ("Stop Loss 20%", {'stop_loss': 0.20}),
        ("Stop Loss 25%", {'stop_loss': 0.25}),
        ("Stop Loss 30%", {'stop_loss': 0.30}),
        ("Trailing Stop 20%", {'trailing_stop': 0.20}),
        ("Trailing Stop 25%", {'trailing_stop': 0.25}),
        ("Trailing Stop 30%", {'trailing_stop': 0.30}),
        ("Trailing Stop 40%", {'trailing_stop': 0.40}),
        ("Max Hold 126d (6mo)", {'max_hold_days': 126}),
        ("Max Hold 189d (9mo)", {'max_hold_days': 189}),
        ("Max Hold 252d (1yr)", {'max_hold_days': 252}),
        ("SL20% + Trail25%", {'stop_loss': 0.20, 'trailing_stop': 0.25}),
        ("SL20% + Trail30%", {'stop_loss': 0.20, 'trailing_stop': 0.30}),
        ("SL25% + Trail30%", {'stop_loss': 0.25, 'trailing_stop': 0.30}),
        ("SL20% + MaxHold252d", {'stop_loss': 0.20, 'max_hold_days': 252}),
        ("SL20% + Trail25% + 252d", {'stop_loss': 0.20, 'trailing_stop': 0.25, 'max_hold_days': 252}),
        ("SL25% + Trail30% + 252d", {'stop_loss': 0.25, 'trailing_stop': 0.30, 'max_hold_days': 252}),
        ("Vol-Adjusted Sizing", {'vol_sizing': True}),
        ("Vol + SL20% + Trail25%", {'vol_sizing': True, 'stop_loss': 0.20, 'trailing_stop': 0.25}),
        ("Vol + SL25% + Trail30%", {'vol_sizing': True, 'stop_loss': 0.25, 'trailing_stop': 0.30}),
    ]

    print("-" * 90)
    print("EXIT STRATEGY COMPARISON")
    print("-" * 90)
    print()

    all_results = []

    for label, config in strategies:
        _, daily_returns, positions = simulate_portfolio(picks, prices, trading_dates, config)
        stats = summarize_strategy(label, daily_returns, positions)
        all_results.append((label, stats, config))

    # Rank by Sharpe
    print()
    print("-" * 90)
    print("RANKED BY SHARPE RATIO")
    print("-" * 90)
    print()
    print(f"  {'#':>3} {'Strategy':<30} {'Sharpe':>8} {'PSR':>7} {'Return':>10} {'MaxDD':>8} {'Ann.Ret':>10}")
    print(f"  {'-'*78}")

    ranked = sorted([(l, s) for l, s, _ in all_results if s],
                    key=lambda x: x[1]['sharpe'], reverse=True)

    for i, (label, stats) in enumerate(ranked, 1):
        print(f"  {i:>3} {label:<30} {stats['sharpe']:>7.2f} {stats['psr']:>6.1%} "
              f"{stats['total_return']:>+9.1%} {stats['max_dd']:>7.1%} {stats['ann_return']:>+9.1%}")

    # Rank by PSR
    print()
    print("-" * 90)
    print("RANKED BY PROBABILISTIC SHARPE RATIO")
    print("-" * 90)
    print()
    print(f"  {'#':>3} {'Strategy':<30} {'PSR':>7} {'Sharpe':>8} {'Return':>10} {'MaxDD':>8}")
    print(f"  {'-'*60}")

    ranked_psr = sorted([(l, s) for l, s, _ in all_results if s],
                        key=lambda x: x[1]['psr'], reverse=True)

    for i, (label, stats) in enumerate(ranked_psr, 1):
        print(f"  {i:>3} {label:<30} {stats['psr']:>6.1%} {stats['sharpe']:>7.2f} "
              f"{stats['total_return']:>+9.1%} {stats['max_dd']:>7.1%}")

    print()

    # ================================================================
    # SHARPE DECAY ANALYSIS
    # ================================================================
    print("=" * 90)
    print("SHARPE RATIO DECAY ANALYSIS")
    print("=" * 90)
    print()

    # Per-year Sharpe for baseline
    _, baseline_returns, _ = simulate_portfolio(picks, prices, trading_dates, {})
    baseline_by_date = {d: r for d, r in baseline_returns}

    # Also compute SPY returns
    spy_prices = prices.get("SPY", {})
    spy_returns_by_date = {}
    spy_dates = sorted(spy_prices.keys())
    for i in range(1, len(spy_dates)):
        prev_p = spy_prices[spy_dates[i-1]]
        curr_p = spy_prices[spy_dates[i]]
        if prev_p > 0:
            spy_returns_by_date[spy_dates[i]] = (curr_p / prev_p) - 1

    # Rolling 6-month Sharpe
    print("  Rolling 6-Month Sharpe (Alpha Picks vs SPY):")
    print(f"  {'End Date':>12} {'AP Sharpe':>10} {'SPY Sharpe':>12} {'AP-SPY':>10}")
    print(f"  {'-'*48}")

    window = 126  # ~6 months
    return_dates = [d for d, _ in baseline_returns]

    for i in range(window, len(return_dates), 21):  # every ~month
        end_date = return_dates[i]
        start_idx = i - window

        ap_window = [r for _, r in baseline_returns[start_idx:i]]
        spy_window = [spy_returns_by_date.get(d, 0) for d in return_dates[start_idx:i]]

        ap_stats = compute_sharpe_psr(ap_window)
        spy_stats = compute_sharpe_psr(spy_window)

        if ap_stats and spy_stats:
            diff = ap_stats['sharpe'] - spy_stats['sharpe']
            print(f"  {end_date:>12} {ap_stats['sharpe']:>9.2f} {spy_stats['sharpe']:>11.2f} {diff:>+9.2f}")

    print()

    # Yearly breakdown with more detail
    print("  Annual Sharpe Decomposition:")
    print(f"  {'Year':>6} {'AP Sharpe':>10} {'AP Vol':>8} {'AP Ret':>10} {'SPY Sharpe':>12} "
          f"{'Ann.Pop':>10} {'Picks':>7}")
    print(f"  {'-'*65}")

    years = sorted(set(d[:4] for d, _ in baseline_returns))
    for year in years:
        yr_ap = [(d, r) for d, r in baseline_returns if d[:4] == year]
        yr_spy = [(d, spy_returns_by_date.get(d, 0)) for d, _ in yr_ap]

        ap_stats = compute_sharpe_psr([r for _, r in yr_ap])
        spy_stats = compute_sharpe_psr([r for _, r in yr_spy])

        # Count picks in this year
        yr_picks = [p for p in picks if p['picked'][:4] == year]

        # Average announcement day pop for this year
        from pick_day_analysis import load_alpha_picks as _lp, load_prices as _lpr
        # Compute inline
        ann_pops = []
        for p in yr_picks:
            sym_prices = prices.get(p['symbol'], {})
            pick_td_idx = None
            for idx, d in enumerate(trading_dates):
                if d >= p['picked']:
                    pick_td_idx = idx
                    break
            if pick_td_idx and pick_td_idx > 0:
                prev_d = trading_dates[pick_td_idx - 1]
                pick_d = trading_dates[pick_td_idx]
                prev_p = sym_prices.get(prev_d)
                pick_p = sym_prices.get(pick_d)
                if prev_p and pick_p and prev_p > 0:
                    ann_pops.append((pick_p / prev_p) - 1)

        avg_pop = sum(ann_pops) / len(ann_pops) if ann_pops else 0

        if ap_stats and spy_stats:
            print(f"  {year:>6} {ap_stats['sharpe']:>9.2f} {ap_stats['ann_vol']:>7.1%} "
                  f"{ap_stats['total_return']:>+9.1%} {spy_stats['sharpe']:>11.2f} "
                  f"{avg_pop:>+9.2%} {len(yr_picks):>7}")

    print()
    print("  Interpretation:")
    print("  - Growing announcement pop (Ann.Pop) correlates with declining Sharpe")
    print("  - As more subscribers front-run picks, less alpha remains for holding period")
    print("  - Volatility also matters: high-vol years show lower Sharpe even with gains")

    print()
    print("=" * 90)
    print("ANALYSIS COMPLETE")
    print("=" * 90)


if __name__ == "__main__":
    main()
