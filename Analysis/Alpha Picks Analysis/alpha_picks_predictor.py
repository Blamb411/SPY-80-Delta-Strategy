"""
Alpha Picks Predictor
======================
Predict the next SA Alpha Pick by scoring candidates against the
historical profile of past picks.

The typical Alpha Pick has:
  - Strong Buy rating for 75+ consecutive days
  - Market cap $2B-$100B (70% of picks)
  - Strong recent momentum: 3m >+15% (91%), 6m >+30% (95%)
  - Sector bias: Industrials (20%), IT (17%), Consumer Disc (14%)
  - Not picked in the last ~6 months
  - Stock price typically $20-$200

Usage:
    python alpha_picks_predictor.py                          # Use sa_strong_buys.csv
    python alpha_picks_predictor.py --file tickers.csv       # Custom ticker list
    python alpha_picks_predictor.py --top 15                 # Show top 15
"""

import os
import sys
import csv
import time
import sqlite3
import argparse
from datetime import datetime, timedelta
from typing import List, Dict, Optional

from massive import RESTClient

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)
API_KEY_FILE = os.path.join(PROJECT_DIR, "Massive backtesting", "api_key.txt")
EXCEL_FILE = os.path.join(BASE_DIR, "ProQuant History 1_29_2026.xlsx")
PRICE_CACHE_DB = os.path.join(BASE_DIR, "price_cache.db")
ANALYSIS_CACHE_DB = os.path.join(BASE_DIR, "alpha_picks_analysis.db")
SA_STRONG_BUYS_CSV = os.path.join(BASE_DIR, "quant_model", "data", "sa_strong_buys.csv")

API_DELAY = 0.15

# Predictive weights (from historical analysis)
SECTOR_WEIGHTS = {
    "Information Technology": 1.5,
    "Industrials": 1.4,
    "Consumer Discretionary": 1.2,
    "Financials": 1.1,
    "Energy": 1.0,
    "Communication Services": 1.0,
    "Materials": 0.9,
    "Health Care": 0.8,
    "Consumer Staples": 0.8,
    "Utilities": 0.5,
    "Real Estate": 0.5,
}

# Market cap scoring brackets
MCAP_SCORE = [
    (500e6, 2e9, 0.6),     # Small: below sweet spot
    (2e9, 10e9, 1.0),      # Mid: sweet spot
    (10e9, 100e9, 0.9),    # Large: still good
    (100e9, float("inf"), 0.5),  # Mega: less likely
]


def load_api_key():
    with open(API_KEY_FILE) as f:
        return f.read().strip()


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_strong_buys_csv(filepath: str) -> List[str]:
    """Load ticker symbols from SA Strong Buys CSV."""
    symbols = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sym = None
            for col in ["Symbol", "Ticker", "symbol", "ticker"]:
                if col in row:
                    sym = row[col].strip().upper()
                    break
            if not sym:
                # Use first column
                first_key = list(row.keys())[0]
                sym = row[first_key].strip().upper()
            if sym and (sym.isalpha() or "." in sym):
                symbols.append(sym)
    return symbols


def load_recent_picks() -> Dict[str, str]:
    """Load recent Alpha Picks to exclude (picked in last 6 months)."""
    try:
        import openpyxl
    except ImportError:
        return {}

    if not os.path.exists(EXCEL_FILE):
        return {}

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    recent = {}
    cutoff = (datetime.now() - timedelta(days=180)).strftime("%Y-%m-%d")

    if "AlphaPicks" in wb.sheetnames:
        ws = wb["AlphaPicks"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row or not row[1] or not row[2]:
                continue
            symbol = str(row[1]).strip().rstrip("*")
            pick_date = row[2]
            if hasattr(pick_date, "strftime"):
                pick_date = pick_date.strftime("%Y-%m-%d")
            else:
                pick_date = str(pick_date).strip()
            if pick_date >= cutoff:
                recent[symbol] = pick_date

    wb.close()
    return recent


# ---------------------------------------------------------------------------
# Ticker details and price data
# ---------------------------------------------------------------------------

def get_ticker_details(client, symbol) -> Optional[Dict]:
    """Get ticker details from cache or API."""
    conn = sqlite3.connect(ANALYSIS_CACHE_DB)
    row = conn.execute(
        "SELECT name, market_cap, sic_code, sic_description, sector, industry "
        "FROM ticker_details WHERE symbol=?",
        (symbol,),
    ).fetchone()

    if row:
        conn.close()
        return {
            "name": row[0], "market_cap": row[1],
            "sic_code": row[2], "sic_description": row[3],
            "sector": row[4], "industry": row[5],
        }

    # Fetch from API
    try:
        details = client.get_ticker_details(symbol)
        if details:
            from alpha_picks_analysis import sic_to_sector, SIC_SECTOR_OVERRIDES

            name = getattr(details, "name", None)
            mcap = getattr(details, "market_cap", None)
            sic = getattr(details, "sic_code", None)
            sic_desc = getattr(details, "sic_description", None)

            if symbol in SIC_SECTOR_OVERRIDES:
                sector = SIC_SECTOR_OVERRIDES[symbol]
            else:
                sector = sic_to_sector(sic) if sic else None
            industry = sic_desc

            conn.execute(
                "INSERT OR REPLACE INTO ticker_details VALUES (?,?,?,?,?,?,?,?,?)",
                (symbol, name, mcap, sic, sic_desc, sector, industry,
                 getattr(details, "list_date", None),
                 datetime.now().isoformat()),
            )
            conn.commit()
            conn.close()

            return {
                "name": name, "market_cap": mcap,
                "sic_code": sic, "sic_description": sic_desc,
                "sector": sector, "industry": industry,
            }
    except Exception:
        pass

    time.sleep(API_DELAY)
    conn.close()
    return None


def get_momentum(client, symbol, as_of_date=None) -> Dict:
    """Compute momentum metrics from price data."""
    if as_of_date is None:
        as_of_date = datetime.now().strftime("%Y-%m-%d")

    conn = sqlite3.connect(PRICE_CACHE_DB)

    # Get current price
    upper = (datetime.strptime(as_of_date, "%Y-%m-%d") + timedelta(days=5)).strftime("%Y-%m-%d")
    lower = (datetime.strptime(as_of_date, "%Y-%m-%d") - timedelta(days=5)).strftime("%Y-%m-%d")

    row = conn.execute(
        "SELECT close, date FROM daily_prices WHERE symbol=? AND date>=? AND date<=? "
        "ORDER BY date DESC LIMIT 1",
        (symbol, lower, upper),
    ).fetchone()

    if not row:
        # Try fetching from API
        fetch_key = f"momentum:{symbol}"
        try:
            start = (datetime.strptime(as_of_date, "%Y-%m-%d") - timedelta(days=400)).strftime("%Y-%m-%d")
            aggs = list(client.list_aggs(
                ticker=symbol, multiplier=1, timespan="day",
                from_=start, to=as_of_date, limit=5000,
            ))
            for a in aggs:
                if isinstance(a.timestamp, (int, float)):
                    bd = datetime.utcfromtimestamp(a.timestamp / 1000).strftime("%Y-%m-%d")
                else:
                    bd = str(a.timestamp)
                conn.execute(
                    "INSERT OR IGNORE INTO daily_prices (symbol, date, open, high, low, close, volume) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (symbol, bd, a.open, a.high, a.low, a.close, a.volume),
                )
            conn.commit()
            time.sleep(API_DELAY)
        except Exception:
            conn.close()
            return {}

        row = conn.execute(
            "SELECT close, date FROM daily_prices WHERE symbol=? AND date>=? AND date<=? "
            "ORDER BY date DESC LIMIT 1",
            (symbol, lower, upper),
        ).fetchone()

    if not row:
        conn.close()
        return {}

    current_price = row[0]
    result = {"current_price": current_price}

    # Compute returns for each period
    for label, cal_days in [("1m", 30), ("3m", 92), ("6m", 183), ("12m", 366)]:
        past_date = (datetime.strptime(as_of_date, "%Y-%m-%d") - timedelta(days=cal_days)).strftime("%Y-%m-%d")
        past_upper = (datetime.strptime(past_date, "%Y-%m-%d") + timedelta(days=10)).strftime("%Y-%m-%d")

        past_row = conn.execute(
            "SELECT close FROM daily_prices WHERE symbol=? AND date>=? AND date<=? "
            "ORDER BY date LIMIT 1",
            (symbol, past_date, past_upper),
        ).fetchone()

        if past_row and past_row[0] and past_row[0] > 0:
            ret = (current_price / past_row[0] - 1) * 100
            result[f"return_{label}"] = round(ret, 1)

    conn.close()
    return result


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

def score_candidate(symbol: str, details: Dict, momentum: Dict,
                    recent_picks: Dict) -> Dict:
    """Score a candidate stock on how likely it is to be the next Alpha Pick."""
    score = 0.0
    reasons = []
    disqualified = False
    disq_reasons = []

    # --- Momentum score (0-40 points) ---
    ret_3m = momentum.get("return_3m")
    ret_6m = momentum.get("return_6m")
    ret_12m = momentum.get("return_12m")
    ret_1m = momentum.get("return_1m")

    if ret_3m is not None:
        if ret_3m >= 30:
            score += 15
            reasons.append(f"3m +{ret_3m:.0f}% (strong)")
        elif ret_3m >= 15:
            score += 10
            reasons.append(f"3m +{ret_3m:.0f}% (good)")
        elif ret_3m >= 0:
            score += 3
            reasons.append(f"3m +{ret_3m:.0f}% (weak)")
        else:
            score -= 5
            reasons.append(f"3m {ret_3m:.0f}% (negative)")

    if ret_6m is not None:
        if ret_6m >= 60:
            score += 15
            reasons.append(f"6m +{ret_6m:.0f}% (strong)")
        elif ret_6m >= 30:
            score += 10
            reasons.append(f"6m +{ret_6m:.0f}% (good)")
        elif ret_6m >= 0:
            score += 3
        else:
            score -= 5

    if ret_12m is not None:
        if ret_12m >= 80:
            score += 10
        elif ret_12m >= 40:
            score += 5

    # --- Sector score (0-15 points) ---
    sector = details.get("sector") if details else None
    sector_weight = SECTOR_WEIGHTS.get(sector, 0.7) if sector else 0.7
    sector_score = 10 * sector_weight
    score += sector_score
    if sector:
        reasons.append(f"Sector: {sector} ({sector_weight:.1f}x)")

    # --- Market cap score (0-15 points) ---
    mcap = details.get("market_cap") if details else None
    if mcap:
        mcap_multiplier = 0.5  # default
        for low, high, mult in MCAP_SCORE:
            if low <= mcap < high:
                mcap_multiplier = mult
                break

        mcap_score = 15 * mcap_multiplier
        score += mcap_score

        if mcap >= 1e9:
            reasons.append(f"MCap: ${mcap/1e9:.1f}B ({mcap_multiplier:.1f}x)")
        else:
            reasons.append(f"MCap: ${mcap/1e6:.0f}M ({mcap_multiplier:.1f}x)")

        # Hard filter: too small
        if mcap < 500e6:
            disqualified = True
            disq_reasons.append(f"MCap too small: ${mcap/1e6:.0f}M")
    else:
        score += 5  # neutral if unknown

    # --- Price range score (0-5 points) ---
    price = momentum.get("current_price")
    if price:
        if 20 <= price <= 200:
            score += 5
        elif 10 <= price < 20:
            score += 2
        elif price > 200:
            score += 3

    # --- Recency penalty ---
    if symbol in recent_picks:
        score -= 20
        reasons.append(f"Recently picked ({recent_picks[symbol]})")

    # --- Disqualification checks ---
    if ret_3m is not None and ret_3m < -10:
        disqualified = True
        disq_reasons.append(f"3m return too negative: {ret_3m:.0f}%")

    return {
        "symbol": symbol,
        "score": round(score, 1),
        "disqualified": disqualified,
        "disq_reasons": disq_reasons,
        "reasons": reasons,
        "sector": sector or "Unknown",
        "market_cap": mcap,
        "current_price": price,
        "return_1m": ret_1m,
        "return_3m": ret_3m,
        "return_6m": ret_6m,
        "return_12m": ret_12m,
        "name": details.get("name", "") if details else "",
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Predict next Alpha Pick")
    parser.add_argument("--file", default=None,
                        help="CSV file with candidate tickers")
    parser.add_argument("--symbols", default=None,
                        help="Comma-separated symbols")
    parser.add_argument("--top", type=int, default=10,
                        help="Number of top candidates to show")
    parser.add_argument("--all", action="store_true",
                        help="Show all candidates, not just top N")
    parser.add_argument("--date", default=None,
                        help="Score as of date (YYYY-MM-DD, default=today)")
    args = parser.parse_args()

    as_of = args.date or datetime.now().strftime("%Y-%m-%d")

    print(f"\n{'=' * 90}")
    print(f" ALPHA PICKS PREDICTOR — Candidates for Next Pick")
    print(f" As of: {as_of}")
    print(f"{'=' * 90}")

    # Load candidate symbols
    if args.symbols:
        symbols = [s.strip().upper() for s in args.symbols.split(",")]
        print(f"\n  Scoring {len(symbols)} specified symbols")
    elif args.file:
        symbols = load_strong_buys_csv(args.file)
        print(f"\n  Loaded {len(symbols)} symbols from {os.path.basename(args.file)}")
    elif os.path.exists(SA_STRONG_BUYS_CSV):
        symbols = load_strong_buys_csv(SA_STRONG_BUYS_CSV)
        print(f"\n  Loaded {len(symbols)} symbols from SA Strong Buys")
    else:
        print("\n  ERROR: No ticker source. Use --file or --symbols, "
              "or place sa_strong_buys.csv in quant_model/data/")
        return

    # Load recent picks to exclude
    recent_picks = load_recent_picks()
    if recent_picks:
        print(f"  Recent picks (last 6 months): {len(recent_picks)} "
              f"({', '.join(list(recent_picks.keys())[:10])}...)")

    # Initialize
    client = RESTClient(api_key=load_api_key())

    # Score each candidate
    print(f"\n  Scoring candidates...\n")
    candidates = []

    for i, symbol in enumerate(symbols, 1):
        if i % 25 == 0 or i == len(symbols):
            print(f"    [{i}/{len(symbols)}] processed...", flush=True)

        details = get_ticker_details(client, symbol)
        momentum = get_momentum(client, symbol, as_of)

        if not momentum.get("current_price"):
            continue

        result = score_candidate(symbol, details, momentum, recent_picks)
        candidates.append(result)

    # Sort by score
    qualified = [c for c in candidates if not c["disqualified"]]
    disqualified = [c for c in candidates if c["disqualified"]]
    qualified.sort(key=lambda x: x["score"], reverse=True)

    # Display results
    show_count = len(qualified) if args.all else min(args.top, len(qualified))

    print(f"\n{'=' * 90}")
    print(f" TOP {show_count} CANDIDATES")
    print(f"{'=' * 90}")
    print(f"\n{'Rank':<5} {'Symbol':<7} {'Score':>6} {'Sector':<22} "
          f"{'MCap':>8} {'Price':>8} {'1m%':>6} {'3m%':>6} {'6m%':>6} {'12m%':>7}")
    print("-" * 90)

    for i, c in enumerate(qualified[:show_count], 1):
        mcap_str = f"${c['market_cap']/1e9:.1f}B" if c.get("market_cap") else "N/A"
        price_str = f"${c['current_price']:.0f}" if c.get("current_price") else "N/A"

        r1 = f"{c['return_1m']:>5.0f}%" if c.get("return_1m") is not None else "  N/A"
        r3 = f"{c['return_3m']:>5.0f}%" if c.get("return_3m") is not None else "  N/A"
        r6 = f"{c['return_6m']:>5.0f}%" if c.get("return_6m") is not None else "  N/A"
        r12 = f"{c['return_12m']:>5.0f}%" if c.get("return_12m") is not None else "   N/A"

        recent_flag = " *" if c["symbol"] in recent_picks else ""

        print(f"{i:<5} {c['symbol']:<7} {c['score']:>6.1f} {c['sector']:<22} "
              f"{mcap_str:>8} {price_str:>8} {r1} {r3} {r6} {r12}{recent_flag}")

    # Sector distribution of top candidates
    print(f"\n--- Sector Distribution (top {show_count}) ---")
    from collections import Counter
    sector_dist = Counter(c["sector"] for c in qualified[:show_count])
    for sector, count in sector_dist.most_common():
        print(f"  {sector:<25} {count}")

    # Market cap distribution
    print(f"\n--- Market Cap Distribution (top {show_count}) ---")
    mcap_buckets = {"Small ($500M-2B)": 0, "Mid ($2B-10B)": 0,
                    "Large ($10B-100B)": 0, "Mega (>$100B)": 0, "Unknown": 0}
    for c in qualified[:show_count]:
        m = c.get("market_cap")
        if not m:
            mcap_buckets["Unknown"] += 1
        elif m < 2e9:
            mcap_buckets["Small ($500M-2B)"] += 1
        elif m < 10e9:
            mcap_buckets["Mid ($2B-10B)"] += 1
        elif m < 100e9:
            mcap_buckets["Large ($10B-100B)"] += 1
        else:
            mcap_buckets["Mega (>$100B)"] += 1

    for bucket, count in mcap_buckets.items():
        if count > 0:
            print(f"  {bucket:<20} {count}")

    # Show recently picked stocks that are still Strong Buy
    recently_picked_still_sb = [c for c in candidates
                                if c["symbol"] in recent_picks and not c["disqualified"]]
    if recently_picked_still_sb:
        print(f"\n--- Recently Picked (still Strong Buy, may be re-picked) ---")
        for c in sorted(recently_picked_still_sb, key=lambda x: x["score"], reverse=True)[:5]:
            r3 = f"{c['return_3m']:>5.0f}%" if c.get("return_3m") is not None else "  N/A"
            print(f"  {c['symbol']:<7} Score: {c['score']:>5.1f}  "
                  f"Last picked: {recent_picks[c['symbol']]}  3m: {r3}")

    # Disqualified summary
    if disqualified:
        print(f"\n--- Disqualified ({len(disqualified)}) ---")
        for c in disqualified[:10]:
            print(f"  {c['symbol']:<7} {', '.join(c['disq_reasons'])}")
        if len(disqualified) > 10:
            print(f"  ... and {len(disqualified) - 10} more")

    # --- Option pricing for top candidates ---
    print(f"\n{'=' * 90}")
    print(f" OPTION PRICING — Top {min(show_count, 20)} Candidates (30d ATM Calls)")
    print(f"{'=' * 90}")

    top_for_options = qualified[:min(show_count, 20)]

    print(f"\n  Fetching ATM call prices...\n")
    print(f"  {'Symbol':<7} {'Stock$':>8} {'Strike':>8} {'Exp Date':>11} "
          f"{'Call$':>7} {'Cost':>7} {'Prem%':>6} {'IV Proxy':>9}")
    print(f"  {'-' * 75}")

    for c in top_for_options:
        sym = c["symbol"]
        price = c.get("current_price")
        if not price:
            continue

        # Find 30d ATM call
        entry_date = as_of
        dt = datetime.strptime(entry_date, "%Y-%m-%d").date()
        exp_start = (dt + timedelta(days=20)).isoformat()
        exp_end = (dt + timedelta(days=45)).isoformat()

        try:
            contracts = list(client.list_options_contracts(
                underlying_ticker=sym,
                as_of=entry_date,
                expiration_date_gte=exp_start,
                expiration_date_lte=exp_end,
                contract_type="call",
                limit=100,
            ))
            time.sleep(API_DELAY)
        except Exception:
            contracts = []

        if not contracts:
            print(f"  {sym:<7} {price:>8.2f}  — no options chain")
            continue

        # Find closest to ATM
        best = None
        best_diff = float("inf")
        for ct in contracts:
            diff = abs(ct.strike_price - price)
            if diff < best_diff:
                best_diff = diff
                best = ct

        if not best:
            continue

        # Get current option price
        opt_ticker = best.ticker
        exp_date = str(best.expiration_date)
        strike = best.strike_price

        try:
            # Fetch recent option bar
            bar_start = (datetime.strptime(as_of, "%Y-%m-%d") - timedelta(days=5)).strftime("%Y-%m-%d")
            opt_bars = list(client.list_aggs(
                ticker=opt_ticker, multiplier=1, timespan="day",
                from_=bar_start, to=as_of, limit=5,
            ))
            time.sleep(API_DELAY)
        except Exception:
            opt_bars = []

        if opt_bars:
            opt_price = opt_bars[-1].close
            cost = opt_price * 100
            premium_pct = opt_price / price * 100
            # Annualized IV proxy: premium% * sqrt(365/DTE)
            dte = (datetime.strptime(exp_date, "%Y-%m-%d").date() - dt).days
            if dte > 0:
                iv_proxy = premium_pct * (365 / dte) ** 0.5
            else:
                iv_proxy = 0

            print(f"  {sym:<7} {price:>8.2f} {strike:>8.2f} {exp_date:>11} "
                  f"{opt_price:>7.2f} {cost:>7.0f} {premium_pct:>5.1f}% "
                  f"~{iv_proxy:>5.0f}%ann")
        else:
            print(f"  {sym:<7} {price:>8.2f} {strike:>8.2f} {exp_date:>11}  — no price data")

    print(f"\n  Prem% = call price / stock price (lower = cheaper leverage)")
    print(f"  IV Proxy = annualized premium (higher = more expensive options)")
    print(f"\n  NOTE: Check SA for Strong Buy duration (75+ days required).")
    print(f"  Look at each candidate's quant rating history on Seeking Alpha.")

    # Summary
    print(f"\n{'=' * 90}")
    print(f" SUMMARY")
    print(f"{'=' * 90}")
    print(f"\n  Candidates scored:    {len(candidates)}")
    print(f"  Qualified:            {len(qualified)}")
    print(f"  Disqualified:         {len(disqualified)}")
    print(f"  Recently picked:      {len(recent_picks)}")
    print(f"\n  Next announcement dates: 1st and 15th of each month")
    print(f"  Strategy: Buy ATM calls on top {min(args.top, 10)} candidates")
    print(f"            before the announcement. Hold for 14 days.")
    print(f"            Exit immediately on 75% loss (stop-loss).")


if __name__ == "__main__":
    main()
