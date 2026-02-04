"""
Alpha Picks — Predictive Factor Analysis
==========================================
Analyze historical Alpha Picks to identify characteristics that predict
which stocks SA will select. Goal: narrow candidates to ~10 before each
announcement date.

Usage:
    python alpha_picks_analysis.py
"""

import os
import sys
import sqlite3
import math
from datetime import datetime, timedelta
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required.")
    sys.exit(1)

from massive import RESTClient

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)
API_KEY_FILE = os.path.join(PROJECT_DIR, "Massive backtesting", "api_key.txt")
EXCEL_FILE = os.path.join(BASE_DIR, "ProQuant History 1_29_2026.xlsx")
PRICE_CACHE_DB = os.path.join(BASE_DIR, "price_cache.db")
OPTIONS_CACHE_DB = os.path.join(BASE_DIR, "options_picks_cache.db")
ANALYSIS_CACHE_DB = os.path.join(BASE_DIR, "alpha_picks_analysis.db")

API_DELAY = 0.15


def load_api_key():
    with open(API_KEY_FILE) as f:
        return f.read().strip()


# ---------------------------------------------------------------------------
# Analysis cache (for ticker details / sector lookups)
# ---------------------------------------------------------------------------

def init_analysis_cache():
    conn = sqlite3.connect(ANALYSIS_CACHE_DB)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS ticker_details (
            symbol          TEXT PRIMARY KEY,
            name            TEXT,
            market_cap      REAL,
            sic_code        TEXT,
            sic_description TEXT,
            sector          TEXT,
            industry        TEXT,
            list_date       TEXT,
            fetched_at      TEXT
        );
        CREATE TABLE IF NOT EXISTS daily_prices (
            symbol   TEXT NOT NULL,
            bar_date TEXT NOT NULL,
            close    REAL,
            PRIMARY KEY (symbol, bar_date)
        );
    """)
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Load Alpha Picks from Excel (full data including returns)
# ---------------------------------------------------------------------------

def load_alpha_picks_full():
    """Load Alpha Picks with all available columns."""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    picks = []

    if "AlphaPicks" not in wb.sheetnames:
        print("  ERROR: No AlphaPicks sheet found")
        wb.close()
        return []

    ws = wb["AlphaPicks"]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or len(row) < 3:
            continue
        if not row[1] or not row[2]:
            continue

        symbol = str(row[1]).strip().rstrip("*")
        if not symbol or len(symbol) > 6:
            continue

        pick_date = row[2]
        if hasattr(pick_date, "strftime"):
            pick_date = pick_date.strftime("%Y-%m-%d")
        else:
            pick_date = str(pick_date).strip()

        # Validate date format
        try:
            datetime.strptime(pick_date, "%Y-%m-%d")
        except ValueError:
            continue

        buy_price = float(row[3]) if len(row) >= 4 and row[3] else None

        # Closed date
        closed = row[4] if len(row) >= 5 else None
        if hasattr(closed, "strftime"):
            closed = closed.strftime("%Y-%m-%d")
        elif closed and str(closed).strip() == "-":
            closed = None
        elif closed:
            closed = str(closed).strip()

        sell_price = None
        if len(row) >= 6 and row[5] and str(row[5]).strip() != "-":
            try:
                sell_price = float(row[5])
            except (ValueError, TypeError):
                pass

        pick_return = None
        if len(row) >= 7 and row[6] is not None:
            try:
                pick_return = float(row[6])
            except (ValueError, TypeError):
                pass

        spy_return = None
        if len(row) >= 8 and row[7] is not None:
            try:
                spy_return = float(row[7])
            except (ValueError, TypeError):
                pass

        alpha = None
        if len(row) >= 9 and row[8] is not None:
            try:
                alpha = float(row[8])
            except (ValueError, TypeError):
                pass

        picks.append({
            "symbol": symbol,
            "pick_date": pick_date,
            "buy_price": buy_price,
            "closed": closed,
            "sell_price": sell_price,
            "return": pick_return,
            "spy_return": spy_return,
            "alpha": alpha,
        })

    wb.close()

    # Deduplicate
    seen = set()
    unique = []
    for p in picks:
        key = (p["symbol"], p["pick_date"])
        if key not in seen:
            seen.add(key)
            unique.append(p)

    return unique


# ---------------------------------------------------------------------------
# Price data helpers
# ---------------------------------------------------------------------------

def get_close_price(symbol, target_date, conn_price, conn_opts):
    """Get closing price from price_cache or options_picks_cache."""
    # Try price_cache first (more complete)
    upper = (datetime.strptime(target_date, "%Y-%m-%d") + timedelta(days=5)).strftime("%Y-%m-%d")
    row = conn_price.execute(
        "SELECT close FROM daily_prices WHERE symbol=? AND date>=? AND date<=? ORDER BY date LIMIT 1",
        (symbol, target_date, upper),
    ).fetchone()
    if row:
        return row[0]

    # Try options cache
    row = conn_opts.execute(
        "SELECT close FROM stock_prices WHERE symbol=? AND bar_date>=? AND bar_date<=? ORDER BY bar_date LIMIT 1",
        (symbol, target_date, upper),
    ).fetchone()
    if row:
        return row[0]

    return None


def get_close_before(symbol, target_date, days_back, conn_price, conn_opts):
    """Get closing price approximately days_back trading days before target_date."""
    cal_days = int(days_back * 1.5)  # approximate calendar days
    start = (datetime.strptime(target_date, "%Y-%m-%d") - timedelta(days=cal_days)).strftime("%Y-%m-%d")
    end = (datetime.strptime(target_date, "%Y-%m-%d") - timedelta(days=max(cal_days - 10, 0))).strftime("%Y-%m-%d")

    # Try price_cache
    row = conn_price.execute(
        "SELECT close FROM daily_prices WHERE symbol=? AND date>=? AND date<=? ORDER BY date DESC LIMIT 1",
        (symbol, start, end),
    ).fetchone()
    if row:
        return row[0]

    # Try options cache
    row = conn_opts.execute(
        "SELECT close FROM stock_prices WHERE symbol=? AND bar_date>=? AND bar_date<=? ORDER BY bar_date DESC LIMIT 1",
        (symbol, start, end),
    ).fetchone()
    if row:
        return row[0]

    return None


def compute_momentum(symbol, pick_date, conn_price, conn_opts):
    """Compute 1m, 3m, 6m, 12m momentum before pick date."""
    current = get_close_price(symbol, pick_date, conn_price, conn_opts)
    if not current:
        return {}

    result = {"price_at_pick": current}

    for label, days in [("1m", 21), ("3m", 63), ("6m", 126), ("12m", 252)]:
        past = get_close_before(symbol, pick_date, days, conn_price, conn_opts)
        if past and past > 0:
            ret = (current / past - 1) * 100
            result[f"return_{label}"] = round(ret, 1)

    return result


def compute_announcement_pop(symbol, pick_date, conn_price, conn_opts):
    """Compute stock return on announcement day and day after."""
    # Day before pick (or closest trading day)
    prev_start = (datetime.strptime(pick_date, "%Y-%m-%d") - timedelta(days=5)).strftime("%Y-%m-%d")
    prev_end = (datetime.strptime(pick_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")

    row = conn_price.execute(
        "SELECT close FROM daily_prices WHERE symbol=? AND date>=? AND date<=? ORDER BY date DESC LIMIT 1",
        (symbol, prev_start, prev_end),
    ).fetchone()
    if not row:
        row = conn_opts.execute(
            "SELECT close FROM stock_prices WHERE symbol=? AND bar_date>=? AND bar_date<=? ORDER BY bar_date DESC LIMIT 1",
            (symbol, prev_start, prev_end),
        ).fetchone()
    prev_close = row[0] if row else None

    pick_close = get_close_price(symbol, pick_date, conn_price, conn_opts)

    # Day after
    next_day = (datetime.strptime(pick_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    next_close = get_close_price(symbol, next_day, conn_price, conn_opts)

    result = {}
    if prev_close and prev_close > 0:
        if pick_close:
            result["day0_return"] = round((pick_close / prev_close - 1) * 100, 2)
        if next_close:
            result["day1_return"] = round((next_close / prev_close - 1) * 100, 2)
    return result


# ---------------------------------------------------------------------------
# Sector lookup via Polygon
# ---------------------------------------------------------------------------

def fetch_ticker_details(client, symbols):
    """Fetch ticker details (sector, market cap, etc.) from Polygon."""
    import time

    conn = sqlite3.connect(ANALYSIS_CACHE_DB)
    results = {}

    for symbol in symbols:
        # Check cache
        row = conn.execute(
            "SELECT symbol, name, market_cap, sic_code, sic_description, sector, industry FROM ticker_details WHERE symbol=?",
            (symbol,),
        ).fetchone()

        if row:
            results[symbol] = {
                "name": row[1], "market_cap": row[2],
                "sic_code": row[3], "sic_description": row[4],
                "sector": row[5], "industry": row[6],
            }
            continue

        # Fetch from API
        try:
            details = client.get_ticker_details(symbol)
            if details:
                name = getattr(details, "name", None)
                mcap = getattr(details, "market_cap", None)
                sic = getattr(details, "sic_code", None)
                sic_desc = getattr(details, "sic_description", None)

                # Map SIC to GICS-like sector
                sector = sic_to_sector(sic) if sic else None
                industry = sic_desc

                conn.execute(
                    "INSERT OR REPLACE INTO ticker_details VALUES (?,?,?,?,?,?,?,?,?)",
                    (symbol, name, mcap, sic, sic_desc, sector, industry,
                     getattr(details, "list_date", None),
                     datetime.now().isoformat()),
                )
                conn.commit()

                results[symbol] = {
                    "name": name, "market_cap": mcap,
                    "sic_code": sic, "sic_description": sic_desc,
                    "sector": sector, "industry": industry,
                }
        except Exception as e:
            pass

        time.sleep(API_DELAY)

    conn.close()
    return results


def sic_to_sector(sic_code):
    """Map SIC code to GICS-like sector.

    Specific ranges MUST come before broad ranges to avoid
    misclassifying tech/pharma/energy as generic 'Industrials'.
    """
    if not sic_code:
        return None
    try:
        sic = int(sic_code)
    except (ValueError, TypeError):
        return None

    # --- Specific ranges first (override broad Manufacturing/Services) ---

    # Energy
    if 1300 <= sic <= 1399:
        return "Energy"       # Oil/gas extraction
    if 2900 <= sic <= 2999:
        return "Energy"       # Petroleum refining

    # Health Care / Pharma / Biotech
    if 2830 <= sic <= 2836:
        return "Health Care"  # Pharmaceutical preparations
    if 2800 <= sic <= 2899:
        return "Health Care"  # Chemicals/pharma/biotech
    if 3841 <= sic <= 3851:
        return "Health Care"  # Medical instruments/devices
    if 5912 == sic:
        return "Health Care"  # Drug stores
    if 8000 <= sic <= 8099:
        return "Health Care"  # Health services
    if 8700 <= sic <= 8749:
        return "Health Care"  # Engineering/research services (biotech often here)

    # Information Technology
    if 3559 == sic:
        return "Information Technology"  # Semiconductor equipment
    if 3570 <= sic <= 3579:
        return "Information Technology"  # Computer hardware
    if 3660 <= sic <= 3679:
        return "Information Technology"  # Semiconductors/electronic components
    if 3674 == sic:
        return "Information Technology"  # Semiconductors
    if 3600 <= sic <= 3699:
        return "Information Technology"  # Electronic equipment
    if 3810 <= sic <= 3849:
        return "Information Technology"  # Instruments (measuring/testing)
    if 3825 == sic:
        return "Information Technology"  # Instruments for measuring
    if 7370 <= sic <= 7379:
        return "Information Technology"  # Computer services/software
    if 7372 == sic:
        return "Information Technology"  # Prepackaged software
    if 5045 == sic:
        return "Information Technology"  # Computer equipment wholesale
    if 5065 == sic:
        return "Information Technology"  # Electronic parts wholesale
    if 3690 == sic:
        return "Information Technology"  # Electronic components NEC
    if 3827 == sic:
        return "Information Technology"  # Optical instruments

    # Communication Services
    if 4800 <= sic <= 4899:
        return "Communication Services"  # Telecom
    if 4810 <= sic <= 4813:
        return "Communication Services"  # Telephone
    if 4840 <= sic <= 4841:
        return "Communication Services"  # Cable TV
    if 7810 <= sic <= 7819:
        return "Communication Services"  # Motion pictures
    if 7812 == sic:
        return "Communication Services"  # Motion picture production
    if 7311 == sic:
        return "Communication Services"  # Advertising services

    # Utilities
    if 4900 <= sic <= 4999:
        return "Utilities"

    # --- Broad ranges ---

    # Mining/Materials
    if 1000 <= sic <= 1499:
        return "Materials"

    # Construction
    if 1500 <= sic <= 1799:
        return "Industrials"

    # Agriculture
    if 100 <= sic <= 999:
        return "Consumer Staples"

    # Manufacturing (broad — after specific tech/pharma/energy carved out)
    if 2000 <= sic <= 2111:
        return "Consumer Staples"  # Food/tobacco
    if 2200 <= sic <= 2399:
        return "Consumer Discretionary"  # Textiles/apparel
    if 2500 <= sic <= 2599:
        return "Consumer Discretionary"  # Furniture
    if 2000 <= sic <= 3999:
        return "Industrials"  # Remaining manufacturing

    # Transportation
    if 4000 <= sic <= 4799:
        return "Industrials"

    # Wholesale
    if 5000 <= sic <= 5199:
        return "Consumer Staples"

    # Retail
    if 5200 <= sic <= 5999:
        return "Consumer Discretionary"

    # Finance/Insurance/Real Estate
    if 6000 <= sic <= 6499:
        return "Financials"
    if 6500 <= sic <= 6599:
        return "Real Estate"
    if 6700 <= sic <= 6799:
        return "Financials"  # Holding companies

    # Services (broad — after specific IT/comms/health carved out)
    if 7000 <= sic <= 7099:
        return "Consumer Discretionary"  # Hotels
    if 7200 <= sic <= 7299:
        return "Consumer Discretionary"  # Personal services
    if 7500 <= sic <= 7549:
        return "Consumer Discretionary"  # Auto repair/services
    if 7900 <= sic <= 7999:
        return "Consumer Discretionary"  # Amusement/recreation
    if 8200 <= sic <= 8299:
        return "Consumer Discretionary"  # Education
    if 5800 <= sic <= 5899:
        return "Consumer Discretionary"  # Eating/drinking
    if 7000 <= sic <= 8999:
        return "Industrials"  # Remaining services

    return None


# More precise SIC mapping with overrides
SIC_SECTOR_OVERRIDES = {
    # Tech companies often miscategorized
    "AAPL": "Information Technology", "MSFT": "Information Technology",
    "GOOGL": "Communication Services", "META": "Communication Services",
    "AMZN": "Consumer Discretionary", "NVDA": "Information Technology",
    "CRM": "Information Technology", "UBER": "Information Technology",
    "OKTA": "Information Technology", "TWLO": "Communication Services",
    "CRDO": "Information Technology", "QTWO": "Information Technology",
    "ZETA": "Information Technology", "GCT": "Communication Services",
    "ACLS": "Information Technology", "HLIT": "Information Technology",
    "APP": "Information Technology", "TGLS": "Information Technology",
    "PERI": "Communication Services", "TMUS": "Communication Services",
    # Financials
    "WFC": "Financials", "SYF": "Financials", "LC": "Financials",
    "JXN": "Financials", "ALL": "Financials", "BRK.B": "Financials",
    "RGA": "Financials", "PYPL": "Financials", "EZPW": "Financials",
    # Consumer Discretionary
    "CCL": "Consumer Discretionary", "RCL": "Consumer Discretionary",
    "ANF": "Consumer Discretionary", "EAT": "Consumer Discretionary",
    "W": "Consumer Discretionary", "GM": "Consumer Discretionary",
    "LRN": "Consumer Discretionary",
    # Health Care
    "INCY": "Health Care", "ARQT": "Health Care",
    # Energy
    "XOM": "Energy", "CVX": "Energy", "COP": "Energy",
    "VLO": "Energy", "MPC": "Energy", "DINO": "Energy",
    "PARR": "Energy", "NUE": "Materials", "LTHM": "Materials",
    # Industrials
    "URI": "Industrials", "TEX": "Industrials", "CAH": "Health Care",
    "BLBD": "Industrials", "SKYW": "Industrials", "STRL": "Industrials",
    "MOD": "Industrials", "POWL": "Industrials", "AGX": "Industrials",
    "ATGE": "Consumer Discretionary", "BXC": "Industrials",
    # Materials / Mining
    "NEM": "Materials", "CDE": "Materials", "KGC": "Materials",
    "SSRM": "Materials", "ARCH": "Energy", "AMR": "Energy",
    # Real Estate / Other
    "CMCSA": "Communication Services",
    "PEP": "Consumer Staples", "UNFI": "Consumer Staples",
    "CAAP": "Industrials", "MFC": "Financials",
    "MHO": "Consumer Discretionary", "GRBK": "Consumer Discretionary",
    "AMPH": "Health Care", "ITRN": "Information Technology",
    "PPC": "Consumer Staples", "DXPE": "Industrials",
    "WLDN": "Industrials", "SFM": "Consumer Staples",
    "TTMI": "Information Technology", "TA": "Consumer Discretionary",
    "SU": "Energy", "CLS": "Information Technology",
    "B": "Industrials", "TIGO": "Communication Services",
    "MU": "Information Technology", "ASC": "Energy",
}


# ---------------------------------------------------------------------------
# Main Analysis
# ---------------------------------------------------------------------------

def main():
    print(f"\n{'=' * 80}")
    print(f" ALPHA PICKS — PREDICTIVE FACTOR ANALYSIS")
    print(f"{'=' * 80}")

    # Load picks
    picks = load_alpha_picks_full()
    print(f"\n  Loaded {len(picks)} unique Alpha Picks")

    if not picks:
        return

    # Initialize
    init_analysis_cache()
    client = RESTClient(api_key=load_api_key())

    # Open price databases
    conn_price = sqlite3.connect(PRICE_CACHE_DB)
    conn_opts = sqlite3.connect(OPTIONS_CACHE_DB)

    # Get ticker details for sectors
    symbols = list(set(p["symbol"] for p in picks))
    print(f"  Fetching ticker details for {len(symbols)} unique symbols...")
    details = fetch_ticker_details(client, symbols)

    # Enrich picks with sector and momentum data
    print(f"  Computing momentum and announcement returns...\n")

    for p in picks:
        sym = p["symbol"]

        # Sector
        if sym in SIC_SECTOR_OVERRIDES:
            p["sector"] = SIC_SECTOR_OVERRIDES[sym]
        elif sym in details and details[sym].get("sector"):
            p["sector"] = details[sym]["sector"]
        else:
            p["sector"] = "Unknown"

        # Market cap
        if sym in details and details[sym].get("market_cap"):
            p["market_cap"] = details[sym]["market_cap"]
        else:
            p["market_cap"] = None

        # Momentum
        mom = compute_momentum(sym, p["pick_date"], conn_price, conn_opts)
        p.update(mom)

        # Announcement pop
        pop = compute_announcement_pop(sym, p["pick_date"], conn_price, conn_opts)
        p.update(pop)

    conn_price.close()
    conn_opts.close()

    # -----------------------------------------------------------------------
    # Analysis 1: Basic Statistics
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 1. BASIC STATISTICS")
    print(f"{'=' * 80}")

    returns = [p["return"] for p in picks if p["return"] is not None]
    spy_returns = [p["spy_return"] for p in picks if p["spy_return"] is not None]
    alphas = [p["alpha"] for p in picks if p["alpha"] is not None]

    print(f"\n  Total picks: {len(picks)}")
    print(f"  Date range: {picks[-1]['pick_date']} to {picks[0]['pick_date']}")
    print(f"  Still open: {sum(1 for p in picks if p['closed'] is None)}")
    print(f"  Closed:     {sum(1 for p in picks if p['closed'] is not None)}")

    if returns:
        avg_ret = sum(returns) / len(returns) * 100
        median_ret = sorted(returns)[len(returns) // 2] * 100
        win_rate = sum(1 for r in returns if r > 0) / len(returns) * 100
        avg_spy = sum(spy_returns) / len(spy_returns) * 100 if spy_returns else 0
        avg_alpha = sum(alphas) / len(alphas) * 100 if alphas else 0

        print(f"\n  Average return:  {avg_ret:>7.1f}%")
        print(f"  Median return:   {median_ret:>7.1f}%")
        print(f"  Win rate:        {win_rate:>7.1f}%")
        print(f"  Avg SPY return:  {avg_spy:>7.1f}%")
        print(f"  Avg alpha:       {avg_alpha:>7.1f}%")

    # -----------------------------------------------------------------------
    # Analysis 2: Sector Distribution
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 2. SECTOR DISTRIBUTION")
    print(f"{'=' * 80}")

    sector_counts = Counter(p["sector"] for p in picks)
    total = len(picks)

    print(f"\n  {'Sector':<28} {'Count':>5} {'%':>6}  {'Avg Return':>10}")
    print(f"  {'-' * 55}")

    for sector, count in sector_counts.most_common():
        pct = count / total * 100
        sector_returns = [p["return"] for p in picks
                         if p["sector"] == sector and p["return"] is not None]
        avg_r = sum(sector_returns) / len(sector_returns) * 100 if sector_returns else 0
        print(f"  {sector:<28} {count:>5} {pct:>5.1f}%  {avg_r:>9.1f}%")

    # -----------------------------------------------------------------------
    # Analysis 3: Market Cap Distribution
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 3. MARKET CAP DISTRIBUTION (current)")
    print(f"{'=' * 80}")

    mcaps = [(p["symbol"], p["market_cap"]) for p in picks if p.get("market_cap")]
    if mcaps:
        caps = [m[1] for m in mcaps]
        caps_sorted = sorted(caps)
        print(f"\n  Stocks with market cap data: {len(mcaps)}")
        print(f"  Min:      ${min(caps)/1e9:>8.1f}B")
        print(f"  25th pct: ${caps_sorted[len(caps_sorted)//4]/1e9:>8.1f}B")
        print(f"  Median:   ${caps_sorted[len(caps_sorted)//2]/1e9:>8.1f}B")
        print(f"  75th pct: ${caps_sorted[3*len(caps_sorted)//4]/1e9:>8.1f}B")
        print(f"  Max:      ${max(caps)/1e9:>8.1f}B")

        # Buckets
        buckets = {"Micro (<$500M)": 0, "Small ($500M-2B)": 0,
                   "Mid ($2B-10B)": 0, "Large ($10B-100B)": 0,
                   "Mega (>$100B)": 0}
        for cap in caps:
            if cap < 500e6:
                buckets["Micro (<$500M)"] += 1
            elif cap < 2e9:
                buckets["Small ($500M-2B)"] += 1
            elif cap < 10e9:
                buckets["Mid ($2B-10B)"] += 1
            elif cap < 100e9:
                buckets["Large ($10B-100B)"] += 1
            else:
                buckets["Mega (>$100B)"] += 1

        print(f"\n  {'Bucket':<20} {'Count':>5} {'%':>6}")
        print(f"  {'-' * 35}")
        for bucket, count in buckets.items():
            pct = count / len(caps) * 100
            print(f"  {bucket:<20} {count:>5} {pct:>5.1f}%")

    # -----------------------------------------------------------------------
    # Analysis 4: Buy Price Distribution
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 4. BUY PRICE AT TIME OF PICK")
    print(f"{'=' * 80}")

    prices = [p["buy_price"] for p in picks if p.get("buy_price")]
    if prices:
        prices_sorted = sorted(prices)
        print(f"\n  Min:    ${min(prices):>8.2f}")
        print(f"  25th:   ${prices_sorted[len(prices_sorted)//4]:>8.2f}")
        print(f"  Median: ${prices_sorted[len(prices_sorted)//2]:>8.2f}")
        print(f"  75th:   ${prices_sorted[3*len(prices_sorted)//4]:>8.2f}")
        print(f"  Max:    ${max(prices):>8.2f}")

        # Buckets
        pbuckets = {"<$20": 0, "$20-50": 0, "$50-100": 0,
                    "$100-200": 0, "$200+": 0}
        for px in prices:
            if px < 20:
                pbuckets["<$20"] += 1
            elif px < 50:
                pbuckets["$20-50"] += 1
            elif px < 100:
                pbuckets["$50-100"] += 1
            elif px < 200:
                pbuckets["$100-200"] += 1
            else:
                pbuckets["$200+"] += 1

        print(f"\n  {'Bucket':<12} {'Count':>5} {'%':>6}")
        print(f"  {'-' * 27}")
        for bucket, count in pbuckets.items():
            pct = count / len(prices) * 100
            print(f"  {bucket:<12} {count:>5} {pct:>5.1f}%")

    # -----------------------------------------------------------------------
    # Analysis 5: Pre-Pick Momentum
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 5. PRE-PICK MOMENTUM (returns before being picked)")
    print(f"{'=' * 80}")

    for period in ["1m", "3m", "6m", "12m"]:
        key = f"return_{period}"
        vals = [p[key] for p in picks if key in p]
        if vals:
            avg = sum(vals) / len(vals)
            median = sorted(vals)[len(vals) // 2]
            positive = sum(1 for v in vals if v > 0) / len(vals) * 100
            print(f"\n  {period} pre-pick return (n={len(vals)}):")
            print(f"    Average: {avg:>7.1f}%    Median: {median:>7.1f}%    "
                  f"% Positive: {positive:.0f}%")

    # Momentum profile of typical pick
    mom_profiles = []
    for p in picks:
        if all(f"return_{t}" in p for t in ["1m", "3m", "6m"]):
            mom_profiles.append(p)

    if mom_profiles:
        print(f"\n  Momentum Profile of Typical Pick (n={len(mom_profiles)}):")
        print(f"  {'Symbol':<7} {'Pick Date':<12} {'1m%':>7} {'3m%':>7} {'6m%':>7} {'12m%':>7}")
        print(f"  {'-' * 55}")

        # Show picks sorted by 3m momentum to see the pattern
        mom_profiles.sort(key=lambda x: x.get("return_3m", 0), reverse=True)
        for p in mom_profiles[:15]:
            r1 = p.get("return_1m", "")
            r3 = p.get("return_3m", "")
            r6 = p.get("return_6m", "")
            r12 = p.get("return_12m", "")
            r1s = f"{r1:>6.1f}%" if isinstance(r1, (int, float)) else f"{'N/A':>7}"
            r3s = f"{r3:>6.1f}%" if isinstance(r3, (int, float)) else f"{'N/A':>7}"
            r6s = f"{r6:>6.1f}%" if isinstance(r6, (int, float)) else f"{'N/A':>7}"
            r12s = f"{r12:>6.1f}%" if isinstance(r12, (int, float)) else f"{'N/A':>7}"
            print(f"  {p['symbol']:<7} {p['pick_date']:<12} {r1s} {r3s} {r6s} {r12s}")

        print(f"  ... (showing top 15 by 3m momentum)")

    # -----------------------------------------------------------------------
    # Analysis 6: Announcement Day Pop
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 6. ANNOUNCEMENT DAY RETURNS")
    print(f"{'=' * 80}")

    day0_returns = [(p["symbol"], p["pick_date"], p["day0_return"])
                    for p in picks if "day0_return" in p]
    day1_returns = [(p["symbol"], p["pick_date"], p["day1_return"])
                    for p in picks if "day1_return" in p]

    if day0_returns:
        vals = [r[2] for r in day0_returns]
        avg = sum(vals) / len(vals)
        median = sorted(vals)[len(vals) // 2]
        positive = sum(1 for v in vals if v > 0) / len(vals) * 100
        print(f"\n  Pick day return (n={len(vals)}):")
        print(f"    Average: {avg:>6.2f}%    Median: {median:>6.2f}%    "
              f"% Positive: {positive:.0f}%")

    if day1_returns:
        vals = [r[2] for r in day1_returns]
        avg = sum(vals) / len(vals)
        median = sorted(vals)[len(vals) // 2]
        positive = sum(1 for v in vals if v > 0) / len(vals) * 100
        print(f"  Day after pick return (n={len(vals)}):")
        print(f"    Average: {avg:>6.2f}%    Median: {median:>6.2f}%    "
              f"% Positive: {positive:.0f}%")

    # Show biggest announcement pops
    if day0_returns:
        day0_returns.sort(key=lambda x: x[2], reverse=True)
        print(f"\n  Biggest announcement day pops:")
        for sym, dt, ret in day0_returns[:10]:
            print(f"    {sym:<7} {dt}  {ret:>+6.2f}%")
        print(f"\n  Biggest announcement day drops:")
        for sym, dt, ret in day0_returns[-5:]:
            print(f"    {sym:<7} {dt}  {ret:>+6.2f}%")

    # -----------------------------------------------------------------------
    # Analysis 7: Timing Patterns
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 7. TIMING PATTERNS")
    print(f"{'=' * 80}")

    # Picks per year
    year_counts = Counter(p["pick_date"][:4] for p in picks)
    print(f"\n  Picks per year:")
    for year in sorted(year_counts.keys()):
        print(f"    {year}: {year_counts[year]}")

    # Day of month distribution
    day_counts = Counter(int(p["pick_date"][8:10]) for p in picks)
    print(f"\n  Day of month distribution:")
    first_half = sum(v for k, v in day_counts.items() if k <= 15)
    second_half = sum(v for k, v in day_counts.items() if k > 15)
    print(f"    1st-15th: {first_half}    16th-31st: {second_half}")

    # -----------------------------------------------------------------------
    # Analysis 8: Repeat Picks
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 8. REPEAT PICKS (stocks picked multiple times)")
    print(f"{'=' * 80}")

    symbol_counts = Counter(p["symbol"] for p in picks)
    repeats = {s: c for s, c in symbol_counts.items() if c > 1}

    if repeats:
        print(f"\n  {len(repeats)} stocks picked more than once:")
        for sym, count in sorted(repeats.items(), key=lambda x: -x[1]):
            dates = [p["pick_date"] for p in picks if p["symbol"] == sym]
            print(f"    {sym:<7} x{count}  dates: {', '.join(dates)}")
    else:
        print(f"\n  No repeat picks found")

    # -----------------------------------------------------------------------
    # Analysis 9: Sector Rotation
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 9. SECTOR ROTATION (consecutive pick sectors)")
    print(f"{'=' * 80}")

    # Group picks by announcement date
    picks_by_date = defaultdict(list)
    for p in picks:
        picks_by_date[p["pick_date"]].append(p)

    dates_sorted = sorted(picks_by_date.keys())
    print(f"\n  {'Date':<12} {'Pick 1':<7} {'Sector 1':<25} {'Pick 2':<7} {'Sector 2':<25}")
    print(f"  {'-' * 80}")

    same_sector_count = 0
    for dt in dates_sorted:
        date_picks = picks_by_date[dt]
        if len(date_picks) >= 2:
            s1 = date_picks[0]["sector"]
            s2 = date_picks[1]["sector"]
            if s1 == s2:
                same_sector_count += 1
            print(f"  {dt:<12} {date_picks[0]['symbol']:<7} {s1:<25} "
                  f"{date_picks[1]['symbol']:<7} {s2:<25}")
        elif len(date_picks) == 1:
            print(f"  {dt:<12} {date_picks[0]['symbol']:<7} {date_picks[0]['sector']:<25}")

    paired_dates = sum(1 for dt in dates_sorted if len(picks_by_date[dt]) >= 2)
    if paired_dates > 0:
        print(f"\n  Same sector in pair: {same_sector_count}/{paired_dates} "
              f"({same_sector_count/paired_dates*100:.0f}%)")

    # -----------------------------------------------------------------------
    # Analysis 10: Predictive Summary
    # -----------------------------------------------------------------------
    print(f"\n{'=' * 80}")
    print(f" 10. PREDICTIVE FACTOR SUMMARY")
    print(f"{'=' * 80}")

    print(f"""
  Based on the analysis above, the typical Alpha Pick has:

  SECTOR:     Most common sectors and their frequencies
  MARKET CAP: Typical size range
  MOMENTUM:   Pre-pick price action pattern
  TIMING:     2 picks per announcement (1st and 15th)
  REPEATS:    Whether stocks get re-picked

  Use these characteristics to narrow candidates before each announcement.
  See detailed numbers in the sections above.
""")

    print(f"\nDone. Analysis cache: {ANALYSIS_CACHE_DB}")


if __name__ == "__main__":
    main()
