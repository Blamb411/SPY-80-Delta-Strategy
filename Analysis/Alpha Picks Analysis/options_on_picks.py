"""
Options on Alpha Picks — Historical Analysis
===============================================
Backtest buying ATM calls on Alpha Picks and ProQuant stocks
the day after announcement.

Usage:
    python options_on_picks.py                        # Alpha Picks only
    python options_on_picks.py --portfolio proquant    # ProQuant only
    python options_on_picks.py --portfolio both         # Both portfolios
    python options_on_picks.py --report                # Save results to file
    python options_on_picks.py --symbols AAPL,MSFT     # Specific tickers only
"""

import os
import sys
import time
import sqlite3
import argparse
from datetime import datetime, timedelta, date
from typing import List, Dict, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

from massive import RESTClient

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)
API_KEY_FILE = os.path.join(PROJECT_DIR, "Massive backtesting", "api_key.txt")
EXCEL_FILE = os.path.join(BASE_DIR, "ProQuant History 1_29_2026.xlsx")
CACHE_DB = os.path.join(BASE_DIR, "options_picks_cache.db")
API_DELAY = 0.15

# DTE targets for call options
DTE_TARGETS = [
    {"label": "30d", "min_dte": 20, "max_dte": 45, "target": 30},
    {"label": "60d", "min_dte": 50, "max_dte": 75, "target": 60},
    {"label": "90d", "min_dte": 80, "max_dte": 105, "target": 90},
]

# Fixed holding periods (days) for early exit analysis
HOLDING_PERIODS = [14, 30]


def load_api_key() -> str:
    with open(API_KEY_FILE) as f:
        return f.read().strip()


# ---------------------------------------------------------------------------
# Cache database
# ---------------------------------------------------------------------------

def init_cache():
    conn = sqlite3.connect(CACHE_DB)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS stock_prices (
            symbol   TEXT NOT NULL,
            bar_date TEXT NOT NULL,
            open     REAL,
            high     REAL,
            low      REAL,
            close    REAL,
            volume   REAL,
            PRIMARY KEY (symbol, bar_date)
        );
        CREATE TABLE IF NOT EXISTS options_contracts (
            option_ticker   TEXT NOT NULL,
            underlying      TEXT NOT NULL,
            as_of_date      TEXT,
            expiration_date TEXT,
            strike_price    REAL,
            contract_type   TEXT,
            PRIMARY KEY (option_ticker, as_of_date)
        );
        CREATE TABLE IF NOT EXISTS option_bars (
            option_ticker TEXT NOT NULL,
            bar_date      TEXT NOT NULL,
            open          REAL,
            high          REAL,
            low           REAL,
            close         REAL,
            volume        REAL,
            PRIMARY KEY (option_ticker, bar_date)
        );
        CREATE TABLE IF NOT EXISTS fetch_log (
            fetch_key    TEXT PRIMARY KEY,
            last_fetched TEXT
        );
    """)
    conn.commit()
    conn.close()


def _is_fetched(key: str) -> bool:
    conn = sqlite3.connect(CACHE_DB)
    row = conn.execute("SELECT 1 FROM fetch_log WHERE fetch_key=?", (key,)).fetchone()
    conn.close()
    return row is not None


def _mark_fetched(key: str):
    conn = sqlite3.connect(CACHE_DB)
    conn.execute(
        "INSERT OR REPLACE INTO fetch_log (fetch_key, last_fetched) VALUES (?, ?)",
        (key, datetime.now().isoformat()),
    )
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Excel loaders
# ---------------------------------------------------------------------------

def load_alpha_picks() -> List[Dict]:
    """Load Alpha Picks from Excel. Returns list of {symbol, pick_date, buy_price}."""
    if not os.path.exists(EXCEL_FILE):
        print(f"  ERROR: Excel file not found: {EXCEL_FILE}")
        return []

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    picks = []

    if "AlphaPicks" in wb.sheetnames:
        ws = wb["AlphaPicks"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row and len(row) >= 3 and row[1] and row[2]:
                symbol = str(row[1]).strip().rstrip("*")
                pick_date = row[2]
                if hasattr(pick_date, "strftime"):
                    pick_date = pick_date.strftime("%Y-%m-%d")
                else:
                    pick_date = str(pick_date).strip()

                buy_price = float(row[3]) if len(row) >= 4 and row[3] else None

                picks.append({
                    "symbol": symbol,
                    "pick_date": pick_date,
                    "buy_price": buy_price,
                    "portfolio": "AlphaPicks",
                })

    wb.close()

    # Deduplicate by (symbol, pick_date)
    seen = set()
    unique_picks = []
    for p in picks:
        key = (p["symbol"], p["pick_date"])
        if key not in seen:
            seen.add(key)
            unique_picks.append(p)

    return unique_picks


def load_proquant_picks() -> List[Dict]:
    """Load ProQuant new additions from Excel.

    Treats the first appearance of each symbol as a 'pick'.
    """
    if not os.path.exists(EXCEL_FILE):
        return []

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    picks = []
    seen_symbols = set()

    if "ProQuant" in wb.sheetnames:
        ws = wb["ProQuant"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row or not row[0] or not row[1]:
                continue

            symbol = str(row[0]).strip()
            rebal_date = row[1]
            if hasattr(rebal_date, "strftime"):
                rebal_date = rebal_date.strftime("%Y-%m-%d")
            else:
                rebal_date = str(rebal_date).strip()

            price = float(row[6]) if len(row) >= 7 and row[6] else None

            # Only count first appearance as a "pick"
            if symbol not in seen_symbols:
                seen_symbols.add(symbol)
                picks.append({
                    "symbol": symbol,
                    "pick_date": rebal_date,
                    "buy_price": price,
                    "portfolio": "ProQuant",
                })

    wb.close()
    return picks


# ---------------------------------------------------------------------------
# Options analyzer
# ---------------------------------------------------------------------------

class OptionsAnalyzer:
    def __init__(self):
        self.client = RESTClient(api_key=load_api_key())
        init_cache()

    def _ts_to_date(self, ts) -> str:
        if isinstance(ts, (int, float)):
            return datetime.utcfromtimestamp(ts / 1000).strftime("%Y-%m-%d")
        return str(ts)

    def _next_business_day(self, date_str: str) -> str:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        dt += timedelta(days=1)
        while dt.weekday() >= 5:
            dt += timedelta(days=1)
        return dt.strftime("%Y-%m-%d")

    # --- Stock prices ---

    def get_stock_price(self, symbol: str, target_date: str) -> Optional[float]:
        """Get stock closing price on or just after target_date (within 10 days)."""
        upper = (datetime.strptime(target_date, "%Y-%m-%d") + timedelta(days=10)).strftime("%Y-%m-%d")
        conn = sqlite3.connect(CACHE_DB)
        row = conn.execute(
            "SELECT close FROM stock_prices WHERE symbol=? AND bar_date>=? AND bar_date<=? ORDER BY bar_date LIMIT 1",
            (symbol, target_date, upper),
        ).fetchone()
        conn.close()
        if row:
            return row[0]

        # Fetch a week of bars from the API
        end = upper
        fetch_key = f"stock:{symbol}:{target_date}"

        if not _is_fetched(fetch_key):
            try:
                aggs = list(self.client.list_aggs(
                    ticker=symbol, multiplier=1, timespan="day",
                    from_=target_date, to=end, limit=10,
                ))
                conn = sqlite3.connect(CACHE_DB)
                for a in aggs:
                    bd = self._ts_to_date(a.timestamp)
                    conn.execute(
                        "INSERT OR IGNORE INTO stock_prices VALUES (?,?,?,?,?,?,?)",
                        (symbol, bd, a.open, a.high, a.low, a.close, a.volume),
                    )
                conn.commit()
                conn.close()
            except Exception as e:
                print(f"    WARNING: No stock data for {symbol} near {target_date}: {e}")

            time.sleep(API_DELAY)
            _mark_fetched(fetch_key)

        conn = sqlite3.connect(CACHE_DB)
        row = conn.execute(
            "SELECT close FROM stock_prices WHERE symbol=? AND bar_date>=? AND bar_date<=? ORDER BY bar_date LIMIT 1",
            (symbol, target_date, upper),
        ).fetchone()
        conn.close()
        return row[0] if row else None

    # --- Options contracts ---

    def find_call_contracts(self, symbol: str, entry_date: str, dte_cfg: Dict) -> List[Dict]:
        """Find available call contracts for a given DTE range."""
        dt = datetime.strptime(entry_date, "%Y-%m-%d").date()
        exp_start = (dt + timedelta(days=dte_cfg["min_dte"])).isoformat()
        exp_end = (dt + timedelta(days=dte_cfg["max_dte"])).isoformat()

        fetch_key = f"contracts:{symbol}:{entry_date}:{dte_cfg['label']}"

        # Check cache
        conn = sqlite3.connect(CACHE_DB)
        conn.row_factory = sqlite3.Row
        cached = conn.execute(
            """SELECT * FROM options_contracts
               WHERE underlying=? AND as_of_date=? AND contract_type='call'
               AND expiration_date>=? AND expiration_date<=?""",
            (symbol, entry_date, exp_start, exp_end),
        ).fetchall()
        conn.close()

        if cached:
            return [dict(r) for r in cached]

        if _is_fetched(fetch_key):
            return []

        # Fetch from Polygon
        contracts = []
        try:
            results = list(self.client.list_options_contracts(
                underlying_ticker=symbol,
                as_of=entry_date,
                expiration_date_gte=exp_start,
                expiration_date_lte=exp_end,
                contract_type="call",
                limit=1000,
            ))
            conn = sqlite3.connect(CACHE_DB)
            for c in results:
                conn.execute(
                    "INSERT OR IGNORE INTO options_contracts VALUES (?,?,?,?,?,?)",
                    (c.ticker, symbol, entry_date, c.expiration_date,
                     c.strike_price, c.contract_type),
                )
                contracts.append({
                    "option_ticker": c.ticker,
                    "underlying": symbol,
                    "as_of_date": entry_date,
                    "expiration_date": str(c.expiration_date),
                    "strike_price": c.strike_price,
                    "contract_type": c.contract_type,
                })
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"    WARNING: No options chain for {symbol} on {entry_date}: {e}")

        time.sleep(API_DELAY)
        _mark_fetched(fetch_key)
        return contracts

    def select_atm_contract(self, contracts: List[Dict], stock_price: float,
                            dte_target: int, entry_date: str,
                            otm_pct: float = 0.0) -> Optional[Dict]:
        """Select the contract closest to target strike and closest to target DTE.

        otm_pct: 0 = ATM, 5 = 5% OTM (strike 5% above stock price), etc.
        """
        if not contracts:
            return None

        target_strike = stock_price * (1 + otm_pct / 100)
        entry_dt = datetime.strptime(entry_date, "%Y-%m-%d").date()
        best = None
        best_score = float("inf")

        for c in contracts:
            strike_diff = abs(c["strike_price"] - target_strike) / stock_price
            exp_dt = datetime.strptime(str(c["expiration_date"]), "%Y-%m-%d").date()
            dte = (exp_dt - entry_dt).days
            dte_diff = abs(dte - dte_target) / max(dte_target, 1)

            # Weight strike closeness more heavily than DTE closeness
            score = strike_diff * 3 + dte_diff
            if score < best_score:
                best_score = score
                best = c

        # Skip if strike is too far from stock price (likely stock split issue)
        if best and abs(best["strike_price"] - stock_price) / stock_price > 0.50:
            return None

        return best

    # --- Option bars ---

    def fetch_option_bars(self, option_ticker: str, start: str, end: str) -> List[Dict]:
        """Fetch and cache daily bars for an option contract."""
        fetch_key = f"opt_bars:{option_ticker}"

        if not _is_fetched(fetch_key):
            try:
                aggs = list(self.client.list_aggs(
                    ticker=option_ticker, multiplier=1, timespan="day",
                    from_=start, to=end, limit=5000,
                ))
                conn = sqlite3.connect(CACHE_DB)
                for a in aggs:
                    bd = self._ts_to_date(a.timestamp)
                    conn.execute(
                        "INSERT OR IGNORE INTO option_bars VALUES (?,?,?,?,?,?,?)",
                        (option_ticker, bd, a.open, a.high, a.low, a.close, a.volume),
                    )
                conn.commit()
                conn.close()
            except Exception as e:
                pass  # silently skip — no bars means illiquid or delisted

            time.sleep(API_DELAY)
            _mark_fetched(fetch_key)

        conn = sqlite3.connect(CACHE_DB)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM option_bars WHERE option_ticker=? AND bar_date>=? AND bar_date<=? ORDER BY bar_date",
            (option_ticker, start, end),
        ).fetchall()
        conn.close()
        return [dict(r) for r in rows]

    # --- Main analysis ---

    def analyze_pick(self, pick: Dict, otm_pct: float = 0.0,
                     dte_filter: str = None) -> Dict:
        """Analyze a single pick across all DTE targets."""
        symbol = pick["symbol"]
        pick_date = pick["pick_date"]

        entry_date = self._next_business_day(pick_date)

        stock_price = self.get_stock_price(symbol, entry_date)
        if not stock_price:
            return {**pick, "entry_date": entry_date, "stock_price": None,
                    "positions": [], "error": "No stock price data"}

        result = {
            **pick,
            "entry_date": entry_date,
            "stock_price": round(stock_price, 2),
            "positions": [],
        }

        dte_targets = DTE_TARGETS
        if dte_filter:
            dte_targets = [d for d in DTE_TARGETS if d["label"] == dte_filter]

        for dte_cfg in dte_targets:
            contracts = self.find_call_contracts(symbol, entry_date, dte_cfg)
            contract = self.select_atm_contract(
                contracts, stock_price, dte_cfg["target"], entry_date,
                otm_pct=otm_pct,
            )

            if not contract:
                result["positions"].append({
                    "dte_label": dte_cfg["label"],
                    "error": "No contracts",
                })
                continue

            opt_ticker = contract["option_ticker"]
            exp_date = str(contract["expiration_date"])
            strike = contract["strike_price"]

            entry_dt = datetime.strptime(entry_date, "%Y-%m-%d").date()
            exp_dt = datetime.strptime(exp_date, "%Y-%m-%d").date()
            dte = (exp_dt - entry_dt).days

            # Fetch option bars for the full life of the contract
            bars = self.fetch_option_bars(opt_ticker, entry_date, exp_date)

            if not bars:
                result["positions"].append({
                    "dte_label": dte_cfg["label"],
                    "strike": strike,
                    "expiration": exp_date,
                    "dte": dte,
                    "error": "No option price data",
                })
                continue

            # Entry price = close on first available bar
            entry_price = bars[0]["close"]
            if not entry_price or entry_price <= 0:
                result["positions"].append({
                    "dte_label": dte_cfg["label"],
                    "strike": strike,
                    "expiration": exp_date,
                    "error": "Invalid entry price",
                })
                continue

            position = {
                "dte_label": dte_cfg["label"],
                "option_ticker": opt_ticker,
                "strike": strike,
                "expiration": exp_date,
                "dte": dte,
                "entry_price": round(entry_price, 2),
                "cost_per_contract": round(entry_price * 100, 2),
                "exits": {},
            }

            # --- Hold to expiry ---
            last_bar = bars[-1]
            exit_price = last_bar["close"]
            stock_at_expiry = self.get_stock_price(symbol, exp_date)
            intrinsic = max(0, stock_at_expiry - strike) if stock_at_expiry else None

            if exit_price is not None:
                pnl = (exit_price - entry_price) * 100
                ret = (exit_price / entry_price - 1) * 100
                position["exits"]["expiry"] = {
                    "exit_date": last_bar["bar_date"],
                    "exit_price": round(exit_price, 2),
                    "intrinsic": round(intrinsic, 2) if intrinsic is not None else None,
                    "stock_at_expiry": round(stock_at_expiry, 2) if stock_at_expiry else None,
                    "pnl": round(pnl, 2),
                    "return_pct": round(ret, 1),
                }

            # --- Fixed holding periods ---
            for hold_days in HOLDING_PERIODS:
                exit_target = (entry_dt + timedelta(days=hold_days)).isoformat()

                # Don't exit past expiration
                if datetime.strptime(exit_target, "%Y-%m-%d").date() > exp_dt:
                    continue

                # Find closest bar on or after the target exit date
                exit_price = None
                exit_bar_date = None
                for b in bars:
                    if b["bar_date"] >= exit_target:
                        exit_price = b["close"]
                        exit_bar_date = b["bar_date"]
                        break

                if exit_price and exit_price > 0:
                    pnl = (exit_price - entry_price) * 100
                    ret = (exit_price / entry_price - 1) * 100
                    position["exits"][f"{hold_days}d"] = {
                        "exit_date": exit_bar_date,
                        "exit_price": round(exit_price, 2),
                        "pnl": round(pnl, 2),
                        "return_pct": round(ret, 1),
                    }

            # --- 14-day hold with 50% stop-loss ---
            sl_max_hold = 14
            sl_exit_target_dt = entry_dt + timedelta(days=sl_max_hold)
            if sl_exit_target_dt > exp_dt:
                sl_exit_target_dt = exp_dt

            sl_exit_price = None
            sl_exit_date = None
            sl_triggered = False

            for b in bars:
                bar_dt = datetime.strptime(b["bar_date"], "%Y-%m-%d").date()
                if bar_dt > sl_exit_target_dt:
                    break

                bar_close = b["close"]
                if not bar_close or bar_close <= 0:
                    continue

                # Check if stop-loss triggered (75% loss from entry)
                pct_change = (bar_close / entry_price - 1) * 100
                if pct_change <= -75.0:
                    sl_exit_price = bar_close
                    sl_exit_date = b["bar_date"]
                    sl_triggered = True
                    break

                # Track for end-of-period exit
                sl_exit_price = bar_close
                sl_exit_date = b["bar_date"]

            if sl_exit_price and sl_exit_price > 0:
                sl_pnl = (sl_exit_price - entry_price) * 100
                sl_ret = (sl_exit_price / entry_price - 1) * 100
                sl_hold_days = (datetime.strptime(sl_exit_date, "%Y-%m-%d").date() - entry_dt).days
                position["exits"]["14d_sl50"] = {
                    "exit_date": sl_exit_date,
                    "exit_price": round(sl_exit_price, 2),
                    "pnl": round(sl_pnl, 2),
                    "return_pct": round(sl_ret, 1),
                    "stop_loss_triggered": sl_triggered,
                    "hold_days": sl_hold_days,
                }

            # --- Max gain during life ---
            max_price = max(b["close"] for b in bars if b["close"])
            if max_price and max_price > 0:
                max_ret = (max_price / entry_price - 1) * 100
                position["max_price"] = round(max_price, 2)
                position["max_return_pct"] = round(max_ret, 1)

            result["positions"].append(position)

        return result


# ---------------------------------------------------------------------------
# Display and reporting
# ---------------------------------------------------------------------------

def print_results(results: List[Dict], title: str):
    """Print formatted results table."""
    print(f"\n{'=' * 100}")
    print(f" {title}")
    print(f"{'=' * 100}")

    for dte_cfg in DTE_TARGETS:
        label = dte_cfg["label"]
        print(f"\n--- {label.upper()} ATM Calls ---")
        print(f"{'Symbol':<7} {'Pick Date':<12} {'Stock$':>8} {'Strike':>8} "
              f"{'Entry$':>8} {'Cost':>8} "
              f"{'14d%':>7} {'SL50%':>7} {'30d%':>7} {'Exp%':>7} {'Max%':>7}")
        print("-" * 110)

        returns_14d = []
        returns_30d = []
        returns_exp = []
        returns_sl50 = []
        sl_triggered_count = 0
        valid_count = 0

        for r in results:
            pos = None
            for p in r.get("positions", []):
                if p.get("dte_label") == label and "error" not in p:
                    pos = p
                    break

            if not pos:
                continue

            valid_count += 1
            exits = pos.get("exits", {})
            r14 = exits.get("14d", {}).get("return_pct", "")
            rsl = exits.get("14d_sl50", {}).get("return_pct", "")
            r30 = exits.get("30d", {}).get("return_pct", "")
            rexp = exits.get("expiry", {}).get("return_pct", "")
            rmax = pos.get("max_return_pct", "")

            if isinstance(r14, (int, float)):
                returns_14d.append(r14)
            if isinstance(rsl, (int, float)):
                returns_sl50.append(rsl)
                if exits.get("14d_sl50", {}).get("stop_loss_triggered"):
                    sl_triggered_count += 1
            if isinstance(r30, (int, float)):
                returns_30d.append(r30)
            if isinstance(rexp, (int, float)):
                returns_exp.append(rexp)

            r14_str = f"{r14:>6.1f}%" if isinstance(r14, (int, float)) else f"{'N/A':>7}"
            rsl_str = f"{rsl:>6.1f}%" if isinstance(rsl, (int, float)) else f"{'N/A':>7}"
            if isinstance(rsl, (int, float)) and exits.get("14d_sl50", {}).get("stop_loss_triggered"):
                rsl_str = f"{rsl:>5.1f}%!"
            r30_str = f"{r30:>6.1f}%" if isinstance(r30, (int, float)) else f"{'N/A':>7}"
            rexp_str = f"{rexp:>6.1f}%" if isinstance(rexp, (int, float)) else f"{'N/A':>7}"
            rmax_str = f"{rmax:>6.1f}%" if isinstance(rmax, (int, float)) else f"{'N/A':>7}"

            print(f"{r['symbol']:<7} {r['pick_date']:<12} "
                  f"{r['stock_price']:>8.2f} {pos['strike']:>8.2f} "
                  f"{pos['entry_price']:>8.2f} {pos['cost_per_contract']:>8.0f} "
                  f"{r14_str} {rsl_str} {r30_str} {rexp_str} {rmax_str}")

        # Summary stats
        if valid_count > 0:
            print(f"\n  Positions with data: {valid_count}")

            for name, rets in [("14-day hold", returns_14d),
                               ("14d + 50% SL", returns_sl50),
                               ("30-day hold", returns_30d),
                               ("Hold to expiry", returns_exp)]:
                if rets:
                    avg = sum(rets) / len(rets)
                    wins = sum(1 for x in rets if x > 0)
                    median = sorted(rets)[len(rets) // 2]
                    print(f"  {name:18s}: Avg {avg:>7.1f}%  Median {median:>7.1f}%  "
                          f"Win Rate {wins}/{len(rets)} ({wins/len(rets)*100:.0f}%)")

            if returns_sl50:
                print(f"  Stop-losses hit:     {sl_triggered_count}/{len(returns_sl50)} "
                      f"({sl_triggered_count/len(returns_sl50)*100:.0f}%)")


def print_iv_analysis(results: List[Dict]):
    """Analyze whether entry premium (IV proxy) correlates with trade returns."""
    print(f"\n{'=' * 100}")
    print(" IV / PREMIUM ANALYSIS — Does Option Cost Affect Returns?")
    print(f"{'=' * 100}")

    # Collect trades with premium% and 14d return
    for dte_cfg in DTE_TARGETS:
        label = dte_cfg["label"]
        trades = []

        for r in results:
            stock_price = r.get("stock_price")
            if not stock_price or stock_price <= 0:
                continue

            for p in r.get("positions", []):
                if p.get("dte_label") != label or "error" in p:
                    continue

                entry_price = p.get("entry_price")
                if not entry_price or entry_price <= 0:
                    continue

                premium_pct = entry_price / stock_price * 100
                dte = p.get("dte", 30)
                # Annualized IV proxy
                iv_proxy = premium_pct * (365 / max(dte, 1)) ** 0.5

                exits = p.get("exits", {})
                ret_14d = exits.get("14d", {}).get("return_pct")
                ret_exp = exits.get("expiry", {}).get("return_pct")

                if ret_14d is not None:
                    trades.append({
                        "symbol": r["symbol"],
                        "pick_date": r["pick_date"],
                        "stock_price": stock_price,
                        "entry_price": entry_price,
                        "premium_pct": round(premium_pct, 2),
                        "iv_proxy": round(iv_proxy, 1),
                        "return_14d": ret_14d,
                        "return_exp": ret_exp,
                    })

        if len(trades) < 10:
            continue

        print(f"\n--- {label.upper()} ATM Calls — Premium vs 14-Day Return ({len(trades)} trades) ---")

        # Sort by premium% and split into terciles
        trades.sort(key=lambda x: x["premium_pct"])
        n = len(trades)
        tercile_size = n // 3

        bins = [
            ("Low Premium", trades[:tercile_size]),
            ("Mid Premium", trades[tercile_size:2*tercile_size]),
            ("High Premium", trades[2*tercile_size:]),
        ]

        print(f"\n  {'Bin':<16} {'Count':>5} {'Prem% Range':>14} "
              f"{'Avg Ret':>8} {'Median':>8} {'Win%':>6} {'Avg IV':>8}")
        print(f"  {'-' * 68}")

        for bin_name, bin_trades in bins:
            if not bin_trades:
                continue
            prems = [t["premium_pct"] for t in bin_trades]
            rets = [t["return_14d"] for t in bin_trades]
            ivs = [t["iv_proxy"] for t in bin_trades]
            avg_ret = sum(rets) / len(rets)
            median_ret = sorted(rets)[len(rets) // 2]
            wins = sum(1 for r in rets if r > 0)
            win_pct = wins / len(rets) * 100
            avg_iv = sum(ivs) / len(ivs)

            print(f"  {bin_name:<16} {len(bin_trades):>5} "
                  f"{min(prems):>5.1f}%-{max(prems):>4.1f}% "
                  f"{avg_ret:>7.1f}% {median_ret:>7.1f}% {win_pct:>5.0f}% "
                  f"~{avg_iv:>4.0f}%ann")

        # Also show individual trades sorted by premium
        print(f"\n  Individual trades (sorted by premium %):")
        print(f"  {'Symbol':<7} {'Pick Date':<12} {'Stock$':>8} {'OptPx':>7} "
              f"{'Prem%':>6} {'~IV':>6} {'14dRet':>7}")
        print(f"  {'-' * 60}")

        for t in trades:
            ret_str = f"{t['return_14d']:>6.1f}%"
            print(f"  {t['symbol']:<7} {t['pick_date']:<12} "
                  f"{t['stock_price']:>8.2f} {t['entry_price']:>7.2f} "
                  f"{t['premium_pct']:>5.1f}% ~{t['iv_proxy']:>4.0f}% {ret_str}")

        # Correlation summary
        prems_all = [t["premium_pct"] for t in trades]
        rets_all = [t["return_14d"] for t in trades]

        # Simple correlation coefficient
        n_t = len(trades)
        mean_p = sum(prems_all) / n_t
        mean_r = sum(rets_all) / n_t
        cov = sum((p - mean_p) * (r - mean_r) for p, r in zip(prems_all, rets_all)) / n_t
        std_p = (sum((p - mean_p)**2 for p in prems_all) / n_t) ** 0.5
        std_r = (sum((r - mean_r)**2 for r in rets_all) / n_t) ** 0.5

        if std_p > 0 and std_r > 0:
            corr = cov / (std_p * std_r)
            print(f"\n  Correlation (Premium% vs 14d Return): {corr:+.3f}")
            if abs(corr) < 0.1:
                print(f"  >> Very weak correlation -- premium level barely matters")
            elif corr < -0.3:
                print(f"  >> Moderate negative -- lower premium tends to produce better returns")
            elif corr < -0.1:
                print(f"  >> Slight negative -- lower premium slightly favors returns")
            elif corr > 0.3:
                print(f"  >> Moderate positive -- higher premium stocks actually do better")
            elif corr > 0.1:
                print(f"  >> Slight positive -- higher premium stocks slightly better")


def print_summary(results: List[Dict]):
    """Print overall summary across all DTE targets."""
    print(f"\n{'=' * 100}")
    print(" OVERALL SUMMARY")
    print(f"{'=' * 100}")

    total = len(results)
    no_data = sum(1 for r in results if r.get("error"))
    with_data = total - no_data

    print(f"\n  Total picks analyzed: {total}")
    print(f"  With options data:    {with_data}")
    print(f"  No data available:    {no_data}")

    if no_data > 0:
        no_data_symbols = [r["symbol"] for r in results if r.get("error")]
        print(f"  Symbols without data: {', '.join(no_data_symbols[:20])}"
              + ("..." if no_data > 20 else ""))

    # Best and worst per DTE
    for dte_cfg in DTE_TARGETS:
        label = dte_cfg["label"]
        expiry_returns = []
        for r in results:
            for p in r.get("positions", []):
                if p.get("dte_label") == label and "expiry" in p.get("exits", {}):
                    expiry_returns.append(
                        (r["symbol"], r["pick_date"],
                         p["exits"]["expiry"]["return_pct"])
                    )

        if expiry_returns:
            expiry_returns.sort(key=lambda x: x[2], reverse=True)
            print(f"\n  {label} Calls — Hold to Expiry:")
            print(f"    Best:  {expiry_returns[0][0]:<7} ({expiry_returns[0][1]}) "
                  f"+{expiry_returns[0][2]:.1f}%")
            print(f"    Worst: {expiry_returns[-1][0]:<7} ({expiry_returns[-1][1]}) "
                  f"{expiry_returns[-1][2]:.1f}%")


def save_results_csv(results: List[Dict], filepath: str):
    """Save results to a CSV file."""
    import csv

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Portfolio", "Symbol", "Pick Date", "Entry Date", "Stock Price",
            "DTE Label", "Strike", "DTE", "Entry Price", "Cost Per Contract",
            "14d Return%", "14d+SL50 Return%", "SL Triggered", "SL Hold Days",
            "30d Return%", "Expiry Return%", "Max Return%",
            "Expiry PnL", "Stock at Expiry",
        ])

        for r in results:
            for p in r.get("positions", []):
                if "error" in p:
                    continue
                exits = p.get("exits", {})
                sl_exit = exits.get("14d_sl50", {})
                writer.writerow([
                    r.get("portfolio", ""),
                    r["symbol"],
                    r["pick_date"],
                    r.get("entry_date", ""),
                    r.get("stock_price", ""),
                    p.get("dte_label", ""),
                    p.get("strike", ""),
                    p.get("dte", ""),
                    p.get("entry_price", ""),
                    p.get("cost_per_contract", ""),
                    exits.get("14d", {}).get("return_pct", ""),
                    sl_exit.get("return_pct", ""),
                    sl_exit.get("stop_loss_triggered", ""),
                    sl_exit.get("hold_days", ""),
                    exits.get("30d", {}).get("return_pct", ""),
                    exits.get("expiry", {}).get("return_pct", ""),
                    p.get("max_return_pct", ""),
                    exits.get("expiry", {}).get("pnl", ""),
                    exits.get("expiry", {}).get("stock_at_expiry", ""),
                ])

    print(f"\nResults saved to {filepath}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Options analysis on SA picks")
    parser.add_argument("--portfolio", default="alpha",
                        choices=["alpha", "proquant", "both"],
                        help="Which portfolio to analyze")
    parser.add_argument("--symbols", default=None,
                        help="Comma-separated symbols to filter")
    parser.add_argument("--report", action="store_true",
                        help="Save results to CSV")
    parser.add_argument("--limit", type=int, default=None,
                        help="Limit number of picks to analyze (for testing)")
    parser.add_argument("--otm", type=float, default=0.0,
                        help="OTM percentage (0=ATM, 5=5%% OTM, 10=10%% OTM)")
    parser.add_argument("--dte-only", default=None,
                        help="Only analyze one DTE target (e.g. '30d')")
    args = parser.parse_args()

    otm_label = f"{args.otm:.0f}% OTM" if args.otm > 0 else "ATM"
    dte_label = f" ({args.dte_only} only)" if args.dte_only else ""

    print(f"\n{'=' * 100}")
    print(f" OPTIONS ON SA PICKS — {otm_label} CALLS{dte_label}")
    print(f"{'=' * 100}")

    # Load picks
    all_picks = []

    if args.portfolio in ("alpha", "both"):
        alpha = load_alpha_picks()
        print(f"\n  Alpha Picks loaded: {len(alpha)}")
        all_picks.extend(alpha)

    if args.portfolio in ("proquant", "both"):
        proquant = load_proquant_picks()
        print(f"  ProQuant picks loaded: {len(proquant)}")
        all_picks.extend(proquant)

    # Filter by symbols if specified
    if args.symbols:
        filter_syms = {s.strip().upper() for s in args.symbols.split(",")}
        all_picks = [p for p in all_picks if p["symbol"] in filter_syms]
        print(f"  Filtered to {len(all_picks)} picks for: {', '.join(filter_syms)}")

    if args.limit:
        all_picks = all_picks[:args.limit]
        print(f"  Limited to first {args.limit} picks")

    if not all_picks:
        print("  No picks to analyze.")
        return

    # Analyze
    analyzer = OptionsAnalyzer()
    results = []

    print(f"\n  Analyzing {len(all_picks)} picks...\n")
    for i, pick in enumerate(all_picks, 1):
        sym = pick["symbol"]
        dt = pick["pick_date"]
        port = pick.get("portfolio", "")
        print(f"  [{i}/{len(all_picks)}] {sym:<7} {dt}  ({port})", end="", flush=True)

        result = analyzer.analyze_pick(pick, otm_pct=args.otm,
                                       dte_filter=args.dte_only)
        results.append(result)

        # Show quick status
        positions_ok = sum(1 for p in result.get("positions", []) if "error" not in p)
        if result.get("error"):
            print(f"  — {result['error']}")
        elif positions_ok == 0:
            print(f"  — no options data")
        else:
            print(f"  — {positions_ok} DTE windows")

    # Display results
    for portfolio in ("AlphaPicks", "ProQuant"):
        port_results = [r for r in results if r.get("portfolio") == portfolio]
        if port_results:
            print_results(port_results, f"{portfolio} — ATM Call Options Analysis")

    print_summary(results)

    # IV analysis (only for Alpha Picks with enough data)
    alpha_results = [r for r in results if r.get("portfolio") == "AlphaPicks"]
    if alpha_results and len(alpha_results) >= 10:
        print_iv_analysis(alpha_results)

    # Save report
    if args.report:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = os.path.join(BASE_DIR, f"options_picks_results_{ts}.csv")
        save_results_csv(results, csv_path)

    print(f"\nDone. Cache: {CACHE_DB}")


if __name__ == "__main__":
    main()
