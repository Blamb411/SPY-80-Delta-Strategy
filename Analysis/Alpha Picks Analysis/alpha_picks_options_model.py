#!/usr/bin/env python3
"""
Alpha Picks Options P&L Model (Black-Scholes Synthetic Pricing)
================================================================
Models the P&L of buying call options on Alpha Picks stocks the day
before announcement and holding for various periods.

Uses synthetic Black-Scholes pricing with estimated IV rather than
historical options data, giving consistent coverage across all picks.

Includes rule-based exit analysis: stop-loss, profit target, max hold.

Usage:
    python alpha_picks_options_model.py
"""

import os
import sys
import math
import sqlite3
from datetime import datetime
from collections import defaultdict

import openpyxl

# Add parent dir so we can import backtest.black_scholes
_this_dir = os.path.dirname(os.path.abspath(__file__))
_project_dir = os.path.dirname(_this_dir)
sys.path.insert(0, _project_dir)

from backtest.black_scholes import black_scholes_price, find_strike_for_delta

XLSX_PATH = os.path.join(_this_dir, "ProQuant History 1_29_2026.xlsx")
DB_PATH = os.path.join(_this_dir, "price_cache.db")

WINDOWS = [1, 10, 30, 60]          # trading days after announcement
DELTAS = [0.80, 0.50]              # call deltas to test
DTE_DAYS = 60                      # days-to-expiration at entry
RATE = 0.05                        # risk-free rate
IV_LOOKBACK = 30                   # trailing trading days for realized vol
IV_PREMIUM = 1.2                   # IV / realized vol ratio
IV_FLOOR = 0.15
IV_CAP = 1.00

# Rule-based exit parameters
STOP_LOSSES = [None, -0.30, -0.50, -0.70]
PROFIT_TARGETS = [None, 0.50, 1.00, 2.00]
MAX_HOLDS = [10, 30, 60]
MAX_SIM_DAYS = 60                   # max trading days to simulate daily prices


# =====================================================================
# Data loading (mirrors announcement_momentum.py)
# =====================================================================

def load_alpha_picks():
    """Load Alpha Picks from Excel."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["AlphaPicks"]
    picks = []
    seen = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] is None or row[2] is None:
            continue
        symbol = str(row[1]).strip()
        pick_date = row[2]
        if isinstance(pick_date, datetime):
            date_str = pick_date.strftime("%Y-%m-%d")
        else:
            date_str = str(pick_date)[:10]
        buy_price = float(row[3]) if row[3] else None
        key = (symbol, date_str)
        if key in seen:
            continue
        seen.add(key)
        picks.append({
            "symbol": symbol,
            "pick_date": date_str,
            "buy_price": buy_price,
        })
    return picks


def load_prices_from_cache():
    """Load all prices from SQLite cache."""
    prices = defaultdict(dict)
    if not os.path.exists(DB_PATH):
        return prices
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT symbol, date, close FROM daily_prices ORDER BY symbol, date")
    for symbol, date, close in c.fetchall():
        prices[symbol][date] = close
    conn.close()
    return prices


def get_trading_dates(prices):
    """Get sorted list of all trading dates from SPY."""
    if "SPY" in prices:
        return sorted(prices["SPY"].keys())
    all_dates = set()
    for sym_prices in prices.values():
        all_dates.update(sym_prices.keys())
    return sorted(all_dates)


def find_trading_date(trading_dates, target, direction="on_or_before"):
    """Find the nearest trading date to target."""
    if target in trading_dates:
        return target
    for i, d in enumerate(trading_dates):
        if d > target:
            if direction == "on_or_before":
                return trading_dates[i - 1] if i > 0 else None
            else:
                return d
    if direction == "on_or_before":
        return trading_dates[-1] if trading_dates else None
    return None


def offset_trading_days(trading_dates, base_date, offset):
    """Move N trading days from base_date."""
    if base_date not in trading_dates:
        base_date = find_trading_date(trading_dates, base_date, "on_or_after")
        if not base_date:
            return None
    idx = trading_dates.index(base_date)
    target_idx = idx + offset
    if 0 <= target_idx < len(trading_dates):
        return trading_dates[target_idx]
    return None


# =====================================================================
# IV estimation
# =====================================================================

def estimate_iv(prices_dict, symbol, date_str, trading_dates, lookback=IV_LOOKBACK):
    """
    Estimate implied volatility from trailing realized volatility.

    realized_vol = std(daily log returns) * sqrt(252)
    estimated_iv = realized_vol * IV_PREMIUM

    Returns IV as a decimal (e.g. 0.30 for 30%), or None if insufficient data.
    """
    if symbol not in prices_dict:
        return None

    sym_prices = prices_dict[symbol]

    if date_str not in trading_dates:
        date_str = find_trading_date(trading_dates, date_str, "on_or_before")
        if not date_str:
            return None

    idx = trading_dates.index(date_str)
    start_idx = idx - lookback
    if start_idx < 0:
        return None

    closes = []
    for i in range(start_idx, idx + 1):
        d = trading_dates[i]
        if d in sym_prices:
            closes.append(sym_prices[d])

    if len(closes) < lookback // 2:
        return None

    log_returns = []
    for i in range(1, len(closes)):
        if closes[i] > 0 and closes[i - 1] > 0:
            log_returns.append(math.log(closes[i] / closes[i - 1]))

    if len(log_returns) < 5:
        return None

    mean_ret = sum(log_returns) / len(log_returns)
    variance = sum((r - mean_ret) ** 2 for r in log_returns) / (len(log_returns) - 1)
    realized_vol = math.sqrt(variance * 252)

    iv = realized_vol * IV_PREMIUM
    iv = max(IV_FLOOR, min(IV_CAP, iv))
    return iv


# =====================================================================
# Options pricing helpers
# =====================================================================

def price_call_at_delta(spot, iv, dte_days, target_delta, rate=RATE):
    """
    Find the call strike at target_delta and price it.
    Returns (strike, call_price) or (None, None).
    """
    t_years = dte_days / 365.0
    strike = find_strike_for_delta(spot, t_years, rate, iv, target_delta, 'C')
    if strike is None:
        return None, None
    price = black_scholes_price(spot, strike, t_years, rate, iv, 'C')
    if price is None or price <= 0:
        return None, None
    return strike, price


def reprice_call(spot, strike, remaining_dte_days, iv, rate=RATE):
    """
    Re-price an existing call at a later date.
    Returns call price or None.
    """
    t_years = remaining_dte_days / 365.0
    if t_years <= 0:
        return max(0.0, spot - strike)
    return black_scholes_price(spot, strike, t_years, rate, iv, 'C')


# =====================================================================
# Per-pick modelling
# =====================================================================

def model_pick_options(pick, prices, trading_dates, deltas=DELTAS, windows=WINDOWS,
                       entry_offset=-1):
    """
    Model call option P&L for one Alpha Pick.

    Args:
        entry_offset: -1 = enter at T-1 close (day before announcement)
                       0 = enter at T0 close (announcement day close)

    Returns a dict with results for each delta and window, or None if data missing.
    """
    sym = pick["symbol"]
    pick_date = pick["pick_date"]

    if sym not in prices or not prices[sym]:
        return None

    # T-1: trading day before announcement
    t_minus1 = find_trading_date(trading_dates, pick_date, "on_or_before")
    if t_minus1 == pick_date:
        idx = trading_dates.index(t_minus1)
        t_minus1 = trading_dates[idx - 1] if idx > 0 else None
    if not t_minus1:
        return None

    # T0: announcement day
    t0 = find_trading_date(trading_dates, pick_date, "on_or_after")
    if not t0:
        return None

    # Entry date depends on entry_offset
    if entry_offset == 0:
        entry_date = t0
    else:
        entry_date = t_minus1

    spot_entry = prices[sym].get(entry_date)
    if not spot_entry:
        return None

    spy_entry = prices.get("SPY", {}).get(entry_date)

    iv = estimate_iv(prices, sym, entry_date, trading_dates)
    if iv is None:
        return None

    result = {
        "symbol": sym,
        "pick_date": pick_date,
        "t_minus1": t_minus1,
        "t0": t0,
        "entry_date": entry_date,
        "entry_offset": entry_offset,
        "spot_entry": spot_entry,
        "iv": iv,
        "spy_entry": spy_entry,
    }

    for delta in deltas:
        dk = f"d{int(delta * 100)}"

        strike, entry_price = price_call_at_delta(spot_entry, iv, DTE_DAYS, delta)
        if strike is None:
            result[f"{dk}_strike"] = None
            result[f"{dk}_entry"] = None
            continue

        result[f"{dk}_strike"] = strike
        result[f"{dk}_entry"] = entry_price

        for w in windows:
            t_exit = offset_trading_days(trading_dates, t0, w)
            if not t_exit:
                result[f"{dk}_pnl_{w}d"] = None
                result[f"{dk}_ret_{w}d"] = None
                result[f"{dk}_stock_ret_{w}d"] = None
                result[f"{dk}_spy_ret_{w}d"] = None
                continue

            spot_exit = prices[sym].get(t_exit)
            if not spot_exit:
                result[f"{dk}_pnl_{w}d"] = None
                result[f"{dk}_ret_{w}d"] = None
                result[f"{dk}_stock_ret_{w}d"] = None
                result[f"{dk}_spy_ret_{w}d"] = None
                continue

            calendar_days_elapsed = w * 365.0 / 252.0
            remaining_dte = max(0, DTE_DAYS - calendar_days_elapsed)

            exit_price = reprice_call(spot_exit, strike, remaining_dte, iv)
            if exit_price is None:
                result[f"{dk}_pnl_{w}d"] = None
                result[f"{dk}_ret_{w}d"] = None
            else:
                pnl_per_contract = (exit_price - entry_price) * 100
                ret_pct = (exit_price / entry_price - 1) if entry_price > 0 else None
                result[f"{dk}_pnl_{w}d"] = pnl_per_contract
                result[f"{dk}_ret_{w}d"] = ret_pct

            result[f"{dk}_stock_ret_{w}d"] = (spot_exit / spot_entry - 1)

            spy_exit = prices.get("SPY", {}).get(t_exit)
            if spy_entry and spy_exit:
                result[f"{dk}_spy_ret_{w}d"] = (spy_exit / spy_entry - 1)
            else:
                result[f"{dk}_spy_ret_{w}d"] = None

    return result


# =====================================================================
# Daily option price series (for rule-based exits)
# =====================================================================

def compute_daily_option_prices(result, prices, trading_dates):
    """
    Compute option price at each trading day from T0+1 through T0+MAX_SIM_DAYS.
    Adds '{dk}_daily' list to the result dict for each delta key.
    """
    sym = result["symbol"]
    t0 = result["t0"]
    iv = result["iv"]
    spot_entry = result["spot_entry"]

    for dk in ["d80", "d50"]:
        strike = result.get(f"{dk}_strike")
        entry_price = result.get(f"{dk}_entry")
        if strike is None or entry_price is None:
            result[f"{dk}_daily"] = None
            continue

        daily = []
        for day in range(1, MAX_SIM_DAYS + 1):
            t_date = offset_trading_days(trading_dates, t0, day)
            if not t_date:
                break

            spot = prices[sym].get(t_date)
            if spot is None:
                continue  # skip missing days, don't break

            calendar_elapsed = day * 365.0 / 252.0
            remaining_dte = max(0, DTE_DAYS - calendar_elapsed)

            opt_price = reprice_call(spot, strike, remaining_dte, iv)
            if opt_price is None:
                continue

            opt_return = (opt_price / entry_price - 1) if entry_price > 0 else 0
            stock_return = (spot / spot_entry - 1) if spot_entry > 0 else 0

            daily.append({
                "day": day,
                "date": t_date,
                "spot": spot,
                "opt_price": opt_price,
                "opt_return": opt_return,
                "stock_return": stock_return,
            })

        result[f"{dk}_daily"] = daily


def apply_rules(daily_series, stop_loss, profit_target, max_hold):
    """
    Simulate a trade with stop-loss, profit-target, and max-hold rules.

    Args:
        daily_series: list of daily price dicts from compute_daily_option_prices
        stop_loss: negative decimal (e.g. -0.30) or None
        profit_target: positive decimal (e.g. 1.00) or None
        max_hold: max trading days to hold

    Returns dict with exit info, or None if no data.
    """
    if not daily_series:
        return None

    last_valid = None

    for entry in daily_series:
        day = entry["day"]
        ret = entry["opt_return"]

        if day > max_hold:
            break  # past max hold window, exit at last_valid

        last_valid = entry

        # Stop-loss check
        if stop_loss is not None and ret <= stop_loss:
            return {
                "exit_day": day,
                "exit_date": entry["date"],
                "exit_reason": "stop_loss",
                "return_pct": ret,
                "exit_price": entry["opt_price"],
                "spot_exit": entry["spot"],
                "stock_return": entry["stock_return"],
            }

        # Profit-target check
        if profit_target is not None and ret >= profit_target:
            return {
                "exit_day": day,
                "exit_date": entry["date"],
                "exit_reason": "profit_target",
                "return_pct": ret,
                "exit_price": entry["opt_price"],
                "spot_exit": entry["spot"],
                "stock_return": entry["stock_return"],
            }

    # Exited loop: max_hold reached or ran out of data
    if last_valid:
        return {
            "exit_day": last_valid["day"],
            "exit_date": last_valid["date"],
            "exit_reason": "max_hold",
            "return_pct": last_valid["opt_return"],
            "exit_price": last_valid["opt_price"],
            "spot_exit": last_valid["spot"],
            "stock_return": last_valid["stock_return"],
        }

    return None


# =====================================================================
# Reporting — original fixed-window analysis
# =====================================================================

def print_per_pick_detail(results, delta_key="d80"):
    """Section 1: Per-pick detail table."""
    print()
    print("=" * 160)
    label = "80-Delta ITM" if delta_key == "d80" else "50-Delta ATM"
    print(f"PER-PICK DETAIL — {label} Calls, {DTE_DAYS} DTE")
    print("=" * 160)

    header = (f"{'Symbol':<8} | {'Pick Date':<12} | {'Spot':>8} | {'IV':>5}"
              f" | {'Strike':>8} | {'Call$':>7}")
    for w in WINDOWS:
        header += f" | {'T+' + str(w) + ' Ret':>9}"
    print(header)
    print("-" * 160)

    for r in sorted(results, key=lambda x: x["pick_date"]):
        strike = r.get(f"{delta_key}_strike")
        entry = r.get(f"{delta_key}_entry")
        if strike is None or entry is None:
            continue

        line = (f"{r['symbol']:<8} | {r['pick_date']:<12}"
                f" | {r['spot_entry']:>8.2f} | {r['iv']:>4.0%}"
                f" | {strike:>8.2f} | {entry:>7.2f}")

        for w in WINDOWS:
            ret = r.get(f"{delta_key}_ret_{w}d")
            if ret is not None:
                line += f" | {ret:>+8.1%}"
            else:
                line += f" |      ---"
        print(line)


def compute_stats(values):
    """Compute mean, median, win rate, min, max for a list of values."""
    if not values:
        return None
    n = len(values)
    mean = sum(values) / n
    sorted_v = sorted(values)
    median = sorted_v[n // 2]
    win_rate = sum(1 for v in values if v > 0) / n
    return {
        "n": n,
        "mean": mean,
        "median": median,
        "win_rate": win_rate,
        "min": min(values),
        "max": max(values),
    }


def print_summary(results):
    """Section 2: Summary statistics for both deltas."""
    print()
    print("=" * 140)
    print("SUMMARY STATISTICS (Fixed-Window, No Trading Rules)")
    print("=" * 140)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        print(f"\n--- {label} Calls ---")
        print(f"{'Window':<12} | {'N':>4} | {'Mean':>8} | {'Median':>8}"
              f" | {'Win%':>6} | {'Min':>8} | {'Max':>8}"
              f" | {'Stock':>8} | {'SPY':>8} | {'Alpha':>8}")
        print("-" * 120)

        for w in WINDOWS:
            rets = [r[f"{dk}_ret_{w}d"] for r in results
                    if r.get(f"{dk}_ret_{w}d") is not None]
            stock_rets = [r[f"{dk}_stock_ret_{w}d"] for r in results
                         if r.get(f"{dk}_stock_ret_{w}d") is not None
                         and r.get(f"{dk}_ret_{w}d") is not None]
            spy_rets = [r[f"{dk}_spy_ret_{w}d"] for r in results
                        if r.get(f"{dk}_spy_ret_{w}d") is not None
                        and r.get(f"{dk}_ret_{w}d") is not None]

            s = compute_stats(rets)
            if not s:
                continue

            stock_mean = sum(stock_rets) / len(stock_rets) if stock_rets else 0
            spy_mean = sum(spy_rets) / len(spy_rets) if spy_rets else 0
            alpha = stock_mean - spy_mean

            print(f"T+{w:<9} | {s['n']:>4} | {s['mean']:>+7.2%} | {s['median']:>+7.2%}"
                  f" | {s['win_rate']:>5.1%} | {s['min']:>+7.1%} | {s['max']:>+7.1%}"
                  f" | {stock_mean:>+7.2%} | {spy_mean:>+7.2%} | {alpha:>+7.2%}")


def print_comparison(results):
    """Section 3: Side-by-side delta comparison."""
    print()
    print("=" * 120)
    print("DELTA COMPARISON: 80-Delta ITM vs 50-Delta ATM (Fixed-Window)")
    print("=" * 120)

    print(f"{'Window':<12} | {'--- 80-Delta ITM ---':>40} | {'--- 50-Delta ATM ---':>40}")
    print(f"{'':12} | {'N':>4} {'Mean':>8} {'Med':>8} {'Win%':>6}"
          f" | {'N':>4} {'Mean':>8} {'Med':>8} {'Win%':>6}")
    print("-" * 100)

    for w in WINDOWS:
        parts = [f"T+{w:<9}"]
        for dk in ["d80", "d50"]:
            rets = [r[f"{dk}_ret_{w}d"] for r in results
                    if r.get(f"{dk}_ret_{w}d") is not None]
            s = compute_stats(rets)
            if s:
                parts.append(f" | {s['n']:>4} {s['mean']:>+7.2%} {s['median']:>+7.2%} {s['win_rate']:>5.1%}")
            else:
                parts.append(f" |  --- no data ---")
        print("".join(parts))


def print_capital_analysis(results):
    """Section 4: Capital required and total P&L."""
    print()
    print("=" * 120)
    print("CAPITAL ANALYSIS — 1 Contract Per Pick (Fixed-Window)")
    print("=" * 120)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        print(f"\n--- {label} ---")

        entries = [(r[f"{dk}_entry"], r) for r in results
                   if r.get(f"{dk}_entry") is not None]
        if not entries:
            print("  No valid entries.")
            continue

        total_capital = sum(e * 100 for e, _ in entries)
        n_picks = len(entries)
        avg_cost = total_capital / n_picks

        print(f"  Picks with valid pricing: {n_picks}")
        print(f"  Total capital (1 contract each): ${total_capital:,.0f}")
        print(f"  Avg cost per contract: ${avg_cost:,.0f}")
        print()

        print(f"  {'Window':<12} | {'Total P&L':>12} | {'Total Ret':>10} | {'Avg P&L':>10}")
        print(f"  {'-'*55}")

        for w in WINDOWS:
            pnls = [r[f"{dk}_pnl_{w}d"] for r in results
                    if r.get(f"{dk}_pnl_{w}d") is not None]
            if not pnls:
                continue
            total_pnl = sum(pnls)
            total_ret = total_pnl / total_capital if total_capital > 0 else 0
            avg_pnl = total_pnl / len(pnls)
            print(f"  T+{w:<9} | ${total_pnl:>+11,.0f} | {total_ret:>+9.2%} | ${avg_pnl:>+9,.0f}")


def print_best_worst(results, delta_key="d80", window=30):
    """Section 5: Top 5 and bottom 5 picks by options return."""
    print()
    print("=" * 120)
    label = "80-Delta ITM" if delta_key == "d80" else "50-Delta ATM"
    print(f"TOP 5 AND BOTTOM 5 — {label} Calls at T+{window}")
    print("=" * 120)

    ret_key = f"{delta_key}_ret_{window}d"
    valid = [r for r in results if r.get(ret_key) is not None]
    if not valid:
        print("  No data for this window.")
        return

    ranked = sorted(valid, key=lambda x: x[ret_key], reverse=True)

    header = (f"{'Symbol':<8} | {'Pick Date':<12} | {'Spot':>8} | {'IV':>5}"
              f" | {'Call$':>7} | {'Opt Ret':>9} | {'Stock Ret':>9}")
    for section_label, subset in [("TOP 5", ranked[:5]), ("BOTTOM 5", ranked[-5:])]:
        print(f"\n{section_label}:")
        print(header)
        print("-" * 80)
        for r in subset:
            entry = r.get(f"{delta_key}_entry", 0)
            ret = r.get(ret_key, 0)
            stock_ret = r.get(f"{delta_key}_stock_ret_{window}d", 0) or 0
            print(f"{r['symbol']:<8} | {r['pick_date']:<12}"
                  f" | {r['spot_entry']:>8.2f} | {r['iv']:>4.0%}"
                  f" | {entry:>7.2f} | {ret:>+8.1%} | {stock_ret:>+8.1%}")


# =====================================================================
# Reporting — Rule-based exit analysis
# =====================================================================

def fmt_sl(sl):
    return "None" if sl is None else f"{sl:+.0%}"

def fmt_pt(pt):
    return "None" if pt is None else f"+{pt:.0%}"

def fmt_rule(sl, pt, mh):
    return f"SL={fmt_sl(sl):>5} PT={fmt_pt(pt):>5} MH={mh:>2}d"


def run_rule_analysis(results):
    """
    Run all rule combinations on both deltas.
    Returns: {delta_key: [(rule_params, stats_dict), ...]}
    """
    all_stats = {}

    for dk in ["d80", "d50"]:
        rule_results = []

        for mh in MAX_HOLDS:
            for sl in STOP_LOSSES:
                for pt in PROFIT_TARGETS:
                    trade_returns = []
                    trade_days = []
                    exit_reasons = {"stop_loss": 0, "profit_target": 0, "max_hold": 0}
                    n_trades = 0

                    for r in results:
                        daily = r.get(f"{dk}_daily")
                        if daily is None:
                            continue

                        exit_info = apply_rules(daily, sl, pt, mh)
                        if exit_info is None:
                            continue

                        n_trades += 1
                        trade_returns.append(exit_info["return_pct"])
                        trade_days.append(exit_info["exit_day"])
                        reason = exit_info["exit_reason"]
                        if reason in exit_reasons:
                            exit_reasons[reason] += 1

                    if not trade_returns:
                        continue

                    s = compute_stats(trade_returns)
                    s["avg_days"] = sum(trade_days) / len(trade_days)
                    s["exit_reasons"] = exit_reasons
                    s["pct_sl"] = exit_reasons["stop_loss"] / n_trades if n_trades else 0
                    s["pct_pt"] = exit_reasons["profit_target"] / n_trades if n_trades else 0
                    s["pct_mh"] = exit_reasons["max_hold"] / n_trades if n_trades else 0

                    # Std dev for risk-adjusted metrics
                    mean = s["mean"]
                    var = sum((r - mean) ** 2 for r in trade_returns) / len(trade_returns)
                    s["std"] = math.sqrt(var) if var > 0 else 0.001
                    s["sharpe_like"] = mean / s["std"] if s["std"] > 0 else 0

                    rule_results.append(((sl, pt, mh), s))

        all_stats[dk] = rule_results

    return all_stats


def print_rule_tables(all_stats):
    """Print full rule-based exit analysis tables."""
    print()
    print("=" * 160)
    print("RULE-BASED EXIT ANALYSIS")
    print("=" * 160)
    print("  For each rule combination, the option is checked daily.")
    print("  Exit triggers: stop-loss hit, profit-target hit, or max-hold reached.")
    print()

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        print()
        print("=" * 150)
        print(f"  {label} CALLS — All Rule Combinations")
        print("=" * 150)

        rules = all_stats.get(dk, [])
        if not rules:
            print("  No data.")
            continue

        # Sort by median return descending
        rules_sorted = sorted(rules, key=lambda x: x[1]["median"], reverse=True)

        print(f"  {'Stop-Loss':>10} | {'Profit-Tgt':>10} | {'MaxHold':>7}"
              f" | {'N':>4} | {'Mean':>8} | {'Median':>8} | {'Win%':>6}"
              f" | {'Min':>8} | {'Max':>8} | {'AvgDays':>7}"
              f" | {'%SL':>5} | {'%PT':>5} | {'%MH':>5}"
              f" | {'Sharpe':>6}")
        print(f"  {'-' * 140}")

        for (sl, pt, mh), s in rules_sorted:
            print(f"  {fmt_sl(sl):>10} | {fmt_pt(pt):>10} | {mh:>5}d"
                  f" | {s['n']:>4} | {s['mean']:>+7.2%} | {s['median']:>+7.2%} | {s['win_rate']:>5.1%}"
                  f" | {s['min']:>+7.1%} | {s['max']:>+7.1%} | {s['avg_days']:>6.1f}"
                  f" | {s['pct_sl']:>4.0%} | {s['pct_pt']:>4.0%} | {s['pct_mh']:>4.0%}"
                  f" | {s['sharpe_like']:>5.2f}")


def print_rule_highlights(all_stats):
    """Print the best rule sets by various metrics."""
    print()
    print("=" * 150)
    print("BEST RULE SETS — HIGHLIGHTED")
    print("=" * 150)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        rules = all_stats.get(dk, [])
        if not rules:
            continue

        print(f"\n--- {label} ---")

        # Best by median
        by_median = sorted(rules, key=lambda x: x[1]["median"], reverse=True)
        # Best by win rate
        by_winrate = sorted(rules, key=lambda x: x[1]["win_rate"], reverse=True)
        # Best by Sharpe-like
        by_sharpe = sorted(rules, key=lambda x: x[1]["sharpe_like"], reverse=True)
        # Best by mean
        by_mean = sorted(rules, key=lambda x: x[1]["mean"], reverse=True)

        categories = [
            ("Highest Median Return", by_median),
            ("Highest Win Rate", by_winrate),
            ("Best Risk-Adjusted (Sharpe)", by_sharpe),
            ("Highest Mean Return", by_mean),
        ]

        for cat_label, ranked in categories:
            (sl, pt, mh), s = ranked[0]
            print(f"\n  {cat_label}:")
            print(f"    Rule: SL={fmt_sl(sl)}, PT={fmt_pt(pt)}, MaxHold={mh}d")
            print(f"    N={s['n']}  Mean={s['mean']:+.2%}  Median={s['median']:+.2%}"
                  f"  Win%={s['win_rate']:.1%}  AvgDays={s['avg_days']:.1f}")
            print(f"    Exits: {s['pct_sl']:.0%} stop-loss, {s['pct_pt']:.0%} profit-target,"
                  f" {s['pct_mh']:.0%} max-hold")

        # No-rules baseline (None SL, None PT, max hold 60)
        baseline = None
        for (sl, pt, mh), s in rules:
            if sl is None and pt is None and mh == 60:
                baseline = s
                break

        if baseline:
            best_rule, best_s = by_sharpe[0]
            print(f"\n  === No-Rules Baseline (MH=60d only) vs Best Risk-Adjusted ===")
            print(f"  {'Metric':<15} | {'No Rules':>10} | {'Best':>10} | {'Diff':>10}")
            print(f"  {'-' * 50}")
            metrics = [
                ("Mean", "mean"),
                ("Median", "median"),
                ("Win%", "win_rate"),
                ("AvgDays", "avg_days"),
            ]
            for label, key in metrics:
                b_val = baseline.get(key, 0)
                r_val = best_s.get(key, 0)
                diff = r_val - b_val
                if key in ("mean", "median", "win_rate"):
                    print(f"  {label:<15} | {b_val:>+9.2%} | {r_val:>+9.2%} | {diff:>+9.2%}")
                else:
                    print(f"  {label:<15} | {b_val:>10.1f} | {r_val:>10.1f} | {diff:>+10.1f}")


def print_entry_comparison(results_tm1, results_t0):
    """Compare T-1 entry vs T0 entry side by side."""
    print()
    print("=" * 150)
    print("ENTRY TIMING COMPARISON: T-1 Close (Pre-Announcement) vs T0 Close (Post-Announcement)")
    print("=" * 150)
    print("  T-1 = buy day before announcement (requires prediction or leak)")
    print("  T0  = buy at announcement day close (actionable on public info)")
    print()

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        print(f"\n  --- {label} Calls (Fixed-Window, No Rules) ---")
        print(f"  {'Window':<10} | {'--- T-1 Entry ---':>35} | {'--- T0 Entry ---':>35} | {'Diff':>8}")
        print(f"  {'':10} | {'N':>4} {'Mean':>8} {'Med':>8} {'Win%':>6}"
              f" | {'N':>4} {'Mean':>8} {'Med':>8} {'Win%':>6} | {'Med':>8}")
        print(f"  {'-' * 110}")

        for w in WINDOWS:
            rets_tm1 = [r[f"{dk}_ret_{w}d"] for r in results_tm1
                        if r.get(f"{dk}_ret_{w}d") is not None]
            rets_t0 = [r[f"{dk}_ret_{w}d"] for r in results_t0
                       if r.get(f"{dk}_ret_{w}d") is not None]

            s1 = compute_stats(rets_tm1)
            s0 = compute_stats(rets_t0)

            if s1 and s0:
                med_diff = s0["median"] - s1["median"]
                print(f"  T+{w:<8}"
                      f" | {s1['n']:>4} {s1['mean']:>+7.2%} {s1['median']:>+7.2%} {s1['win_rate']:>5.1%}"
                      f" | {s0['n']:>4} {s0['mean']:>+7.2%} {s0['median']:>+7.2%} {s0['win_rate']:>5.1%}"
                      f" | {med_diff:>+7.2%}")

    # Rule-based comparison: best rules for each entry
    print()
    print(f"\n  --- Best Rule Sets Comparison ---")
    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        print(f"\n  {label}:")
        print(f"  {'Metric':<20} | {'T-1 Best':>35} | {'T0 Best':>35}")


def print_entry_comparison_rules(stats_tm1, stats_t0):
    """Compare best rule sets between T-1 and T0 entry."""
    print()
    print("=" * 150)
    print("BEST RULES COMPARISON: T-1 Entry vs T0 Entry")
    print("=" * 150)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        rules_tm1 = stats_tm1.get(dk, [])
        rules_t0 = stats_t0.get(dk, [])
        if not rules_tm1 or not rules_t0:
            continue

        print(f"\n  --- {label} ---")

        categories = [
            ("Highest Median", "median"),
            ("Highest Win Rate", "win_rate"),
            ("Best Sharpe", "sharpe_like"),
        ]

        for cat_label, sort_key in categories:
            best_tm1 = sorted(rules_tm1, key=lambda x: x[1][sort_key], reverse=True)[0]
            best_t0 = sorted(rules_t0, key=lambda x: x[1][sort_key], reverse=True)[0]

            (sl1, pt1, mh1), s1 = best_tm1
            (sl0, pt0, mh0), s0 = best_t0

            print(f"\n  {cat_label}:")
            print(f"    T-1: SL={fmt_sl(sl1)}, PT={fmt_pt(pt1)}, MH={mh1}d"
                  f"  -> Mean={s1['mean']:+.2%}  Med={s1['median']:+.2%}"
                  f"  Win%={s1['win_rate']:.1%}  Days={s1['avg_days']:.0f}")
            print(f"    T0:  SL={fmt_sl(sl0)}, PT={fmt_pt(pt0)}, MH={mh0}d"
                  f"  -> Mean={s0['mean']:+.2%}  Med={s0['median']:+.2%}"
                  f"  Win%={s0['win_rate']:.1%}  Days={s0['avg_days']:.0f}")


def print_rule_grid(all_stats):
    """Print compact grid view: SL vs PT for each max-hold and delta."""
    print()
    print("=" * 150)
    print("RULE GRID — Median Return by SL x PT (compact view)")
    print("=" * 150)

    for dk, label in [("d80", "80-Delta ITM"), ("d50", "50-Delta ATM")]:
        rules = all_stats.get(dk, [])
        if not rules:
            continue

        # Build lookup
        lookup = {}
        for (sl, pt, mh), s in rules:
            lookup[(sl, pt, mh)] = s

        for mh in MAX_HOLDS:
            print(f"\n  {label} — Max Hold: {mh} days")
            # Header: PT values
            sl_pt_label = "SL \\ PT"
            pt_header = f"  {sl_pt_label:<10}"
            for pt in PROFIT_TARGETS:
                pt_header += f" | {fmt_pt(pt):>10}"
            print(pt_header)
            print(f"  {'-' * (12 + 13 * len(PROFIT_TARGETS))}")

            for sl in STOP_LOSSES:
                row = f"  {fmt_sl(sl):<10}"
                for pt in PROFIT_TARGETS:
                    s = lookup.get((sl, pt, mh))
                    if s:
                        med = s["median"]
                        wr = s["win_rate"]
                        row += f" | {med:>+5.0%} {wr:>4.0%}w"
                    else:
                        row += f" |        ---"
                print(row)


# =====================================================================
# Main
# =====================================================================

def main():
    print("=" * 80)
    print("ALPHA PICKS OPTIONS P&L MODEL (Black-Scholes Synthetic)")
    print("=" * 80)
    print(f"  DTE at entry: {DTE_DAYS} days")
    print(f"  Deltas: {DELTAS}")
    print(f"  Hold windows: {WINDOWS} trading days")
    print(f"  Risk-free rate: {RATE:.0%}")
    print(f"  IV estimation: {IV_LOOKBACK}-day realized vol x {IV_PREMIUM}")
    print(f"  IV floor/cap: {IV_FLOOR:.0%} / {IV_CAP:.0%}")
    print()

    # Load data
    print("Loading Alpha Picks...")
    picks = load_alpha_picks()
    print(f"  {len(picks)} picks loaded")

    print("Loading price cache...")
    prices = load_prices_from_cache()
    print(f"  {len(prices)} symbols in cache")

    trading_dates = get_trading_dates(prices)
    print(f"  Trading dates: {trading_dates[0]} to {trading_dates[-1]}")

    # =================================================================
    # Model each pick — T-1 entry (pre-announcement)
    # =================================================================
    print("\nModelling options P&L (T-1 entry)...")
    results_tm1 = []
    skipped_tm1 = 0
    for pick in picks:
        r = model_pick_options(pick, prices, trading_dates, entry_offset=-1)
        if r:
            results_tm1.append(r)
        else:
            skipped_tm1 += 1
    print(f"  {len(results_tm1)} picks analyzed, {skipped_tm1} skipped")

    print("Computing daily option prices (T-1 entry)...")
    for r in results_tm1:
        compute_daily_option_prices(r, prices, trading_dates)
    print("  Done.")

    # =================================================================
    # Model each pick — T0 entry (announcement day close)
    # =================================================================
    print("\nModelling options P&L (T0 entry)...")
    results_t0 = []
    skipped_t0 = 0
    for pick in picks:
        r = model_pick_options(pick, prices, trading_dates, entry_offset=0)
        if r:
            results_t0.append(r)
        else:
            skipped_t0 += 1
    print(f"  {len(results_t0)} picks analyzed, {skipped_t0} skipped")

    print("Computing daily option prices (T0 entry)...")
    for r in results_t0:
        compute_daily_option_prices(r, prices, trading_dates)
    print("  Done.")

    # ===================================================================
    # PART 1: T-1 Entry — Fixed-window analysis
    # ===================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 1: T-1 ENTRY (Pre-Announcement)")
    print("#" * 160)

    print_per_pick_detail(results_tm1, "d80")
    print_per_pick_detail(results_tm1, "d50")
    print_summary(results_tm1)
    print_comparison(results_tm1)
    print_capital_analysis(results_tm1)
    print_best_worst(results_tm1, "d80", 30)
    print_best_worst(results_tm1, "d50", 30)

    # ===================================================================
    # PART 2: T-1 Entry — Rule-based exit analysis
    # ===================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 2: T-1 ENTRY — RULE-BASED EXIT ANALYSIS")
    print("#" * 160)
    print()
    print(f"  Stop-loss levels:   {[fmt_sl(s) for s in STOP_LOSSES]}")
    print(f"  Profit targets:     {[fmt_pt(p) for p in PROFIT_TARGETS]}")
    print(f"  Max hold periods:   {MAX_HOLDS} trading days")
    print(f"  Combinations:       {len(STOP_LOSSES) * len(PROFIT_TARGETS) * len(MAX_HOLDS)} per delta")

    stats_tm1 = run_rule_analysis(results_tm1)

    print_rule_grid(stats_tm1)
    print_rule_highlights(stats_tm1)

    # ===================================================================
    # PART 3: T0 Entry — Fixed-window analysis
    # ===================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 3: T0 ENTRY (Announcement Day Close — Actionable on Public Info)")
    print("#" * 160)

    print_per_pick_detail(results_t0, "d80")
    print_per_pick_detail(results_t0, "d50")
    print_summary(results_t0)
    print_capital_analysis(results_t0)
    print_best_worst(results_t0, "d80", 30)
    print_best_worst(results_t0, "d50", 30)

    # ===================================================================
    # PART 4: T0 Entry — Rule-based exit analysis
    # ===================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 4: T0 ENTRY — RULE-BASED EXIT ANALYSIS")
    print("#" * 160)
    print()
    print(f"  Stop-loss levels:   {[fmt_sl(s) for s in STOP_LOSSES]}")
    print(f"  Profit targets:     {[fmt_pt(p) for p in PROFIT_TARGETS]}")
    print(f"  Max hold periods:   {MAX_HOLDS} trading days")

    stats_t0 = run_rule_analysis(results_t0)

    print_rule_grid(stats_t0)
    print_rule_highlights(stats_t0)

    # ===================================================================
    # PART 5: Entry timing comparison (T-1 vs T0)
    # ===================================================================
    print("\n\n")
    print("#" * 160)
    print("##  PART 5: ENTRY TIMING COMPARISON — T-1 vs T0")
    print("#" * 160)

    print_entry_comparison(results_tm1, results_t0)
    print_entry_comparison_rules(stats_tm1, stats_t0)

    # Closing notes
    print()
    print("=" * 80)
    print("NOTES:")
    print("  - Option prices are synthetic (Black-Scholes), not historical market prices")
    print("  - IV estimated from trailing 30-day realized vol x 1.2")
    print("  - Same IV used at entry and exit (no IV crush / expansion modelled)")
    print("  - No bid/ask spread applied (mid-price execution assumed)")
    print("  - T-1 entry: buy at close the day before announcement")
    print("  - T0 entry: buy at close on announcement day (public info)")
    print("  - Forward windows (T+1, T+10, etc.) measured from T0 in both cases")
    print("  - 60 DTE option expires ~42 trading days in; max hold 60d includes expiry")
    print("=" * 80)


if __name__ == "__main__":
    main()
