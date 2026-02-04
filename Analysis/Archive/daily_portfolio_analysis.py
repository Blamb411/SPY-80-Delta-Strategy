"""
Daily Portfolio Simulation & Risk Analysis
=============================================
Simulates both Alpha Picks and ProQuant portfolios day-by-day
using actual daily prices from the Massive API.

Alpha Picks: Equal dollar investment ($1000) per pick, buy-and-hold
ProQuant: Weekly rebalanced equal-weight portfolio (~30 stocks)
SPY: Buy-and-hold benchmark

Computes: daily returns, Sharpe, PSR, drawdown, rolling stats
"""

import openpyxl
import sqlite3
import math
import os
from collections import defaultdict, OrderedDict
from datetime import datetime, timedelta


DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "price_cache.db")
XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "ProQuant History 1_29_2026.xlsx")

INVESTMENT_PER_PICK = 1000  # dollars per Alpha Pick


# ============================================================
# PRICE DATA ACCESS
# ============================================================

def load_all_prices(conn):
    """Load all prices into memory for fast access."""
    cur = conn.execute("SELECT symbol, date, close FROM daily_prices ORDER BY symbol, date")
    prices = defaultdict(dict)  # symbol -> {date: close}
    for symbol, date, close in cur.fetchall():
        prices[symbol][date] = close
    return prices


def get_trading_dates(prices, start_date, end_date):
    """Get sorted list of trading dates from SPY data."""
    spy = prices.get("SPY", {})
    dates = sorted(d for d in spy.keys() if start_date <= d <= end_date)
    return dates


# ============================================================
# ALPHA PICKS SIMULATION
# ============================================================

def load_alpha_picks():
    """Load Alpha Picks from Excel."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['AlphaPicks']
    picks = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        symbol = row[1]
        date = row[2]
        if not symbol or not date:
            continue

        date_str = date.strftime("%Y-%m-%d") if hasattr(date, 'strftime') else str(date)[:10]

        closed = None
        if row[4] and str(row[4]).strip() not in ('-', 'None', ''):
            closed = row[4].strftime("%Y-%m-%d") if hasattr(row[4], 'strftime') else str(row[4])[:10]

        picks.append({
            'symbol': symbol,
            'picked': date_str,
            'buy_price': float(row[3]) if row[3] else 0.0,
            'closed': closed,
        })

    # Deduplicate: if same symbol + same pick date appears twice,
    # keep both as separate positions (one closed, one open)
    picks.sort(key=lambda x: x['picked'])
    return picks


def simulate_alpha_picks(picks, prices, trading_dates):
    """
    Simulate Alpha Picks portfolio day-by-day.

    Each pick gets $INVESTMENT_PER_PICK at entry.
    Shares = investment / buy_price.
    Position value = shares * current_price.
    """
    # Build position list
    positions = []
    for i, p in enumerate(picks):
        symbol = p['symbol']
        # Map SMCI* to SMCI for price lookup
        price_symbol = "SMCI*" if symbol == "SMCI*" else symbol

        # Get the actual entry price from market data (closest date)
        entry_price = p['buy_price']
        if entry_price > 0:
            shares = INVESTMENT_PER_PICK / entry_price
        else:
            shares = 0

        positions.append({
            'id': i,
            'symbol': price_symbol,
            'entry_date': p['picked'],
            'exit_date': p['closed'],  # None if still open
            'shares': shares,
            'entry_price': entry_price,
        })

    # Day-by-day simulation
    daily_values = []
    daily_returns = []
    prev_value = 0

    for date in trading_dates:
        total_value = 0
        active_count = 0

        for pos in positions:
            # Check if position is active on this date
            if date < pos['entry_date']:
                continue
            if pos['exit_date'] and date > pos['exit_date']:
                continue

            # Get current price
            symbol_prices = prices.get(pos['symbol'], {})
            current_price = symbol_prices.get(date)

            if current_price is None:
                # Try to find the most recent price before this date
                available = sorted(d for d in symbol_prices.keys() if d <= date)
                if available:
                    current_price = symbol_prices[available[-1]]
                else:
                    # Use entry price as fallback
                    current_price = pos['entry_price']

            total_value += pos['shares'] * current_price
            active_count += 1

        if total_value > 0:
            if prev_value > 0:
                daily_ret = (total_value / prev_value) - 1
                daily_returns.append((date, daily_ret))
            daily_values.append((date, total_value, active_count))
            prev_value = total_value
        elif prev_value > 0:
            # No active positions but had some before
            daily_values.append((date, 0, 0))
            prev_value = 0

    return daily_values, daily_returns


# ============================================================
# PROQUANT SIMULATION
# ============================================================

def load_proquant():
    """Load ProQuant weekly data."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['ProQuant']
    weekly = defaultdict(list)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        symbol, date, action, start_w, new_w, change_w, price = row
        if not symbol or not date:
            continue
        date_str = date.strftime("%Y-%m-%d") if hasattr(date, 'strftime') else str(date)[:10]
        weekly[date_str].append({
            'symbol': symbol,
            'action': action,
            'new_weight': float(new_w) if new_w else 0.0,
            'price': float(price) if price else 0.0,
        })

    return OrderedDict(sorted(weekly.items()))


def simulate_proquant(weekly_data, prices, trading_dates):
    """
    Simulate ProQuant portfolio day-by-day.

    Uses a weight-based approach: on rebalance dates, record the portfolio
    weights. Between rebalances, compute daily returns as the weighted
    average of individual stock daily returns. This avoids price-level
    mismatches between Excel rebalance prices and API close prices.
    """
    rebalance_dates = sorted(weekly_data.keys())
    start_date = rebalance_dates[0]

    # Filter trading dates to ProQuant period
    pq_dates = [d for d in trading_dates if d >= start_date]
    if not pq_dates:
        return [], []

    # Build weight snapshots: for each rebalance date, record {symbol: weight}
    weight_snapshots = {}
    for rb_date, entries in weekly_data.items():
        weights = {}
        for e in entries:
            if e['new_weight'] > 0:
                weights[e['symbol']] = e['new_weight']
        # Normalize weights to sum to 1
        total_w = sum(weights.values())
        if total_w > 0:
            weights = {s: w / total_w for s, w in weights.items()}
        weight_snapshots[rb_date] = weights

    # Day-by-day simulation using weighted returns
    daily_values = []
    daily_returns = []

    # Track which rebalance weights are active
    current_weights = {}
    portfolio_value = 10000.0
    rb_idx = 0

    for i, date in enumerate(pq_dates):
        # Check if we rebalance today
        while rb_idx < len(rebalance_dates) and rebalance_dates[rb_idx] <= date:
            current_weights = weight_snapshots[rebalance_dates[rb_idx]]
            rb_idx += 1

        if not current_weights:
            continue

        # Compute weighted daily return from individual stock returns
        if i == 0:
            # First day: no return, just record value
            daily_values.append((date, portfolio_value))
            continue

        prev_date = pq_dates[i - 1]
        port_return = 0.0
        weight_covered = 0.0

        for sym, weight in current_weights.items():
            sym_prices = prices.get(sym, {})
            p_today = sym_prices.get(date)
            p_prev = sym_prices.get(prev_date)

            if p_today is not None and p_prev is not None and p_prev > 0:
                stock_ret = (p_today / p_prev) - 1
                port_return += weight * stock_ret
                weight_covered += weight
            else:
                # Stock has no price data for this day - assume flat
                weight_covered += weight

        # Scale return if some weight wasn't covered
        if weight_covered > 0 and weight_covered < 0.9:
            port_return = port_return / weight_covered

        portfolio_value *= (1 + port_return)
        daily_returns.append((date, port_return))
        daily_values.append((date, portfolio_value))

    return daily_values, daily_returns


# ============================================================
# SPY BENCHMARK
# ============================================================

def simulate_spy(prices, trading_dates, start_date):
    """Simple buy-and-hold SPY from start_date."""
    spy_prices = prices.get("SPY", {})
    dates = [d for d in trading_dates if d >= start_date]

    daily_returns = []
    prev_price = None

    for date in dates:
        p = spy_prices.get(date)
        if p is None:
            continue
        if prev_price is not None and prev_price > 0:
            daily_returns.append((date, (p / prev_price) - 1))
        prev_price = p

    return daily_returns


# ============================================================
# RISK METRICS
# ============================================================

def norm_cdf(x):
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def compute_metrics(daily_returns, label="Portfolio"):
    """Comprehensive risk-adjusted metrics from daily returns."""
    returns = [r for _, r in daily_returns]
    n = len(returns)

    if n < 10:
        print(f"  {label}: only {n} observations, insufficient for analysis")
        return None

    mu = sum(returns) / n
    var = sum((r - mu) ** 2 for r in returns) / (n - 1)
    sigma = math.sqrt(var)

    if sigma == 0:
        print(f"  {label}: zero variance")
        return None

    # Annualized
    ann_return = (1 + mu) ** 252 - 1
    ann_vol = sigma * math.sqrt(252)
    sharpe = (mu / sigma) * math.sqrt(252)

    # Compounded total return
    total = 1.0
    for r in returns:
        total *= (1 + r)
    total_return = total - 1

    # Max drawdown
    cumulative = 1.0
    peak = 1.0
    max_dd = 0
    dd_start = None
    dd_end = None
    current_dd_start = daily_returns[0][0]

    for date, r in daily_returns:
        cumulative *= (1 + r)
        if cumulative > peak:
            peak = cumulative
            current_dd_start = date
        dd = (peak - cumulative) / peak
        if dd > max_dd:
            max_dd = dd
            dd_start = current_dd_start
            dd_end = date

    # Skewness
    if n > 2:
        skew = (n / ((n - 1) * (n - 2))) * sum(((r - mu) / sigma) ** 3 for r in returns)
    else:
        skew = 0

    # Excess kurtosis
    if n > 3:
        raw_kurt = (n * (n + 1)) / ((n - 1) * (n - 2) * (n - 3))
        raw_kurt *= sum(((r - mu) / sigma) ** 4 for r in returns)
        kurt = raw_kurt - (3 * (n - 1) ** 2) / ((n - 2) * (n - 3))
    else:
        kurt = 0

    # Non-annualized Sharpe for PSR
    sr_hat = mu / sigma

    # PSR
    denom_sq = 1 - skew * sr_hat + (kurt / 4) * sr_hat ** 2
    psr = {}
    for sr_star_ann in [0.0, 0.5, 1.0, 1.5, 2.0]:
        sr_star = sr_star_ann / math.sqrt(252)
        if denom_sq > 0:
            z = (sr_hat - sr_star) * math.sqrt(n - 1) / math.sqrt(denom_sq)
            psr[sr_star_ann] = norm_cdf(z)
        else:
            psr[sr_star_ann] = 0.5

    # Min TRL
    z_95 = 1.645
    if sr_hat != 0:
        min_trl = max(1, math.ceil(
            z_95 ** 2 * (1 - skew * sr_hat + (kurt / 4) * sr_hat ** 2)
            / (sr_hat ** 2)
        ) + 1)
    else:
        min_trl = float("inf")

    # Calmar ratio
    calmar = ann_return / max_dd if max_dd > 0 else 0

    # Sortino ratio (downside deviation)
    downside = [r for r in returns if r < 0]
    if downside:
        downside_var = sum(r ** 2 for r in downside) / len(downside)
        downside_dev = math.sqrt(downside_var) * math.sqrt(252)
        sortino = ann_return / downside_dev if downside_dev > 0 else 0
    else:
        sortino = float("inf")

    # Win/loss stats
    up_days = [r for r in returns if r > 0]
    down_days = [r for r in returns if r < 0]

    return {
        'n': n,
        'total_return': total_return,
        'ann_return': ann_return,
        'ann_vol': ann_vol,
        'sharpe': sharpe,
        'sortino': sortino,
        'calmar': calmar,
        'max_dd': max_dd,
        'dd_start': dd_start,
        'dd_end': dd_end,
        'skewness': skew,
        'kurtosis': kurt,
        'psr': psr,
        'min_trl': min_trl,
        'up_days': len(up_days),
        'down_days': len(down_days),
        'flat_days': n - len(up_days) - len(down_days),
        'avg_up': sum(up_days) / len(up_days) if up_days else 0,
        'avg_down': sum(down_days) / len(down_days) if down_days else 0,
        'best_day': max(returns),
        'worst_day': min(returns),
    }


def print_metrics(label, m):
    """Print formatted metrics."""
    if m is None:
        return

    print(f"  {label}")
    print(f"  {'='*70}")
    print(f"    Trading Days:         {m['n']}")
    print(f"    Total Return:         {m['total_return']:>+.2%}")
    print(f"    Annualized Return:    {m['ann_return']:>+.2%}")
    print(f"    Annualized Volatility:{m['ann_vol']:>.2%}")
    print(f"    Max Drawdown:         {m['max_dd']:>.2%}  ({m['dd_start']} to {m['dd_end']})")
    print()
    print(f"    Sharpe Ratio:         {m['sharpe']:>.4f}")
    print(f"    Sortino Ratio:        {m['sortino']:>.4f}")
    print(f"    Calmar Ratio:         {m['calmar']:>.4f}")
    print()
    print(f"    Skewness:             {m['skewness']:>.4f}")
    print(f"    Excess Kurtosis:      {m['kurtosis']:>.4f}")
    print()
    print(f"    Up Days:   {m['up_days']:>5}  ({m['up_days']/m['n']:.1%})  Avg: {m['avg_up']:>+.3%}")
    print(f"    Down Days: {m['down_days']:>5}  ({m['down_days']/m['n']:.1%})  Avg: {m['avg_down']:>+.3%}")
    print(f"    Best Day:  {m['best_day']:>+.3%}")
    print(f"    Worst Day: {m['worst_day']:>+.3%}")
    print()
    print(f"    Probabilistic Sharpe Ratio (PSR):")
    for sr_star, p in sorted(m['psr'].items()):
        marker = ""
        if sr_star == 0.0:
            marker = " <- prob of positive Sharpe"
        elif sr_star == 1.0:
            marker = " (good benchmark)"
        elif sr_star == 2.0:
            marker = " (excellent benchmark)"
        print(f"      PSR(SR* = {sr_star:.1f}): {p:>7.1%}{marker}")
    print()
    print(f"    Min Track Record (95% conf SR>0): {m['min_trl']} days")
    print(f"    Current observations:              {m['n']} days  "
          f"{'SUFFICIENT' if m['n'] >= m['min_trl'] else 'INSUFFICIENT'}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 90)
    print("SEEKING ALPHA — DAILY PORTFOLIO SIMULATION & RISK ANALYSIS")
    print("=" * 90)
    print()

    conn = sqlite3.connect(DB_PATH)
    prices = load_all_prices(conn)
    conn.close()

    print(f"Loaded prices for {len(prices)} symbols")
    print()

    # Get all trading dates from SPY
    all_trading_dates = get_trading_dates(prices, "2022-06-01", "2026-01-29")
    print(f"Trading days available: {len(all_trading_dates)} ({all_trading_dates[0]} to {all_trading_dates[-1]})")
    print()

    # ================================================================
    # ALPHA PICKS
    # ================================================================
    print("=" * 90)
    print("ALPHA PICKS — DAILY PORTFOLIO SIMULATION")
    print("=" * 90)
    print(f"  Investment per pick: ${INVESTMENT_PER_PICK:,.0f}")
    print()

    picks = load_alpha_picks()
    ap_values, ap_returns = simulate_alpha_picks(picks, prices, all_trading_dates)

    if ap_values:
        first_val = ap_values[0]
        last_val = ap_values[-1]
        print(f"  Simulation period:  {first_val[0]} to {last_val[0]}")
        print(f"  Starting value:     ${first_val[1]:>12,.2f}  ({first_val[2]} positions)")
        print(f"  Ending value:       ${last_val[1]:>12,.2f}  ({last_val[2]} positions)")
        total_invested = INVESTMENT_PER_PICK * len(picks)
        print(f"  Total invested:     ${total_invested:>12,.0f}  ({len(picks)} picks)")
        print(f"  Final value:        ${last_val[1]:>12,.2f}")
        print(f"  Total profit:       ${last_val[1] - total_invested:>12,.2f}")
        print()

        # Track peak active positions
        max_positions = max(v[2] for v in ap_values)
        print(f"  Max concurrent positions: {max_positions}")
        print()

    print("-" * 90)
    ap_metrics = compute_metrics(ap_returns, "Alpha Picks")
    print_metrics("ALPHA PICKS", ap_metrics)
    print()

    # SPY benchmark over same period as Alpha Picks
    ap_start = picks[0]['picked'] if picks else "2022-07-01"
    spy_ap_returns = simulate_spy(prices, all_trading_dates, ap_start)
    print("-" * 90)
    spy_ap_metrics = compute_metrics(spy_ap_returns, "SPY (Alpha Picks period)")
    print_metrics("SPY BENCHMARK (Alpha Picks period)", spy_ap_metrics)
    print()

    # ================================================================
    # PROQUANT
    # ================================================================
    print("=" * 90)
    print("PROQUANT — DAILY PORTFOLIO SIMULATION")
    print("=" * 90)
    print()

    weekly_data = load_proquant()
    pq_values, pq_returns = simulate_proquant(weekly_data, prices, all_trading_dates)

    if pq_values:
        first_val = pq_values[0]
        last_val = pq_values[-1]
        print(f"  Simulation period:  {first_val[0]} to {last_val[0]}")
        print(f"  Starting value:     ${first_val[1]:>12,.2f}")
        print(f"  Ending value:       ${last_val[1]:>12,.2f}")
        print(f"  Total return:       {(last_val[1] / first_val[1]) - 1:>+.2%}")
        print()

    print("-" * 90)
    pq_metrics = compute_metrics(pq_returns, "ProQuant")
    print_metrics("PROQUANT", pq_metrics)
    print()

    # SPY over same period as ProQuant
    pq_start = list(weekly_data.keys())[0] if weekly_data else "2025-05-30"
    spy_pq_returns = simulate_spy(prices, all_trading_dates, pq_start)
    print("-" * 90)
    spy_pq_metrics = compute_metrics(spy_pq_returns, "SPY (ProQuant period)")
    print_metrics("SPY BENCHMARK (ProQuant period)", spy_pq_metrics)
    print()

    # ================================================================
    # SIDE-BY-SIDE COMPARISON
    # ================================================================
    print("=" * 90)
    print("SIDE-BY-SIDE COMPARISON")
    print("=" * 90)
    print()

    headers = ["Metric", "Alpha Picks", "SPY (AP period)", "ProQuant", "SPY (PQ period)"]
    metrics_list = [ap_metrics, spy_ap_metrics, pq_metrics, spy_pq_metrics]

    def fmt(val, fmt_str):
        if val is None:
            return "N/A"
        return fmt_str.format(val)

    rows = [
        ("Trading Days", [str(m['n']) if m else "N/A" for m in metrics_list]),
        ("Total Return", [fmt(m['total_return'] if m else None, "{:+.2%}") for m in metrics_list]),
        ("Ann. Return", [fmt(m['ann_return'] if m else None, "{:+.2%}") for m in metrics_list]),
        ("Ann. Volatility", [fmt(m['ann_vol'] if m else None, "{:.2%}") for m in metrics_list]),
        ("Sharpe Ratio", [fmt(m['sharpe'] if m else None, "{:.4f}") for m in metrics_list]),
        ("Sortino Ratio", [fmt(m['sortino'] if m else None, "{:.4f}") for m in metrics_list]),
        ("Calmar Ratio", [fmt(m['calmar'] if m else None, "{:.4f}") for m in metrics_list]),
        ("Max Drawdown", [fmt(m['max_dd'] if m else None, "{:.2%}") for m in metrics_list]),
        ("Skewness", [fmt(m['skewness'] if m else None, "{:.4f}") for m in metrics_list]),
        ("Excess Kurtosis", [fmt(m['kurtosis'] if m else None, "{:.4f}") for m in metrics_list]),
        ("PSR(SR>0)", [fmt(m['psr'][0.0] if m else None, "{:.1%}") for m in metrics_list]),
        ("PSR(SR>1)", [fmt(m['psr'][1.0] if m else None, "{:.1%}") for m in metrics_list]),
        ("PSR(SR>2)", [fmt(m['psr'][2.0] if m else None, "{:.1%}") for m in metrics_list]),
    ]

    # Print table
    col_widths = [18, 15, 17, 15, 17]
    header_line = "  ".join(h.rjust(w) for h, w in zip(headers, col_widths))
    print(f"  {header_line}")
    print(f"  {'-' * len(header_line)}")
    for label, values in rows:
        row_str = label.ljust(col_widths[0])
        for val, w in zip(values, col_widths[1:]):
            row_str += "  " + val.rjust(w)
        print(f"  {row_str}")

    print()

    # ================================================================
    # ALPHA PICKS: Overlapping period with ProQuant
    # ================================================================
    if pq_metrics and ap_metrics:
        print("=" * 90)
        print("OVERLAPPING PERIOD COMPARISON (ProQuant start to present)")
        print("=" * 90)
        print()

        pq_start_date = pq_values[0][0] if pq_values else "2025-05-30"

        # Filter Alpha Picks returns to ProQuant period
        ap_overlap = [(d, r) for d, r in ap_returns if d >= pq_start_date]
        spy_overlap = [(d, r) for d, r in spy_ap_returns if d >= pq_start_date]

        if len(ap_overlap) > 10:
            ap_overlap_m = compute_metrics(ap_overlap, "Alpha Picks (overlap)")
            pq_overlap_m = pq_metrics  # already only covers this period
            spy_overlap_m = compute_metrics(spy_overlap, "SPY (overlap)")

            overlap_metrics = [ap_overlap_m, pq_overlap_m, spy_overlap_m]
            overlap_headers = ["Metric", "Alpha Picks", "ProQuant", "SPY"]
            overlap_widths = [18, 15, 15, 15]

            header_line = "  ".join(h.rjust(w) for h, w in zip(overlap_headers, overlap_widths))
            print(f"  {header_line}")
            print(f"  {'-' * len(header_line)}")

            overlap_rows = [
                ("Days", [str(m['n']) if m else "N/A" for m in overlap_metrics]),
                ("Total Return", [fmt(m['total_return'] if m else None, "{:+.2%}") for m in overlap_metrics]),
                ("Ann. Return", [fmt(m['ann_return'] if m else None, "{:+.2%}") for m in overlap_metrics]),
                ("Sharpe", [fmt(m['sharpe'] if m else None, "{:.4f}") for m in overlap_metrics]),
                ("Sortino", [fmt(m['sortino'] if m else None, "{:.4f}") for m in overlap_metrics]),
                ("Max Drawdown", [fmt(m['max_dd'] if m else None, "{:.2%}") for m in overlap_metrics]),
                ("PSR(SR>0)", [fmt(m['psr'][0.0] if m else None, "{:.1%}") for m in overlap_metrics]),
                ("PSR(SR>1)", [fmt(m['psr'][1.0] if m else None, "{:.1%}") for m in overlap_metrics]),
            ]

            for label, values in overlap_rows:
                row_str = label.ljust(overlap_widths[0])
                for val, w in zip(values, overlap_widths[1:]):
                    row_str += "  " + val.rjust(w)
                print(f"  {row_str}")

        print()

    # ================================================================
    # MONTHLY RETURNS TABLE
    # ================================================================
    print("=" * 90)
    print("ALPHA PICKS — MONTHLY RETURNS")
    print("=" * 90)

    monthly_pnl = defaultdict(float)
    monthly_start = {}

    cum = 1.0
    for date, ret in ap_returns:
        month = date[:7]  # YYYY-MM
        if month not in monthly_start:
            monthly_start[month] = cum
        cum *= (1 + ret)

    months = sorted(monthly_start.keys())
    for i, month in enumerate(months):
        if i < len(months) - 1:
            next_start = monthly_start[months[i + 1]]
        else:
            next_start = cum
        monthly_ret = (next_start / monthly_start[month]) - 1
        monthly_pnl[month] = monthly_ret

    # Print by year
    years = sorted(set(m[:4] for m in months))
    print(f"  {'Year':>6}", end="")
    for m in range(1, 13):
        print(f"  {datetime(2000, m, 1).strftime('%b'):>7}", end="")
    print(f"  {'Annual':>8}")
    print(f"  {'-'*110}")

    for year in years:
        print(f"  {year:>6}", end="")
        year_cum = 1.0
        for m in range(1, 13):
            key = f"{year}-{m:02d}"
            if key in monthly_pnl:
                ret = monthly_pnl[key]
                year_cum *= (1 + ret)
                print(f"  {ret:>+6.1%}", end="")
            else:
                print(f"  {'':>7}", end="")
        print(f"  {year_cum - 1:>+7.1%}")

    print()

    # ================================================================
    # ALPHA PICKS BY YEAR
    # ================================================================
    print("=" * 90)
    print("ALPHA PICKS — ANNUAL RISK METRICS")
    print("=" * 90)
    print()

    for year in years:
        year_returns = [(d, r) for d, r in ap_returns if d[:4] == year]
        if len(year_returns) < 20:
            continue
        ym = compute_metrics(year_returns, year)
        if ym:
            print(f"  {year}: Return {ym['total_return']:>+.2%}  "
                  f"Sharpe {ym['sharpe']:>.2f}  "
                  f"Sortino {ym['sortino']:>.2f}  "
                  f"MaxDD {ym['max_dd']:>.2%}  "
                  f"PSR(>0) {ym['psr'][0.0]:>.0%}")

    print()
    print("=" * 90)
    print("ANALYSIS COMPLETE")
    print("=" * 90)


if __name__ == "__main__":
    main()
