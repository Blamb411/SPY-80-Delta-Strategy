#!/usr/bin/env python3
"""
Alpha Picks Announcement & Post-Announcement Momentum Analysis
================================================================
Measures stock price performance at:
  - Announcement day (T-1 close to T close)
  - T-1 to T+1 (overnight/next-day pop)
  - T to T+10, T+30, T+60, T+120 trading days
  - SPY benchmark over same periods

Usage:
    python announcement_momentum.py
"""

import os
import sys
import sqlite3
from datetime import datetime, timedelta
from collections import defaultdict

import openpyxl

# Try to import yfinance for filling price gaps
try:
    import yfinance as yf
    HAS_YF = True
except ImportError:
    HAS_YF = False

_this_dir = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(_this_dir, "ProQuant History 1_29_2026.xlsx")
DB_PATH = os.path.join(_this_dir, "price_cache.db")

WINDOWS = [1, 10, 30, 60, 120]  # trading days after announcement


def load_alpha_picks():
    """Load Alpha Picks from Excel."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["AlphaPicks"]
    picks = []
    seen = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] is None or row[2] is None:
            continue
        symbol = str(row[1]).strip()
        pick_date = row[2]
        if isinstance(pick_date, datetime):
            date_str = pick_date.strftime("%Y-%m-%d")
        else:
            date_str = str(pick_date)[:10]
        buy_price = float(row[3]) if row[3] else None
        key = (symbol, date_str)
        if key in seen:
            continue
        seen.add(key)
        picks.append({
            "symbol": symbol,
            "pick_date": date_str,
            "buy_price": buy_price,
        })
    return picks


def load_prices_from_cache():
    """Load all prices from SQLite cache."""
    prices = defaultdict(dict)
    if not os.path.exists(DB_PATH):
        return prices
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT symbol, date, close FROM daily_prices ORDER BY symbol, date")
    for symbol, date, close in c.fetchall():
        prices[symbol][date] = close
    conn.close()
    return prices


def fetch_missing_prices(symbols, start_date="2022-06-01", end_date="2026-01-31"):
    """Fetch prices for symbols missing from cache using yfinance."""
    if not HAS_YF:
        print("WARNING: yfinance not available, some tickers may be missing")
        return {}
    fetched = {}
    for sym in symbols:
        try:
            ticker = yf.Ticker(sym)
            df = ticker.history(start=start_date, end=end_date, auto_adjust=True)
            if df.empty:
                print(f"  WARNING: No yfinance data for {sym}")
                continue
            fetched[sym] = {}
            for idx, row in df.iterrows():
                date_str = idx.strftime("%Y-%m-%d")
                fetched[sym][date_str] = float(row["Close"])
        except Exception as e:
            print(f"  WARNING: Failed to fetch {sym}: {e}")
    return fetched


def get_trading_dates(prices):
    """Get sorted list of all trading dates from SPY."""
    if "SPY" in prices:
        return sorted(prices["SPY"].keys())
    # Fallback: use all dates across all symbols
    all_dates = set()
    for sym_prices in prices.values():
        all_dates.update(sym_prices.keys())
    return sorted(all_dates)


def find_trading_date(trading_dates, target, direction="on_or_before"):
    """Find the nearest trading date to target."""
    if target in trading_dates:
        return target
    for i, d in enumerate(trading_dates):
        if d > target:
            if direction == "on_or_before":
                return trading_dates[i - 1] if i > 0 else None
            else:  # on_or_after
                return d
    if direction == "on_or_before":
        return trading_dates[-1] if trading_dates else None
    return None


def offset_trading_days(trading_dates, base_date, offset):
    """Move N trading days from base_date."""
    if base_date not in trading_dates:
        base_date = find_trading_date(trading_dates, base_date, "on_or_after")
        if not base_date:
            return None
    idx = trading_dates.index(base_date)
    target_idx = idx + offset
    if 0 <= target_idx < len(trading_dates):
        return trading_dates[target_idx]
    return None


def main():
    print("Loading Alpha Picks...")
    picks = load_alpha_picks()
    print(f"  {len(picks)} picks loaded")

    print("Loading price cache...")
    prices = load_prices_from_cache()
    print(f"  {len(prices)} symbols in cache")

    # Find symbols needing data
    all_symbols = set(p["symbol"] for p in picks)
    all_symbols.add("SPY")
    missing = []
    for sym in all_symbols:
        if sym not in prices or len(prices[sym]) < 100:
            missing.append(sym)

    if missing:
        print(f"Fetching {len(missing)} missing symbols via yfinance: {missing}")
        fetched = fetch_missing_prices(missing)
        for sym, sym_prices in fetched.items():
            prices[sym].update(sym_prices)
        # Save to cache
        if fetched:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("""CREATE TABLE IF NOT EXISTS daily_prices (
                symbol TEXT, date TEXT, open REAL, high REAL, low REAL,
                close REAL, volume REAL,
                PRIMARY KEY (symbol, date))""")
            for sym, sym_prices in fetched.items():
                for date, close in sym_prices.items():
                    c.execute("""INSERT OR REPLACE INTO daily_prices
                                (symbol, date, close) VALUES (?, ?, ?)""",
                              (sym, date, close))
            conn.commit()
            conn.close()
            print(f"  Cached {sum(len(v) for v in fetched.values())} price records")

    trading_dates = get_trading_dates(prices)
    print(f"  Trading dates: {trading_dates[0]} to {trading_dates[-1]}")

    # ===================================================================
    # Analyze each pick
    # ===================================================================
    results = []
    for pick in picks:
        sym = pick["symbol"]
        pick_date = pick["pick_date"]

        if sym not in prices or not prices[sym]:
            print(f"  SKIP {sym} — no price data")
            continue

        sym_dates = sorted(prices[sym].keys())

        # T-1: trading day before announcement
        t_minus1 = find_trading_date(trading_dates, pick_date, "on_or_before")
        if t_minus1 == pick_date:
            # pick_date is a trading day, go one before
            idx = trading_dates.index(t_minus1)
            t_minus1 = trading_dates[idx - 1] if idx > 0 else None

        # T: announcement day (or next trading day if not a trading day)
        t0 = find_trading_date(trading_dates, pick_date, "on_or_after")

        if not t_minus1 or not t0:
            print(f"  SKIP {sym} {pick_date} — can't find T-1/T0")
            continue

        # Get prices
        p_minus1 = prices[sym].get(t_minus1)
        p_t0 = prices[sym].get(t0)

        if not p_minus1 or not p_t0:
            print(f"  SKIP {sym} {pick_date} — missing T-1 or T0 price")
            continue

        # SPY prices for same dates
        spy_minus1 = prices["SPY"].get(t_minus1)
        spy_t0 = prices["SPY"].get(t0)

        result = {
            "symbol": sym,
            "pick_date": pick_date,
            "t_minus1": t_minus1,
            "t0": t0,
            "p_minus1": p_minus1,
            "p_t0": p_t0,
            "ann_return": (p_t0 / p_minus1 - 1),
            "spy_ann_return": (spy_t0 / spy_minus1 - 1) if spy_minus1 and spy_t0 else None,
        }

        # Forward windows
        for w in WINDOWS:
            t_plus = offset_trading_days(trading_dates, t0, w)
            if t_plus and t_plus in prices[sym]:
                p_plus = prices[sym][t_plus]
                result[f"ret_{w}d"] = (p_plus / p_minus1 - 1)
                result[f"date_{w}d"] = t_plus
                # SPY comparison
                spy_plus = prices["SPY"].get(t_plus)
                if spy_minus1 and spy_plus:
                    result[f"spy_{w}d"] = (spy_plus / spy_minus1 - 1)
                else:
                    result[f"spy_{w}d"] = None
            else:
                result[f"ret_{w}d"] = None
                result[f"spy_{w}d"] = None
                result[f"date_{w}d"] = None

        results.append(result)

    # ===================================================================
    # SECTION 1: Individual Pick Detail
    # ===================================================================
    print()
    print("=" * 160)
    print("ALPHA PICKS: ANNOUNCEMENT & POST-ANNOUNCEMENT RETURNS (from T-1 close)")
    print("=" * 160)

    header = (f"{'Symbol':<8} | {'Pick Date':<12} | {'Ann Day':>8}"
              f" | {'T+1':>8} | {'T+10':>8} | {'T+30':>8}"
              f" | {'T+60':>8} | {'T+120':>8}")
    print(header)
    print("-" * 160)

    for r in sorted(results, key=lambda x: x["pick_date"]):
        ann = f"{r['ann_return']:>+7.1%}" if r["ann_return"] is not None else "   ---"
        parts = [f"{r['symbol']:<8} | {r['pick_date']:<12} | {ann}"]
        for w in WINDOWS:
            ret = r.get(f"ret_{w}d")
            if ret is not None:
                parts.append(f" | {ret:>+7.1%}")
            else:
                parts.append(f" |    ---")
        print("".join(parts))

    # ===================================================================
    # SECTION 2: Summary Statistics
    # ===================================================================
    print()
    print("=" * 120)
    print("SUMMARY STATISTICS")
    print("=" * 120)

    # Collect returns for each window
    windows_labels = [("ann_return", "Ann Day (T-1 to T)")]
    for w in WINDOWS:
        windows_labels.append((f"ret_{w}d", f"T-1 to T+{w}"))

    print()
    print(f"{'Window':<20} | {'N':>4} | {'Mean':>8} | {'Median':>8}"
          f" | {'Win%':>6} | {'Min':>8} | {'Max':>8}"
          f" | {'SPY Mean':>9} | {'Alpha':>8}")
    print("-" * 120)

    for key, label in windows_labels:
        spy_key = "spy_ann_return" if key == "ann_return" else f"spy_{key.split('_')[1]}"
        vals = [r[key] for r in results if r.get(key) is not None]
        spy_vals = [r[spy_key] for r in results if r.get(spy_key) is not None
                    and r.get(key) is not None]

        if not vals:
            continue

        n = len(vals)
        mean_ret = sum(vals) / n
        sorted_vals = sorted(vals)
        median_ret = sorted_vals[n // 2]
        win_pct = sum(1 for v in vals if v > 0) / n
        min_ret = min(vals)
        max_ret = max(vals)
        spy_mean = sum(spy_vals) / len(spy_vals) if spy_vals else 0
        alpha = mean_ret - spy_mean

        print(f"{label:<20} | {n:>4} | {mean_ret:>+7.2%} | {median_ret:>+7.2%}"
              f" | {win_pct:>5.1%} | {min_ret:>+7.1%} | {max_ret:>+7.1%}"
              f" | {spy_mean:>+8.2%} | {alpha:>+7.2%}")

    print("=" * 120)

    # ===================================================================
    # SECTION 3: By Year
    # ===================================================================
    print()
    print("=" * 120)
    print("ANNOUNCEMENT DAY RETURNS BY YEAR")
    print("=" * 120)

    by_year = defaultdict(list)
    for r in results:
        yr = r["pick_date"][:4]
        if r["ann_return"] is not None:
            by_year[yr].append(r["ann_return"])

    print(f"{'Year':<6} | {'N':>4} | {'Mean':>8} | {'Median':>8}"
          f" | {'Win%':>6} | {'Min':>8} | {'Max':>8}")
    print("-" * 70)
    for yr in sorted(by_year.keys()):
        vals = by_year[yr]
        n = len(vals)
        mean_ret = sum(vals) / n
        sorted_vals = sorted(vals)
        median_ret = sorted_vals[n // 2]
        win_pct = sum(1 for v in vals if v > 0) / n
        print(f"{yr:<6} | {n:>4} | {mean_ret:>+7.2%} | {median_ret:>+7.2%}"
              f" | {win_pct:>5.1%} | {min(vals):>+7.1%} | {max(vals):>+7.1%}")

    # ===================================================================
    # SECTION 4: Winners vs Losers at Each Window
    # ===================================================================
    print()
    print("=" * 120)
    print("DO ANNOUNCEMENT WINNERS KEEP WINNING?")
    print("=" * 120)
    print("(Split picks into those with positive vs negative announcement day returns)")
    print()

    ann_positive = [r for r in results if r.get("ann_return") is not None and r["ann_return"] > 0]
    ann_negative = [r for r in results if r.get("ann_return") is not None and r["ann_return"] <= 0]

    print(f"{'Window':<20} | {'Ann Winners (N)':>15} {'Mean':>8}"
          f" | {'Ann Losers (N)':>15} {'Mean':>8}")
    print("-" * 90)
    for key, label in windows_labels:
        pos_vals = [r[key] for r in ann_positive if r.get(key) is not None]
        neg_vals = [r[key] for r in ann_negative if r.get(key) is not None]
        pos_mean = sum(pos_vals) / len(pos_vals) if pos_vals else 0
        neg_mean = sum(neg_vals) / len(neg_vals) if neg_vals else 0
        print(f"{label:<20} | {len(pos_vals):>15} {pos_mean:>+7.2%}"
              f" | {len(neg_vals):>15} {neg_mean:>+7.2%}")

    # ===================================================================
    # SECTION 5: Top and Bottom Picks at Each Window
    # ===================================================================
    print()
    print("=" * 120)
    print("TOP 5 AND BOTTOM 5 PICKS BY ANNOUNCEMENT DAY RETURN")
    print("=" * 120)

    sorted_by_ann = sorted([r for r in results if r.get("ann_return") is not None],
                           key=lambda x: x["ann_return"], reverse=True)

    print("\nTOP 5:")
    print(f"{'Symbol':<8} | {'Pick Date':<12} | {'Ann Day':>8}"
          f" | {'T+10':>8} | {'T+30':>8} | {'T+60':>8} | {'T+120':>8}")
    print("-" * 90)
    for r in sorted_by_ann[:5]:
        parts = [f"{r['symbol']:<8} | {r['pick_date']:<12} | {r['ann_return']:>+7.1%}"]
        for w in [10, 30, 60, 120]:
            ret = r.get(f"ret_{w}d")
            parts.append(f" | {ret:>+7.1%}" if ret is not None else " |    ---")
        print("".join(parts))

    print("\nBOTTOM 5:")
    print(f"{'Symbol':<8} | {'Pick Date':<12} | {'Ann Day':>8}"
          f" | {'T+10':>8} | {'T+30':>8} | {'T+60':>8} | {'T+120':>8}")
    print("-" * 90)
    for r in sorted_by_ann[-5:]:
        parts = [f"{r['symbol']:<8} | {r['pick_date']:<12} | {r['ann_return']:>+7.1%}"]
        for w in [10, 30, 60, 120]:
            ret = r.get(f"ret_{w}d")
            parts.append(f" | {ret:>+7.1%}" if ret is not None else " |    ---")
        print("".join(parts))

    print()
    print("=" * 120)
    print("NOTES:")
    print("  - All returns measured from T-1 close (day before announcement)")
    print("  - 'Ann Day' = T-1 close to T close (announcement day effect)")
    print("  - T+N = N trading days after announcement")
    print("  - Alpha = pick return minus SPY return over same period")
    print("  - Win% = percentage of picks with positive return")
    print("=" * 120)


if __name__ == "__main__":
    main()
