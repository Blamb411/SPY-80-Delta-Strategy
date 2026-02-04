#!/usr/bin/env python3
"""
Alpha Picks Prediction Accuracy Breakeven Model
=================================================
Answers: "How accurately do I need to predict Alpha Picks to break even
on buying 80-delta calls?"

Approach:
  1. "Hit" distribution: actual ThetaData returns from our 87 validated picks
  2. "Miss" distribution: estimated option returns on random non-pick stocks
     sampled from the same dates (30-day stock returns + B-S option modeling)
  3. Blend at each accuracy level (50%-90%) via Monte Carlo
  4. Report expected P&L, median return, win rate, and breakeven accuracy

Focuses exclusively on 80-delta ITM calls (per earlier finding that
50-delta ATM is not viable with real execution costs).

Usage:
    python prediction_breakeven.py
"""

import os
import sys
import math
import random
import sqlite3
from datetime import datetime
from collections import defaultdict

import openpyxl

_this_dir = os.path.dirname(os.path.abspath(__file__))
_project_dir = os.path.dirname(_this_dir)
sys.path.insert(0, _project_dir)

from backtest.black_scholes import black_scholes_price, find_strike_for_delta

XLSX_PATH = os.path.join(_this_dir, "ProQuant History 1_29_2026.xlsx")
STOCK_DB_PATH = os.path.join(_this_dir, "price_cache.db")

# Option parameters (80-delta only)
DTE_DAYS = 60
RATE = 0.05
IV_LOOKBACK = 30
IV_PREMIUM = 1.2
IV_FLOOR = 0.15
IV_CAP = 1.00
TARGET_DELTA = 0.80

# Rule-based exits
PROFIT_TARGET = 0.50    # +50%
MAX_HOLD = 60           # trading days
MAX_SIM_DAYS = 60

# Execution cost: apply bid-ask penalty to entry/exit
# From ThetaData results: median 80-delta spread is ~15%
# Mid-price = 0% penalty, 25% slippage = ~4% round-trip cost
SPREAD_PENALTY_PCT = 0.04  # 4% round-trip cost on options (conservative)

# Monte Carlo parameters
MC_TRIALS = 5000
ACCURACY_LEVELS = [0.40, 0.45, 0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90]

# Number of predictions per simulation year
# Alpha Picks come every ~2 weeks = ~26 per year
# Model: 1 prediction per pick cycle
PICKS_PER_YEAR = 26

random.seed(42)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_alpha_picks():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["AlphaPicks"]
    picks = []
    seen = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] is None or row[2] is None:
            continue
        symbol = str(row[1]).strip()
        pick_date = row[2]
        if isinstance(pick_date, datetime):
            date_str = pick_date.strftime("%Y-%m-%d")
        else:
            date_str = str(pick_date)[:10]
        key = (symbol, date_str)
        if key in seen:
            continue
        seen.add(key)
        picks.append({"symbol": symbol, "pick_date": date_str})
    return picks


def load_prices():
    prices = defaultdict(dict)
    conn = sqlite3.connect(STOCK_DB_PATH)
    c = conn.cursor()
    c.execute("SELECT symbol, date, close FROM daily_prices ORDER BY symbol, date")
    for sym, dt, close in c.fetchall():
        prices[sym][dt] = close
    conn.close()
    return prices


def get_trading_dates(prices):
    if "SPY" in prices:
        return sorted(prices["SPY"].keys())
    all_dates = set()
    for sym_prices in prices.values():
        all_dates.update(sym_prices.keys())
    return sorted(all_dates)


def find_trading_date(trading_dates, target, direction="on_or_before"):
    if target in trading_dates:
        return target
    for i, d in enumerate(trading_dates):
        if d > target:
            if direction == "on_or_before":
                return trading_dates[i - 1] if i > 0 else None
            else:
                return d
    return trading_dates[-1] if direction == "on_or_before" and trading_dates else None


def offset_trading_days(trading_dates, base_date, offset):
    if base_date not in trading_dates:
        base_date = find_trading_date(trading_dates, base_date, "on_or_after")
        if not base_date:
            return None
    idx = trading_dates.index(base_date)
    target_idx = idx + offset
    if 0 <= target_idx < len(trading_dates):
        return trading_dates[target_idx]
    return None


# ---------------------------------------------------------------------------
# IV estimation
# ---------------------------------------------------------------------------

def estimate_iv(prices_dict, symbol, date_str, trading_dates):
    if symbol not in prices_dict:
        return None
    sym_prices = prices_dict[symbol]
    if date_str not in trading_dates:
        date_str = find_trading_date(trading_dates, date_str, "on_or_before")
        if not date_str:
            return None
    idx = trading_dates.index(date_str)
    start_idx = idx - IV_LOOKBACK
    if start_idx < 0:
        return None
    closes = []
    for i in range(start_idx, idx + 1):
        d = trading_dates[i]
        if d in sym_prices:
            closes.append(sym_prices[d])
    if len(closes) < IV_LOOKBACK // 2:
        return None
    log_returns = []
    for i in range(1, len(closes)):
        if closes[i] > 0 and closes[i - 1] > 0:
            log_returns.append(math.log(closes[i] / closes[i - 1]))
    if len(log_returns) < 5:
        return None
    mean_ret = sum(log_returns) / len(log_returns)
    variance = sum((r - mean_ret) ** 2 for r in log_returns) / (len(log_returns) - 1)
    realized_vol = math.sqrt(variance * 252)
    iv = realized_vol * IV_PREMIUM
    return max(IV_FLOOR, min(IV_CAP, iv))


# ---------------------------------------------------------------------------
# Option return computation (synthetic, for miss distribution)
# ---------------------------------------------------------------------------

def compute_option_return_fixed_window(spot_entry, spot_exit, iv, dte_entry=DTE_DAYS,
                                        days_held_trading=30):
    """Compute 80-delta call return for a fixed holding period."""
    t_entry = dte_entry / 365.0
    strike = find_strike_for_delta(spot_entry, t_entry, RATE, iv, TARGET_DELTA, 'C')
    if strike is None:
        return None

    entry_price = black_scholes_price(spot_entry, strike, t_entry, RATE, iv, 'C')
    if entry_price is None or entry_price <= 0:
        return None

    calendar_days = days_held_trading * 365.0 / 252.0
    remaining_dte = max(0, dte_entry - calendar_days)
    t_exit = remaining_dte / 365.0

    if t_exit <= 0:
        exit_price = max(0.0, spot_exit - strike)
    else:
        exit_price = black_scholes_price(spot_exit, strike, t_exit, RATE, iv, 'C')
        if exit_price is None:
            exit_price = max(0.0, spot_exit - strike)

    # Apply spread penalty
    entry_cost = entry_price * (1 + SPREAD_PENALTY_PCT / 2)
    exit_proceeds = exit_price * (1 - SPREAD_PENALTY_PCT / 2)

    ret = (exit_proceeds / entry_cost - 1) if entry_cost > 0 else None
    return ret


def compute_option_return_with_rules(spot_entry, daily_spots, iv, dte_entry=DTE_DAYS):
    """
    Compute 80-delta call return with PT=+50%, MH=60d rules.
    daily_spots: list of (day_number, spot_price) from day 1 onward.
    """
    t_entry = dte_entry / 365.0
    strike = find_strike_for_delta(spot_entry, t_entry, RATE, iv, TARGET_DELTA, 'C')
    if strike is None:
        return None, None

    entry_price = black_scholes_price(spot_entry, strike, t_entry, RATE, iv, 'C')
    if entry_price is None or entry_price <= 0:
        return None, None

    entry_cost = entry_price * (1 + SPREAD_PENALTY_PCT / 2)

    for day, spot in daily_spots:
        if day > MAX_HOLD:
            break

        calendar_days = day * 365.0 / 252.0
        remaining_dte = max(0, dte_entry - calendar_days)
        t = remaining_dte / 365.0

        if t <= 0:
            opt_price = max(0.0, spot - strike)
        else:
            opt_price = black_scholes_price(spot, strike, t, RATE, iv, 'C')
            if opt_price is None:
                opt_price = max(0.0, spot - strike)

        exit_proceeds = opt_price * (1 - SPREAD_PENALTY_PCT / 2)
        ret = (exit_proceeds / entry_cost - 1) if entry_cost > 0 else 0

        # Check profit target
        if ret >= PROFIT_TARGET:
            return ret, "profit_target"

    # Max hold reached — exit at last available price
    if daily_spots:
        last_day, last_spot = daily_spots[-1] if daily_spots[-1][0] <= MAX_HOLD else daily_spots[-2] if len(daily_spots) > 1 else daily_spots[-1]
        day = min(last_day, MAX_HOLD)
        calendar_days = day * 365.0 / 252.0
        remaining_dte = max(0, dte_entry - calendar_days)
        t = remaining_dte / 365.0
        if t <= 0:
            opt_price = max(0.0, last_spot - strike)
        else:
            opt_price = black_scholes_price(last_spot, strike, t, RATE, iv, 'C')
            if opt_price is None:
                opt_price = max(0.0, last_spot - strike)
        exit_proceeds = opt_price * (1 - SPREAD_PENALTY_PCT / 2)
        ret = (exit_proceeds / entry_cost - 1) if entry_cost > 0 else 0
        return ret, "max_hold"

    return None, None


# ---------------------------------------------------------------------------
# Build the "hit" and "miss" distributions
# ---------------------------------------------------------------------------

def build_hit_distribution(picks, prices, trading_dates):
    """
    Build the return distribution for correct predictions.
    Uses the same B-S synthetic model with spread penalty for consistency
    with the miss distribution (both use same methodology).

    Also loads ThetaData results where available for comparison.
    """
    hit_returns_fixed = []    # fixed 30-day hold
    hit_returns_rules = []    # PT=+50%, MH=60d

    for pick in picks:
        sym = pick["symbol"]
        pick_date = pick["pick_date"]

        if sym not in prices or not prices[sym]:
            continue

        # T0: announcement day
        t0 = find_trading_date(trading_dates, pick_date, "on_or_after")
        if not t0:
            continue
        spot_entry = prices[sym].get(t0)
        if not spot_entry:
            continue

        iv = estimate_iv(prices, sym, t0, trading_dates)
        if iv is None:
            continue

        # Fixed 30-day return
        t30 = offset_trading_days(trading_dates, t0, 30)
        if t30:
            spot_30 = prices[sym].get(t30)
            if spot_30:
                ret = compute_option_return_fixed_window(spot_entry, spot_30, iv, DTE_DAYS, 30)
                if ret is not None:
                    hit_returns_fixed.append(ret)

        # Rule-based return (daily series)
        daily_spots = []
        for day in range(1, MAX_SIM_DAYS + 1):
            t_date = offset_trading_days(trading_dates, t0, day)
            if not t_date:
                break
            spot = prices[sym].get(t_date)
            if spot is not None:
                daily_spots.append((day, spot))

        if daily_spots:
            ret, reason = compute_option_return_with_rules(spot_entry, daily_spots, iv)
            if ret is not None:
                hit_returns_rules.append(ret)

    return hit_returns_fixed, hit_returns_rules


def build_miss_distribution(picks, prices, trading_dates):
    """
    Build the return distribution for incorrect predictions.
    For each pick date, sample random non-pick stocks and compute
    what an 80-delta call would have returned.
    """
    pick_symbols = {p["symbol"] for p in picks}
    all_symbols = [s for s in prices.keys() if s not in pick_symbols and s != "SPY"]

    miss_returns_fixed = []
    miss_returns_rules = []

    # For each pick date, sample several non-pick stocks
    samples_per_date = 10  # sample 10 random stocks per pick date
    pick_dates = sorted(set(p["pick_date"] for p in picks))

    for pick_date in pick_dates:
        t0 = find_trading_date(trading_dates, pick_date, "on_or_after")
        if not t0:
            continue

        # Find symbols with price data around this date
        eligible = []
        for sym in all_symbols:
            if t0 in prices[sym]:
                eligible.append(sym)

        if not eligible:
            continue

        # Sample
        sample_size = min(samples_per_date, len(eligible))
        sampled = random.sample(eligible, sample_size)

        for sym in sampled:
            spot_entry = prices[sym].get(t0)
            if not spot_entry or spot_entry <= 0:
                continue

            iv = estimate_iv(prices, sym, t0, trading_dates)
            if iv is None:
                continue

            # Fixed 30-day return
            t30 = offset_trading_days(trading_dates, t0, 30)
            if t30:
                spot_30 = prices[sym].get(t30)
                if spot_30:
                    ret = compute_option_return_fixed_window(spot_entry, spot_30, iv, DTE_DAYS, 30)
                    if ret is not None:
                        miss_returns_fixed.append(ret)

            # Rule-based return
            daily_spots = []
            for day in range(1, MAX_SIM_DAYS + 1):
                t_date = offset_trading_days(trading_dates, t0, day)
                if not t_date:
                    break
                spot = prices[sym].get(t_date)
                if spot is not None:
                    daily_spots.append((day, spot))

            if daily_spots:
                ret, reason = compute_option_return_with_rules(spot_entry, daily_spots, iv)
                if ret is not None:
                    miss_returns_rules.append(ret)

    return miss_returns_fixed, miss_returns_rules


# ---------------------------------------------------------------------------
# Monte Carlo blending
# ---------------------------------------------------------------------------

def compute_stats(values):
    if not values:
        return None
    n = len(values)
    mean = sum(values) / n
    sorted_v = sorted(values)
    median = sorted_v[n // 2]
    win_rate = sum(1 for v in values if v > 0) / n
    return {"n": n, "mean": mean, "median": median, "win_rate": win_rate,
            "min": min(values), "max": max(values)}


def run_monte_carlo(hit_returns, miss_returns, accuracy, n_trades=26, n_trials=MC_TRIALS):
    """
    Simulate n_trials portfolios, each with n_trades predictions at given accuracy.

    Returns distribution of portfolio-level metrics.
    """
    if not hit_returns or not miss_returns:
        return None

    portfolio_means = []
    portfolio_medians = []
    portfolio_totals = []  # total P&L assuming $1000 per trade

    for _ in range(n_trials):
        trade_returns = []
        for _ in range(n_trades):
            if random.random() < accuracy:
                # Correct prediction — sample from hit distribution
                trade_returns.append(random.choice(hit_returns))
            else:
                # Wrong prediction — sample from miss distribution
                trade_returns.append(random.choice(miss_returns))

        if trade_returns:
            portfolio_mean = sum(trade_returns) / len(trade_returns)
            sorted_rets = sorted(trade_returns)
            portfolio_median = sorted_rets[len(sorted_rets) // 2]
            # Assuming $1000 capital per trade
            total_pnl = sum(r * 1000 for r in trade_returns)

            portfolio_means.append(portfolio_mean)
            portfolio_medians.append(portfolio_median)
            portfolio_totals.append(total_pnl)

    return {
        "mean_of_means": sum(portfolio_means) / len(portfolio_means),
        "median_of_means": sorted(portfolio_means)[len(portfolio_means) // 2],
        "mean_of_medians": sum(portfolio_medians) / len(portfolio_medians),
        "mean_total_pnl": sum(portfolio_totals) / len(portfolio_totals),
        "median_total_pnl": sorted(portfolio_totals)[len(portfolio_totals) // 2],
        "pct_profitable": sum(1 for t in portfolio_totals if t > 0) / len(portfolio_totals),
        "p5_total": sorted(portfolio_totals)[int(0.05 * len(portfolio_totals))],
        "p25_total": sorted(portfolio_totals)[int(0.25 * len(portfolio_totals))],
        "p75_total": sorted(portfolio_totals)[int(0.75 * len(portfolio_totals))],
        "p95_total": sorted(portfolio_totals)[int(0.95 * len(portfolio_totals))],
    }


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def print_distribution_stats(label, returns):
    """Print summary of a return distribution."""
    s = compute_stats(returns)
    if not s:
        print(f"  {label}: No data")
        return
    print(f"  {label}:")
    print(f"    N={s['n']}  Mean={s['mean']:+.2%}  Median={s['median']:+.2%}"
          f"  Win%={s['win_rate']:.1%}  Range=[{s['min']:+.1%}, {s['max']:+.1%}]")


def print_breakeven_table(results_fixed, results_rules):
    """Print the main breakeven analysis table."""

    print()
    print("=" * 150)
    print("PREDICTION ACCURACY BREAKEVEN — Fixed 30-Day Hold")
    print("=" * 150)
    print(f"  Assumptions: 26 trades/year, $1,000 capital per trade, {SPREAD_PENALTY_PCT:.0%} round-trip spread cost")
    print(f"  At each accuracy level, {MC_TRIALS:,} Monte Carlo portfolios simulated")
    print()
    print(f"  {'Accuracy':>8} | {'Mean Ret':>8} | {'Med Ret':>8} | {'Year P&L':>10} | {'Med P&L':>10}"
          f" | {'%Profit':>8} | {'P5 P&L':>10} | {'P25 P&L':>10}"
          f" | {'P75 P&L':>10} | {'P95 P&L':>10}")
    print(f"  {'-' * 140}")

    for acc in ACCURACY_LEVELS:
        r = results_fixed.get(acc)
        if not r:
            continue
        print(f"  {acc:>7.0%} | {r['mean_of_means']:>+7.2%} | {r['median_of_means']:>+7.2%}"
              f" | ${r['mean_total_pnl']:>+9,.0f} | ${r['median_total_pnl']:>+9,.0f}"
              f" | {r['pct_profitable']:>7.1%}"
              f" | ${r['p5_total']:>+9,.0f} | ${r['p25_total']:>+9,.0f}"
              f" | ${r['p75_total']:>+9,.0f} | ${r['p95_total']:>+9,.0f}")

    # Find breakeven
    for acc in ACCURACY_LEVELS:
        r = results_fixed.get(acc)
        if r and r['mean_of_means'] > 0:
            print(f"\n  >>> BREAKEVEN (mean > 0): ~{acc:.0%} accuracy")
            break
    for acc in ACCURACY_LEVELS:
        r = results_fixed.get(acc)
        if r and r['median_of_means'] > 0:
            print(f"  >>> BREAKEVEN (median > 0): ~{acc:.0%} accuracy")
            break
    for acc in ACCURACY_LEVELS:
        r = results_fixed.get(acc)
        if r and r['pct_profitable'] > 0.5:
            print(f"  >>> 50%+ PROFITABLE YEARS: ~{acc:.0%} accuracy")
            break

    print()
    print("=" * 150)
    print("PREDICTION ACCURACY BREAKEVEN — Rule-Based (PT=+50%, MH=60d, No Stop-Loss)")
    print("=" * 150)
    print(f"  Assumptions: 26 trades/year, $1,000 capital per trade, {SPREAD_PENALTY_PCT:.0%} round-trip spread cost")
    print()
    print(f"  {'Accuracy':>8} | {'Mean Ret':>8} | {'Med Ret':>8} | {'Year P&L':>10} | {'Med P&L':>10}"
          f" | {'%Profit':>8} | {'P5 P&L':>10} | {'P25 P&L':>10}"
          f" | {'P75 P&L':>10} | {'P95 P&L':>10}")
    print(f"  {'-' * 140}")

    for acc in ACCURACY_LEVELS:
        r = results_rules.get(acc)
        if not r:
            continue
        print(f"  {acc:>7.0%} | {r['mean_of_means']:>+7.2%} | {r['median_of_means']:>+7.2%}"
              f" | ${r['mean_total_pnl']:>+9,.0f} | ${r['median_total_pnl']:>+9,.0f}"
              f" | {r['pct_profitable']:>7.1%}"
              f" | ${r['p5_total']:>+9,.0f} | ${r['p25_total']:>+9,.0f}"
              f" | ${r['p75_total']:>+9,.0f} | ${r['p95_total']:>+9,.0f}")

    for acc in ACCURACY_LEVELS:
        r = results_rules.get(acc)
        if r and r['mean_of_means'] > 0:
            print(f"\n  >>> BREAKEVEN (mean > 0): ~{acc:.0%} accuracy")
            break
    for acc in ACCURACY_LEVELS:
        r = results_rules.get(acc)
        if r and r['median_of_means'] > 0:
            print(f"  >>> BREAKEVEN (median > 0): ~{acc:.0%} accuracy")
            break
    for acc in ACCURACY_LEVELS:
        r = results_rules.get(acc)
        if r and r['pct_profitable'] > 0.5:
            print(f"  >>> 50%+ PROFITABLE YEARS: ~{acc:.0%} accuracy")
            break


def print_sensitivity(hit_fixed, miss_fixed, hit_rules, miss_rules):
    """Test sensitivity to different capital assumptions."""
    print()
    print("=" * 150)
    print("SENSITIVITY: TRADES PER YEAR")
    print("=" * 150)
    print("  What if you only trade when confidence is high? Fewer trades, higher accuracy.")
    print()

    # Scenario: trade less frequently but with higher accuracy
    scenarios = [
        (26, "All picks (26/yr)"),
        (13, "Every other pick (13/yr)"),
        (6, "Selective (6/yr)"),
    ]

    for n_trades, label in scenarios:
        print(f"\n  --- {label} ---")
        print(f"  {'Accuracy':>8} | {'Fixed Mean':>10} | {'Fixed Med':>10} | {'Fixed $':>10}"
              f" | {'Rules Mean':>10} | {'Rules Med':>10} | {'Rules $':>10}")
        print(f"  {'-' * 80}")

        for acc in [0.50, 0.60, 0.70, 0.80, 0.90]:
            rf = run_monte_carlo(hit_fixed, miss_fixed, acc, n_trades)
            rr = run_monte_carlo(hit_rules, miss_rules, acc, n_trades)
            if rf and rr:
                print(f"  {acc:>7.0%} | {rf['mean_of_means']:>+9.2%} | {rf['median_of_means']:>+9.2%}"
                      f" | ${rf['mean_total_pnl']:>+9,.0f}"
                      f" | {rr['mean_of_means']:>+9.2%} | {rr['median_of_means']:>+9.2%}"
                      f" | ${rr['mean_total_pnl']:>+9,.0f}")


def print_cost_of_being_wrong(miss_fixed, miss_rules):
    """Show what happens on wrong predictions."""
    print()
    print("=" * 120)
    print("COST OF BEING WRONG — What a miss looks like")
    print("=" * 120)

    s_fixed = compute_stats(miss_fixed)
    s_rules = compute_stats(miss_rules)

    if s_fixed:
        print(f"\n  Fixed 30-day hold:")
        print(f"    Mean={s_fixed['mean']:+.2%}  Median={s_fixed['median']:+.2%}"
              f"  Win%={s_fixed['win_rate']:.1%}")
        print(f"    Worst={s_fixed['min']:+.1%}  Best={s_fixed['max']:+.1%}")
        # Distribution buckets
        buckets = [(-1.0, -0.50), (-0.50, -0.20), (-0.20, 0.0), (0.0, 0.20), (0.20, 0.50), (0.50, 10.0)]
        print(f"    Distribution:")
        for lo, hi in buckets:
            count = sum(1 for r in miss_fixed if lo <= r < hi)
            pct = count / len(miss_fixed) if miss_fixed else 0
            if lo <= -0.50:
                label = f"    < -50%"
            elif hi >= 10:
                label = f"    > +50%"
            else:
                label = f"    {lo:+.0%} to {hi:+.0%}"
            bar = "#" * int(pct * 50)
            print(f"      {label:<15} {count:>4} ({pct:>5.1%}) {bar}")

    if s_rules:
        print(f"\n  Rule-based (PT=+50%, MH=60d):")
        print(f"    Mean={s_rules['mean']:+.2%}  Median={s_rules['median']:+.2%}"
              f"  Win%={s_rules['win_rate']:.1%}")
        print(f"    Worst={s_rules['min']:+.1%}  Best={s_rules['max']:+.1%}")
        # How many hit PT vs max hold
        pt_count = sum(1 for r in miss_rules if r >= PROFIT_TARGET * 0.95)  # approximate
        print(f"    ~{pt_count} of {len(miss_rules)} non-picks hit +50% PT ({pt_count/len(miss_rules):.1%})")


def print_stock_vs_options(picks, prices, trading_dates, hit_fixed, miss_fixed):
    """Compare options strategy to just buying stock."""
    print()
    print("=" * 150)
    print("STOCK-ONLY COMPARISON — What if you just buy shares instead of options?")
    print("=" * 150)

    # Compute stock return distribution for picks (hits) at T+30
    hit_stock = []
    miss_stock = []
    pick_symbols = {p["symbol"] for p in picks}
    all_symbols = [s for s in prices.keys() if s not in pick_symbols and s != "SPY"]

    for pick in picks:
        sym = pick["symbol"]
        pick_date = pick["pick_date"]
        if sym not in prices:
            continue
        t0 = find_trading_date(trading_dates, pick_date, "on_or_after")
        if not t0:
            continue
        spot0 = prices[sym].get(t0)
        t30 = offset_trading_days(trading_dates, t0, 30)
        if not t30:
            continue
        spot30 = prices[sym].get(t30)
        if spot0 and spot30:
            hit_stock.append(spot30 / spot0 - 1)

    # Miss: random stocks at pick dates
    pick_dates = sorted(set(p["pick_date"] for p in picks))
    for pick_date in pick_dates:
        t0 = find_trading_date(trading_dates, pick_date, "on_or_after")
        if not t0:
            continue
        eligible = [s for s in all_symbols if t0 in prices[s]]
        if not eligible:
            continue
        for sym in random.sample(eligible, min(10, len(eligible))):
            spot0 = prices[sym].get(t0)
            t30 = offset_trading_days(trading_dates, t0, 30)
            if not t30:
                continue
            spot30 = prices[sym].get(t30)
            if spot0 and spot30:
                miss_stock.append(spot30 / spot0 - 1)

    print(f"\n  Stock return distributions (30-day hold):")
    print_distribution_stats("Correct picks (stock)", hit_stock)
    print_distribution_stats("Wrong picks (stock)", miss_stock)

    # Run MC for stock-only at various accuracy levels
    print(f"\n  {'Accuracy':>8} | {'Stock Mean':>10} | {'Stock Med':>10} | {'Stock $':>10}"
          f" | {'Opt Mean':>10} | {'Opt Med':>10} | {'Opt $':>10}"
          f" | {'Leverage':>8}")
    print(f"  {'-' * 95}")

    for acc in [0.50, 0.60, 0.70, 0.80, 0.90]:
        # Stock MC
        stock_means = []
        stock_totals = []
        opt_means = []
        opt_totals = []
        for _ in range(MC_TRIALS):
            s_rets = []
            o_rets = []
            for _ in range(26):
                if random.random() < acc:
                    s_rets.append(random.choice(hit_stock) if hit_stock else 0)
                    o_rets.append(random.choice(hit_fixed) if hit_fixed else 0)
                else:
                    s_rets.append(random.choice(miss_stock) if miss_stock else 0)
                    o_rets.append(random.choice(miss_fixed) if miss_fixed else 0)
            stock_means.append(sum(s_rets) / len(s_rets))
            stock_totals.append(sum(r * 1000 for r in s_rets))
            opt_means.append(sum(o_rets) / len(o_rets))
            opt_totals.append(sum(r * 1000 for r in o_rets))

        s_mean = sum(stock_means) / len(stock_means)
        s_total = sum(stock_totals) / len(stock_totals)
        o_mean = sum(opt_means) / len(opt_means)
        o_total = sum(opt_totals) / len(opt_totals)
        leverage = o_mean / s_mean if abs(s_mean) > 0.001 else 0

        print(f"  {acc:>7.0%} | {s_mean:>+9.2%} | {sorted(stock_means)[len(stock_means)//2]:>+9.2%}"
              f" | ${s_total:>+9,.0f}"
              f" | {o_mean:>+9.2%} | {sorted(opt_means)[len(opt_means)//2]:>+9.2%}"
              f" | ${o_total:>+9,.0f}"
              f" | {leverage:>7.1f}x")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 80)
    print("ALPHA PICKS PREDICTION ACCURACY BREAKEVEN MODEL")
    print("=" * 80)
    print(f"  Option: 80-delta ITM call, {DTE_DAYS} DTE")
    print(f"  Spread cost: {SPREAD_PENALTY_PCT:.0%} round-trip")
    print(f"  Rule-based: PT=+{PROFIT_TARGET:.0%}, MH={MAX_HOLD}d, no stop-loss")
    print(f"  Monte Carlo: {MC_TRIALS:,} trials per accuracy level")
    print(f"  Trades per year: {PICKS_PER_YEAR}")
    print()

    print("Loading data...")
    picks = load_alpha_picks()
    prices = load_prices()
    trading_dates = get_trading_dates(prices)
    print(f"  {len(picks)} picks, {len(prices)} symbols, {len(trading_dates)} trading dates")

    # Build distributions
    print("\nBuilding HIT distribution (correct predictions = actual Alpha Picks)...")
    hit_fixed, hit_rules = build_hit_distribution(picks, prices, trading_dates)
    print(f"  Fixed 30-day: {len(hit_fixed)} samples")
    print(f"  Rule-based:   {len(hit_rules)} samples")
    print_distribution_stats("Hits (fixed 30d)", hit_fixed)
    print_distribution_stats("Hits (rules)", hit_rules)

    print("\nBuilding MISS distribution (wrong predictions = random non-pick stocks)...")
    miss_fixed, miss_rules = build_miss_distribution(picks, prices, trading_dates)
    print(f"  Fixed 30-day: {len(miss_fixed)} samples")
    print(f"  Rule-based:   {len(miss_rules)} samples")
    print_distribution_stats("Miss (fixed 30d)", miss_fixed)
    print_distribution_stats("Miss (rules)", miss_rules)

    if not hit_fixed or not miss_fixed:
        print("\nInsufficient data to run Monte Carlo. Exiting.")
        return

    # Run Monte Carlo at each accuracy level
    print(f"\nRunning Monte Carlo ({MC_TRIALS:,} trials x {len(ACCURACY_LEVELS)} accuracy levels)...")

    results_fixed = {}
    results_rules = {}
    for acc in ACCURACY_LEVELS:
        results_fixed[acc] = run_monte_carlo(hit_fixed, miss_fixed, acc)
        if hit_rules and miss_rules:
            results_rules[acc] = run_monte_carlo(hit_rules, miss_rules, acc)
        print(f"  {acc:.0%} done")

    # Print results
    print_breakeven_table(results_fixed, results_rules)
    print_cost_of_being_wrong(miss_fixed, miss_rules)
    print_sensitivity(hit_fixed, miss_fixed, hit_rules, miss_rules)
    print_stock_vs_options(picks, prices, trading_dates, hit_fixed, miss_fixed)

    # Final summary
    print()
    print("=" * 80)
    print("KEY TAKEAWAYS")
    print("=" * 80)

    for acc in ACCURACY_LEVELS:
        r = results_rules.get(acc)
        if r and r['mean_of_means'] > 0:
            print(f"  - Rule-based strategy breaks even at ~{acc:.0%} prediction accuracy")
            break

    for acc in ACCURACY_LEVELS:
        r = results_rules.get(acc)
        if r and r['pct_profitable'] > 0.80:
            print(f"  - 80%+ chance of profitable year at ~{acc:.0%} accuracy")
            break

    print(f"  - Each wrong prediction costs approximately "
          f"{sum(miss_rules)/len(miss_rules):+.1%} on average" if miss_rules else "")
    print(f"  - Each correct prediction returns approximately "
          f"{sum(hit_rules)/len(hit_rules):+.1%} on average" if hit_rules else "")

    if hit_rules and miss_rules:
        hit_mean = sum(hit_rules) / len(hit_rules)
        miss_mean = sum(miss_rules) / len(miss_rules)
        if abs(hit_mean - miss_mean) > 0:
            be = -miss_mean / (hit_mean - miss_mean)
            print(f"  - Analytical breakeven (mean): {be:.1%}")

    print()


if __name__ == "__main__":
    main()
