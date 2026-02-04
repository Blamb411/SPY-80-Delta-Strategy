#!/usr/bin/env python3
"""
Alpha Picks Prediction Accuracy Breakeven Model (v2)
=====================================================
Compares three "miss" distributions:
  1. SPY — you buy SPY calls when your prediction is wrong
  2. Random S&P 500 stock — you buy calls on a random index member
  3. Random broader stock — downloaded fresh via yfinance

Hit distribution: actual Alpha Picks (from synthetic model, consistent method)

All option returns computed via Black-Scholes with spread penalty for
apples-to-apples comparison between hit and miss distributions.

Usage:
    python prediction_breakeven_v2.py
"""

import os
import sys
import math
import random
import sqlite3
from datetime import datetime, timedelta
from collections import defaultdict

import openpyxl

_this_dir = os.path.dirname(os.path.abspath(__file__))
_project_dir = os.path.dirname(_this_dir)
sys.path.insert(0, _project_dir)

from backtest.black_scholes import black_scholes_price, find_strike_for_delta

XLSX_PATH = os.path.join(_this_dir, "ProQuant History 1_29_2026.xlsx")
STOCK_DB_PATH = os.path.join(_this_dir, "price_cache.db")

# --- Option parameters (80-delta only) ---
DTE_DAYS = 60
RATE = 0.05
TARGET_DELTA = 0.80
IV_LOOKBACK = 30
IV_PREMIUM = 1.2
IV_FLOOR = 0.15
IV_CAP = 1.00

# --- Spread penalties per scenario ---
# From ThetaData: 80-delta median spread ~15% for individual stocks, ~3% for SPY
SPREAD_COST_SPY = 0.02       # 2% round-trip for SPY options
SPREAD_COST_STOCK = 0.04     # 4% round-trip for individual stock options
SPREAD_COST_PICK = 0.04      # same for picks (they're individual stocks too)

# --- Rule-based exit ---
PROFIT_TARGET = 0.50
MAX_HOLD = 60
MAX_SIM_DAYS = 60

# --- Monte Carlo ---
MC_TRIALS = 5000
ACCURACY_LEVELS = [0.30, 0.35, 0.40, 0.45, 0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90]
PICKS_PER_YEAR = 26

# --- Random broader stocks to download ---
# Mix of mid-cap, small-cap, various sectors — NOT in S&P 500
BROADER_TICKERS = [
    # Mid-cap / small-cap names across sectors
    "BILL", "DUOL", "GTLB", "GLOB", "CWAN", "BRZE", "FRSH", "ESTC",
    "PCOR", "DOCS", "RKLB", "ASAN", "ZI", "CFLT", "MNDY", "S",
    "IONQ", "SOUN", "JOBY", "STEM", "CHPT", "QS", "RIVN", "LCID",
    "SOFI", "UPST", "AFRM", "HOOD", "COIN", "MARA", "RIOT", "CLSK",
    "CELH", "CAVA", "BROS", "DT", "GLBE", "CORT", "ARCT", "RXRX",
    "SMMT", "GPCR", "VERA", "TGTX", "ITCI", "PCVX", "ALNY", "MRNA",
    "NVCR", "FATE", "BMRN", "CRNX", "XENE", "RARE", "PRTA", "DCPH",
    "SPR", "TDG", "HEI", "AXON", "TTC", "RBC", "WMS", "SITE",
    "TREX", "POOL", "WSO", "FERG", "WFRD", "CHX", "OII", "PTEN",
    "RIG", "VAL", "HP", "LBRT", "CTRA", "PR", "CHRD", "SM",
    "CNX", "RRC", "EQT", "AR", "MTDR", "VNOM", "FANG", "DKNG",
    "PENN", "FLUT", "RSI", "GENI", "BALY", "IGT", "EVRI", "AGS",
    "LUMN", "TRMB", "KEYS", "ZBRA", "TER", "MKSI", "NOVT", "COHR",
    "LITE", "VIAV", "CIEN", "CALX", "EXTR", "COMM", "SATS", "GILT",
    "AMC", "CNK", "IMAX", "LGF.A", "WBD", "PARA", "FOXA", "NWSA",
    "OWL", "STEP", "ARES", "TPG", "KKR", "APO", "BAM", "BX",
]

random.seed(42)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_alpha_picks():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["AlphaPicks"]
    picks = []
    seen = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] is None or row[2] is None:
            continue
        symbol = str(row[1]).strip()
        pick_date = row[2]
        if isinstance(pick_date, datetime):
            date_str = pick_date.strftime("%Y-%m-%d")
        else:
            date_str = str(pick_date)[:10]
        key = (symbol, date_str)
        if key in seen:
            continue
        seen.add(key)
        picks.append({"symbol": symbol, "pick_date": date_str})
    return picks


def load_prices():
    prices = defaultdict(dict)
    conn = sqlite3.connect(STOCK_DB_PATH)
    c = conn.cursor()
    c.execute("SELECT symbol, date, close FROM daily_prices ORDER BY symbol, date")
    for sym, dt, close in c.fetchall():
        prices[sym][dt] = close
    conn.close()
    return prices


def download_broader_stocks(existing_prices):
    """Download prices for broader stock universe via yfinance."""
    try:
        import yfinance as yf
    except ImportError:
        print("  yfinance not available — skipping broader stock download")
        return existing_prices

    # Only download tickers we don't already have
    to_download = [t for t in BROADER_TICKERS if t not in existing_prices]
    if not to_download:
        print(f"  All {len(BROADER_TICKERS)} broader stocks already in cache")
        return existing_prices

    print(f"  Downloading {len(to_download)} broader stocks via yfinance...")
    # Download in batches
    batch_size = 20
    downloaded = 0
    for i in range(0, len(to_download), batch_size):
        batch = to_download[i:i + batch_size]
        tickers_str = " ".join(batch)
        try:
            data = yf.download(tickers_str, start="2022-06-01", end="2026-01-31",
                               progress=False, group_by="ticker", threads=True)
            if data.empty:
                continue
            for ticker in batch:
                try:
                    if len(batch) == 1:
                        ticker_data = data
                    else:
                        ticker_data = data[ticker] if ticker in data.columns.get_level_values(0) else None
                    if ticker_data is None or ticker_data.empty:
                        continue
                    for idx, row in ticker_data.iterrows():
                        d = idx.strftime("%Y-%m-%d")
                        close = row.get("Close")
                        if close is not None and not math.isnan(close) and close > 0:
                            existing_prices[ticker][d] = float(close)
                            downloaded += 1
                except Exception:
                    continue
        except Exception as e:
            print(f"    Error downloading batch: {e}")
            continue

    new_symbols = sum(1 for t in to_download if t in existing_prices and existing_prices[t])
    print(f"  Downloaded data for {new_symbols} new symbols")
    return existing_prices


def get_trading_dates(prices):
    if "SPY" in prices:
        return sorted(prices["SPY"].keys())
    all_dates = set()
    for sym_prices in prices.values():
        all_dates.update(sym_prices.keys())
    return sorted(all_dates)


def find_trading_date(trading_dates, target, direction="on_or_before"):
    if target in trading_dates:
        return target
    for i, d in enumerate(trading_dates):
        if d > target:
            if direction == "on_or_before":
                return trading_dates[i - 1] if i > 0 else None
            else:
                return d
    return trading_dates[-1] if direction == "on_or_before" and trading_dates else None


def offset_trading_days(trading_dates, base_date, offset):
    if base_date not in trading_dates:
        base_date = find_trading_date(trading_dates, base_date, "on_or_after")
        if not base_date:
            return None
    idx = trading_dates.index(base_date)
    target_idx = idx + offset
    if 0 <= target_idx < len(trading_dates):
        return trading_dates[target_idx]
    return None


# ---------------------------------------------------------------------------
# IV estimation
# ---------------------------------------------------------------------------

def estimate_iv(prices_dict, symbol, date_str, trading_dates):
    if symbol not in prices_dict:
        return None
    sym_prices = prices_dict[symbol]
    if date_str not in trading_dates:
        date_str = find_trading_date(trading_dates, date_str, "on_or_before")
        if not date_str:
            return None
    idx = trading_dates.index(date_str)
    start_idx = idx - IV_LOOKBACK
    if start_idx < 0:
        return None
    closes = []
    for i in range(start_idx, idx + 1):
        d = trading_dates[i]
        if d in sym_prices:
            closes.append(sym_prices[d])
    if len(closes) < IV_LOOKBACK // 2:
        return None
    log_returns = []
    for i in range(1, len(closes)):
        if closes[i] > 0 and closes[i - 1] > 0:
            log_returns.append(math.log(closes[i] / closes[i - 1]))
    if len(log_returns) < 5:
        return None
    mean_ret = sum(log_returns) / len(log_returns)
    variance = sum((r - mean_ret) ** 2 for r in log_returns) / (len(log_returns) - 1)
    realized_vol = math.sqrt(variance * 252)
    iv = realized_vol * IV_PREMIUM
    return max(IV_FLOOR, min(IV_CAP, iv))


# ---------------------------------------------------------------------------
# Option return computation
# ---------------------------------------------------------------------------

def compute_daily_option_returns(spot_entry, daily_spots, iv, spread_cost):
    """
    Compute daily option return series for rule-based exits.
    Returns list of (day, return_pct) or None.
    """
    t_entry = DTE_DAYS / 365.0
    strike = find_strike_for_delta(spot_entry, t_entry, RATE, iv, TARGET_DELTA, 'C')
    if strike is None:
        return None

    entry_price = black_scholes_price(spot_entry, strike, t_entry, RATE, iv, 'C')
    if entry_price is None or entry_price <= 0:
        return None

    entry_cost = entry_price * (1 + spread_cost / 2)

    series = []
    for day, spot in daily_spots:
        calendar_days = day * 365.0 / 252.0
        remaining_dte = max(0, DTE_DAYS - calendar_days)
        t = remaining_dte / 365.0

        if t <= 0:
            opt_price = max(0.0, spot - strike)
        else:
            opt_price = black_scholes_price(spot, strike, t, RATE, iv, 'C')
            if opt_price is None:
                opt_price = max(0.0, spot - strike)

        exit_proceeds = opt_price * (1 - spread_cost / 2)
        ret = (exit_proceeds / entry_cost - 1) if entry_cost > 0 else 0
        series.append((day, ret))

    return series


def apply_rules(daily_returns):
    """Apply PT=+50%, MH=60d, no stop-loss."""
    if not daily_returns:
        return None
    last_ret = None
    for day, ret in daily_returns:
        if day > MAX_HOLD:
            break
        last_ret = ret
        if ret >= PROFIT_TARGET:
            return {"return": ret, "reason": "profit_target", "day": day}
    if last_ret is not None:
        return {"return": last_ret, "reason": "max_hold", "day": min(day, MAX_HOLD) if daily_returns else MAX_HOLD}
    return None


def compute_fixed_30d_return(spot_entry, spot_exit, iv, spread_cost):
    """Compute 80-delta call return for a 30-day hold."""
    t_entry = DTE_DAYS / 365.0
    strike = find_strike_for_delta(spot_entry, t_entry, RATE, iv, TARGET_DELTA, 'C')
    if strike is None:
        return None
    entry_price = black_scholes_price(spot_entry, strike, t_entry, RATE, iv, 'C')
    if entry_price is None or entry_price <= 0:
        return None
    calendar_days = 30 * 365.0 / 252.0
    remaining_dte = max(0, DTE_DAYS - calendar_days)
    t_exit = remaining_dte / 365.0
    if t_exit <= 0:
        exit_price = max(0.0, spot_exit - strike)
    else:
        exit_price = black_scholes_price(spot_exit, strike, t_exit, RATE, iv, 'C')
        if exit_price is None:
            exit_price = max(0.0, spot_exit - strike)
    entry_cost = entry_price * (1 + spread_cost / 2)
    exit_proceeds = exit_price * (1 - spread_cost / 2)
    return (exit_proceeds / entry_cost - 1) if entry_cost > 0 else None


# ---------------------------------------------------------------------------
# Build distributions
# ---------------------------------------------------------------------------

def build_distribution(symbols_and_dates, prices, trading_dates, spread_cost, label=""):
    """
    Build return distributions for a set of (symbol, entry_date) pairs.
    Returns (fixed_30d_returns, rule_returns, stock_returns, details).
    """
    fixed_returns = []
    rule_returns = []
    stock_returns = []
    details = []
    skipped = 0

    for sym, entry_date in symbols_and_dates:
        if sym not in prices or not prices[sym]:
            skipped += 1
            continue

        t0 = find_trading_date(trading_dates, entry_date, "on_or_after")
        if not t0:
            skipped += 1
            continue

        spot_entry = prices[sym].get(t0)
        if not spot_entry or spot_entry <= 0:
            skipped += 1
            continue

        iv = estimate_iv(prices, sym, t0, trading_dates)
        if iv is None:
            skipped += 1
            continue

        # Fixed 30-day
        t30 = offset_trading_days(trading_dates, t0, 30)
        if t30:
            spot_30 = prices[sym].get(t30)
            if spot_30:
                ret = compute_fixed_30d_return(spot_entry, spot_30, iv, spread_cost)
                if ret is not None:
                    fixed_returns.append(ret)
                stock_ret = spot_30 / spot_entry - 1
                stock_returns.append(stock_ret)

        # Rule-based (daily series)
        daily_spots = []
        for day in range(1, MAX_SIM_DAYS + 1):
            t_date = offset_trading_days(trading_dates, t0, day)
            if not t_date:
                break
            spot = prices[sym].get(t_date)
            if spot is not None:
                daily_spots.append((day, spot))

        if daily_spots:
            daily_rets = compute_daily_option_returns(spot_entry, daily_spots, iv, spread_cost)
            if daily_rets:
                result = apply_rules(daily_rets)
                if result:
                    rule_returns.append(result["return"])
                    details.append({
                        "symbol": sym, "date": entry_date, "iv": iv,
                        "spot": spot_entry, "return": result["return"],
                        "reason": result["reason"], "day": result["day"],
                    })

    if label:
        print(f"  {label}: {len(fixed_returns)} fixed, {len(rule_returns)} rule-based"
              f" ({skipped} skipped)")

    return fixed_returns, rule_returns, stock_returns, details


def build_spy_miss(pick_dates, prices, trading_dates):
    """Build miss distribution using SPY calls."""
    pairs = [("SPY", d) for d in pick_dates]
    return build_distribution(pairs, prices, trading_dates, SPREAD_COST_SPY, "SPY miss")


def build_sp500_miss(pick_dates, pick_symbols, prices, trading_dates, samples_per_date=10):
    """Build miss distribution using random S&P 500 stocks (excluding picks)."""
    # Get S&P 500 symbols from cache that aren't picks
    non_pick = [s for s in prices.keys()
                if s not in pick_symbols and s != "SPY"
                and len(prices[s]) >= 100]

    pairs = []
    for d in pick_dates:
        t0 = find_trading_date(trading_dates, d, "on_or_after")
        if not t0:
            continue
        eligible = [s for s in non_pick if t0 in prices[s]]
        if eligible:
            sample_size = min(samples_per_date, len(eligible))
            for sym in random.sample(eligible, sample_size):
                pairs.append((sym, d))

    return build_distribution(pairs, prices, trading_dates, SPREAD_COST_STOCK, "S&P 500 miss")


def build_broader_miss(pick_dates, prices, trading_dates, samples_per_date=5):
    """Build miss distribution using broader stock universe."""
    broader = [s for s in BROADER_TICKERS
               if s in prices and len(prices[s]) >= 30]

    pairs = []
    for d in pick_dates:
        t0 = find_trading_date(trading_dates, d, "on_or_after")
        if not t0:
            continue
        eligible = [s for s in broader if t0 in prices[s]]
        if eligible:
            sample_size = min(samples_per_date, len(eligible))
            for sym in random.sample(eligible, sample_size):
                pairs.append((sym, d))

    return build_distribution(pairs, prices, trading_dates, SPREAD_COST_STOCK, "Broader miss")


# ---------------------------------------------------------------------------
# Monte Carlo
# ---------------------------------------------------------------------------

def compute_stats(values):
    if not values:
        return None
    n = len(values)
    mean = sum(values) / n
    sorted_v = sorted(values)
    median = sorted_v[n // 2]
    win_rate = sum(1 for v in values if v > 0) / n
    return {"n": n, "mean": mean, "median": median, "win_rate": win_rate,
            "min": min(values), "max": max(values)}


def run_monte_carlo(hit_returns, miss_returns, accuracy, n_trades=26, n_trials=MC_TRIALS):
    if not hit_returns or not miss_returns:
        return None
    portfolio_means = []
    portfolio_totals = []
    for _ in range(n_trials):
        trade_returns = []
        for _ in range(n_trades):
            if random.random() < accuracy:
                trade_returns.append(random.choice(hit_returns))
            else:
                trade_returns.append(random.choice(miss_returns))
        mean_ret = sum(trade_returns) / len(trade_returns)
        total_pnl = sum(r * 1000 for r in trade_returns)
        portfolio_means.append(mean_ret)
        portfolio_totals.append(total_pnl)

    sorted_means = sorted(portfolio_means)
    sorted_totals = sorted(portfolio_totals)
    n = len(portfolio_totals)
    return {
        "mean_ret": sum(portfolio_means) / n,
        "median_ret": sorted_means[n // 2],
        "mean_pnl": sum(portfolio_totals) / n,
        "median_pnl": sorted_totals[n // 2],
        "pct_profitable": sum(1 for t in portfolio_totals if t > 0) / n,
        "p5": sorted_totals[int(0.05 * n)],
        "p25": sorted_totals[int(0.25 * n)],
        "p75": sorted_totals[int(0.75 * n)],
        "p95": sorted_totals[int(0.95 * n)],
    }


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def print_dist_summary(label, fixed, rules, stock):
    s_f = compute_stats(fixed)
    s_r = compute_stats(rules)
    s_s = compute_stats(stock)
    print(f"\n  {label}:")
    if s_f:
        print(f"    Fixed 30d:  N={s_f['n']:>4}  Mean={s_f['mean']:>+7.2%}  Med={s_f['median']:>+7.2%}"
              f"  Win%={s_f['win_rate']:>5.1%}  Range=[{s_f['min']:>+.0%}, {s_f['max']:>+.0%}]")
    if s_r:
        print(f"    Rules:      N={s_r['n']:>4}  Mean={s_r['mean']:>+7.2%}  Med={s_r['median']:>+7.2%}"
              f"  Win%={s_r['win_rate']:>5.1%}  Range=[{s_r['min']:>+.0%}, {s_r['max']:>+.0%}]")
    if s_s:
        print(f"    Stock only: N={s_s['n']:>4}  Mean={s_s['mean']:>+7.2%}  Med={s_s['median']:>+7.2%}"
              f"  Win%={s_s['win_rate']:>5.1%}")

    # PT hit rate for rules
    if rules:
        pt_hits = sum(1 for r in rules if r >= PROFIT_TARGET * 0.95)
        print(f"    PT hit rate: {pt_hits}/{len(rules)} ({pt_hits/len(rules):.1%})")


def print_breakeven_table(label, hit_rets, miss_rets, mode="rules"):
    print()
    print(f"  --- {label} ---")
    mode_label = "Rule-Based (PT=+50%, MH=60d)" if mode == "rules" else "Fixed 30-Day Hold"
    print(f"  {mode_label}")
    print(f"  {'Acc':>5} | {'Mean':>7} | {'Med':>7} | {'Year$':>9} | {'Med$':>9}"
          f" | {'%Prof':>6} | {'P5$':>9} | {'P25$':>9} | {'P75$':>9} | {'P95$':>9}")
    print(f"  {'-' * 105}")

    breakeven_mean = None
    breakeven_med = None
    breakeven_prof = None

    for acc in ACCURACY_LEVELS:
        r = run_monte_carlo(hit_rets, miss_rets, acc)
        if not r:
            continue
        print(f"  {acc:>4.0%} | {r['mean_ret']:>+6.2%} | {r['median_ret']:>+6.2%}"
              f" | ${r['mean_pnl']:>+8,.0f} | ${r['median_pnl']:>+8,.0f}"
              f" | {r['pct_profitable']:>5.1%}"
              f" | ${r['p5']:>+8,.0f} | ${r['p25']:>+8,.0f}"
              f" | ${r['p75']:>+8,.0f} | ${r['p95']:>+8,.0f}")

        if breakeven_mean is None and r['mean_ret'] > 0:
            breakeven_mean = acc
        if breakeven_med is None and r['median_ret'] > 0:
            breakeven_med = acc
        if breakeven_prof is None and r['pct_profitable'] > 0.5:
            breakeven_prof = acc

    if breakeven_mean is not None:
        print(f"\n  Breakeven (mean > 0): ~{breakeven_mean:.0%}")
    else:
        print(f"\n  Breakeven (mean > 0): NOT REACHED (strategy loses at all accuracy levels)")
    if breakeven_med is not None:
        print(f"  Breakeven (median > 0): ~{breakeven_med:.0%}")
    if breakeven_prof is not None:
        print(f"  50%+ profitable years: ~{breakeven_prof:.0%}")

    # Analytical breakeven
    if hit_rets and miss_rets:
        h_mean = sum(hit_rets) / len(hit_rets)
        m_mean = sum(miss_rets) / len(miss_rets)
        if abs(h_mean - m_mean) > 0.001:
            be = -m_mean / (h_mean - m_mean)
            print(f"  Analytical breakeven: {be:.1%}  (hit mean={h_mean:+.2%}, miss mean={m_mean:+.2%})")
        else:
            print(f"  Analytical breakeven: N/A (hit and miss means are equal:"
                  f" hit={h_mean:+.2%}, miss={m_mean:+.2%})")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 100)
    print("PREDICTION ACCURACY BREAKEVEN MODEL v2")
    print("Three Miss Distributions: SPY / S&P 500 Random / Broader Random")
    print("=" * 100)
    print(f"  Option: 80-delta ITM call, {DTE_DAYS} DTE")
    print(f"  Spread cost: SPY={SPREAD_COST_SPY:.0%}, stocks={SPREAD_COST_STOCK:.0%} round-trip")
    print(f"  Rules: PT=+{PROFIT_TARGET:.0%}, MH={MAX_HOLD}d, no stop-loss")
    print(f"  Monte Carlo: {MC_TRIALS:,} trials, {PICKS_PER_YEAR} trades/year")
    print()

    # Load data
    print("Loading data...")
    picks = load_alpha_picks()
    prices = load_prices()

    # Download broader stocks
    print("Downloading broader stock universe...")
    prices = download_broader_stocks(prices)

    trading_dates = get_trading_dates(prices)
    print(f"  {len(picks)} picks, {len(prices)} symbols, {len(trading_dates)} trading dates")

    pick_symbols = {p["symbol"] for p in picks}
    pick_dates = sorted(set(p["pick_date"] for p in picks))

    # ===================================================================
    # Build all distributions
    # ===================================================================
    print()
    print("=" * 100)
    print("BUILDING RETURN DISTRIBUTIONS")
    print("=" * 100)

    # Hit: Alpha Picks at T0
    hit_pairs = [(p["symbol"], p["pick_date"]) for p in picks]
    hit_fixed, hit_rules, hit_stock, hit_details = build_distribution(
        hit_pairs, prices, trading_dates, SPREAD_COST_PICK, "Alpha Picks (hits)")

    # Miss 1: SPY
    spy_fixed, spy_rules, spy_stock, _ = build_spy_miss(pick_dates, prices, trading_dates)

    # Miss 2: Random S&P 500 stocks (excluding picks)
    sp500_fixed, sp500_rules, sp500_stock, _ = build_sp500_miss(
        pick_dates, pick_symbols, prices, trading_dates, samples_per_date=10)

    # Miss 3: Broader random stocks
    broad_fixed, broad_rules, broad_stock, _ = build_broader_miss(
        pick_dates, prices, trading_dates, samples_per_date=5)

    # ===================================================================
    # Distribution summaries
    # ===================================================================
    print()
    print("=" * 100)
    print("DISTRIBUTION SUMMARIES")
    print("=" * 100)

    print_dist_summary("HITS: Alpha Picks (T0 entry)", hit_fixed, hit_rules, hit_stock)
    print_dist_summary("MISS 1: SPY calls", spy_fixed, spy_rules, spy_stock)
    print_dist_summary("MISS 2: Random S&P 500 stocks", sp500_fixed, sp500_rules, sp500_stock)
    print_dist_summary("MISS 3: Random broader stocks", broad_fixed, broad_rules, broad_stock)

    # ===================================================================
    # Key comparison: hit vs miss
    # ===================================================================
    print()
    print("=" * 100)
    print("HIT vs MISS COMPARISON")
    print("=" * 100)

    print("\n  Rule-based returns (PT=+50%, MH=60d):")
    print(f"  {'Distribution':<30} | {'N':>4} | {'Mean':>8} | {'Median':>8} | {'Win%':>6} | {'PT Hit%':>8}")
    print(f"  {'-' * 80}")
    for label, rets in [("Alpha Picks (hits)", hit_rules),
                         ("SPY (miss)", spy_rules),
                         ("S&P 500 random (miss)", sp500_rules),
                         ("Broader random (miss)", broad_rules)]:
        s = compute_stats(rets)
        if s:
            pt_hits = sum(1 for r in rets if r >= PROFIT_TARGET * 0.95)
            print(f"  {label:<30} | {s['n']:>4} | {s['mean']:>+7.2%} | {s['median']:>+7.2%}"
                  f" | {s['win_rate']:>5.1%} | {pt_hits/len(rets):>7.1%}")

    print("\n  Fixed 30-day stock returns (no options):")
    print(f"  {'Distribution':<30} | {'N':>4} | {'Mean':>8} | {'Median':>8} | {'Win%':>6}")
    print(f"  {'-' * 65}")
    for label, rets in [("Alpha Picks", hit_stock),
                         ("SPY", spy_stock),
                         ("S&P 500 random", sp500_stock),
                         ("Broader random", broad_stock)]:
        s = compute_stats(rets)
        if s:
            print(f"  {label:<30} | {s['n']:>4} | {s['mean']:>+7.2%} | {s['median']:>+7.2%}"
                  f" | {s['win_rate']:>5.1%}")

    # ===================================================================
    # Breakeven tables
    # ===================================================================
    print()
    print("=" * 100)
    print("BREAKEVEN ANALYSIS — Rule-Based (PT=+50%, MH=60d)")
    print("=" * 100)
    print("  $1,000 per trade, 26 trades/year")

    if hit_rules and spy_rules:
        print_breakeven_table("Miss = SPY Calls", hit_rules, spy_rules, "rules")
    if hit_rules and sp500_rules:
        print_breakeven_table("Miss = Random S&P 500 Stock Calls", hit_rules, sp500_rules, "rules")
    if hit_rules and broad_rules:
        print_breakeven_table("Miss = Random Broader Stock Calls", hit_rules, broad_rules, "rules")

    print()
    print("=" * 100)
    print("BREAKEVEN ANALYSIS — Fixed 30-Day Hold")
    print("=" * 100)

    if hit_fixed and spy_fixed:
        print_breakeven_table("Miss = SPY Calls", hit_fixed, spy_fixed, "fixed")
    if hit_fixed and sp500_fixed:
        print_breakeven_table("Miss = Random S&P 500 Stock Calls", hit_fixed, sp500_fixed, "fixed")
    if hit_fixed and broad_fixed:
        print_breakeven_table("Miss = Random Broader Stock Calls", hit_fixed, broad_fixed, "fixed")

    # ===================================================================
    # Stock-only comparison
    # ===================================================================
    print()
    print("=" * 100)
    print("STOCK-ONLY BREAKEVEN (No Options)")
    print("=" * 100)
    print("  What if you just buy shares instead of calls?")

    for miss_label, miss_stock_rets in [("SPY", spy_stock),
                                         ("S&P 500 random", sp500_stock),
                                         ("Broader random", broad_stock)]:
        if hit_stock and miss_stock_rets:
            print(f"\n  --- Miss = {miss_label} ---")
            print(f"  {'Acc':>5} | {'Stock Mean':>10} | {'Stock Med':>10} | {'Year$':>9}")
            print(f"  {'-' * 45}")
            for acc in [0.40, 0.50, 0.60, 0.70, 0.80, 0.90]:
                r = run_monte_carlo(hit_stock, miss_stock_rets, acc)
                if r:
                    print(f"  {acc:>4.0%} | {r['mean_ret']:>+9.2%} | {r['median_ret']:>+9.2%}"
                          f" | ${r['mean_pnl']:>+8,.0f}")

    # ===================================================================
    # Final summary
    # ===================================================================
    print()
    print("=" * 100)
    print("SUMMARY")
    print("=" * 100)

    for miss_label, miss_r, miss_f in [("SPY", spy_rules, spy_fixed),
                                        ("S&P 500", sp500_rules, sp500_fixed),
                                        ("Broader", broad_rules, broad_fixed)]:
        print(f"\n  Miss = {miss_label}:")
        if hit_rules and miss_r:
            h = sum(hit_rules) / len(hit_rules)
            m = sum(miss_r) / len(miss_r)
            edge = h - m
            if abs(edge) > 0.001:
                be = -m / edge
                print(f"    Rule-based: hit mean={h:+.2%}, miss mean={m:+.2%},"
                      f" edge={edge:+.2%}, breakeven={be:.1%}")
            else:
                print(f"    Rule-based: hit mean={h:+.2%}, miss mean={m:+.2%}, NO EDGE")
        if hit_fixed and miss_f:
            h = sum(hit_fixed) / len(hit_fixed)
            m = sum(miss_f) / len(miss_f)
            edge = h - m
            if abs(edge) > 0.001:
                be = -m / edge
                print(f"    Fixed 30d:  hit mean={h:+.2%}, miss mean={m:+.2%},"
                      f" edge={edge:+.2%}, breakeven={be:.1%}")
            else:
                print(f"    Fixed 30d:  hit mean={h:+.2%}, miss mean={m:+.2%}, NO EDGE")

    print()


if __name__ == "__main__":
    main()
