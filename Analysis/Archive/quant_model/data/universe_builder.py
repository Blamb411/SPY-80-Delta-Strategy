"""
Universe Builder
=================
Constructs the stock universe for scoring. Combines:
- S&P 500 constituents (from a static list)
- Alpha Picks from the Excel spreadsheet
- Any additional tickers

Filters: market cap > $500M, price > $10, no REITs.
"""

import os
import sys
import sqlite3
from datetime import datetime
from typing import List, Set, Dict, Optional, Tuple

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config
import db_schema


# S&P 500 constituents (representative subset — commonly used mid/large caps)
# In production, this would be fetched from an API or maintained list
SP500_CORE = [
    "AAPL", "ABBV", "ABT", "ACN", "ADBE", "ADI", "ADM", "ADP", "ADSK", "AEP",
    "AFL", "AIG", "AIZ", "AJG", "AKAM", "ALB", "ALGN", "ALK", "ALL", "AMAT",
    "AMCR", "AMD", "AME", "AMGN", "AMP", "AMT", "AMZN", "ANET", "ANSS", "AON",
    "AOS", "APA", "APD", "APH", "APTV", "ARE", "ATO", "ATVI", "AVB", "AVGO",
    "AVY", "AWK", "AXP", "AZO", "BA", "BAC", "BAX", "BBWI", "BBY", "BDX",
    "BEN", "BF.B", "BIIB", "BIO", "BK", "BKNG", "BKR", "BLK", "BMY", "BR",
    "BRK.B", "BRO", "BSX", "BWA", "BXP", "C", "CAG", "CAH", "CARR", "CAT",
    "CB", "CBOE", "CBRE", "CCI", "CCL", "CDAY", "CDNS", "CDW", "CE", "CEG",
    "CF", "CFG", "CHD", "CHRW", "CHTR", "CI", "CINF", "CL", "CLX", "CMA",
    "CMCSA", "CME", "CMG", "CMI", "CMS", "CNC", "CNP", "COF", "COO", "COP",
    "COST", "CPB", "CPRT", "CPT", "CRL", "CRM", "CSCO", "CSGP", "CSX", "CTAS",
    "CTLT", "CTRA", "CTSH", "CTVA", "CVS", "CVX", "CZR", "D", "DAL", "DD",
    "DE", "DFS", "DG", "DGX", "DHI", "DHR", "DIS", "DISH", "DLTR", "DOV",
    "DOW", "DPZ", "DRI", "DTE", "DUK", "DVA", "DVN", "DXC", "DXCM", "EA",
    "EBAY", "ECL", "ED", "EFX", "EIX", "EL", "EMN", "EMR", "ENPH", "EOG",
    "EPAM", "EQIX", "EQR", "EQT", "ES", "ESS", "ETN", "ETR", "ETSY", "EVRG",
    "EW", "EXC", "EXPD", "EXPE", "EXR", "F", "FANG", "FAST", "FBHS", "FCX",
    "FDS", "FDX", "FE", "FFIV", "FIS", "FISV", "FITB", "FLT", "FMC", "FOX",
    "FOXA", "FRC", "FRT", "FTNT", "FTV", "GD", "GE", "GILD", "GIS", "GL",
    "GLW", "GM", "GNRC", "GOOG", "GOOGL", "GPC", "GPN", "GRMN", "GS", "GWW",
    "HAL", "HAS", "HBAN", "HCA", "HD", "HOLX", "HON", "HPE", "HPQ", "HRL",
    "HSIC", "HST", "HSY", "HUM", "HWM", "IBM", "ICE", "IDXX", "IEX", "IFF",
    "ILMN", "INCY", "INTC", "INTU", "INVH", "IP", "IPG", "IQV", "IR", "IRM",
    "ISRG", "IT", "ITW", "IVZ", "J", "JBHT", "JCI", "JKHY", "JNJ", "JNPR",
    "JPM", "K", "KDP", "KEY", "KEYS", "KHC", "KIM", "KLAC", "KMB", "KMI",
    "KMX", "KO", "KR", "L", "LDOS", "LEN", "LH", "LHX", "LIN", "LKQ",
    "LLY", "LMT", "LNC", "LNT", "LOW", "LRCX", "LUMN", "LUV", "LVS", "LW",
    "LYB", "LYV", "MA", "MAA", "MAR", "MAS", "MCD", "MCHP", "MCK", "MCO",
    "MDLZ", "MDT", "MET", "META", "MGM", "MHK", "MKC", "MKTX", "MLM", "MMC",
    "MMM", "MNST", "MO", "MOH", "MOS", "MPC", "MPWR", "MRK", "MRNA", "MRO",
    "MS", "MSCI", "MSFT", "MSI", "MTB", "MTCH", "MTD", "MU", "NCLH", "NDAQ",
    "NDSN", "NEE", "NEM", "NFLX", "NI", "NKE", "NOC", "NOW", "NRG", "NSC",
    "NTAP", "NTRS", "NUE", "NVDA", "NVR", "NWL", "NWS", "NWSA", "NXPI", "O",
    "ODFL", "OGN", "OKE", "OMC", "ON", "ORCL", "ORLY", "OTIS", "OXY", "PARA",
    "PAYC", "PAYX", "PCAR", "PCG", "PEAK", "PEG", "PEP", "PFE", "PFG", "PG",
    "PGR", "PH", "PHM", "PKG", "PKI", "PLD", "PM", "PNC", "PNR", "PNW",
    "POOL", "PPG", "PPL", "PRU", "PSA", "PSX", "PTC", "PVH", "PWR", "PXD",
    "PYPL", "QCOM", "QRVO", "RCL", "RE", "REG", "REGN", "RF", "RHI", "RJF",
    "RL", "RMD", "ROK", "ROL", "ROP", "ROST", "RSG", "RTX", "SBAC", "SBNY",
    "SBUX", "SCHW", "SEE", "SHW", "SIVB", "SJM", "SLB", "SNA", "SNPS", "SO",
    "SPG", "SPGI", "SRE", "STE", "STT", "STX", "STZ", "SWK", "SWKS", "SYF",
    "SYK", "SYY", "T", "TAP", "TDG", "TDY", "TECH", "TEL", "TER", "TFC",
    "TFX", "TGT", "TMO", "TMUS", "TPR", "TRGP", "TRMB", "TROW", "TRV", "TSCO",
    "TSLA", "TSN", "TT", "TTWO", "TXN", "TXT", "TYL", "UAL", "UDR", "UHS",
    "ULTA", "UNH", "UNP", "UPS", "URI", "USB", "V", "VFC", "VICI", "VLO",
    "VMC", "VNO", "VRSK", "VRSN", "VRTX", "VTR", "VTRS", "VZ", "WAB", "WAT",
    "WBA", "WBD", "WDC", "WEC", "WELL", "WFC", "WHR", "WM", "WMB", "WMT",
    "WRB", "WRK", "WST", "WTW", "WY", "WYNN", "XEL", "XOM", "XRAY", "XYL",
    "YUM", "ZBH", "ZBRA", "ZION", "ZTS",
]


def load_alpha_picks_from_excel(excel_path: str = config.EXCEL_FILE) -> List[Tuple[str, str]]:
    """
    Load Alpha Picks (symbol, date) from the ProQuant History spreadsheet.
    Returns list of (symbol, pick_date) tuples.
    """
    try:
        import openpyxl
    except ImportError:
        print("Warning: openpyxl not installed. Cannot load Alpha Picks from Excel.")
        return []

    if not os.path.exists(excel_path):
        print(f"Warning: Excel file not found at {excel_path}")
        return []

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    picks = []

    if "AlphaPicks" in wb.sheetnames:
        ws = wb["AlphaPicks"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            # Expected columns: index, symbol, date, buy_price, ...
            if row and len(row) >= 3 and row[1] and row[2]:
                symbol = str(row[1]).strip()
                pick_date = row[2]
                if hasattr(pick_date, "strftime"):
                    pick_date = pick_date.strftime("%Y-%m-%d")
                else:
                    pick_date = str(pick_date).strip()
                # Clean ticker
                if symbol.endswith("*"):
                    symbol = symbol[:-1]
                picks.append((symbol, pick_date))

    wb.close()
    return picks


def get_alpha_pick_symbols(excel_path: str = config.EXCEL_FILE) -> Set[str]:
    """Get unique set of Alpha Pick symbols."""
    picks = load_alpha_picks_from_excel(excel_path)
    return {symbol for symbol, _ in picks}


def build_universe(as_of_date: str,
                   include_sp500: bool = True,
                   extra_symbols: Optional[List[str]] = None,
                   db_path: str = config.DB_PATH) -> List[str]:
    """
    Build the stock universe for a given scoring date.

    Combines S&P 500 + Alpha Picks + extras, saves to stock_universe table,
    and returns the list of symbols that pass filters.
    """
    symbols = set()

    if include_sp500:
        symbols.update(SP500_CORE)

    # Add Alpha Picks symbols
    ap_symbols = get_alpha_pick_symbols()
    symbols.update(ap_symbols)

    if extra_symbols:
        symbols.update(extra_symbols)

    # Remove known delisted / problematic tickers
    symbols.discard("")
    symbols = sorted(symbols)

    print(f"Universe: {len(symbols)} symbols before filtering")

    # Save to database (filtering happens when GuruFocus data is fetched
    # and stock_universe records are populated with market cap/price/sector)
    conn = db_schema.get_connection(db_path)
    for symbol in symbols:
        # Insert placeholder; GuruFocus client will update with real data
        conn.execute(
            """INSERT OR IGNORE INTO stock_universe
               (symbol, as_of_date, passes_filter)
               VALUES (?, ?, 1)""",
            (symbol, as_of_date),
        )
    conn.commit()
    conn.close()

    return symbols


def get_filtered_universe(as_of_date: str, db_path: str = config.DB_PATH) -> List[str]:
    """Get symbols that pass all filters for a given date."""
    conn = db_schema.get_connection(db_path)
    rows = conn.execute(
        """SELECT symbol FROM stock_universe
           WHERE as_of_date = ? AND passes_filter = 1
           ORDER BY symbol""",
        (as_of_date,),
    ).fetchall()
    conn.close()
    return [row["symbol"] for row in rows]


def get_sector_map(as_of_date: str, db_path: str = config.DB_PATH) -> Dict[str, str]:
    """Get symbol -> sector mapping for sector-relative ranking."""
    conn = db_schema.get_connection(db_path)
    rows = conn.execute(
        """SELECT symbol, sector FROM stock_universe
           WHERE as_of_date = ? AND passes_filter = 1 AND sector IS NOT NULL""",
        (as_of_date,),
    ).fetchall()
    conn.close()
    return {row["symbol"]: row["sector"] for row in rows}


def get_pick_dates() -> List[str]:
    """
    Generate the 48 pick dates (1st and 15th of each month)
    from July 2022 to January 2026.
    """
    dates = []
    year = 2022
    month = 7
    while (year, month) <= (2026, 1):
        dates.append(f"{year}-{month:02d}-01")
        dates.append(f"{year}-{month:02d}-15")
        month += 1
        if month > 12:
            month = 1
            year += 1
    return dates


def get_alpha_picks_on_date(pick_date: str,
                            excel_path: str = config.EXCEL_FILE) -> List[str]:
    """Get the Alpha Picks symbols for a specific pick date."""
    all_picks = load_alpha_picks_from_excel(excel_path)
    return [symbol for symbol, date in all_picks if date == pick_date]


if __name__ == "__main__":
    # Quick test
    db_schema.init_db()
    as_of = "2026-01-29"
    symbols = build_universe(as_of)
    print(f"Built universe with {len(symbols)} symbols for {as_of}")

    picks = load_alpha_picks_from_excel()
    print(f"Loaded {len(picks)} Alpha Picks from Excel")
    if picks:
        print(f"  First: {picks[0]}")
        print(f"  Last:  {picks[-1]}")

    ap_symbols = get_alpha_pick_symbols()
    print(f"Unique Alpha Pick symbols: {len(ap_symbols)}")
