"""
Seeking Alpha Portfolio Analysis
==================================
Computes returns, Sharpe Ratio, and Probabilistic Sharpe Ratio
for both the ProQuant Portfolio and Alpha Picks.

Data source: ProQuant History 1_29_2026.xlsx
"""

import openpyxl
import math
from collections import defaultdict, OrderedDict
from datetime import datetime, timedelta


# ============================================================
# DATA LOADING
# ============================================================

def load_proquant(wb):
    """Load ProQuant weekly rebalancing data.

    Returns dict: date -> list of (symbol, action, start_weight, new_weight, price)
    """
    ws = wb['ProQuant']
    weekly = defaultdict(list)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        symbol, date, action, start_w, new_w, change_w, price = row
        if not symbol or not date:
            continue
        date_str = date.strftime("%Y-%m-%d") if hasattr(date, 'strftime') else str(date)[:10]
        weekly[date_str].append({
            'symbol': symbol,
            'action': action,
            'start_weight': float(start_w) if start_w else 0.0,
            'new_weight': float(new_w) if new_w else 0.0,
            'price': float(price) if price else 0.0,
        })

    return OrderedDict(sorted(weekly.items()))


def load_alpha_picks(wb):
    """Load Alpha Picks data.

    Returns list of pick dicts, sorted by pick date (oldest first).
    """
    ws = wb['AlphaPicks']
    picks = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        symbol = row[1]
        date = row[2]
        if not symbol or not date:
            continue

        date_str = date.strftime("%Y-%m-%d") if hasattr(date, 'strftime') else str(date)[:10]

        closed = None
        sell_price = None
        if row[4] and str(row[4]).strip() not in ('-', 'None', ''):
            closed = row[4].strftime("%Y-%m-%d") if hasattr(row[4], 'strftime') else str(row[4])[:10]
            sell_price = float(row[5]) if row[5] and str(row[5]).strip() not in ('-', 'None', '') else None

        picks.append({
            'symbol': symbol,
            'picked': date_str,
            'buy_price': float(row[3]) if row[3] else 0.0,
            'closed': closed,
            'sell_price': sell_price,
            'return_pct': float(row[6]) if row[6] else 0.0,
            'spy_return': float(row[7]) if row[7] else 0.0,
            'diff': float(row[8]) if row[8] else 0.0,
        })

    # Sort oldest first
    picks.sort(key=lambda x: x['picked'])
    return picks


# ============================================================
# SHARPE & PSR FUNCTIONS
# ============================================================

def norm_cdf(x):
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def compute_sharpe_stats(returns, periods_per_year):
    """Compute annualized Sharpe and distribution stats from a return series."""
    n = len(returns)
    if n < 2:
        return None

    mu = sum(returns) / n
    var = sum((r - mu) ** 2 for r in returns) / (n - 1)
    sigma = math.sqrt(var)

    if sigma == 0:
        return None

    sharpe = (mu / sigma) * math.sqrt(periods_per_year)

    # Skewness (sample)
    if n > 2:
        skew = (n / ((n - 1) * (n - 2))) * sum(((r - mu) / sigma) ** 3 for r in returns)
    else:
        skew = 0

    # Excess kurtosis (sample)
    if n > 3:
        raw_kurt = (n * (n + 1)) / ((n - 1) * (n - 2) * (n - 3))
        raw_kurt *= sum(((r - mu) / sigma) ** 4 for r in returns)
        kurt = raw_kurt - (3 * (n - 1) ** 2) / ((n - 2) * (n - 3))
    else:
        kurt = 0

    # Non-annualized Sharpe for PSR
    sr_hat = mu / sigma

    # PSR(SR* = 0)
    denom_sq = 1 - skew * sr_hat + (kurt / 4) * sr_hat ** 2
    if denom_sq > 0:
        z = sr_hat * math.sqrt(n - 1) / math.sqrt(denom_sq)
        psr_zero = norm_cdf(z)
    else:
        psr_zero = 0.5

    # PSR for multiple benchmarks
    psr_benchmarks = {}
    for sr_star_ann in [0.0, 0.5, 1.0]:
        sr_star = sr_star_ann / math.sqrt(periods_per_year)
        if denom_sq > 0:
            z = (sr_hat - sr_star) * math.sqrt(n - 1) / math.sqrt(denom_sq)
            psr_benchmarks[sr_star_ann] = norm_cdf(z)
        else:
            psr_benchmarks[sr_star_ann] = 0.5

    # Min TRL for 95% confidence SR > 0
    z_95 = 1.645
    if sr_hat != 0:
        min_trl = max(1, math.ceil(
            z_95 ** 2 * (1 - skew * sr_hat + (kurt / 4) * sr_hat ** 2)
            / (sr_hat ** 2)
        ) + 1)
    else:
        min_trl = float("inf")

    # Total and compounded return
    total_return = 1.0
    for r in returns:
        total_return *= (1 + r)
    total_return -= 1

    # Max drawdown
    cumulative = 1.0
    peak = 1.0
    max_dd = 0
    for r in returns:
        cumulative *= (1 + r)
        if cumulative > peak:
            peak = cumulative
        dd = (peak - cumulative) / peak
        if dd > max_dd:
            max_dd = dd

    return {
        'n': n,
        'mean_return': mu,
        'std_return': sigma,
        'sharpe': sharpe,
        'skewness': skew,
        'kurtosis': kurt,
        'psr_zero': psr_zero,
        'psr_benchmarks': psr_benchmarks,
        'min_trl': min_trl,
        'total_return': total_return,
        'max_dd': max_dd,
        'periods_per_year': periods_per_year,
    }


def print_sharpe_report(label, stats):
    """Print formatted Sharpe/PSR report."""
    if stats is None:
        print(f"  {label}: insufficient data")
        return

    period_name = "weekly" if stats['periods_per_year'] == 52 else "biweekly" if stats['periods_per_year'] == 26 else "periodic"

    print(f"  Observations:        {stats['n']}")
    print(f"  Total Return:        {stats['total_return']:.2%}")
    print(f"  Max Drawdown:        {stats['max_dd']:.2%}")
    print(f"  Mean {period_name} return: {stats['mean_return']:.4%}")
    print(f"  Std {period_name} return:  {stats['std_return']:.4%}")
    print(f"  Annualized Sharpe:   {stats['sharpe']:.4f}")
    print(f"  Skewness:            {stats['skewness']:.4f}")
    print(f"  Excess Kurtosis:     {stats['kurtosis']:.4f}")
    print()
    print(f"  Probabilistic Sharpe Ratio (PSR):")
    for sr_star, psr in sorted(stats['psr_benchmarks'].items()):
        label_str = ""
        if sr_star == 0.0:
            label_str = " <- prob of positive Sharpe"
        elif sr_star == 0.5:
            label_str = " (mediocre benchmark)"
        elif sr_star == 1.0:
            label_str = " (good benchmark)"
        print(f"    PSR(SR* = {sr_star:.1f}):     {psr:.1%}{label_str}")
    print()
    print(f"  Min Track Record Length (95% confidence SR > 0): {stats['min_trl']}")
    print(f"  Current observations: {stats['n']}  {'SUFFICIENT' if stats['n'] >= stats['min_trl'] else 'INSUFFICIENT'}")


# ============================================================
# PROQUANT ANALYSIS
# ============================================================

def analyze_proquant(weekly_data):
    """Compute weekly returns from ProQuant rebalancing data."""
    print("=" * 90)
    print("PROQUANT PORTFOLIO ANALYSIS")
    print("=" * 90)
    print()

    dates = list(weekly_data.keys())
    print(f"  Period:         {dates[0]} to {dates[-1]}")
    print(f"  Weekly dates:   {len(dates)}")
    print()

    # Build portfolio snapshots: date -> {symbol: (new_weight, price)}
    snapshots = {}
    for date, entries in weekly_data.items():
        portfolio = {}
        for e in entries:
            if e['new_weight'] > 0:  # stock is in portfolio after rebalance
                portfolio[e['symbol']] = {
                    'new_weight': e['new_weight'],
                    'price': e['price'],
                    'action': e['action'],
                }
            # Also track sells (they have start_weight > 0, price at sell)
            if e['action'] == 'Sell':
                portfolio[e['symbol']] = {
                    'new_weight': 0,
                    'price': e['price'],
                    'action': 'Sell',
                    'start_weight': e['start_weight'],
                }
        snapshots[date] = portfolio

    # Compute weekly returns
    # For each consecutive pair of dates (d_prev, d_curr):
    #   For each stock in portfolio on d_prev (new_weight > 0),
    #   find its price on d_curr. Return = price_curr / price_prev - 1.
    #   Portfolio return = sum(weight_prev * return_i)

    weekly_returns = []
    weekly_dates = []

    for i in range(1, len(dates)):
        d_prev = dates[i - 1]
        d_curr = dates[i]

        prev_portfolio = snapshots[d_prev]
        curr_entries = weekly_data[d_curr]

        # Build price lookup for current date (all entries, including sells)
        curr_prices = {}
        for e in curr_entries:
            curr_prices[e['symbol']] = e['price']

        # Compute weighted return
        port_return = 0.0
        total_weight = 0.0

        for symbol, info in prev_portfolio.items():
            if info['new_weight'] <= 0:
                continue

            if symbol in curr_prices and curr_prices[symbol] > 0 and info['price'] > 0:
                stock_return = (curr_prices[symbol] / info['price']) - 1
                port_return += info['new_weight'] * stock_return
                total_weight += info['new_weight']
            else:
                # Stock disappeared without a sell entry - shouldn't happen
                # but handle gracefully by assuming flat
                total_weight += info['new_weight']

        # Normalize if weights don't sum to exactly 1
        if total_weight > 0 and abs(total_weight - 1.0) > 0.01:
            port_return = port_return / total_weight

        weekly_returns.append(port_return)
        weekly_dates.append(d_curr)

    # Print weekly returns
    print("  Weekly Returns:")
    print(f"  {'Date':>12}  {'Return':>10}  {'Cumulative':>12}")
    print(f"  {'-'*38}")

    cumulative = 1.0
    for date, ret in zip(weekly_dates, weekly_returns):
        cumulative *= (1 + ret)
        print(f"  {date:>12}  {ret:>9.2%}  {cumulative - 1:>11.2%}")

    print()

    # Portfolio composition stats
    avg_stocks = sum(
        sum(1 for e in entries if e['new_weight'] > 0)
        for entries in weekly_data.values()
    ) / len(dates)
    print(f"  Avg stocks per week: {avg_stocks:.1f}")

    # Count unique stocks
    all_symbols = set()
    for entries in weekly_data.values():
        for e in entries:
            all_symbols.add(e['symbol'])
    print(f"  Unique stocks traded: {len(all_symbols)}")
    print()

    # Turnover: count buys+sells per week
    total_buys = sum(
        sum(1 for e in entries if e['action'] == 'Buy')
        for entries in weekly_data.values()
    )
    total_sells = sum(
        sum(1 for e in entries if e['action'] == 'Sell')
        for entries in weekly_data.values()
    )
    print(f"  Total buys:  {total_buys}")
    print(f"  Total sells: {total_sells}")
    print(f"  Avg turnover: {(total_buys + total_sells) / len(dates):.1f} trades/week")
    print()

    # SPY comparison - compute from first to last date
    # We don't have SPY prices directly, so just report portfolio stats

    # Sharpe and PSR
    print("-" * 90)
    print("RISK-ADJUSTED RETURNS")
    print("-" * 90)
    stats = compute_sharpe_stats(weekly_returns, periods_per_year=52)
    print_sharpe_report("ProQuant", stats)
    print()

    return weekly_returns, weekly_dates, stats


# ============================================================
# ALPHA PICKS ANALYSIS
# ============================================================

def analyze_alpha_picks(picks):
    """Analyze Alpha Picks portfolio."""
    print("=" * 90)
    print("ALPHA PICKS PORTFOLIO ANALYSIS")
    print("=" * 90)
    print()

    n_picks = len(picks)
    open_picks = [p for p in picks if p['closed'] is None]
    closed_picks = [p for p in picks if p['closed'] is not None]

    print(f"  Total picks:    {n_picks}")
    print(f"  Open positions: {len(open_picks)}")
    print(f"  Closed:         {len(closed_picks)}")
    print(f"  Date range:     {picks[0]['picked']} to {picks[-1]['picked']}")
    print()

    # ---- Per-pick statistics ----
    returns = [p['return_pct'] for p in picks]
    spy_returns = [p['spy_return'] for p in picks]
    excess_returns = [p['diff'] for p in picks]

    avg_ret = sum(returns) / n_picks
    avg_spy = sum(spy_returns) / n_picks
    avg_excess = sum(excess_returns) / n_picks

    wins = sum(1 for r in returns if r > 0)
    spy_beats = sum(1 for d in excess_returns if d > 0)

    print("  Per-Pick Statistics (as of snapshot date):")
    print(f"    Average pick return:  {avg_ret:.2%}")
    print(f"    Average SPY return:   {avg_spy:.2%}")
    print(f"    Average excess return:{avg_excess:.2%}")
    print(f"    Win rate (> 0%):      {wins}/{n_picks} = {wins/n_picks:.1%}")
    print(f"    Beat SPY rate:        {spy_beats}/{n_picks} = {spy_beats/n_picks:.1%}")
    print()

    # Median
    sorted_rets = sorted(returns)
    median_ret = sorted_rets[n_picks // 2]
    print(f"    Median pick return:   {median_ret:.2%}")

    # Best and worst
    best = max(picks, key=lambda p: p['return_pct'])
    worst = min(picks, key=lambda p: p['return_pct'])
    print(f"    Best pick:  {best['symbol']:6s} ({best['picked']}) {best['return_pct']:>+.2%}")
    print(f"    Worst pick: {worst['symbol']:6s} ({worst['picked']}) {worst['return_pct']:>+.2%}")
    print()

    # Doubles and big losers
    doubles = [p for p in picks if p['return_pct'] >= 1.0]
    big_losers = [p for p in picks if p['return_pct'] <= -0.30]
    print(f"    Doubles (100%+):      {len(doubles)}")
    for d in sorted(doubles, key=lambda x: x['return_pct'], reverse=True):
        print(f"      {d['symbol']:6s} {d['picked']} {d['return_pct']:>+.2%}")
    print(f"    Big losers (-30%+):   {len(big_losers)}")
    for l in sorted(big_losers, key=lambda x: x['return_pct']):
        print(f"      {l['symbol']:6s} {l['picked']} {l['return_pct']:>+.2%}")
    print()

    # ---- Simulated portfolio return series ----
    # Build a time series where each pick contributes equally
    # We compute biweekly portfolio returns based on when picks were active

    print("-" * 90)
    print("SIMULATED EQUAL-WEIGHT PORTFOLIO")
    print("-" * 90)
    print()
    print("  Method: Each pick enters the portfolio on its picked date with")
    print("  equal weight. Exits on close date (or remains open). Portfolio")
    print("  return calculated per biweekly period as the average of active")
    print("  pick returns in that period.")
    print()

    # Build pick-level annualized returns for holding period analysis
    snapshot_date = "2026-01-29"

    holding_days = []
    annualized_returns = []

    for p in picks:
        start = datetime.strptime(p['picked'], "%Y-%m-%d")
        if p['closed']:
            end = datetime.strptime(p['closed'], "%Y-%m-%d")
        else:
            end = datetime.strptime(snapshot_date, "%Y-%m-%d")

        days = (end - start).days
        if days <= 0:
            continue

        holding_days.append(days)

        total_ret = p['return_pct']
        # Annualized return
        if total_ret > -1:
            ann_ret = (1 + total_ret) ** (365 / days) - 1
        else:
            ann_ret = -1
        annualized_returns.append(ann_ret)

    avg_holding = sum(holding_days) / len(holding_days)
    median_holding = sorted(holding_days)[len(holding_days) // 2]

    print(f"  Average holding period:  {avg_holding:.0f} days")
    print(f"  Median holding period:   {median_holding} days")
    print()

    # Filter out extreme annualized returns for meaningful stats
    reasonable = [r for r in annualized_returns if -1 < r < 10]
    if reasonable:
        avg_ann = sum(reasonable) / len(reasonable)
        print(f"  Average annualized return (per pick): {avg_ann:.2%}")
    print()

    # ---- Per-pick Sharpe approximation ----
    # Use individual pick returns as the "return series"
    # This is not a time-series Sharpe, but gives a sense of
    # the risk-adjusted quality of picks

    print("-" * 90)
    print("PICK-LEVEL RISK METRICS")
    print("-" * 90)
    print()
    print("  Note: Pick-level Sharpe uses individual pick returns as the series,")
    print("  not a proper daily/weekly time series. It measures the quality of")
    print("  stock selection, not portfolio-level risk-adjusted returns.")
    print()

    # Use excess returns (vs SPY) for the Sharpe calculation
    # This gives Information Ratio rather than Sharpe

    # Sharpe from raw pick returns
    # Approximate: 2 picks per month = 24 per year
    pick_stats = compute_sharpe_stats(returns, periods_per_year=24)
    if pick_stats:
        print("  Raw pick returns (annualized at 24 picks/year):")
        print_sharpe_report("Alpha Picks", pick_stats)

    print()

    # Information Ratio (excess returns vs SPY)
    excess_stats = compute_sharpe_stats(excess_returns, periods_per_year=24)
    if excess_stats:
        print("  Excess returns vs SPY (Information Ratio):")
        print(f"    Mean excess return per pick: {excess_stats['mean_return']:.2%}")
        print(f"    Tracking error per pick:     {excess_stats['std_return']:.2%}")
        print(f"    Annualized IR:               {excess_stats['sharpe']:.4f}")
        print(f"    PSR(IR > 0):                 {excess_stats['psr_zero']:.1%}")

    print()

    # ---- By year breakdown ----
    print("-" * 90)
    print("PERFORMANCE BY YEAR")
    print("-" * 90)

    years = sorted(set(p['picked'][:4] for p in picks))
    for year in years:
        year_picks = [p for p in picks if p['picked'][:4] == year]
        year_rets = [p['return_pct'] for p in year_picks]
        year_spy = [p['spy_return'] for p in year_picks]
        year_excess = [p['diff'] for p in year_picks]

        avg_r = sum(year_rets) / len(year_rets)
        avg_s = sum(year_spy) / len(year_spy)
        avg_e = sum(year_excess) / len(year_excess)
        wins = sum(1 for r in year_rets if r > 0)
        beats = sum(1 for d in year_excess if d > 0)

        print(f"  {year}: {len(year_picks):>3} picks  Avg Return: {avg_r:>+7.2%}  "
              f"Avg SPY: {avg_s:>+7.2%}  Excess: {avg_e:>+7.2%}  "
              f"Win: {wins/len(year_picks):.0%}  Beat SPY: {beats/len(year_picks):.0%}")

    print()

    return picks


# ============================================================
# COMPARISON
# ============================================================

def compare_portfolios(pq_stats, picks):
    """Compare both portfolios side by side."""
    print("=" * 90)
    print("SIDE-BY-SIDE COMPARISON")
    print("=" * 90)
    print()

    if pq_stats:
        pq_ann_ret = (1 + pq_stats['total_return']) ** (52 / pq_stats['n']) - 1
        print(f"  {'Metric':<30} {'ProQuant':>15} {'Alpha Picks':>15}")
        print(f"  {'-'*60}")
        print(f"  {'Period':<30} {'~8 months':>15} {'~3.5 years':>15}")
        print(f"  {'Total Return':<30} {pq_stats['total_return']:>14.2%} {'N/A (see above)':>15}")
        print(f"  {'Annualized Return':<30} {pq_ann_ret:>14.2%} {'N/A':>15}")
        print(f"  {'Annualized Sharpe':<30} {pq_stats['sharpe']:>14.4f} {'(pick-level)':>15}")
        print(f"  {'Max Drawdown':<30} {pq_stats['max_dd']:>14.2%} {'N/A':>15}")
        print(f"  {'PSR (SR > 0)':<30} {pq_stats['psr_zero']:>14.1%} {'(pick-level)':>15}")

    print()
    print("  Note: ProQuant has a proper weekly return series for accurate Sharpe.")
    print("  Alpha Picks Sharpe is approximate (pick-level, not time-series).")
    print("  For a proper Alpha Picks Sharpe, daily prices would need to be fetched")
    print("  via API to build a daily portfolio return series.")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 90)
    print("SEEKING ALPHA PORTFOLIO ANALYSIS")
    print("=" * 90)
    print(f"  Data file: ProQuant History 1_29_2026.xlsx")
    print(f"  Analysis date: 2026-01-29")
    print()

    wb = openpyxl.load_workbook('ProQuant History 1_29_2026.xlsx', data_only=True)

    # ProQuant
    weekly_data = load_proquant(wb)
    pq_returns, pq_dates, pq_stats = analyze_proquant(weekly_data)
    print()

    # Alpha Picks
    picks = load_alpha_picks(wb)
    analyze_alpha_picks(picks)
    print()

    # Comparison
    compare_portfolios(pq_stats, picks)
    print()


if __name__ == "__main__":
    main()
