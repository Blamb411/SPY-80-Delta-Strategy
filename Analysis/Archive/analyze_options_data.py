"""Analyze the Data for Claude.xlsx options transaction log."""
import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook("Data for Claude.xlsx", data_only=True)
ws = wb["transactions"]

option_trades = []
for row in ws.iter_rows(min_row=2, values_only=True):
    (trade_date, post_date, settle_date, acct_name, acct_num, acct_type,
     txn_type, desc, cusip, ticker, sec_type, currency, price_usd,
     price_local, quantity, gl_short, gl_short_local, gl_long, gl_long_local,
     amount_usd, amount_local, income, income_local, balance, comm,
     comm_local, tran_code, tran_desc, broker, check_num, tax) = row
    if sec_type == "Option":
        option_trades.append({
            "date": trade_date,
            "type": txn_type,
            "ticker": ticker,
            "desc": str(desc or ""),
            "price": float(price_usd) if price_usd else 0,
            "qty": int(quantity) if quantity else 0,
            "amount": float(amount_usd) if amount_usd else 0,
        })

# Group by contract
by_contract = defaultdict(list)
for t in option_trades:
    by_contract[t["ticker"]].append(t)

print("=" * 130)
print("OPTION TRADE P&L BY CONTRACT")
print("=" * 130)
hdr = (f"{'Contract':<35} | {'Open Date':<12} | {'Open Px':>8} | "
       f"{'Close Date':<12} | {'Close Px':>8} | {'Qty':>5} | "
       f"{'P&L':>12} | {'Status':<8}")
print(hdr)
print("-" * 130)

total_realized = 0
open_positions = []

for contract in sorted(by_contract.keys()):
    trades = sorted(by_contract[contract], key=lambda x: x["date"])
    buys = [t for t in trades if t["type"] == "Buy"]
    sells = [t for t in trades if t["type"] == "Sell"]
    journals = [t for t in trades if t["type"] == "Journal"]

    if buys and sells:
        buy = buys[0]
        sell = sells[0]
        qty = abs(buy["qty"])
        pnl = (sell["price"] - buy["price"]) * qty * 100
        total_realized += pnl
        status = "CLOSED"
        print(f"{contract:<35} | {buy['date'].strftime('%Y-%m-%d'):<12} | "
              f"${buy['price']:>7.2f} | {sell['date'].strftime('%Y-%m-%d'):<12} | "
              f"${sell['price']:>7.2f} | {qty:>5} | ${pnl:>+10,.0f} | {status}")
        # Check for multiple buys/sells (partial closes, rolls)
        if len(buys) > 1 or len(sells) > 1:
            print(f"  ** Multiple legs: {len(buys)} buys, {len(sells)} sells")
            for b in buys:
                print(f"     Buy  {b['date'].strftime('%Y-%m-%d')}  qty={b['qty']}  px=${b['price']:.2f}")
            for s in sells:
                print(f"     Sell {s['date'].strftime('%Y-%m-%d')}  qty={s['qty']}  px=${s['price']:.2f}")
    elif buys and journals:
        buy = buys[0]
        qty = abs(buy["qty"])
        pnl = -buy["price"] * qty * 100
        total_realized += pnl
        status = "EXPIRED"
        print(f"{contract:<35} | {buy['date'].strftime('%Y-%m-%d'):<12} | "
              f"${buy['price']:>7.2f} | {journals[0]['date'].strftime('%Y-%m-%d'):<12} | "
              f"$   0.00 | {qty:>5} | ${pnl:>+10,.0f} | {status}")
    elif buys and not sells:
        buy = buys[0]
        qty = abs(buy["qty"])
        open_positions.append((contract, buy, qty))
        print(f"{contract:<35} | {buy['date'].strftime('%Y-%m-%d'):<12} | "
              f"${buy['price']:>7.2f} | {'---':<12} | {'---':>8} | "
              f"{qty:>5} | {'---':>12} | OPEN")
    elif sells and not buys:
        sell = sells[0]
        qty = abs(sell["qty"])
        print(f"{contract:<35} | {'---':<12} | {'---':>8} | "
              f"{sell['date'].strftime('%Y-%m-%d'):<12} | "
              f"${sell['price']:>7.2f} | {qty:>5} | {'???':>12} | SELL-ONLY")

print("-" * 130)
print(f"Total Realized P&L (closed + expired): ${total_realized:>+,.0f}")
print(f"Open positions: {len(open_positions)}")

if open_positions:
    print("\nOPEN POSITIONS:")
    total_cost = 0
    for contract, buy, qty in open_positions:
        cost = buy["price"] * qty * 100
        total_cost += cost
        print(f"  {contract:<35}  qty={qty:>3}  "
              f"cost_basis=${buy['price']:>.2f}/sh  "
              f"total=${cost:>,.0f}  opened={buy['date'].strftime('%Y-%m-%d')}")
    print(f"  Total capital in open positions: ${total_cost:>,.0f}")

# SGOV
print()
print("=" * 80)
print("SGOV (CASH PARKING) SUMMARY")
print("=" * 80)
sgov_buys = 0
sgov_sells = 0
sgov_divs = 0
sgov_shares = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    ticker = row[9]
    txn_type = row[6]
    amount = float(row[19]) if row[19] else 0
    qty = int(row[14]) if row[14] else 0
    if ticker == "SGOV":
        if txn_type == "Buy":
            sgov_buys += amount
            sgov_shares += qty
        elif txn_type == "Sell":
            sgov_sells += amount
            sgov_shares -= abs(qty)
        elif txn_type == "Dividend":
            sgov_divs += amount

print(f"Total SGOV purchased:  ${abs(sgov_buys):>12,.2f}")
print(f"Total SGOV sold:       ${sgov_sells:>12,.2f}")
print(f"Current SGOV shares:   {sgov_shares:>12}")
print(f"SGOV dividends earned: ${sgov_divs:>12,.2f}")

# Cash flows
print()
print("=" * 80)
print("EXTERNAL CASH FLOWS")
print("=" * 80)
deposits = 0
withdrawals = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    txn_type = row[6]
    amount = float(row[19]) if row[19] else 0
    if txn_type == "BNK":
        if amount > 0:
            deposits += amount
        else:
            withdrawals += amount
print(f"Total deposits:    ${deposits:>12,.2f}")
print(f"Total withdrawals: ${abs(withdrawals):>12,.2f}")
print(f"Net cash flow:     ${deposits + withdrawals:>+12,.2f}")

# Timeline
print()
print("=" * 80)
print("ACCOUNT TIMELINE")
print("=" * 80)
all_dates = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        all_dates.append(row[0])
all_dates.sort()
print(f"First transaction: {all_dates[0].strftime('%Y-%m-%d')}")
print(f"Last transaction:  {all_dates[-1].strftime('%Y-%m-%d')}")
print(f"Total transactions: {len(all_dates)}")
