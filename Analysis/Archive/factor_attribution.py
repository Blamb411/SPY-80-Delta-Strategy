"""
Sector & Factor Attribution Analysis for Alpha Picks
======================================================
Decomposes Alpha Picks returns into:
1. Sector allocation effects
2. Momentum factor
3. Size factor (small vs large cap)
4. Value/Growth tilt
5. Volatility factor

Uses daily price data from Massive API and stock metadata.
For sector classification, we use a hardcoded mapping based on
the actual stocks in the Alpha Picks universe.
"""

import openpyxl
import sqlite3
import os
import math
from collections import defaultdict, OrderedDict
from datetime import datetime, timedelta

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "price_cache.db")
XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "ProQuant History 1_29_2026.xlsx")

# ============================================================
# SECTOR CLASSIFICATION (manual mapping for Alpha Picks stocks)
# ============================================================

SECTOR_MAP = {
    # Technology
    'APP': 'Technology', 'CLS': 'Technology', 'SMCI*': 'Technology',
    'CRDO': 'Technology', 'TWLO': 'Technology', 'OKTA': 'Technology',
    'CRM': 'Technology', 'GOOGL': 'Technology', 'META': 'Technology',
    'HLIT': 'Technology', 'ACLS': 'Technology', 'TGLS': 'Technology',
    'TTMI': 'Technology', 'QTWO': 'Technology', 'MU': 'Technology',
    'LRN': 'Technology', 'ZETA': 'Technology',

    # Industrials
    'STRL': 'Industrials', 'POWL': 'Industrials', 'URI': 'Industrials',
    'TEX': 'Industrials', 'AGX': 'Industrials', 'MOD': 'Industrials',
    'BLBD': 'Industrials', 'DXPE': 'Industrials', 'WLDN': 'Industrials',

    # Consumer Discretionary
    'ANF': 'Consumer Disc', 'EAT': 'Consumer Disc', 'RCL': 'Consumer Disc',
    'CCL': 'Consumer Disc', 'W': 'Consumer Disc', 'GCT': 'Consumer Disc',
    'UBER': 'Consumer Disc', 'EZPW': 'Consumer Disc',

    # Financials
    'SYF': 'Financials', 'ALL': 'Financials', 'WFC': 'Financials',
    'JXN': 'Financials', 'LC': 'Financials', 'MFC': 'Financials',
    'BRK.B': 'Financials', 'RGA': 'Financials',

    # Energy
    'MPC': 'Energy', 'DINO': 'Energy', 'XOM': 'Energy',
    'CVX': 'Energy', 'VLO': 'Energy', 'SU': 'Energy',
    'ARCH': 'Energy', 'COP': 'Energy', 'PARR': 'Energy',
    'AMR': 'Energy',

    # Materials / Mining
    'NUE': 'Materials', 'SSRM': 'Materials', 'KGC': 'Materials',
    'CDE': 'Materials', 'NEM': 'Materials', 'B': 'Materials',

    # Healthcare
    'INCY': 'Healthcare', 'ARQT': 'Healthcare', 'AMPH': 'Healthcare',
    'ATGE': 'Healthcare',

    # Communications
    'TMUS': 'Communications', 'CMCSA': 'Communications',
    'TIGO': 'Communications',

    # Consumer Staples
    'PEP': 'Consumer Staples', 'PPC': 'Consumer Staples',
    'SFM': 'Consumer Staples', 'UNFI': 'Consumer Staples',

    # Real Estate / Infrastructure
    'MHO': 'Homebuilders', 'GRBK': 'Homebuilders',

    # Transportation
    'SKYW': 'Transportation', 'CAAP': 'Transportation',
    'ASC': 'Transportation', 'BXC': 'Transportation',

    # Other / International
    'PERI': 'Technology', 'ITRN': 'Technology',
    'LTHM': 'Materials', 'TA': 'Energy',
    'VISN': 'Technology', 'PYPL': 'Financials',
    'GM': 'Consumer Disc', 'STRL': 'Industrials',
}

# Market cap classification (approximate, based on pick-time market cap)
SIZE_MAP = {
    # Large Cap (>$10B)
    'GOOGL': 'Large', 'META': 'Large', 'BRK.B': 'Large', 'PEP': 'Large',
    'XOM': 'Large', 'CVX': 'Large', 'WFC': 'Large', 'CRM': 'Large',
    'MPC': 'Large', 'TMUS': 'Large', 'ALL': 'Large', 'CMCSA': 'Large',
    'UBER': 'Large', 'NUE': 'Large', 'NEM': 'Large', 'COP': 'Large',
    'PYPL': 'Large', 'MU': 'Large', 'SYF': 'Large', 'RCL': 'Large',
    'URI': 'Large', 'SU': 'Large', 'MFC': 'Large', 'GM': 'Large',
    'INCY': 'Large', 'TWLO': 'Large', 'OKTA': 'Large', 'CCL': 'Large',
    'VLO': 'Large',

    # Mid Cap ($2B-$10B)
    'APP': 'Mid', 'CLS': 'Mid', 'ANF': 'Mid', 'EAT': 'Mid',
    'TEX': 'Mid', 'STRL': 'Mid', 'MOD': 'Mid', 'SFM': 'Mid',
    'SKYW': 'Mid', 'PPC': 'Mid', 'BLBD': 'Mid', 'DINO': 'Mid',
    'RGA': 'Mid', 'ATGE': 'Mid', 'JXN': 'Mid', 'W': 'Mid',
    'KGC': 'Mid', 'CAH': 'Mid', 'CRDO': 'Mid', 'B': 'Mid',
    'TIGO': 'Mid', 'LRN': 'Mid',

    # Small Cap (<$2B)
    'SMCI*': 'Small', 'POWL': 'Small', 'AGX': 'Small', 'HLIT': 'Small',
    'ACLS': 'Small', 'TGLS': 'Small', 'MHO': 'Small', 'GRBK': 'Small',
    'AMPH': 'Small', 'PERI': 'Small', 'ASC': 'Small', 'BXC': 'Small',
    'CAAP': 'Small', 'GCT': 'Small', 'ITRN': 'Small', 'LTHM': 'Small',
    'DXPE': 'Small', 'WLDN': 'Small', 'ARQT': 'Small', 'CDE': 'Small',
    'SSRM': 'Small', 'TTMI': 'Small', 'QTWO': 'Small', 'LC': 'Small',
    'ZETA': 'Small', 'EZPW': 'Small', 'PARR': 'Small', 'UNFI': 'Small',
    'VISN': 'Small', 'ARCH': 'Small', 'AMR': 'Small', 'TA': 'Small',
}


def load_prices():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.execute("SELECT symbol, date, close FROM daily_prices ORDER BY symbol, date")
    prices = defaultdict(dict)
    for symbol, date, close in cur.fetchall():
        prices[symbol][date] = close
    conn.close()
    return prices


def get_trading_dates(prices):
    return sorted(prices.get("SPY", {}).keys())


def load_alpha_picks():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['AlphaPicks']
    picks = []
    seen = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        symbol = row[1]
        date = row[2]
        if not symbol or not date:
            continue
        date_str = date.strftime("%Y-%m-%d") if hasattr(date, 'strftime') else str(date)[:10]
        closed = None
        if row[4] and str(row[4]).strip() not in ('-', 'None', ''):
            closed = row[4].strftime("%Y-%m-%d") if hasattr(row[4], 'strftime') else str(row[4])[:10]
        key = (symbol, date_str)
        if key in seen:
            for p in picks:
                if p['symbol'] == symbol and p['picked'] == date_str:
                    if closed is None and p['closed'] is not None:
                        p['closed'] = None
                    break
            continue
        seen.add(key)
        picks.append({
            'symbol': symbol,
            'picked': date_str,
            'buy_price': float(row[3]) if row[3] else 0.0,
            'closed': closed,
            'return_pct': float(row[6]) if row[6] else 0.0,
            'spy_return': float(row[7]) if row[7] else 0.0,
        })
    picks.sort(key=lambda x: x['picked'])
    return picks


def compute_pre_entry_momentum(sym, entry_date, prices, trading_dates, lookback=60):
    """Compute momentum (return) over lookback days before entry."""
    sym_prices = prices.get(sym, {})
    if not sym_prices:
        return None

    try:
        idx = trading_dates.index(entry_date)
    except ValueError:
        for i, d in enumerate(trading_dates):
            if d >= entry_date:
                idx = i
                break
        else:
            return None

    start_idx = idx - lookback
    if start_idx < 0:
        return None

    start_date = trading_dates[start_idx]
    p_start = sym_prices.get(start_date)
    p_end = sym_prices.get(trading_dates[idx])

    if p_start and p_end and p_start > 0:
        return (p_end / p_start) - 1
    return None


def compute_pre_entry_volatility(sym, entry_date, prices, trading_dates, lookback=20):
    """Compute annualized volatility over lookback days before entry."""
    sym_prices = prices.get(sym, {})
    if not sym_prices:
        return None

    try:
        idx = trading_dates.index(entry_date)
    except ValueError:
        for i, d in enumerate(trading_dates):
            if d >= entry_date:
                idx = i
                break
        else:
            return None

    if idx < lookback + 1:
        return None

    closes = []
    for j in range(idx - lookback, idx + 1):
        p = sym_prices.get(trading_dates[j])
        if p:
            closes.append(p)

    if len(closes) < lookback:
        return None

    rets = [(closes[i] / closes[i-1]) - 1 for i in range(1, len(closes))]
    if not rets:
        return None

    mu = sum(rets) / len(rets)
    var = sum((r - mu)**2 for r in rets) / (len(rets) - 1)
    return math.sqrt(var) * math.sqrt(252)


# ============================================================
# MAIN ANALYSIS
# ============================================================

def main():
    prices = load_prices()
    trading_dates = get_trading_dates(prices)
    picks = load_alpha_picks()

    print("=" * 90)
    print("ALPHA PICKS — SECTOR & FACTOR ATTRIBUTION")
    print("=" * 90)
    print()

    # ================================================================
    # SECTOR ATTRIBUTION
    # ================================================================
    print("=" * 90)
    print("SECTOR ATTRIBUTION")
    print("=" * 90)
    print()

    sector_picks = defaultdict(list)
    for p in picks:
        sector = SECTOR_MAP.get(p['symbol'], 'Unknown')
        sector_picks[sector].append(p)

    print(f"  {'Sector':<18} {'Picks':>6} {'Avg Ret':>10} {'Med Ret':>10} {'Win%':>7} "
          f"{'Avg SPY':>10} {'Excess':>10} {'Beat%':>7}")
    print(f"  {'-'*80}")

    sector_stats = []
    for sector in sorted(sector_picks.keys()):
        sp = sector_picks[sector]
        rets = [p['return_pct'] for p in sp]
        spy_rets = [p['spy_return'] for p in sp]
        n = len(rets)
        avg = sum(rets) / n
        med = sorted(rets)[n // 2]
        avg_spy = sum(spy_rets) / n
        wins = sum(1 for r in rets if r > 0) / n
        beats = sum(1 for r, s in zip(rets, spy_rets) if r > s) / n

        sector_stats.append((sector, n, avg, med, wins, avg_spy, avg - avg_spy, beats))

        print(f"  {sector:<18} {n:>6} {avg:>+9.1%} {med:>+9.1%} {wins:>6.0%} "
              f"{avg_spy:>+9.1%} {avg - avg_spy:>+9.1%} {beats:>6.0%}")

    print()

    # Contribution analysis: how much of total return comes from each sector
    print("  Return Contribution by Sector:")
    total_return_dollars = sum(p['return_pct'] * 1000 for p in picks)

    for sector in sorted(sector_picks.keys(), key=lambda s: -sum(p['return_pct'] for p in sector_picks[s])):
        sp = sector_picks[sector]
        sector_dollars = sum(p['return_pct'] * 1000 for p in sp)
        pct = sector_dollars / total_return_dollars if total_return_dollars != 0 else 0
        print(f"    {sector:<18} ${sector_dollars:>12,.0f} ({pct:>5.1%} of total)")

    print()

    # ================================================================
    # SECTOR PERFORMANCE OVER TIME
    # ================================================================
    print("=" * 90)
    print("SECTOR PERFORMANCE BY YEAR")
    print("=" * 90)
    print()

    years = sorted(set(p['picked'][:4] for p in picks))
    sectors_in_data = sorted(set(SECTOR_MAP.get(p['symbol'], 'Unknown') for p in picks))

    # Header
    print(f"  {'Sector':<18}", end="")
    for year in years:
        print(f"  {year:>10}", end="")
    print()
    print(f"  {'-'*18}", end="")
    for _ in years:
        print(f"  {'-'*10}", end="")
    print()

    for sector in sectors_in_data:
        print(f"  {sector:<18}", end="")
        for year in years:
            yr_picks = [p for p in sector_picks[sector] if p['picked'][:4] == year]
            if yr_picks:
                avg = sum(p['return_pct'] for p in yr_picks) / len(yr_picks)
                print(f"  {avg:>+9.1%}", end="")
            else:
                print(f"  {'—':>10}", end="")
        print()

    print()

    # ================================================================
    # SIZE FACTOR
    # ================================================================
    print("=" * 90)
    print("SIZE FACTOR ATTRIBUTION")
    print("=" * 90)
    print()

    size_picks = defaultdict(list)
    for p in picks:
        size = SIZE_MAP.get(p['symbol'], 'Unknown')
        size_picks[size].append(p)

    print(f"  {'Size':<10} {'Picks':>6} {'Avg Ret':>10} {'Med Ret':>10} {'Win%':>7} "
          f"{'Avg SPY':>10} {'Excess':>10}")
    print(f"  {'-'*65}")

    for size in ['Small', 'Mid', 'Large', 'Unknown']:
        sp = size_picks.get(size, [])
        if not sp:
            continue
        rets = [p['return_pct'] for p in sp]
        spy_rets = [p['spy_return'] for p in sp]
        n = len(rets)
        avg = sum(rets) / n
        med = sorted(rets)[n // 2]
        avg_spy = sum(spy_rets) / n
        wins = sum(1 for r in rets if r > 0) / n

        print(f"  {size:<10} {n:>6} {avg:>+9.1%} {med:>+9.1%} {wins:>6.0%} "
              f"{avg_spy:>+9.1%} {avg - avg_spy:>+9.1%}")

    print()

    # Size by year
    print("  Size Factor by Year (average excess return):")
    print(f"  {'Size':<10}", end="")
    for year in years:
        print(f"  {year:>10}", end="")
    print()
    print(f"  {'-'*10}", end="")
    for _ in years:
        print(f"  {'-'*10}", end="")
    print()

    for size in ['Small', 'Mid', 'Large']:
        print(f"  {size:<10}", end="")
        for year in years:
            yr_picks = [p for p in size_picks.get(size, []) if p['picked'][:4] == year]
            if yr_picks:
                avg_excess = sum(p['return_pct'] - p['spy_return'] for p in yr_picks) / len(yr_picks)
                print(f"  {avg_excess:>+9.1%}", end="")
            else:
                print(f"  {'—':>10}", end="")
        print()

    print()

    # ================================================================
    # MOMENTUM FACTOR
    # ================================================================
    print("=" * 90)
    print("MOMENTUM FACTOR ATTRIBUTION")
    print("=" * 90)
    print()
    print("  Pre-entry momentum: 60-day return before pick date")
    print()

    momentum_data = []
    for p in picks:
        entry_td = None
        for d in trading_dates:
            if d >= p['picked']:
                entry_td = d
                break
        if not entry_td:
            continue

        mom = compute_pre_entry_momentum(p['symbol'], entry_td, prices, trading_dates, 60)
        if mom is not None:
            momentum_data.append({
                'symbol': p['symbol'],
                'picked': p['picked'],
                'momentum': mom,
                'return': p['return_pct'],
                'excess': p['return_pct'] - p['spy_return'],
            })

    if momentum_data:
        # Sort by momentum and split into quintiles
        sorted_by_mom = sorted(momentum_data, key=lambda x: x['momentum'])
        n = len(sorted_by_mom)
        q_size = n // 3

        quintiles = {
            'Low Mom (bottom third)': sorted_by_mom[:q_size],
            'Mid Mom (middle third)': sorted_by_mom[q_size:2*q_size],
            'High Mom (top third)': sorted_by_mom[2*q_size:],
        }

        print(f"  {'Momentum Tercile':<28} {'N':>5} {'Avg Pre-Mom':>12} {'Avg Ret':>10} "
              f"{'Avg Excess':>12} {'Win%':>7}")
        print(f"  {'-'*78}")

        for label, group in quintiles.items():
            avg_mom = sum(d['momentum'] for d in group) / len(group)
            avg_ret = sum(d['return'] for d in group) / len(group)
            avg_exc = sum(d['excess'] for d in group) / len(group)
            wins = sum(1 for d in group if d['return'] > 0) / len(group)

            print(f"  {label:<28} {len(group):>5} {avg_mom:>+11.1%} {avg_ret:>+9.1%} "
                  f"{avg_exc:>+11.1%} {wins:>6.0%}")

        print()

        # Correlation between pre-entry momentum and subsequent return
        moms = [d['momentum'] for d in momentum_data]
        rets = [d['return'] for d in momentum_data]
        n = len(moms)
        mean_m = sum(moms) / n
        mean_r = sum(rets) / n
        cov = sum((m - mean_m) * (r - mean_r) for m, r in zip(moms, rets)) / (n - 1)
        std_m = math.sqrt(sum((m - mean_m)**2 for m in moms) / (n - 1))
        std_r = math.sqrt(sum((r - mean_r)**2 for r in rets) / (n - 1))
        corr = cov / (std_m * std_r) if std_m > 0 and std_r > 0 else 0

        print(f"  Correlation (pre-entry momentum vs subsequent return): {corr:.3f}")
        if corr > 0.1:
            print(f"  -> Positive: momentum stocks tend to keep winning")
        elif corr < -0.1:
            print(f"  -> Negative: high-momentum entries tend to underperform (mean reversion)")
        else:
            print(f"  -> Weak/no relationship between pre-entry momentum and outcome")

    print()

    # ================================================================
    # VOLATILITY FACTOR
    # ================================================================
    print("=" * 90)
    print("VOLATILITY FACTOR ATTRIBUTION")
    print("=" * 90)
    print()
    print("  Pre-entry volatility: 20-day annualized volatility before pick")
    print()

    vol_data = []
    for p in picks:
        entry_td = None
        for d in trading_dates:
            if d >= p['picked']:
                entry_td = d
                break
        if not entry_td:
            continue

        vol = compute_pre_entry_volatility(p['symbol'], entry_td, prices, trading_dates, 20)
        if vol is not None:
            vol_data.append({
                'symbol': p['symbol'],
                'picked': p['picked'],
                'volatility': vol,
                'return': p['return_pct'],
                'excess': p['return_pct'] - p['spy_return'],
            })

    if vol_data:
        sorted_by_vol = sorted(vol_data, key=lambda x: x['volatility'])
        n = len(sorted_by_vol)
        q_size = n // 3

        vol_terciles = {
            'Low Vol (bottom third)': sorted_by_vol[:q_size],
            'Mid Vol (middle third)': sorted_by_vol[q_size:2*q_size],
            'High Vol (top third)': sorted_by_vol[2*q_size:],
        }

        print(f"  {'Volatility Tercile':<28} {'N':>5} {'Avg Vol':>10} {'Avg Ret':>10} "
              f"{'Avg Excess':>12} {'Win%':>7}")
        print(f"  {'-'*76}")

        for label, group in vol_terciles.items():
            avg_vol = sum(d['volatility'] for d in group) / len(group)
            avg_ret = sum(d['return'] for d in group) / len(group)
            avg_exc = sum(d['excess'] for d in group) / len(group)
            wins = sum(1 for d in group if d['return'] > 0) / len(group)

            print(f"  {label:<28} {len(group):>5} {avg_vol:>9.0%} {avg_ret:>+9.1%} "
                  f"{avg_exc:>+11.1%} {wins:>6.0%}")

        print()

        # Correlation
        vols = [d['volatility'] for d in vol_data]
        rets = [d['return'] for d in vol_data]
        n = len(vols)
        mean_v = sum(vols) / n
        mean_r = sum(rets) / n
        cov = sum((v - mean_v) * (r - mean_r) for v, r in zip(vols, rets)) / (n - 1)
        std_v = math.sqrt(sum((v - mean_v)**2 for v in vols) / (n - 1))
        std_r = math.sqrt(sum((r - mean_r)**2 for r in rets) / (n - 1))
        corr = cov / (std_v * std_r) if std_v > 0 and std_r > 0 else 0

        print(f"  Correlation (pre-entry vol vs subsequent return): {corr:.3f}")
        if corr > 0.1:
            print(f"  -> Higher vol stocks tend to return more (risk premium)")
        elif corr < -0.1:
            print(f"  -> Lower vol stocks tend to return more (low-vol anomaly)")
        else:
            print(f"  -> Weak/no relationship")

    print()

    # ================================================================
    # MULTI-FACTOR SUMMARY
    # ================================================================
    print("=" * 90)
    print("MULTI-FACTOR SUMMARY")
    print("=" * 90)
    print()
    print("  Factor                   Avg Excess (top tercile)    Key Finding")
    print("  " + "-" * 72)

    # Sector
    best_sector = max(sector_stats, key=lambda x: x[6])  # highest excess
    worst_sector = min(sector_stats, key=lambda x: x[6])
    print(f"  Sector                   Best: {best_sector[0]:<14} {best_sector[6]:>+.1%}   "
          f"Worst: {worst_sector[0]:<14} {worst_sector[6]:>+.1%}")

    # Size
    for size in ['Small', 'Mid', 'Large']:
        sp = size_picks.get(size, [])
        if sp:
            avg_excess = sum(p['return_pct'] - p['spy_return'] for p in sp) / len(sp)
            if size == 'Small':
                print(f"  Size (Small Cap)         Excess: {avg_excess:>+.1%}")

    # Momentum
    if momentum_data:
        high_mom = sorted(momentum_data, key=lambda x: x['momentum'])[2*(len(momentum_data)//3):]
        avg_exc_high = sum(d['excess'] for d in high_mom) / len(high_mom)
        print(f"  Momentum (High)          Excess: {avg_exc_high:>+.1%}")

    # Volatility
    if vol_data:
        high_vol = sorted(vol_data, key=lambda x: x['volatility'])[2*(len(vol_data)//3):]
        avg_exc_high_vol = sum(d['excess'] for d in high_vol) / len(high_vol)
        print(f"  Volatility (High)        Excess: {avg_exc_high_vol:>+.1%}")

    print()
    print("  Implications for building our own model:")
    print("  - Which factors drive the most excess return?")
    print("  - Can we replicate the selection by screening on these factors?")
    print("  - Which factors are decaying vs persistent?")

    print()
    print("=" * 90)
    print("ANALYSIS COMPLETE")
    print("=" * 90)


if __name__ == "__main__":
    main()
