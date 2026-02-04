"""
Fetch Daily Prices from Massive API
=====================================
Downloads daily OHLCV data for all stocks in the Alpha Picks and
ProQuant portfolios, plus SPY as benchmark.

Stores data in a local SQLite database to avoid redundant API calls.
"""

import os
import sys
import time
import sqlite3
import json
from datetime import datetime, timedelta

# API key
API_KEY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "..", "Massive backtesting", "api_key.txt")
with open(API_KEY_PATH) as f:
    API_KEY = f.read().strip()

from massive import RESTClient
client = RESTClient(api_key=API_KEY)

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "price_cache.db")

# Ticker corrections
TICKER_MAP = {
    "SMCI*": "SMCI",
    "ARCH": "ARCH",  # Arch Resources - keep as is
}


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS daily_prices (
            symbol TEXT NOT NULL,
            date TEXT NOT NULL,
            open REAL,
            high REAL,
            low REAL,
            close REAL,
            volume REAL,
            PRIMARY KEY (symbol, date)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS fetch_log (
            symbol TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            fetched_at TEXT NOT NULL,
            bar_count INTEGER,
            PRIMARY KEY (symbol, start_date, end_date)
        )
    """)
    conn.commit()
    return conn


def is_fetched(conn, symbol, start_date, end_date):
    """Check if we've already fetched this symbol for this range."""
    cur = conn.execute(
        "SELECT 1 FROM fetch_log WHERE symbol = ? AND start_date <= ? AND end_date >= ?",
        (symbol, start_date, end_date)
    )
    return cur.fetchone() is not None


def fetch_symbol(conn, symbol, start_date, end_date):
    """Fetch daily bars for a symbol and cache in DB."""
    api_symbol = TICKER_MAP.get(symbol, symbol)

    if is_fetched(conn, symbol, start_date, end_date):
        return 0

    print(f"  Fetching {symbol} ({start_date} to {end_date}) ...", end=" ", flush=True)

    try:
        aggs = list(client.list_aggs(
            ticker=api_symbol,
            multiplier=1,
            timespan="day",
            from_=start_date,
            to=end_date,
            limit=50000,
        ))
    except Exception as e:
        print(f"ERROR: {e}")
        return 0

    if not aggs:
        print("no data")
        # Still log to avoid retrying
        conn.execute(
            "INSERT OR REPLACE INTO fetch_log VALUES (?, ?, ?, ?, ?)",
            (symbol, start_date, end_date, datetime.utcnow().isoformat(), 0)
        )
        conn.commit()
        return 0

    count = 0
    for bar in aggs:
        date_str = datetime.fromtimestamp(bar.timestamp / 1000).strftime("%Y-%m-%d")
        conn.execute(
            "INSERT OR REPLACE INTO daily_prices VALUES (?, ?, ?, ?, ?, ?, ?)",
            (symbol, date_str, bar.open, bar.high, bar.low, bar.close, bar.volume)
        )
        count += 1

    conn.execute(
        "INSERT OR REPLACE INTO fetch_log VALUES (?, ?, ?, ?, ?)",
        (symbol, start_date, end_date, datetime.utcnow().isoformat(), count)
    )
    conn.commit()

    print(f"{count} bars")
    time.sleep(0.15)  # rate limit
    return count


def get_prices(conn, symbol, start_date=None, end_date=None):
    """Get cached prices for a symbol."""
    query = "SELECT date, close FROM daily_prices WHERE symbol = ?"
    params = [symbol]
    if start_date:
        query += " AND date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
    query += " ORDER BY date"

    cur = conn.execute(query, params)
    return cur.fetchall()


def main():
    import openpyxl

    conn = init_db()
    wb = openpyxl.load_workbook('ProQuant History 1_29_2026.xlsx', data_only=True)

    # Collect all symbols
    pq_symbols = set()
    ws = wb['ProQuant']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0]:
            pq_symbols.add(row[0])

    ap_symbols = set()
    ws = wb['AlphaPicks']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] and row[2]:
            ap_symbols.add(row[1])

    all_symbols = sorted(pq_symbols | ap_symbols | {"SPY"})
    print(f"Total symbols to fetch: {len(all_symbols)}")

    # Date range: from well before first Alpha Pick to today
    start_date = "2022-06-01"
    end_date = "2026-01-29"

    total_bars = 0
    fetched = 0
    skipped = 0

    for i, symbol in enumerate(all_symbols, 1):
        mapped = TICKER_MAP.get(symbol, symbol)
        if is_fetched(conn, symbol, start_date, end_date):
            skipped += 1
            continue

        bars = fetch_symbol(conn, symbol, start_date, end_date)
        total_bars += bars
        fetched += 1

        if i % 20 == 0:
            print(f"  Progress: {i}/{len(all_symbols)} symbols")

    print()
    print(f"Done. Fetched {fetched} symbols ({total_bars} bars). Skipped {skipped} (cached).")

    # Quick verification
    spy_prices = get_prices(conn, "SPY", "2022-07-01", "2026-01-29")
    print(f"SPY price check: {len(spy_prices)} trading days")
    if spy_prices:
        print(f"  First: {spy_prices[0][0]} ${spy_prices[0][1]:.2f}")
        print(f"  Last:  {spy_prices[-1][0]} ${spy_prices[-1][1]:.2f}")

    # Check for symbols with no data
    no_data = []
    for symbol in all_symbols:
        prices = get_prices(conn, symbol)
        if not prices:
            no_data.append(symbol)
    if no_data:
        print(f"\nSymbols with no price data: {', '.join(no_data)}")

    conn.close()


if __name__ == "__main__":
    main()
