"""
Pick-Day Effect Analysis
=========================
Measures the stock price behavior around Alpha Picks announcement dates
and ProQuant portfolio additions to quantify the "announcement effect."

Alpha Picks: disclosed at noon on 1st/15th, portfolio enters next day
ProQuant: enters at Friday close, disclosed Monday morning

Key questions:
1. Is there a measurable pop on the announcement day?
2. How much return accrues in the first N days vs the remainder?
3. Does the timing gap help or hurt investors vs stated returns?
"""

import openpyxl
import sqlite3
import os
import math
from collections import defaultdict
from datetime import datetime, timedelta

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "price_cache.db")
XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "ProQuant History 1_29_2026.xlsx")


def load_prices():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.execute("SELECT symbol, date, open, high, low, close FROM daily_prices ORDER BY symbol, date")
    prices = defaultdict(dict)  # symbol -> {date: {open, high, low, close}}
    for symbol, date, o, h, l, c in cur.fetchall():
        prices[symbol][date] = {'open': o, 'high': h, 'low': l, 'close': c}
    conn.close()
    return prices


def get_trading_dates(prices):
    """Get sorted trading dates from SPY."""
    return sorted(prices.get("SPY", {}).keys())


def find_nearest_date(trading_dates, target, direction="on_or_after"):
    """Find the nearest trading date to target."""
    for d in trading_dates:
        if direction == "on_or_after" and d >= target:
            return d
        if direction == "on_or_before" and d <= target:
            pass  # keep going
    # For on_or_before, return the last one we saw
    if direction == "on_or_before":
        result = None
        for d in trading_dates:
            if d <= target:
                result = d
            else:
                break
        return result
    return None


def find_date_offset(trading_dates, base_date, offset):
    """Find the trading date that is `offset` trading days from base_date."""
    try:
        idx = trading_dates.index(base_date)
    except ValueError:
        # Find nearest
        nearest = find_nearest_date(trading_dates, base_date, "on_or_after")
        if nearest is None:
            return None
        idx = trading_dates.index(nearest)

    target_idx = idx + offset
    if 0 <= target_idx < len(trading_dates):
        return trading_dates[target_idx]
    return None


def load_alpha_picks():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['AlphaPicks']
    picks = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        symbol = row[1]
        date = row[2]
        if not symbol or not date:
            continue
        date_str = date.strftime("%Y-%m-%d") if hasattr(date, 'strftime') else str(date)[:10]
        closed = None
        if row[4] and str(row[4]).strip() not in ('-', 'None', ''):
            closed = row[4].strftime("%Y-%m-%d") if hasattr(row[4], 'strftime') else str(row[4])[:10]
        picks.append({
            'symbol': symbol,
            'picked': date_str,
            'buy_price': float(row[3]) if row[3] else 0.0,
            'closed': closed,
            'return_pct': float(row[6]) if row[6] else 0.0,
        })
    picks.sort(key=lambda x: x['picked'])
    return picks


def load_proquant():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['ProQuant']
    weekly = defaultdict(list)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        symbol, date, action, start_w, new_w, change_w, price = row
        if not symbol or not date:
            continue
        date_str = date.strftime("%Y-%m-%d") if hasattr(date, 'strftime') else str(date)[:10]
        weekly[date_str].append({
            'symbol': symbol,
            'action': action,
            'price': float(price) if price else 0.0,
        })
    return dict(sorted(weekly.items()))


# ============================================================
# ALPHA PICKS: TIMING VERIFICATION
# ============================================================

def verify_alpha_picks_timing(picks, prices, trading_dates):
    """Verify when the purchase price matches: pick day close, next day open/close."""
    print("=" * 90)
    print("ALPHA PICKS — ENTRY TIMING VERIFICATION")
    print("=" * 90)
    print()
    print("  Hypothesis: picks disclosed at noon on 1st/15th, portfolio enters next day")
    print("  Test: compare SA buy price to close on pick day, open/close on next day")
    print()

    matches = {'pick_day_close': 0, 'next_day_open': 0, 'next_day_close': 0,
               'prev_day_close': 0, 'other': 0, 'no_data': 0}
    details = []

    for p in picks:
        sym = p['symbol']
        pick_date = p['picked']
        buy_price = p['buy_price']

        sym_prices = prices.get(sym, {})
        if not sym_prices:
            matches['no_data'] += 1
            continue

        # Get prices around pick date
        pick_td = find_nearest_date(trading_dates, pick_date, "on_or_after")
        prev_td = find_date_offset(trading_dates, pick_date, -1)
        next_td = find_date_offset(trading_dates, pick_date, 1)

        pick_data = sym_prices.get(pick_td, {})
        prev_data = sym_prices.get(prev_td, {}) if prev_td else {}
        next_data = sym_prices.get(next_td, {}) if next_td else {}

        pick_close = pick_data.get('close')
        next_open = next_data.get('open')
        next_close = next_data.get('close')
        prev_close = prev_data.get('close')

        # Check which price matches (within 1%)
        matched = "other"
        if pick_close and abs(buy_price - pick_close) / buy_price < 0.01:
            matched = "pick_day_close"
        elif next_open and abs(buy_price - next_open) / buy_price < 0.01:
            matched = "next_day_open"
        elif next_close and abs(buy_price - next_close) / buy_price < 0.01:
            matched = "next_day_close"
        elif prev_close and abs(buy_price - prev_close) / buy_price < 0.01:
            matched = "prev_day_close"

        matches[matched] += 1

        details.append({
            'symbol': sym,
            'pick_date': pick_date,
            'buy_price': buy_price,
            'pick_close': pick_close,
            'next_open': next_open,
            'next_close': next_close,
            'prev_close': prev_close,
            'matched': matched,
        })

    total = sum(matches.values())
    print(f"  Price match analysis ({total} picks):")
    for key, count in sorted(matches.items(), key=lambda x: -x[1]):
        print(f"    {key:<20} {count:>4} ({count/total:.1%})")
    print()

    # Show some examples of "other" matches
    others = [d for d in details if d['matched'] == 'other']
    if others:
        print(f"  Examples of unmatched prices (first 10):")
        print(f"  {'Symbol':<8} {'Pick Date':>12} {'Buy$':>10} {'PickClose':>10} {'NextOpen':>10} {'NextClose':>10} {'PrevClose':>10}")
        print(f"  {'-'*72}")
        for d in others[:10]:
            print(f"  {d['symbol']:<8} {d['pick_date']:>12} "
                  f"${d['buy_price']:>8.2f} "
                  f"${d['pick_close'] or 0:>8.2f} "
                  f"${d['next_open'] or 0:>8.2f} "
                  f"${d['next_close'] or 0:>8.2f} "
                  f"${d['prev_close'] or 0:>8.2f}")
    print()

    return details


# ============================================================
# ALPHA PICKS: PICK-DAY ALPHA MEASUREMENT
# ============================================================

def measure_pick_day_alpha(picks, prices, trading_dates):
    """Measure returns at various intervals after the pick announcement."""
    print("=" * 90)
    print("ALPHA PICKS — PICK-DAY ALPHA ANALYSIS")
    print("=" * 90)
    print()
    print("  Measures average return at various intervals after pick announcement.")
    print("  T=0 is the pick date (1st or 15th). Returns measured from T-1 close.")
    print()

    # Windows to measure (in trading days from pick date)
    windows = [
        ("T-1 to T close", -1, 0),         # Day before to pick day close
        ("T-1 to T+1 close", -1, 1),       # Overnight after announcement
        ("T close to T+1 close", 0, 1),     # Next day (SA entry day)
        ("T+1 to T+5", 1, 5),              # First week after entry
        ("T+1 to T+10", 1, 10),            # First two weeks
        ("T+1 to T+20", 1, 20),            # First month
        ("T+1 to T+60", 1, 60),            # First quarter
        ("T+1 to T+120", 1, 120),          # First half year
        ("T+1 to T+252", 1, 252),          # First year
    ]

    results = {w[0]: [] for w in windows}
    spy_results = {w[0]: [] for w in windows}

    for p in picks:
        sym = p['symbol']
        pick_date = p['picked']

        sym_prices = prices.get(sym, {})
        spy_prices = prices.get("SPY", {})
        if not sym_prices:
            continue

        pick_td = find_nearest_date(trading_dates, pick_date, "on_or_after")
        if not pick_td:
            continue

        for label, start_offset, end_offset in windows:
            start_date = find_date_offset(trading_dates, pick_td, start_offset)
            end_date = find_date_offset(trading_dates, pick_td, end_offset)

            if not start_date or not end_date:
                continue

            start_price = sym_prices.get(start_date, {}).get('close')
            end_price = sym_prices.get(end_date, {}).get('close')

            spy_start = spy_prices.get(start_date, {}).get('close')
            spy_end = spy_prices.get(end_date, {}).get('close')

            if start_price and end_price and start_price > 0:
                ret = (end_price / start_price) - 1
                results[label].append(ret)

            if spy_start and spy_end and spy_start > 0:
                spy_ret = (spy_end / spy_start) - 1
                spy_results[label].append(spy_ret)

    print(f"  {'Window':<25} {'N':>5} {'Avg Ret':>10} {'Med Ret':>10} {'Avg SPY':>10} {'Excess':>10} {'%>0':>7} {'%>SPY':>7}")
    print(f"  {'-'*85}")

    for label, _, _ in windows:
        rets = results[label]
        spy_rets = spy_results[label]

        if not rets:
            continue

        n = len(rets)
        avg = sum(rets) / n
        med = sorted(rets)[n // 2]
        avg_spy = sum(spy_rets) / len(spy_rets) if spy_rets else 0
        excess = avg - avg_spy
        pct_pos = sum(1 for r in rets if r > 0) / n
        pct_beat = sum(1 for r, s in zip(rets, spy_rets) if r > s) / n if spy_rets else 0

        marker = ""
        if label == "T-1 to T close":
            marker = "  ** announcement day"
        elif label == "T close to T+1 close":
            marker = "  ** SA entry day"

        print(f"  {label:<25} {n:>5} {avg:>+9.2%} {med:>+9.2%} {avg_spy:>+9.2%} {excess:>+9.2%} {pct_pos:>6.0%} {pct_beat:>6.0%}{marker}")

    print()

    # Breakdown: announcement day return distribution
    ann_day_rets = results.get("T-1 to T close", [])
    if ann_day_rets:
        print("  Announcement Day (T-1 to T close) Distribution:")
        buckets = [(-1, -0.05), (-0.05, -0.02), (-0.02, 0), (0, 0.02), (0.02, 0.05), (0.05, 0.10), (0.10, 1)]
        for lo, hi in buckets:
            count = sum(1 for r in ann_day_rets if lo <= r < hi)
            pct = count / len(ann_day_rets)
            bar = '#' * int(pct * 40)
            print(f"    {lo:>+6.0%} to {hi:>+5.0%}: {count:>3} ({pct:>5.1%}) {bar}")
        print()

    # Per-year pick-day alpha
    print("  Announcement Day Alpha by Year:")
    year_alphas = defaultdict(list)
    year_spy = defaultdict(list)

    for i, p in enumerate(picks):
        year = p['picked'][:4]
        ann_rets = results.get("T-1 to T close", [])
        spy_rets = spy_results.get("T-1 to T close", [])
        if i < len(ann_rets):
            year_alphas[year].append(ann_rets[i])
        if i < len(spy_rets):
            year_spy[year].append(spy_rets[i])

    for year in sorted(year_alphas.keys()):
        rets = year_alphas[year]
        spy_r = year_spy.get(year, [])
        avg = sum(rets) / len(rets)
        avg_s = sum(spy_r) / len(spy_r) if spy_r else 0
        print(f"    {year}: {len(rets):>3} picks  Avg announcement return: {avg:>+.2%}  SPY: {avg_s:>+.2%}  Excess: {avg - avg_s:>+.2%}")
    print()

    return results


# ============================================================
# ALPHA PICKS: INVESTOR TIMING ANALYSIS
# ============================================================

def investor_timing_analysis(picks, prices, trading_dates):
    """Compare returns for different entry timings."""
    print("=" * 90)
    print("ALPHA PICKS — INVESTOR TIMING COMPARISON")
    print("=" * 90)
    print()
    print("  Compares total pick returns under different entry timing:")
    print("    1. SA stated: enters at stated buy price (next day after announcement)")
    print("    2. Noon buyer: enters at pick-day close (proxy for noon buyer)")
    print("    3. Pre-market buyer: enters at next-day open")
    print("    4. Late buyer: enters at T+5 close")
    print()

    scenarios = {
        'sa_stated': [],    # SA's stated buy price
        'pick_close': [],   # Buy at pick day close
        'next_open': [],    # Buy at next day open
        'next_close': [],   # Buy at next day close
        't5_close': [],     # Buy at T+5
    }

    snapshot_date = "2026-01-29"

    for p in picks:
        sym = p['symbol']
        pick_date = p['picked']
        exit_date = p['closed'] or snapshot_date

        sym_prices = prices.get(sym, {})
        if not sym_prices:
            continue

        # Get exit price
        exit_td = find_nearest_date(trading_dates, exit_date, "on_or_before")
        if not exit_td:
            continue
        exit_price = sym_prices.get(exit_td, {}).get('close')
        if not exit_price:
            continue

        # Get various entry prices
        pick_td = find_nearest_date(trading_dates, pick_date, "on_or_after")
        next_td = find_date_offset(trading_dates, pick_td, 1) if pick_td else None
        t5_td = find_date_offset(trading_dates, pick_td, 5) if pick_td else None

        pick_close = sym_prices.get(pick_td, {}).get('close') if pick_td else None
        next_open = sym_prices.get(next_td, {}).get('open') if next_td else None
        next_close = sym_prices.get(next_td, {}).get('close') if next_td else None
        t5_close = sym_prices.get(t5_td, {}).get('close') if t5_td else None

        buy_price = p['buy_price']

        if buy_price > 0:
            scenarios['sa_stated'].append((exit_price / buy_price) - 1)
        if pick_close and pick_close > 0:
            scenarios['pick_close'].append((exit_price / pick_close) - 1)
        if next_open and next_open > 0:
            scenarios['next_open'].append((exit_price / next_open) - 1)
        if next_close and next_close > 0:
            scenarios['next_close'].append((exit_price / next_close) - 1)
        if t5_close and t5_close > 0:
            scenarios['t5_close'].append((exit_price / t5_close) - 1)

    labels = {
        'sa_stated': 'SA stated price',
        'pick_close': 'Pick day close (noon buyer proxy)',
        'next_open': 'Next day open',
        'next_close': 'Next day close',
        't5_close': 'T+5 close (wait a week)',
    }

    print(f"  {'Entry Timing':<35} {'N':>5} {'Avg Ret':>10} {'Med Ret':>10} {'Total$':>12}")
    print(f"  {'-'*75}")

    for key in ['pick_close', 'sa_stated', 'next_open', 'next_close', 't5_close']:
        rets = scenarios[key]
        if not rets:
            continue
        n = len(rets)
        avg = sum(rets) / n
        med = sorted(rets)[n // 2]
        # Simulate $1000 per pick
        total = sum(1000 * r for r in rets)

        marker = ""
        if key == 'sa_stated':
            marker = " <- SA portfolio"
        elif key == 'pick_close':
            marker = " <- best for investor"

        print(f"  {labels[key]:<35} {n:>5} {avg:>+9.2%} {med:>+9.2%} ${total:>10,.0f}{marker}")

    print()

    # Cost of delay
    if scenarios['pick_close'] and scenarios['sa_stated']:
        pick_avg = sum(scenarios['pick_close']) / len(scenarios['pick_close'])
        sa_avg = sum(scenarios['sa_stated']) / len(scenarios['sa_stated'])
        diff = pick_avg - sa_avg
        total_diff = sum(1000 * r for r in scenarios['pick_close']) - sum(1000 * r for r in scenarios['sa_stated'])
        print(f"  Cost of 1-day delay (pick close vs SA stated):")
        print(f"    Per-pick average: {diff:>+.2%}")
        print(f"    Total on $1,000/pick portfolio: ${total_diff:>+,.0f}")
    print()


# ============================================================
# PROQUANT: TIMING VERIFICATION
# ============================================================

def verify_proquant_timing(weekly_data, prices, trading_dates):
    """Verify ProQuant entry pricing: Friday close vs Monday open."""
    print("=" * 90)
    print("PROQUANT — ENTRY TIMING VERIFICATION")
    print("=" * 90)
    print()
    print("  Hypothesis: portfolio enters at Friday close, disclosed Monday morning")
    print("  Test: compare portfolio price to Friday close and Monday open/close")
    print()

    # ProQuant dates are the rebalance dates. Check what day of week they fall on.
    day_of_week_counts = defaultdict(int)
    for date_str in weekly_data.keys():
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        day_of_week_counts[dt.strftime("%A")] += 1

    print(f"  Rebalance dates by day of week:")
    for day, count in sorted(day_of_week_counts.items(),
                              key=lambda x: ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'].index(x[0])):
        print(f"    {day}: {count}")
    print()

    # For Buy actions, compare portfolio price to nearby market prices
    matches = {'exact_date_close': 0, 'prev_friday_close': 0, 'monday_open': 0,
               'monday_close': 0, 'other': 0, 'no_data': 0}

    buy_details = []

    for date_str, entries in weekly_data.items():
        buys = [e for e in entries if e['action'] == 'Buy']

        for e in buys:
            sym = e['symbol']
            portfolio_price = e['price']
            sym_prices = prices.get(sym, {})

            if not sym_prices:
                matches['no_data'] += 1
                continue

            dt = datetime.strptime(date_str, "%Y-%m-%d")

            # Find the date itself, previous Friday, and Monday
            exact_data = sym_prices.get(date_str, {})
            exact_close = exact_data.get('close')

            # Previous Friday
            days_back = (dt.weekday() - 4) % 7  # distance to prev Friday
            if days_back == 0 and dt.weekday() != 4:
                days_back = 7
            prev_friday = (dt - timedelta(days=days_back)).strftime("%Y-%m-%d")
            fri_data = sym_prices.get(prev_friday, {})
            fri_close = fri_data.get('close')

            # Next Monday (if rebalance is on a weekend or Monday itself)
            if dt.weekday() == 0:  # Monday
                mon_date = date_str
            else:
                days_fwd = (7 - dt.weekday()) % 7
                if days_fwd == 0:
                    days_fwd = 7
                mon_date = (dt + timedelta(days=days_fwd)).strftime("%Y-%m-%d")

            mon_data = sym_prices.get(mon_date, {})
            mon_open = mon_data.get('open')
            mon_close = mon_data.get('close')

            matched = "other"
            if exact_close and abs(portfolio_price - exact_close) / portfolio_price < 0.01:
                matched = "exact_date_close"
            elif fri_close and abs(portfolio_price - fri_close) / portfolio_price < 0.01:
                matched = "prev_friday_close"
            elif mon_open and abs(portfolio_price - mon_open) / portfolio_price < 0.01:
                matched = "monday_open"
            elif mon_close and abs(portfolio_price - mon_close) / portfolio_price < 0.01:
                matched = "monday_close"

            matches[matched] += 1

            buy_details.append({
                'symbol': sym,
                'date': date_str,
                'portfolio_price': portfolio_price,
                'exact_close': exact_close,
                'fri_close': fri_close,
                'mon_open': mon_open,
                'mon_close': mon_close,
                'matched': matched,
            })

    total = sum(matches.values())
    print(f"  Buy price match analysis ({total} buys):")
    for key, count in sorted(matches.items(), key=lambda x: -x[1]):
        print(f"    {key:<20} {count:>4} ({count/total:.1%})" if total > 0 else f"    {key:<20} {count:>4}")
    print()

    # Show examples
    others = [d for d in buy_details if d['matched'] == 'other']
    if others:
        print(f"  Examples of unmatched Buy prices (first 10):")
        print(f"  {'Sym':<8} {'Date':>12} {'Port$':>10} {'ExactCl':>10} {'FriCl':>10} {'MonOpen':>10} {'MonCl':>10}")
        print(f"  {'-'*72}")
        for d in others[:10]:
            print(f"  {d['symbol']:<8} {d['date']:>12} "
                  f"${d['portfolio_price']:>8.2f} "
                  f"${d['exact_close'] or 0:>8.2f} "
                  f"${d['fri_close'] or 0:>8.2f} "
                  f"${d['mon_open'] or 0:>8.2f} "
                  f"${d['mon_close'] or 0:>8.2f}")
    print()

    # Measure the gap between Friday close and Monday open for Buy stocks
    gaps = []
    for d in buy_details:
        if d['fri_close'] and d['mon_open'] and d['fri_close'] > 0:
            gap = (d['mon_open'] / d['fri_close']) - 1
            gaps.append(gap)

    if gaps:
        avg_gap = sum(gaps) / len(gaps)
        pos_gaps = sum(1 for g in gaps if g > 0)
        print(f"  Friday close to Monday open gap (Buy stocks):")
        print(f"    Avg gap: {avg_gap:>+.2%}")
        print(f"    Positive (stock gapped up): {pos_gaps}/{len(gaps)} ({pos_gaps/len(gaps):.0%})")
        print(f"    -> If stocks gap up Monday, investor pays more than stated entry")
        if avg_gap > 0:
            cost_per_pick = avg_gap * 1000
            print(f"    -> Estimated cost of timing gap: ${cost_per_pick:.2f} per $1,000 invested")
    print()


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 90)
    print("SEEKING ALPHA — PICK-DAY EFFECT ANALYSIS")
    print("=" * 90)
    print()

    prices = load_prices()
    trading_dates = get_trading_dates(prices)
    picks = load_alpha_picks()
    weekly = load_proquant()

    # 1. Verify Alpha Picks timing
    verify_alpha_picks_timing(picks, prices, trading_dates)

    # 2. Measure pick-day alpha
    measure_pick_day_alpha(picks, prices, trading_dates)

    # 3. Investor timing comparison
    investor_timing_analysis(picks, prices, trading_dates)

    # 4. Verify ProQuant timing
    verify_proquant_timing(weekly, prices, trading_dates)

    print("=" * 90)
    print("ANALYSIS COMPLETE")
    print("=" * 90)


if __name__ == "__main__":
    main()
