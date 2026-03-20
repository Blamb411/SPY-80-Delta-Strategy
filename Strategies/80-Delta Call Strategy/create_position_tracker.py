"""
Generate position_tracker.xlsx with five sheets:
  Sheet 1: SPY 80-Delta Call Strategy
  Sheet 2: Put Credit Spreads (PCS) Paper Trades
  Sheet 3: Put Credit Spreads (PCS) Live Trades
  Sheet 4: TSLA Bear Put Debit Spread
  Sheet 5: UPRO Long Position (DD25%/Cool40)

Pulls live prices from yfinance, PCS trades from put_spread_paper.db
and put_spread_live.db, Greeks from ThetaData (with BS fallback).
"""

import math
import os
import sys
import sqlite3
from datetime import date, datetime, timedelta
import numpy as np
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

PCS_DB_PATH = "C:/Users/Admin/Trading/data/put_spread_paper.db"
PCS_LIVE_DB_PATH = "C:/Users/Admin/Trading/data/put_spread_live.db"

# ============================================================================
# TRADE DATA
# ============================================================================

TRADES = [
    {
        "trade_num": 1,
        "entry_date": "2026-02-03",
        "symbol": "SPY",
        "strike": 660,
        "expiration": "2026-06-18",
        "type": "CALL",
        "quantity": 10,
        "entry_price": 51.60,
        "delta_at_entry": 0.74,
        "dte_at_entry": 135,
        "spy_at_entry": 687.0,
        "sma200_at_entry": 640.0,
    },
    {
        "trade_num": 2,
        "entry_date": "2026-02-04",
        "symbol": "SPY",
        "strike": 650,
        "expiration": "2026-05-29",
        "type": "CALL",
        "quantity": 10,
        "entry_price": 55.41,
        "delta_at_entry": 0.80,
        "dte_at_entry": 114,
        "spy_at_entry": 685.0,
        "sma200_at_entry": 640.0,
    },
    {
        "trade_num": 3,
        "entry_date": "2026-02-06",
        "symbol": "SPY",
        "strike": 655,
        "expiration": "2026-05-15",
        "type": "CALL",
        "quantity": 10,
        "entry_price": 49.70,
        "delta_at_entry": 0.76,
        "dte_at_entry": 98,
        "spy_at_entry": 684.0,
        "sma200_at_entry": 640.0,
    },
    {
        "trade_num": 4,
        "entry_date": "2026-02-19",
        "symbol": "SPY",
        "strike": 655,
        "expiration": "2026-06-18",
        "type": "CALL",
        "quantity": 10,
        "entry_price": 51.76,
        "delta_at_entry": 0.78,
        "dte_at_entry": 119,
        "spy_at_entry": 688.0,
        "sma200_at_entry": 641.0,
    },
    {
        "trade_num": 5,
        "entry_date": "2026-03-03",
        "symbol": "SPY",
        "strike": 625,
        "expiration": "2026-06-18",
        "type": "CALL",
        "quantity": 5,
        "entry_price": 71.68,
        "delta_at_entry": 0.80,
        "dte_at_entry": 107,
        "spy_at_entry": 681.5,
        "sma200_at_entry": 653.0,
    },
    {
        "trade_num": 6,
        "entry_date": "2026-03-03",
        "symbol": "QQQ",
        "strike": 550,
        "expiration": "2026-06-18",
        "type": "CALL",
        "quantity": 5,
        "entry_price": 71.47,
        "delta_at_entry": 0.81,
        "dte_at_entry": 107,
        "spy_at_entry": 602.8,
        "sma200_at_entry": 587.0,
    },
    {
        "trade_num": 7,
        "entry_date": "2026-03-17",
        "symbol": "SPY",
        "strike": 620,
        "expiration": "2026-07-17",
        "type": "CALL",
        "quantity": 5,
        "entry_price": 70.07,
        "delta_at_entry": 0.80,
        "dte_at_entry": 122,
        "spy_at_entry": 671.0,
        "sma200_at_entry": 655.0,
    },
    {
        "trade_num": 8,
        "entry_date": "2026-03-17",
        "symbol": "QQQ",
        "strike": 560,
        "expiration": "2026-08-21",
        "type": "CALL",
        "quantity": 5,
        "entry_price": 71.50,
        "delta_at_entry": 0.80,
        "dte_at_entry": 157,
        "spy_at_entry": 595.0,
        "sma200_at_entry": 587.0,
    },
]

TSLA_SPREAD = {
    "symbol": "TSLA", "structure": "Bear Put Debit Spread", "account": "IRA",
    "entry_date": "2026-03-09", "expiration": "2027-01-15",
    "long_strike": 300, "long_qty": 10, "long_entry_price": 28.50,
    "short_strike": 250, "short_qty": 10, "short_entry_price": 16.10,
    "net_debit": 12.40, "tsla_at_entry": 392.20,
}

UPRO_POSITION = {
    "symbol": "UPRO", "account": "IRA", "shares": 1000,
    "entry_price": 109.80, "entry_date": "2026-03-10",
    "strategy": "DD25%/Cool40", "dd_threshold": 0.25, "cooling_period": 40,
    "known_ath": 122.23, "known_ath_date": "2026-01-12",
}

# Bear put spread candidates — watchlist (not yet entered)
# Update status to "ENTERED" and fill in entry fields once trades are placed
BEAR_SPREAD_CANDIDATES = [
    {
        "symbol": "MDLZ", "name": "Mondelez International",
        "long_strike": 55, "short_strike": 27.5, "expiration": "2027-01-15",
        "screener_score": 66.3, "rr_ratio": 7.09, "sector": "Consumer Staples",
        "thesis": "Cocoa hedge trap, volume decline 4.8%, FY26 guidance below Street",
        "status": "WATCHLIST",  # WATCHLIST / ENTERED / CLOSED
        "qty": 0, "net_debit": 0, "entry_date": "",
    },
    {
        "symbol": "KHC", "name": "Kraft Heinz",
        "long_strike": 22.5, "short_strike": 12.5, "expiration": "2027-01-15",
        "screener_score": 65.5, "rr_ratio": 4.21, "sector": "Consumer Staples",
        "thesis": "Revenue declining, split paused, FY26 EPS $2.03 vs Street $2.68",
        "status": "WATCHLIST",
        "qty": 0, "net_debit": 0, "entry_date": "",
    },
    {
        "symbol": "AIG", "name": "American International Group",
        "long_strike": 77.5, "short_strike": 37.5, "expiration": "2027-01-15",
        "screener_score": 65.9, "rr_ratio": 5.32, "sector": "Financials",
        "thesis": "CEO transition, underwriting headwinds, cat exposure",
        "status": "WATCHLIST",
        "qty": 0, "net_debit": 0, "entry_date": "",
    },
    {
        "symbol": "SBUX", "name": "Starbucks",
        "long_strike": 100, "short_strike": 50, "expiration": "2027-01-15",
        "screener_score": 64.4, "rr_ratio": 3.96, "sector": "Consumer Discretionary",
        "thesis": "Consumer weakness, high FwdPE 44x, turnaround uncertainty",
        "status": "WATCHLIST",
        "qty": 0, "net_debit": 0, "entry_date": "",
    },
    {
        "symbol": "CCI", "name": "Crown Castle",
        "long_strike": 87.5, "short_strike": 45, "expiration": "2027-01-15",
        "screener_score": 73.6, "rr_ratio": 3.43, "sector": "Real Estate",
        "thesis": "REIT rate-sensitive, GFV=9, tower growth slowing",
        "status": "WATCHLIST",
        "qty": 0, "net_debit": 0, "entry_date": "",
    },
]

# ============================================================================
# PRICING FUNCTIONS
# ============================================================================

def norm_cdf(x):
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def bs_call_price(spot, strike, dte, iv=0.18, rate=0.045):
    if dte <= 0:
        return max(0, spot - strike)
    t = dte / 365.0
    d1 = (math.log(spot / strike) + (rate + 0.5 * iv**2) * t) / (iv * math.sqrt(t))
    d2 = d1 - iv * math.sqrt(t)
    return spot * norm_cdf(d1) - strike * math.exp(-rate * t) * norm_cdf(d2)


def bs_put_price(spot, strike, dte, iv=0.18, rate=0.045):
    """Black-Scholes put price."""
    if dte <= 0:
        return max(0, strike - spot)
    t = dte / 365.0
    d1 = (math.log(spot / strike) + (rate + 0.5 * iv**2) * t) / (iv * math.sqrt(t))
    d2 = d1 - iv * math.sqrt(t)
    return strike * math.exp(-rate * t) * norm_cdf(-d2) - spot * norm_cdf(-d1)


def estimate_entry_leg_prices(spot, sp_strike, lp_strike, dte, vix, entry_credit):
    """Estimate individual leg prices at entry, scaled to match actual net credit."""
    iv = vix / 100.0
    sp_bs = bs_put_price(spot, sp_strike, dte, iv)
    lp_bs = bs_put_price(spot, lp_strike, dte, iv)
    net_bs = sp_bs - lp_bs
    if net_bs > 0.01:
        scale = entry_credit / net_bs
        return sp_bs * scale, lp_bs * scale
    return entry_credit, 0.0


def bs_delta(spot, strike, dte, iv=0.18, rate=0.045):
    if dte <= 0:
        return 1.0 if spot > strike else 0.0
    t = dte / 365.0
    d1 = (math.log(spot / strike) + (rate + 0.5 * iv**2) * t) / (iv * math.sqrt(t))
    return norm_cdf(d1)


def trading_days_between(start_str, end_date):
    start = np.datetime64(start_str)
    end = np.datetime64(end_date.isoformat())
    return int(np.busday_count(start, end))


def max_hold_date(entry_str, max_days=60):
    """Calculate the date that is max_days trading days after entry."""
    d = datetime.strptime(entry_str, "%Y-%m-%d").date()
    count = 0
    while count < max_days:
        d += timedelta(days=1)
        if np.is_busday(np.datetime64(d.isoformat())):
            count += 1
    return d


def get_historical_closes(symbols, start_str, end_date):
    """Fetch daily close prices from yfinance for a date range.

    Returns {symbol: {date: price}} dict.
    """
    hist_prices = {}
    start_dt = datetime.strptime(start_str, "%Y-%m-%d").date()
    # Fetch a few extra days before start to handle weekends/holidays
    fetch_start = (start_dt - timedelta(days=10)).strftime("%Y-%m-%d")
    fetch_end = (end_date + timedelta(days=1)).strftime("%Y-%m-%d")
    for sym in symbols:
        hist_prices[sym] = {}
        try:
            tk = yf.Ticker(sym)
            df = tk.history(start=fetch_start, end=fetch_end)
            for idx, row in df.iterrows():
                d = idx.date() if hasattr(idx, 'date') else idx
                hist_prices[sym][d] = float(row["Close"])
        except Exception as e:
            print(f"Warning: could not fetch history for {sym}: {e}")
    return hist_prices


def lookup_hist_price(hist_prices, symbol, target_date):
    """Look up close price on or before target_date (handles weekends/holidays)."""
    prices = hist_prices.get(symbol, {})
    if target_date in prices:
        return prices[target_date]
    # Walk backwards up to 10 days to find most recent prior close
    for i in range(1, 11):
        d = target_date - timedelta(days=i)
        if d in prices:
            return prices[d]
    return None


def norm_pdf(x):
    """Standard normal probability density function."""
    return math.exp(-0.5 * x * x) / math.sqrt(2.0 * math.pi)


def bs_put_greeks(spot, strike, dte, iv, rate=0.045):
    """Full Black-Scholes Greeks for a put option."""
    if dte <= 0:
        delta = -1.0 if spot < strike else 0.0
        return {"delta": delta, "gamma": 0, "theta": 0, "vega": 0, "iv": iv}
    t = dte / 365.0
    sqrt_t = math.sqrt(t)
    d1 = (math.log(spot / strike) + (rate + 0.5 * iv**2) * t) / (iv * sqrt_t)
    d2 = d1 - iv * sqrt_t
    delta = norm_cdf(d1) - 1.0
    gamma = norm_pdf(d1) / (spot * iv * sqrt_t)
    theta = (-(spot * norm_pdf(d1) * iv) / (2 * sqrt_t)
             + rate * strike * math.exp(-rate * t) * norm_cdf(-d2)) / 365.0
    vega = spot * norm_pdf(d1) * sqrt_t / 100.0
    return {"delta": delta, "gamma": gamma, "theta": theta, "vega": vega, "iv": iv}


def _get_thetadata_client():
    """Get a connected ThetaData client, or None."""
    try:
        sys.path.insert(0, "C:/Users/Admin/Trading/repos/backtest-infrastructure")
        from thetadata_client import ThetaDataClient
        client = ThetaDataClient()
        if client.connect():
            return client
    except Exception as e:
        print(f"ThetaData connection error: {e}")
    return None


def get_thetadata_greeks(root, expiration, strike, right, query_date):
    """Fetch Greeks from ThetaData Terminal. Returns dict or None on failure."""
    try:
        client = _get_thetadata_client()
        if not client:
            return None
        results = client.get_option_greeks(
            root=root, expiration=expiration, strike=float(strike),
            right=right, start=query_date, end=query_date
        )
        if results:
            return results[-1]
    except Exception as e:
        print(f"ThetaData Greeks error: {e}")
    return None


def get_recommended_trades(symbols, price_map, today):
    """Fetch real expiration/strike/quote data for recommended 80-delta trades.

    Uses ThetaData for expirations, strikes, Greeks, and EOD quotes.
    Falls back to IBKR for live bid/ask if available, then to BS estimates.

    Returns list of dicts with keys:
        symbol, strike, expiration, dte, delta, iv, bid, ask, mid,
        cost_5, cost_10, source
    """
    client = _get_thetadata_client()
    ib = None

    # Try IBKR for live bid/ask
    try:
        from ib_insync import IB, Stock, Option as IBOption
        ib = IB()
        ib.connect("127.0.0.1", 7497, clientId=97, timeout=5)
    except Exception:
        ib = None

    results = []
    target_dte = 120

    for sym in symbols:
        spot = price_map.get(sym, 600.0)
        vix_iv = (get_current_vix() or 20.0) / 100.0
        rec = {"symbol": sym, "source": "BS estimate"}

        # ── Step 1: Get real expirations from ThetaData ──
        expirations = []
        if client:
            try:
                all_exps = client.get_expirations(sym)
                for exp_str in all_exps:
                    exp_date = datetime.strptime(exp_str, "%Y-%m-%d").date()
                    dte = (exp_date - today).days
                    if 90 <= dte <= 150:
                        expirations.append((abs(dte - target_dte), exp_str, exp_date, dte))
            except Exception as e:
                print(f"  ThetaData expirations error for {sym}: {e}")

        if expirations:
            expirations.sort()
            _, best_exp_str, best_exp_date, best_dte = expirations[0]
            print(f"  {sym}: Found {len(expirations)} expirations in 90-150d range, "
                  f"best = {best_exp_str} ({best_dte} DTE)")
        else:
            # Fallback to calculated 3rd Friday
            print(f"  {sym}: No ThetaData expirations, using calculated 3rd Friday")
            candidates = []
            for month_offset in range(3, 7):
                m = (today.month + month_offset - 1) % 12 + 1
                y = today.year + (today.month + month_offset - 1) // 12
                first = date(y, m, 1)
                dow = first.weekday()
                first_fri = first + timedelta(days=(4 - dow) % 7)
                third_fri = first_fri + timedelta(days=14)
                exp_dte = (third_fri - today).days
                if 90 <= exp_dte <= 150:
                    candidates.append((abs(exp_dte - target_dte), third_fri, exp_dte))
            if not candidates:
                continue
            candidates.sort()
            best_exp_date = candidates[0][1]
            best_dte = candidates[0][2]
            best_exp_str = best_exp_date.strftime("%Y-%m-%d")

        # ── Step 2: Find 80-delta strike from real strikes ──
        # BS estimate for target strike
        lo, hi = spot * 0.7, spot * 1.1
        for _ in range(100):
            mid_k = (lo + hi) / 2
            d = bs_delta(spot, mid_k, best_dte, iv=vix_iv)
            if d > 0.80:
                lo = mid_k
            else:
                hi = mid_k
        est_strike = round(mid_k)

        # Snap to real available strikes if possible
        rec_strike = est_strike
        if client:
            try:
                strikes = client.get_strikes(sym, best_exp_str)
                if strikes:
                    rec_strike = min(strikes, key=lambda k: abs(k - est_strike))
                    print(f"  {sym}: Snapped strike ${est_strike} -> ${rec_strike} "
                          f"(from {len(strikes)} available)")
            except Exception as e:
                print(f"  ThetaData strikes error for {sym}: {e}")

        # ── Step 3: Get real Greeks ──
        real_delta = None
        real_iv = None
        today_str = today.strftime("%Y-%m-%d")

        if client:
            try:
                greeks = client.get_option_greeks(
                    root=sym, expiration=best_exp_str,
                    strike=float(rec_strike), right="C",
                    start=today_str, end=today_str)
                if greeks:
                    g = greeks[-1]
                    real_delta = g.get("delta")
                    real_iv = g.get("iv")
                    rec["source"] = "ThetaData"
                    print(f"  {sym}: ThetaData Greeks — delta={real_delta:.3f}, IV={real_iv:.1%}")
            except Exception as e:
                print(f"  ThetaData Greeks error for {sym}: {e}")

        # ── Step 4: Get bid/ask ──
        bid, ask, mid_price = None, None, None

        # Try IBKR first for live quotes
        if ib:
            try:
                exp_ib = best_exp_str.replace("-", "")
                opt = IBOption(sym, exp_ib, rec_strike, "C", "SMART")
                ib.qualifyContracts(opt)
                ticker = ib.reqMktData(opt, '', False, False)
                ib.sleep(2)
                if ticker.bid and ticker.bid > 0:
                    bid = ticker.bid
                if ticker.ask and ticker.ask > 0:
                    ask = ticker.ask
                if bid and ask:
                    mid_price = (bid + ask) / 2
                    rec["source"] = "IBKR live"
                    print(f"  {sym}: IBKR quote — bid=${bid:.2f} ask=${ask:.2f}")
                elif ticker.last and ticker.last > 0:
                    mid_price = ticker.last
                    rec["source"] = "IBKR last"
                ib.cancelMktData(opt)
            except Exception as e:
                print(f"  IBKR quote error for {sym}: {e}")

        # Fallback to ThetaData EOD quotes
        if mid_price is None and client:
            try:
                eod = client.get_option_eod(
                    root=sym, expiration=best_exp_str,
                    strike=float(rec_strike), right="C",
                    start=today_str, end=today_str)
                if eod:
                    row_data = eod[-1]
                    bid = row_data.get("bid")
                    ask = row_data.get("ask")
                    if bid and ask and bid > 0 and ask > 0:
                        mid_price = (bid + ask) / 2
                        if rec["source"] == "BS estimate":
                            rec["source"] = "ThetaData EOD"
                    elif row_data.get("close") and row_data["close"] > 0:
                        mid_price = row_data["close"]
                        if rec["source"] == "BS estimate":
                            rec["source"] = "ThetaData close"
            except Exception as e:
                print(f"  ThetaData EOD error for {sym}: {e}")

        # Final fallback to BS price
        if mid_price is None:
            mid_price = bs_call_price(spot, rec_strike, best_dte, iv=vix_iv)

        if real_delta is None:
            real_delta = bs_delta(spot, rec_strike, best_dte, iv=vix_iv)
        if real_iv is None:
            real_iv = vix_iv

        rec.update({
            "strike": rec_strike,
            "expiration": best_exp_str,
            "dte": best_dte,
            "delta": real_delta,
            "iv": real_iv,
            "bid": bid,
            "ask": ask,
            "mid": mid_price,
            "cost_5": mid_price * 100 * 5,
            "cost_10": mid_price * 100 * 10,
        })
        results.append(rec)

    if ib:
        try:
            ib.disconnect()
        except Exception:
            pass

    return results


def get_tsla_iv_from_yfinance():
    """Get TSLA ATM implied volatility from yfinance. Default 0.55 if fails."""
    try:
        tk = yf.Ticker("TSLA")
        exps = tk.options
        if exps:
            chain = tk.option_chain(exps[0])
            puts = chain.puts
            spot = tk.history(period="5d")["Close"].iloc[-1]
            atm_idx = (puts["strike"] - spot).abs().idxmin()
            iv = puts.loc[atm_idx, "impliedVolatility"]
            if iv and iv > 0:
                print(f"TSLA ATM IV from yfinance: {iv:.1%}")
                return float(iv)
    except Exception as e:
        print(f"TSLA yfinance IV error: {e}")
    print("Using default TSLA IV: 55%")
    return 0.55


# ============================================================================
# GET LIVE DATA
# ============================================================================

def get_price(symbol, fallback=600.0):
    try:
        ticker = yf.Ticker(symbol)
        data = ticker.history(period="5d")
        if not data.empty:
            price = float(data["Close"].iloc[-1])
            print(f"{symbol} price from yfinance: ${price:.2f}")
            return price
    except Exception as e:
        print(f"yfinance error for {symbol}: {e}")
    print(f"Using fallback {symbol} price: ${fallback:.2f}")
    return fallback


# ============================================================================
# BUILD SPREADSHEET
# ============================================================================

def get_styles():
    """Shared styles for both sheets."""
    return {
        "title_font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
        "title_fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "section_font": Font(name="Calibri", size=12, bold=True, color="1F4E79"),
        "header_font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        "header_fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
        "data_font": Font(name="Calibri", size=10),
        "money_fmt": '#,##0.00',
        "money_whole_fmt": '$#,##0',
        "pct_fmt": '0.0%',
        "delta_fmt": '0.00',
        "int_fmt": '0',
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
        "green_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "red_fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "light_blue_fill": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "light_gray_fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        "orange_fill": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "paper_fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        # PCS-specific
        "pcs_title_fill": PatternFill(start_color="4A2545", end_color="4A2545", fill_type="solid"),
        "pcs_header_fill": PatternFill(start_color="7B4F7B", end_color="7B4F7B", fill_type="solid"),
        "pcs_section_font": Font(name="Calibri", size=12, bold=True, color="4A2545"),
        # Bear spread (maroon)
        "bear_title_fill": PatternFill(start_color="7B2020", end_color="7B2020", fill_type="solid"),
        "bear_header_fill": PatternFill(start_color="A03030", end_color="A03030", fill_type="solid"),
        "bear_section_font": Font(name="Calibri", size=12, bold=True, color="7B2020"),
        # UPRO long (dark green)
        "upro_title_fill": PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),
        "upro_header_fill": PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
        "upro_section_font": Font(name="Calibri", size=12, bold=True, color="1B5E20"),
        # Signal fills
        "yellow_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    }


# ============================================================================
# PCS DATA FROM DATABASE
# ============================================================================

def load_pcs_trades(db_path=None):
    """Load all PCS trades from a PCS database."""
    db_path = db_path or PCS_DB_PATH
    if not os.path.exists(db_path):
        return [], {}
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM positions ORDER BY id")
    positions = [dict(r) for r in cur.fetchall()]

    # Get latest daily_log entry for each open position
    latest_logs = {}
    cur.execute("""
        SELECT dl.* FROM daily_log dl
        INNER JOIN (
            SELECT position_id, MAX(created_at) as max_ts
            FROM daily_log WHERE position_id IS NOT NULL
            GROUP BY position_id
        ) latest ON dl.position_id = latest.position_id
            AND dl.created_at = latest.max_ts
    """)
    for r in cur.fetchall():
        latest_logs[r["position_id"]] = dict(r)

    conn.close()
    return positions, latest_logs


# ============================================================================
# SHEET 1: 80-DELTA CALLS
# ============================================================================

def build_80delta_sheet(ws, price_map, sma_map, today):
    s = get_styles()
    spy_price = price_map.get("SPY", 600.0)
    qqq_price = price_map.get("QQQ", 500.0)

    # ========================================================================
    # PRECOMPUTE CUMULATIVE CAPITAL & MARKET VALUE PER ROW
    # ========================================================================
    all_symbols = list({t["symbol"] for t in TRADES})
    entry_dates = [t["entry_date"] for t in TRADES]
    earliest = min(entry_dates)
    hist_prices = get_historical_closes(all_symbols, earliest, today)

    # Sort trades by entry date for cumulative computation
    sorted_trades = sorted(TRADES, key=lambda x: x["entry_date"])
    for i, t in enumerate(sorted_trades):
        entry_dt = datetime.strptime(t["entry_date"], "%Y-%m-%d").date()
        # Cumulative capital: sum of entry costs for all trades entered on or before this row's date
        cum_capital = 0
        for prev in sorted_trades[:i + 1]:
            cum_capital += prev["entry_price"] * 100 * prev["quantity"]
        t["_cum_capital"] = cum_capital

        # Market value: sum of BS call values for all outstanding trades as of this row's date
        cum_mkt = 0
        for prev in sorted_trades[:i + 1]:
            prev_exp = datetime.strptime(prev["expiration"], "%Y-%m-%d").date()
            prev_dte = (prev_exp - entry_dt).days
            sym = prev["symbol"]
            if entry_dt >= today:
                # Use current live price for today's rows
                spot = price_map.get(sym, spy_price)
            else:
                spot = lookup_hist_price(hist_prices, sym, entry_dt)
                if spot is None:
                    spot = price_map.get(sym, spy_price)
            val = bs_call_price(spot, prev["strike"], prev_dte) * 100 * prev["quantity"]
            cum_mkt += val
        t["_cum_mkt_value"] = cum_mkt

    # ========================================================================
    # TITLE ROW
    # ========================================================================
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=27)
    cell = ws.cell(row=row, column=1, value="80-Delta Call Strategy -- Position Tracker")
    cell.font = s["title_font"]
    cell.fill = s["title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=27)
    price_str = "  |  ".join(f"{sym}: ${p:.2f}" for sym, p in price_map.items())
    cell = ws.cell(row=row, column=1,
                   value=f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  {price_str}")
    cell.font = Font(name="Calibri", size=10, italic=True, color="1F4E79")
    cell.alignment = Alignment(horizontal="center")

    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=27)
    sma_parts = []
    for sym in ["SPY", "QQQ"]:
        sma_val = sma_map.get(sym)
        if sma_val:
            diff = price_map.get(sym, 0) - sma_val
            pct = diff / sma_val * 100 if sma_val else 0
            sma_parts.append(f"{sym} SMA200: ${sma_val:.2f} ({pct:+.1f}%)")
        else:
            sma_parts.append(f"{sym} SMA200: N/A")
    cell = ws.cell(row=row, column=1, value="  |  ".join(sma_parts))
    cell.font = Font(name="Calibri", size=10, italic=True, color="1F4E79")
    cell.alignment = Alignment(horizontal="center")

    # ========================================================================
    # STRATEGY CRITERIA SECTION
    # ========================================================================
    row = 5
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="Strategy Criteria")
    cell.font = s["section_font"]

    criteria = [
        ("Entry Rules", ""),
        ("  Signal", "SPY > 200-day SMA"),
        ("  Delta", "70-80 delta at entry"),
        ("  DTE", "90-150 days to expiration"),
        ("  Frequency", "Monthly only (standard expiry)"),
        ("Exit Rules", ""),
        ("  Profit Target", "+50% from entry price"),
        ("  Max Hold", "60 trading days"),
        ("  SMA Breach", "Optional: exit if SPY >2% below SMA200"),
        ("  Stop Loss", "None (per backtest results)"),
        ("Notes", ""),
        ("  Delta Cap", "Monitor total delta exposure (not a hard limit)"),
    ]

    row += 1
    for label, value in criteria:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        if not value:
            c1.font = Font(name="Calibri", size=10, bold=True)
        else:
            c1.font = s["data_font"]
            c1.fill = s["light_gray_fill"]
            c2.font = s["data_font"]
            c2.fill = s["light_gray_fill"]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # ========================================================================
    # MAIN POSITION TABLE
    # ========================================================================
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=24)
    cell = ws.cell(row=row, column=1, value="Open Positions")
    cell.font = s["section_font"]
    row += 1

    headers = [
        "Trade #", "Entry Date", "Symbol", "Strike", "Expiration", "Type", "Qty",
        "Entry Price", "Total Cost",
        "Capital Invested", "Market Value",
        "Delta@Entry", "DTE@Entry", "SPY@Entry", "SMA200@Entry",
        "Current Price", "Current Value",
        "Cur Delta", "Cur DTE",
        "Unreal P&L ($)", "Unreal P&L (%)",
        "Target Price", "% To Target",
        "Days Held", "Max Hold Date", "Days Remaining",
        "Status",
    ]

    header_row = row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = s["thin_border"]
    ws.row_dimensions[row].height = 30

    total_cost = 0
    total_value = 0
    total_delta = 0
    total_contracts = 0

    for t in TRADES:
        row += 1
        sym = t["symbol"]
        spot = price_map.get(sym, spy_price)
        exp_date = datetime.strptime(t["expiration"], "%Y-%m-%d").date()
        dte = (exp_date - today).days
        current_price = bs_call_price(spot, t["strike"], dte)
        current_value = current_price * 100 * t["quantity"]
        entry_cost = t["entry_price"] * 100 * t["quantity"]
        pnl_dollar = current_value - entry_cost
        pnl_pct = (current_price / t["entry_price"] - 1) if t["entry_price"] > 0 else 0
        cur_delta = bs_delta(spot, t["strike"], dte)
        position_delta = cur_delta * t["quantity"] * 100
        profit_target = t["entry_price"] * 1.50
        pct_to_target = (profit_target - current_price) / current_price if current_price > 0 else 0
        days_held = trading_days_between(t["entry_date"], today)
        mhd = max_hold_date(t["entry_date"])
        days_remaining = max(0, 60 - days_held)

        if pnl_pct >= 0.50:
            status = "SELL - TARGET HIT"
        elif days_remaining <= 5:
            status = "REVIEW - MAX HOLD"
        else:
            status = "OPEN"

        total_cost += entry_cost
        total_value += current_value
        total_delta += position_delta
        total_contracts += t["quantity"]

        values = [
            t["trade_num"], t["entry_date"], t["symbol"], t["strike"],
            t["expiration"], t["type"], t["quantity"],
            t["entry_price"], entry_cost,
            t["_cum_capital"], t["_cum_mkt_value"],
            t["delta_at_entry"], t["dte_at_entry"], t["spy_at_entry"],
            t["sma200_at_entry"],
            current_price, current_value, cur_delta, dte,
            pnl_dollar, pnl_pct, profit_target, pct_to_target,
            days_held, mhd.strftime("%Y-%m-%d"), days_remaining, status,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = s["data_font"]
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")

        for c in [8, 16, 22]:
            ws.cell(row=row, column=c).number_format = s["money_fmt"]
        for c in [9, 10, 11, 17, 20]:
            ws.cell(row=row, column=c).number_format = s["money_whole_fmt"]
        for c in [12, 18]:
            ws.cell(row=row, column=c).number_format = s["delta_fmt"]
        for c in [21, 23]:
            ws.cell(row=row, column=c).number_format = s["pct_fmt"]
        for c in [1, 7, 13, 19, 24, 26]:
            ws.cell(row=row, column=c).number_format = s["int_fmt"]
        for c in [4, 14, 15]:
            ws.cell(row=row, column=c).number_format = s["money_whole_fmt"]

        pnl_cell = ws.cell(row=row, column=20)
        pnl_pct_cell = ws.cell(row=row, column=21)
        if pnl_dollar >= 0:
            pnl_cell.fill = s["green_fill"]
            pnl_pct_cell.fill = s["green_fill"]
        else:
            pnl_cell.fill = s["red_fill"]
            pnl_pct_cell.fill = s["red_fill"]

        status_cell = ws.cell(row=row, column=27)
        if "SELL" in status:
            status_cell.fill = s["orange_fill"]
            status_cell.font = Font(name="Calibri", size=10, bold=True)
        elif "REVIEW" in status:
            status_cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            status_cell.font = Font(name="Calibri", size=10, bold=True)

        if t["trade_num"] % 2 == 0:
            for col_idx in range(1, len(values) + 1):
                c = ws.cell(row=row, column=col_idx)
                if c.fill == PatternFill():
                    c.fill = s["light_blue_fill"]

    # ========================================================================
    # PORTFOLIO SUMMARY
    # ========================================================================
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="Portfolio Summary")
    cell.font = s["section_font"]
    row += 1

    total_pnl = total_value - total_cost
    total_pnl_pct = (total_value / total_cost - 1) if total_cost > 0 else 0

    summary_data = [
        ("Open Positions", len(TRADES)),
        ("Total Contracts", total_contracts),
        ("Total Cost Basis", total_cost),
        ("Total Current Value", total_value),
        ("Total Unrealized P&L", total_pnl),
        ("Total P&L %", total_pnl_pct),
        ("Total Delta Exposure", total_delta),
        ("Equivalent SPY Shares", int(total_delta)),
    ]

    for label, value in summary_data:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = Font(name="Calibri", size=10, bold=True)
        c1.fill = s["light_gray_fill"]
        c1.border = s["thin_border"]
        c2.font = s["data_font"]
        c2.border = s["thin_border"]
        c2.alignment = Alignment(horizontal="right")

        if "Cost" in label or "Value" in label or ("P&L" in label and "%" not in label):
            c2.number_format = s["money_whole_fmt"]
        if "P&L %" in label:
            c2.number_format = s["pct_fmt"]
        if "Delta" in label or "Shares" in label:
            c2.number_format = '#,##0'
        if "P&L" in label and "%" not in label:
            c2.fill = s["green_fill"] if value >= 0 else s["red_fill"]

        row += 1

    # ========================================================================
    # RECOMMENDED NEXT TRADES (live data)
    # ========================================================================
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value="Recommended Next Trades (80-Delta Calls)")
    cell.font = s["section_font"]
    row += 1

    rec_headers = ["Symbol", "Strike", "Expiration", "DTE", "Delta",
                   "IV", "Bid", "Ask", "Mid",
                   "Cost (5 ct)", "Cost (10 ct)", "Source"]
    for col_idx, h in enumerate(rec_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = s["thin_border"]

    print("\nFetching recommended trade data...")
    recommendations = get_recommended_trades(["SPY", "QQQ"], price_map, today)

    for rec in recommendations:
        row += 1
        values = [
            rec["symbol"],                                    # 1
            rec["strike"],                                    # 2
            rec["expiration"],                                # 3
            rec["dte"],                                       # 4
            rec["delta"],                                     # 5
            rec["iv"],                                        # 6
            rec["bid"] if rec["bid"] else "-",                # 7
            rec["ask"] if rec["ask"] else "-",                # 8
            rec["mid"],                                       # 9
            rec["cost_5"],                                    # 10
            rec["cost_10"],                                   # 11
            rec["source"],                                    # 12
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = s["data_font"]
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=2).number_format = s["money_whole_fmt"]  # Strike
        ws.cell(row=row, column=5).number_format = s["delta_fmt"]        # Delta
        ws.cell(row=row, column=6).number_format = s["pct_fmt"]          # IV
        for c in [7, 8, 9]:                                               # Bid/Ask/Mid
            if ws.cell(row=row, column=c).value != "-":
                ws.cell(row=row, column=c).number_format = s["money_fmt"]
        ws.cell(row=row, column=10).number_format = s["money_whole_fmt"]  # Cost 5ct
        ws.cell(row=row, column=11).number_format = s["money_whole_fmt"]  # Cost 10ct

    # Column widths
    col_widths = {
        1: 8, 2: 12, 3: 7, 4: 9, 5: 12, 6: 6, 7: 5,
        8: 12, 9: 13, 10: 15, 11: 14,
        12: 11, 13: 11, 14: 11, 15: 13,
        16: 13, 17: 14, 18: 10, 19: 9,
        20: 14, 21: 13, 22: 12, 23: 11,
        24: 10, 25: 13, 26: 14, 27: 18,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    return total_cost, total_value, total_pnl, total_delta, total_contracts


# ============================================================================
# SHEET 2: PUT CREDIT SPREADS (PAPER TRADES)
# ============================================================================

def compute_sma200(ticker_sym):
    """Compute current 200-day SMA for a ticker."""
    try:
        tk = yf.Ticker(ticker_sym)
        hist = tk.history(period="1y")
        if len(hist) >= 200:
            return float(hist["Close"].iloc[-200:].mean())
        elif not hist.empty:
            return float(hist["Close"].mean())
    except Exception:
        pass
    return None


def compute_iv_rank(current_vix, lookback=252):
    """Compute IV rank using VIX 1-year history."""
    try:
        vix_tk = yf.Ticker("^VIX")
        hist = vix_tk.history(period="1y")
        if len(hist) >= 20:
            closes = hist["Close"].dropna().tolist()
            low = min(closes)
            high = max(closes)
            if high > low:
                return (current_vix - low) / (high - low)
    except Exception:
        pass
    return None


def get_current_vix():
    """Get current VIX from yfinance."""
    try:
        vix = yf.Ticker("^VIX")
        hist = vix.history(period="5d")
        if not hist.empty:
            val = float(hist["Close"].iloc[-1])
            print(f"VIX from yfinance: {val:.1f}")
            return val
    except Exception:
        pass
    return None


def build_pcs_sheet(ws, spy_price, today, db_path=None, sheet_title=None):
    s = get_styles()
    positions, latest_logs = load_pcs_trades(db_path=db_path)

    # Get QQQ price
    qqq_price = get_price("QQQ", 500.0)

    iwm_price = get_price("IWM", 250.0)
    spot_map = {"SPY": spy_price, "QQQ": qqq_price, "IWM": iwm_price}

    # Current market data for the "Current" columns
    cur_vix = get_current_vix()
    cur_ivr = compute_iv_rank(cur_vix) if cur_vix else None
    sma200_map = {}
    for sym in ["SPY", "QQQ", "IWM"]:
        sma = compute_sma200(sym)
        if sma:
            sma200_map[sym] = sma
            print(f"{sym} SMA200: ${sma:.2f}")
        else:
            sma200_map[sym] = None

    # ========================================================================
    # TITLE
    # ========================================================================
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=32)
    is_live = (db_path and "live" in db_path.lower())
    default_title = ("Put Credit Spreads -- Live Trades" if is_live
                     else "Put Credit Spreads -- Paper Trades (Account DUA976236)")
    cell = ws.cell(row=row, column=1,
                   value=sheet_title or default_title)
    cell.font = s["title_font"]
    cell.fill = s["pcs_title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

    mode_label = "LIVE TRADING" if is_live else "PAPER TRADING - NOT REAL MONEY"
    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=32)
    cell = ws.cell(row=row, column=1,
                   value=f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
                         f"SPY: ${spy_price:.2f}  |  QQQ: ${qqq_price:.2f}  |  "
                         f"{mode_label}")
    cell.font = Font(name="Calibri", size=10, italic=True, bold=True, color="4A2545")
    cell.alignment = Alignment(horizontal="center")

    # ========================================================================
    # STRATEGY CRITERIA
    # ========================================================================
    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="Strategy Criteria")
    cell.font = s["pcs_section_font"]

    criteria = [
        ("Entry Rules", ""),
        ("  Short Delta", "0.20 (20-delta short put)"),
        ("  Wing Width", "0.75 sigma (volatility-scaled)"),
        ("  DTE", "30 days target (25-45 range)"),
        ("  IV Rank Floor", "15% minimum to enter"),
        ("  Min C/W Ratio", "20% for QQQ, 12% for SPY (credit / wing width)"),
        ("  Trend Filter", "Price > 200-day SMA"),
        ("  Entry Spacing", "5 days minimum between entries"),
        ("  Max Open", "3 positions per ticker"),
        ("Exit Rules", ""),
        ("  Take Profit", "50% of credit received (spread <= 50% of entry credit)"),
        ("  Stop Loss", "3x credit multiplier"),
        ("  Expiration", "Close or let expire if no trigger hit"),
    ]

    row += 1
    for label, value in criteria:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        if not value:
            c1.font = Font(name="Calibri", size=10, bold=True)
        else:
            c1.font = s["data_font"]
            c1.fill = s["light_gray_fill"]
            c2.font = s["data_font"]
            c2.fill = s["light_gray_fill"]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # ========================================================================
    # ALL TRADES TABLE
    # ========================================================================
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=32)
    cell = ws.cell(row=row, column=1, value="All Trades (Paper)")
    cell.font = s["pcs_section_font"]
    row += 1

    headers = [
        "ID", "Status", "Ticker", "Entry Date", "Expiration", "DTE",   # 1-6
        "Short Strike", "Long Strike", "Wing $",                        # 7-9
        "Contracts",                                                     # 10
        "Credit/sh",                                                     # 11
        "SP Rcvd/Ctr", "LP Paid/Ctr",                                   # 12-13
        "TP Target", "SL Trigger",                                       # 14-15
        "Cur Spread", "P&L ($)", "P&L (%)",                             # 16-18
        "Total Credit", "Max Loss",                                      # 19-20
        "Exit Date", "Exit Debit", "Exit Reason",                        # 21-23
        "Spot@Entry", "Cur Spot",                                        # 24-25
        "VIX@Entry", "Cur VIX",                                          # 26-27
        "IVR@Entry", "Cur IVR",                                          # 28-29
        "SMA200@Entry", "Cur SMA200",                                    # 30-31
        "Notes",                                                         # 32
    ]

    header_row = row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["pcs_header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = s["thin_border"]
    ws.row_dimensions[row].height = 30

    # Track summary for open positions
    open_total_credit = 0
    open_total_pnl = 0
    open_total_max_loss = 0
    open_contracts = 0
    n_open = 0

    for p in positions:
        row += 1
        exp_date = datetime.strptime(p["expiration"], "%Y-%m-%d").date()
        dte = (exp_date - today).days
        total_credit = p["entry_credit"] * 100 * p["num_contracts"]
        wing = p["wing_width"]

        # Current spread value and P&L
        log = latest_logs.get(p["id"])
        if p["status"] == "open" and log and log["spread_value"] > 0:
            cur_spread = log["spread_value"]
            pnl_dollar = (p["entry_credit"] - cur_spread) * 100 * p["num_contracts"]
        elif p["status"] == "open":
            # No log data — estimate: assume spread roughly same as entry
            cur_spread = p["entry_credit"]
            pnl_dollar = 0
        else:
            # Cancelled/closed
            cur_spread = None
            pnl_dollar = p.get("pnl") or 0

        pnl_pct = (pnl_dollar / total_credit) if total_credit > 0 and cur_spread is not None else None

        status_display = p["status"].upper()
        if p["status"] == "open":
            open_total_credit += total_credit
            open_total_pnl += pnl_dollar
            open_total_max_loss += p["max_loss"]
            open_contracts += p["num_contracts"]
            n_open += 1

        cur_spot = spot_map.get(p["ticker"])

        # Estimate entry leg prices (scaled BS to match actual net credit)
        entry_date_obj = datetime.strptime(p["entry_date"], "%Y-%m-%d").date()
        dte_at_entry = (exp_date - entry_date_obj).days
        sp_entry, lp_entry = estimate_entry_leg_prices(
            p["spot_at_entry"], p["sp_strike"], p["lp_strike"],
            dte_at_entry, p["vix_at_entry"], p["entry_credit"])
        sp_rcvd_ctr = sp_entry * 100  # per contract
        lp_paid_ctr = lp_entry * 100  # per contract

        # For closed trades, show realized P&L; for open, show unrealized
        if p["status"] == "closed":
            display_pnl = p.get("pnl") or 0
            display_pnl_pct = (display_pnl / total_credit) if total_credit > 0 else 0
            exit_debit = p.get("exit_debit") or p.get("exit_credit") or "-"
        else:
            display_pnl = pnl_dollar if cur_spread is not None else "-"
            display_pnl_pct = pnl_pct if pnl_pct is not None else "-"
            exit_debit = "-"

        values = [
            p["id"],                                                     # 1
            status_display,                                              # 2
            p["ticker"],                                                 # 3
            p["entry_date"],                                             # 4
            p["expiration"],                                             # 5
            dte if p["status"] == "open" else "-",                       # 6
            p["sp_strike"],                                              # 7
            p["lp_strike"],                                              # 8
            wing,                                                        # 9
            p["num_contracts"],                                          # 10
            p["entry_credit"],                                           # 11 Credit/sh
            sp_rcvd_ctr,                                                 # 12 SP Rcvd/Ctr
            lp_paid_ctr,                                                 # 13 LP Paid/Ctr
            p["tp_target"],                                              # 14 TP Target
            p["sl_trigger_debit"],                                       # 15 SL Trigger
            cur_spread if cur_spread is not None else "-",               # 16 Cur Spread
            display_pnl,                                                 # 17 P&L ($)
            display_pnl_pct,                                             # 18 P&L (%)
            total_credit,                                                # 19 Total Credit
            p["max_loss"],                                               # 20 Max Loss
            p.get("exit_date") or "-",                                   # 21 Exit Date
            exit_debit,                                                  # 22 Exit Debit
            p.get("exit_reason") or "-",                                 # 23 Exit Reason
            p["spot_at_entry"],                                          # 24 Spot@Entry
            cur_spot if p["status"] == "open" else "-",                  # 25 Cur Spot
            p["vix_at_entry"],                                           # 26 VIX@Entry
            cur_vix if p["status"] == "open" else "-",                   # 27 Cur VIX
            p["iv_rank"],                                                # 28 IVR@Entry
            cur_ivr if p["status"] == "open" and cur_ivr is not None else "-",  # 29 Cur IVR
            p["sma_value"],                                              # 30 SMA200@Entry
            sma200_map.get(p["ticker"]) if p["status"] == "open" else "-",  # 31 Cur SMA200
            p["notes"] or "",                                            # 32 Notes
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = s["data_font"]
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")

        # Number formats
        # Per-share credit cols: credit/sh(11), tp_target(14), sl_trigger(15), cur_spread(16)
        for c in [11, 14, 15, 16]:
            if ws.cell(row=row, column=c).value != "-":
                ws.cell(row=row, column=c).number_format = s["money_fmt"]
        # Per-contract dollar cols: SP rcvd(12), LP paid(13)
        for c in [12, 13]:
            ws.cell(row=row, column=c).number_format = s["money_whole_fmt"]
        # Total dollar cols: pnl(17), total_credit(19), max_loss(20)
        for c in [17, 19, 20]:
            if ws.cell(row=row, column=c).value != "-":
                ws.cell(row=row, column=c).number_format = s["money_whole_fmt"]
        # Exit debit (22)
        if ws.cell(row=row, column=22).value != "-":
            ws.cell(row=row, column=22).number_format = s["money_fmt"]
        # Strike/wing cols
        for c in [7, 8, 9]:
            ws.cell(row=row, column=c).number_format = s["money_fmt"]
        # Spot cols: spot@entry(24), cur_spot(25)
        for c in [24, 25]:
            if ws.cell(row=row, column=c).value != "-":
                ws.cell(row=row, column=c).number_format = s["money_fmt"]
        # VIX: vix@entry(26), cur_vix(27)
        for c in [26, 27]:
            if ws.cell(row=row, column=c).value != "-":
                ws.cell(row=row, column=c).number_format = '0.0'
        # IVR: ivr@entry(28), cur_ivr(29)
        for c in [28, 29]:
            if ws.cell(row=row, column=c).value != "-":
                ws.cell(row=row, column=c).number_format = s["pct_fmt"]
        # SMA200: sma@entry(30), cur_sma(31)
        for c in [30, 31]:
            if ws.cell(row=row, column=c).value != "-":
                ws.cell(row=row, column=c).number_format = s["money_fmt"]
        # P&L %
        if ws.cell(row=row, column=18).value != "-":
            ws.cell(row=row, column=18).number_format = s["pct_fmt"]

        # Color coding for P&L
        if p["status"] in ("open", "closed"):
            pnl_val = display_pnl if display_pnl != "-" else 0
            if pnl_val != "-":
                pnl_cell = ws.cell(row=row, column=17)
                pnl_pct_cell = ws.cell(row=row, column=18)
                if pnl_val >= 0:
                    pnl_cell.fill = s["green_fill"]
                    pnl_pct_cell.fill = s["green_fill"]
                else:
                    pnl_cell.fill = s["red_fill"]
                    pnl_pct_cell.fill = s["red_fill"]
        if p["status"] == "cancelled":
            # Gray out entire row for cancelled
            gray_font = Font(name="Calibri", size=10, color="999999")
            for col_idx in range(1, len(values) + 1):
                ws.cell(row=row, column=col_idx).font = gray_font

        # Status cell styling
        status_cell = ws.cell(row=row, column=2)
        if p["status"] == "open":
            status_cell.fill = s["green_fill"]
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="006100")
        elif p["status"] == "closed":
            status_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
        elif p["status"] == "cancelled":
            status_cell.fill = s["light_gray_fill"]
            status_cell.font = Font(name="Calibri", size=10, color="999999")

    # ========================================================================
    # PORTFOLIO SUMMARY (open positions only)
    # ========================================================================
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="Open Position Summary (Paper)")
    cell.font = s["pcs_section_font"]
    row += 1

    open_pnl_pct = (open_total_pnl / open_total_credit) if open_total_credit > 0 else 0

    summary_data = [
        ("Open Positions", n_open),
        ("Total Open Contracts", open_contracts),
        ("Total Credit Received", open_total_credit),
        ("Total Unrealized P&L", open_total_pnl),
        ("Total P&L %", open_pnl_pct),
        ("Total Max Risk", open_total_max_loss),
    ]

    for label, value in summary_data:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = Font(name="Calibri", size=10, bold=True)
        c1.fill = s["light_gray_fill"]
        c1.border = s["thin_border"]
        c2.font = s["data_font"]
        c2.border = s["thin_border"]
        c2.alignment = Alignment(horizontal="right")

        if "Credit" in label or "Risk" in label or ("P&L" in label and "%" not in label):
            c2.number_format = s["money_whole_fmt"]
        if "P&L %" in label:
            c2.number_format = s["pct_fmt"]
        if "P&L" in label and "%" not in label:
            c2.fill = s["green_fill"] if value >= 0 else s["red_fill"]

        row += 1

    # ========================================================================
    # CLOSED TRADES SUMMARY
    # ========================================================================
    closed_trades = [p for p in positions if p["status"] == "closed"]
    if closed_trades:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value="Closed Trades Summary")
        cell.font = s["pcs_section_font"]
        row += 1

        closed_pnl = sum(p.get("pnl") or 0 for p in closed_trades)
        closed_credit = sum(
            p["entry_credit"] * 100 * p["num_contracts"] for p in closed_trades)
        wins = sum(1 for p in closed_trades if (p.get("pnl") or 0) > 0)
        win_rate = wins / len(closed_trades) if closed_trades else 0

        closed_summary = [
            ("Closed Trades", len(closed_trades)),
            ("Win Rate", win_rate),
            ("Total Realized P&L", closed_pnl),
            ("Total Credit Collected", closed_credit),
        ]

        for label, value in closed_summary:
            c1 = ws.cell(row=row, column=1, value=label)
            c2 = ws.cell(row=row, column=2, value=value)
            c1.font = Font(name="Calibri", size=10, bold=True)
            c1.fill = s["light_gray_fill"]
            c1.border = s["thin_border"]
            c2.font = s["data_font"]
            c2.border = s["thin_border"]
            c2.alignment = Alignment(horizontal="right")

            if "Credit" in label or ("P&L" in label):
                c2.number_format = s["money_whole_fmt"]
            if "Win Rate" in label:
                c2.number_format = s["pct_fmt"]
            if "P&L" in label:
                c2.fill = s["green_fill"] if value >= 0 else s["red_fill"]

            row += 1

    # Column widths
    col_widths = {
        1: 5, 2: 12, 3: 8, 4: 12, 5: 12, 6: 6,
        7: 12, 8: 12, 9: 8,
        10: 10,
        11: 10,                # Credit/sh
        12: 12, 13: 12,       # SP Rcvd/Ctr, LP Paid/Ctr
        14: 10, 15: 10,       # TP Target, SL Trigger
        16: 11, 17: 14, 18: 13,  # Cur Spread, P&L ($), P&L (%)
        19: 13, 20: 12,       # Total Credit, Max Loss
        21: 12, 22: 10, 23: 14,  # Exit Date, Exit Debit, Exit Reason
        24: 12, 25: 11,       # Spot@Entry, Cur Spot
        26: 10, 27: 9,        # VIX@Entry, Cur VIX
        28: 10, 29: 9,        # IVR@Entry, Cur IVR
        30: 13, 31: 12,       # SMA200@Entry, Cur SMA200
        32: 40,                # Notes
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    return n_open, open_contracts, open_total_credit, open_total_pnl


# ============================================================================
# SHEET 3: TSLA BEAR PUT DEBIT SPREAD
# ============================================================================

def build_tsla_spread_sheet(ws, today):
    s = get_styles()
    t = TSLA_SPREAD
    tsla_price = get_price("TSLA", 250.0)
    exp_date = datetime.strptime(t["expiration"], "%Y-%m-%d").date()
    dte = (exp_date - today).days

    # Spread math
    spread_width = t["long_strike"] - t["short_strike"]  # 50
    max_profit = (spread_width - t["net_debit"]) * t["long_qty"] * 100  # $37,600
    max_loss = t["net_debit"] * t["long_qty"] * 100  # $12,400
    breakeven = t["long_strike"] - t["net_debit"]  # $287.60
    total_cost = t["net_debit"] * t["long_qty"] * 100

    # Get IV for BS pricing
    iv = get_tsla_iv_from_yfinance()

    # Current spread value via BS
    long_put_val = bs_put_price(tsla_price, t["long_strike"], dte, iv)
    short_put_val = bs_put_price(tsla_price, t["short_strike"], dte, iv)
    spread_value = long_put_val - short_put_val
    current_value = spread_value * t["long_qty"] * 100
    pnl = current_value - total_cost
    pnl_pct = pnl / total_cost if total_cost > 0 else 0

    # ========================================================================
    # TITLE
    # ========================================================================
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1,
                   value="TSLA Bear Put Debit Spread -- Position Tracker")
    cell.font = s["title_font"]
    cell.fill = s["bear_title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1,
                   value=f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
                         f"TSLA: ${tsla_price:.2f}  |  Account: {t['account']}")
    cell.font = Font(name="Calibri", size=10, italic=True, color="7B2020")
    cell.alignment = Alignment(horizontal="center")

    # ========================================================================
    # STRATEGY INFO
    # ========================================================================
    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="Strategy Info")
    cell.font = s["bear_section_font"]

    info = [
        ("Structure", t["structure"]),
        ("Long Leg", f"{t['long_qty']}x TSLA ${t['long_strike']}P @ ${t['long_entry_price']:.2f}"),
        ("Short Leg", f"{t['short_qty']}x TSLA ${t['short_strike']}P @ ${t['short_entry_price']:.2f}"),
        ("Spread Width", f"${spread_width:.2f}"),
        ("Net Debit", f"${t['net_debit']:.2f}/share  (${total_cost:,.0f} total)"),
        ("Max Profit", f"${max_profit:,.0f}  (at TSLA <= ${t['short_strike']})"),
        ("Max Loss", f"${max_loss:,.0f}  (at TSLA >= ${t['long_strike']})"),
        ("Breakeven", f"${breakeven:.2f}"),
    ]

    row += 1
    for label, value in info:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = Font(name="Calibri", size=10, bold=True)
        c1.fill = s["light_gray_fill"]
        c1.border = s["thin_border"]
        c2.font = s["data_font"]
        c2.fill = s["light_gray_fill"]
        c2.border = s["thin_border"]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # ========================================================================
    # POSITION TABLE
    # ========================================================================
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value="Position Detail")
    cell.font = s["bear_section_font"]
    row += 1

    headers = [
        "Entry Date", "Expiration", "DTE",
        "Long Leg", "Short Leg", "Net Debit",
        "TSLA @ Entry", "TSLA Current",
        "Spread Value", "Position Value",
        "P&L ($)", "P&L (%)",
    ]
    header_row = row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["bear_header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = s["thin_border"]
    ws.row_dimensions[row].height = 30

    row += 1
    values = [
        t["entry_date"], t["expiration"], dte,
        f"${t['long_strike']}P x{t['long_qty']}", f"${t['short_strike']}P x{t['short_qty']}",
        t["net_debit"],
        t["tsla_at_entry"], tsla_price,
        spread_value, current_value,
        pnl, pnl_pct,
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = s["data_font"]
        cell.border = s["thin_border"]
        cell.alignment = Alignment(horizontal="center")

    # Number formats
    for c in [6, 9]:
        ws.cell(row=row, column=c).number_format = s["money_fmt"]
    for c in [7, 8, 10, 11]:
        ws.cell(row=row, column=c).number_format = s["money_whole_fmt"]
    ws.cell(row=row, column=12).number_format = s["pct_fmt"]
    ws.cell(row=row, column=3).number_format = s["int_fmt"]

    pnl_cell = ws.cell(row=row, column=11)
    pnl_pct_cell = ws.cell(row=row, column=12)
    if pnl >= 0:
        pnl_cell.fill = s["green_fill"]
        pnl_pct_cell.fill = s["green_fill"]
    else:
        pnl_cell.fill = s["red_fill"]
        pnl_pct_cell.fill = s["red_fill"]

    # ========================================================================
    # GREEKS TABLE
    # ========================================================================
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="Net Greeks (x10 contracts x100 multiplier)")
    cell.font = s["bear_section_font"]
    row += 1

    greek_headers = ["Date", "Net Delta", "Net Gamma", "Net Theta ($/day)",
                     "Net Vega", "IV (Long)", "IV (Short)", "Source"]
    for col_idx, h in enumerate(greek_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["bear_header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = s["thin_border"]

    multiplier = t["long_qty"] * 100

    # Inception Greeks — try ThetaData first
    row += 1
    entry_greeks = get_thetadata_greeks("TSLA", t["expiration"], t["long_strike"], "P", t["entry_date"])
    entry_short_greeks = get_thetadata_greeks("TSLA", t["expiration"], t["short_strike"], "P", t["entry_date"])

    if entry_greeks and entry_short_greeks:
        net_delta = (entry_greeks["delta"] - entry_short_greeks["delta"]) * multiplier
        net_gamma = (entry_greeks["gamma"] - entry_short_greeks["gamma"]) * multiplier
        net_theta = (entry_greeks["theta"] - entry_short_greeks["theta"]) * multiplier
        net_vega = (entry_greeks["vega"] - entry_short_greeks["vega"]) * multiplier
        iv_long = entry_greeks["iv"]
        iv_short = entry_short_greeks["iv"]
        source = "ThetaData"
    else:
        # BS fallback using entry price
        entry_dte = (exp_date - datetime.strptime(t["entry_date"], "%Y-%m-%d").date()).days
        lg = bs_put_greeks(t["tsla_at_entry"], t["long_strike"], entry_dte, iv)
        sg = bs_put_greeks(t["tsla_at_entry"], t["short_strike"], entry_dte, iv)
        net_delta = (lg["delta"] - sg["delta"]) * multiplier
        net_gamma = (lg["gamma"] - sg["gamma"]) * multiplier
        net_theta = (lg["theta"] - sg["theta"]) * multiplier
        net_vega = (lg["vega"] - sg["vega"]) * multiplier
        iv_long = iv
        iv_short = iv
        source = "BS estimate"

    inception_vals = [t["entry_date"], net_delta, net_gamma, net_theta,
                      net_vega, iv_long, iv_short, source]
    for col_idx, val in enumerate(inception_vals, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = s["data_font"]
        cell.border = s["thin_border"]
        cell.alignment = Alignment(horizontal="center")
        cell.fill = s["light_gray_fill"]
    for c in [2, 3]:
        ws.cell(row=row, column=c).number_format = '0.0'
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    ws.cell(row=row, column=5).number_format = '0.00'
    for c in [6, 7]:
        ws.cell(row=row, column=c).number_format = '0.1%'

    # Current Greeks
    row += 1
    cur_long_greeks = get_thetadata_greeks("TSLA", t["expiration"], t["long_strike"], "P",
                                           today.strftime("%Y-%m-%d"))
    cur_short_greeks = get_thetadata_greeks("TSLA", t["expiration"], t["short_strike"], "P",
                                            today.strftime("%Y-%m-%d"))

    if cur_long_greeks and cur_short_greeks:
        net_delta = (cur_long_greeks["delta"] - cur_short_greeks["delta"]) * multiplier
        net_gamma = (cur_long_greeks["gamma"] - cur_short_greeks["gamma"]) * multiplier
        net_theta = (cur_long_greeks["theta"] - cur_short_greeks["theta"]) * multiplier
        net_vega = (cur_long_greeks["vega"] - cur_short_greeks["vega"]) * multiplier
        iv_long = cur_long_greeks["iv"]
        iv_short = cur_short_greeks["iv"]
        source = "ThetaData"
    else:
        lg = bs_put_greeks(tsla_price, t["long_strike"], dte, iv)
        sg = bs_put_greeks(tsla_price, t["short_strike"], dte, iv)
        net_delta = (lg["delta"] - sg["delta"]) * multiplier
        net_gamma = (lg["gamma"] - sg["gamma"]) * multiplier
        net_theta = (lg["theta"] - sg["theta"]) * multiplier
        net_vega = (lg["vega"] - sg["vega"]) * multiplier
        iv_long = iv
        iv_short = iv
        source = "BS estimate"

    current_vals = [today.strftime("%Y-%m-%d"), net_delta, net_gamma, net_theta,
                    net_vega, iv_long, iv_short, source]
    for col_idx, val in enumerate(current_vals, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = s["data_font"]
        cell.border = s["thin_border"]
        cell.alignment = Alignment(horizontal="center")
    for c in [2, 3]:
        ws.cell(row=row, column=c).number_format = '0.0'
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    ws.cell(row=row, column=5).number_format = '0.00'
    for c in [6, 7]:
        ws.cell(row=row, column=c).number_format = '0.1%'

    # ========================================================================
    # P&L SUMMARY
    # ========================================================================
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="P&L Summary")
    cell.font = s["bear_section_font"]
    row += 1

    dist_to_be = tsla_price - breakeven
    dist_to_be_pct = dist_to_be / tsla_price if tsla_price > 0 else 0
    dist_to_max = tsla_price - t["short_strike"]
    dist_to_max_pct = dist_to_max / tsla_price if tsla_price > 0 else 0

    summary = [
        ("Entry Cost", f"${total_cost:,.0f}"),
        ("Current Value", f"${current_value:,.0f}"),
        ("P&L ($)", f"${pnl:+,.0f}"),
        ("P&L (%)", f"{pnl_pct:+.1%}"),
        ("Distance to Breakeven ($287.60)", f"${dist_to_be:+,.2f}  ({dist_to_be_pct:+.1%})"),
        ("Distance to Max Profit ($250)", f"${dist_to_max:+,.2f}  ({dist_to_max_pct:+.1%})"),
    ]

    for label, value in summary:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = Font(name="Calibri", size=10, bold=True)
        c1.fill = s["light_gray_fill"]
        c1.border = s["thin_border"]
        c2.font = s["data_font"]
        c2.border = s["thin_border"]
        c2.alignment = Alignment(horizontal="right")
        if "P&L ($)" == label:
            c2.fill = s["green_fill"] if pnl >= 0 else s["red_fill"]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    # Column widths
    col_widths = {1: 16, 2: 16, 3: 8, 4: 16, 5: 16, 6: 12,
                  7: 14, 8: 14, 9: 14, 10: 14, 11: 14, 12: 12}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    return total_cost, current_value, pnl, pnl_pct


# ============================================================================
# SHEET 4: UPRO LONG — DD25%/COOL40
# ============================================================================

def build_upro_sheet(ws, today):
    s = get_styles()
    u = UPRO_POSITION
    upro_price = get_price("UPRO", 100.0)
    cost_basis = u["shares"] * u["entry_price"]
    current_value = u["shares"] * upro_price
    pnl = current_value - cost_basis
    pnl_pct = pnl / cost_basis if cost_basis > 0 else 0

    # ATH calculation — check yfinance history against known ATH
    ath = u["known_ath"]
    ath_date = u["known_ath_date"]
    try:
        tk = yf.Ticker("UPRO")
        hist = tk.history(period="max")
        if not hist.empty:
            yf_max = float(hist["Close"].max())
            if yf_max > ath:
                ath = yf_max
                ath_date = hist["Close"].idxmax()
                if hasattr(ath_date, "strftime"):
                    ath_date = ath_date.strftime("%Y-%m-%d")
                else:
                    ath_date = str(ath_date)[:10]
            print(f"UPRO ATH: ${ath:.2f} ({ath_date})")
    except Exception as e:
        print(f"UPRO ATH fetch error: {e}")

    drawdown = (upro_price / ath) - 1.0
    exit_trigger = ath * (1.0 - u["dd_threshold"])
    dist_to_exit = upro_price - exit_trigger
    dist_to_exit_pct = dist_to_exit / upro_price if upro_price > 0 else 0
    status = "EXIT SIGNAL" if upro_price <= exit_trigger else "IN"

    # ========================================================================
    # TITLE
    # ========================================================================
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1,
                   value="UPRO Long Position -- DD25%/Cool40 Strategy")
    cell.font = s["title_font"]
    cell.fill = s["upro_title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1,
                   value=f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
                         f"UPRO: ${upro_price:.2f}  |  Account: {u['account']}")
    cell.font = Font(name="Calibri", size=10, italic=True, color="1B5E20")
    cell.alignment = Alignment(horizontal="center")

    # ========================================================================
    # STRATEGY RULES
    # ========================================================================
    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="DD25%/Cool40 Strategy Rules")
    cell.font = s["upro_section_font"]

    rules = [
        ("Instrument", "UPRO (3x leveraged S&P 500)"),
        ("Exit Signal", f"UPRO closes >= {u['dd_threshold']:.0%} below all-time high"),
        ("Cooling Period", f"{u['cooling_period']} trading days out of market after exit"),
        ("Re-entry", "Buy UPRO after cooling period expires"),
        ("Note", "COOLING state requires manual update to UPRO_POSITION dict"),
    ]

    row += 1
    for label, value in rules:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = Font(name="Calibri", size=10, bold=True)
        c1.fill = s["light_gray_fill"]
        c1.border = s["thin_border"]
        c2.font = s["data_font"]
        c2.fill = s["light_gray_fill"]
        c2.border = s["thin_border"]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # ========================================================================
    # POSITION TABLE
    # ========================================================================
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="Position Detail")
    cell.font = s["upro_section_font"]
    row += 1

    headers = ["Shares", "Entry Date", "Entry Price", "Cost Basis",
               "Current Price", "Current Value", "P&L ($)", "P&L (%)"]
    header_row = row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["upro_header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = s["thin_border"]
    ws.row_dimensions[row].height = 30

    row += 1
    values = [u["shares"], u["entry_date"], u["entry_price"], cost_basis,
              upro_price, current_value, pnl, pnl_pct]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = s["data_font"]
        cell.border = s["thin_border"]
        cell.alignment = Alignment(horizontal="center")

    ws.cell(row=row, column=1).number_format = '#,##0'
    for c in [3]:
        ws.cell(row=row, column=c).number_format = s["money_fmt"]
    for c in [4, 5, 6, 7]:
        ws.cell(row=row, column=c).number_format = s["money_whole_fmt"]
    ws.cell(row=row, column=8).number_format = s["pct_fmt"]

    pnl_cell = ws.cell(row=row, column=7)
    pnl_pct_cell = ws.cell(row=row, column=8)
    if pnl >= 0:
        pnl_cell.fill = s["green_fill"]
        pnl_pct_cell.fill = s["green_fill"]
    else:
        pnl_cell.fill = s["red_fill"]
        pnl_pct_cell.fill = s["red_fill"]

    # ========================================================================
    # DD25%/COOL40 MONITORING
    # ========================================================================
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="DD25%/Cool40 Monitoring")
    cell.font = s["upro_section_font"]
    row += 1

    monitor_data = [
        ("All-Time High", f"${ath:.2f}"),
        ("ATH Date", str(ath_date)),
        ("Current Drawdown", f"{drawdown:+.1%}"),
        ("Exit Trigger Level", f"${exit_trigger:.2f}  (ATH x 0.75)"),
        ("Distance to Exit ($)", f"${dist_to_exit:+,.2f}"),
        ("Distance to Exit (%)", f"{dist_to_exit_pct:+.1%}"),
        ("Status", status),
    ]

    for label, value in monitor_data:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = Font(name="Calibri", size=10, bold=True)
        c1.fill = s["light_gray_fill"]
        c1.border = s["thin_border"]
        c2.font = s["data_font"]
        c2.border = s["thin_border"]
        c2.alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

        if label == "Status":
            c2.font = Font(name="Calibri", size=12, bold=True,
                           color="006100" if status == "IN" else "9C0006")
            c2.fill = s["green_fill"] if status == "IN" else s["red_fill"]
        if label == "Current Drawdown" and drawdown <= -0.20:
            c2.fill = s["yellow_fill"]
            c2.font = Font(name="Calibri", size=10, bold=True, color="9C6500")

        row += 1

    # Column widths
    col_widths = {1: 20, 2: 16, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 12, 9: 12, 10: 12}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    return u["shares"], cost_basis, current_value, pnl, status


# ============================================================================
# SHEET 5: BEAR PUT SPREAD CANDIDATES
# ============================================================================

def build_bear_candidates_sheet(ws, today):
    """Build a watchlist/tracker sheet for bear put spread candidates."""
    s = get_styles()

    # Fetch live prices for all candidates
    symbols = list({c["symbol"] for c in BEAR_SPREAD_CANDIDATES})
    prices = {}
    for sym in symbols:
        prices[sym] = get_price(sym, 0.0)

    # Get IV for BS pricing
    ivs = {}
    for sym in symbols:
        try:
            tk = yf.Ticker(sym)
            exps = tk.options
            if exps:
                chain = tk.option_chain(exps[0])
                puts = chain.puts
                spot = prices[sym]
                if spot > 0 and not puts.empty:
                    atm_idx = (puts["strike"] - spot).abs().idxmin()
                    iv_val = puts.loc[atm_idx, "impliedVolatility"]
                    if iv_val and iv_val > 0:
                        ivs[sym] = float(iv_val)
                        continue
        except Exception:
            pass
        ivs[sym] = 0.30  # fallback

    # ── TITLE ──
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    cell = ws.cell(row=row, column=1, value="Bear Put Spread Candidates -- Watchlist")
    cell.font = s["title_font"]
    cell.fill = s["bear_title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    cell = ws.cell(row=row, column=1,
                   value=f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
                         f"Screened: 2026-03-11  |  Source: IC-derived scoring + GuruFocus")
    cell.font = Font(name="Calibri", size=10, italic=True, color="7B2020")
    cell.alignment = Alignment(horizontal="center")

    # ── SUMMARY TABLE ──
    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    cell = ws.cell(row=row, column=1, value="Candidate Overview")
    cell.font = s["bear_section_font"]
    row += 1

    headers = [
        "Symbol", "Name", "Sector", "Status",
        "Long Put", "Short Put", "Exp", "Score", "R:R",
        "Current Price", "Spread Value", "Est. Debit",
        "Max Loss/Ct", "Max Profit/Ct",
    ]
    header_row = row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["bear_header_fill"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = s["thin_border"]
    ws.row_dimensions[row].height = 30

    total_est_risk = 0

    for cand in BEAR_SPREAD_CANDIDATES:
        row += 1
        sym = cand["symbol"]
        spot = prices.get(sym, 0)
        iv = ivs.get(sym, 0.30)
        exp_date = datetime.strptime(cand["expiration"], "%Y-%m-%d").date()
        dte = (exp_date - today).days

        # BS spread value
        if spot > 0 and dte > 0:
            long_val = bs_put_price(spot, cand["long_strike"], dte, iv)
            short_val = bs_put_price(spot, cand["short_strike"], dte, iv)
            spread_val = long_val - short_val
        else:
            spread_val = 0

        width = cand["long_strike"] - cand["short_strike"]
        est_debit = spread_val
        max_loss = est_debit * 100
        max_profit = (width - est_debit) * 100

        if cand["status"] == "ENTERED" and cand["net_debit"] > 0:
            est_debit = cand["net_debit"]
            max_loss = cand["net_debit"] * cand["qty"] * 100
            max_profit = (width - cand["net_debit"]) * cand["qty"] * 100

        total_est_risk += max_loss

        values = [
            sym, cand["name"], cand["sector"], cand["status"],
            f"${cand['long_strike']}P", f"${cand['short_strike']}P",
            cand["expiration"], cand["screener_score"], cand["rr_ratio"],
            spot, spread_val, est_debit,
            max_loss, max_profit,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = s["data_font"]
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")

        # Number formats
        ws.cell(row=row, column=8).number_format = '0.0'
        ws.cell(row=row, column=9).number_format = '0.00'
        ws.cell(row=row, column=10).number_format = '$#,##0.00'
        ws.cell(row=row, column=11).number_format = '#,##0.00'
        ws.cell(row=row, column=12).number_format = '#,##0.00'
        ws.cell(row=row, column=13).number_format = '$#,##0'
        ws.cell(row=row, column=14).number_format = '$#,##0'

        # Status coloring
        status_cell = ws.cell(row=row, column=4)
        if cand["status"] == "ENTERED":
            status_cell.fill = s["green_fill"]
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="006100")
        elif cand["status"] == "WATCHLIST":
            status_cell.fill = s["yellow_fill"]
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="9C6500")

    # ── THESIS SECTION ──
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    cell = ws.cell(row=row, column=1, value="Bear Thesis Summary")
    cell.font = s["bear_section_font"]
    row += 1

    for cand in BEAR_SPREAD_CANDIDATES:
        c1 = ws.cell(row=row, column=1, value=cand["symbol"])
        c1.font = Font(name="Calibri", size=10, bold=True)
        c1.fill = s["light_gray_fill"]
        c1.border = s["thin_border"]
        c2 = ws.cell(row=row, column=2, value=cand["thesis"])
        c2.font = s["data_font"]
        c2.border = s["thin_border"]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        row += 1

    # ── PORTFOLIO RISK SUMMARY ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="Portfolio Risk Summary")
    cell.font = s["bear_section_font"]
    row += 1

    # Include TSLA existing position
    tsla_risk = TSLA_SPREAD["net_debit"] * TSLA_SPREAD["long_qty"] * 100

    risk_items = [
        ("TSLA Bear Put Spread (existing)", f"${tsla_risk:,.0f}"),
        ("Candidates (est. 1 contract each)", f"${total_est_risk:,.0f}"),
        ("Total Bear Overlay Risk", f"${tsla_risk + total_est_risk:,.0f}"),
    ]

    for label, value in risk_items:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = Font(name="Calibri", size=10, bold=True)
        c1.fill = s["light_gray_fill"]
        c1.border = s["thin_border"]
        c2.font = s["data_font"]
        c2.border = s["thin_border"]
        c2.alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    # Column widths
    col_widths = {1: 10, 2: 24, 3: 20, 4: 12, 5: 10, 6: 10, 7: 12,
                  8: 8, 9: 8, 10: 14, 11: 14, 12: 12, 13: 14, 14: 14}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    return len(BEAR_SPREAD_CANDIDATES), total_est_risk


# ============================================================================
# IRA EQUITY POSITIONS (Software Thesis + Other Holdings)
# ============================================================================

IRA_POSITIONS = [
    {
        "ticker": "WDAY",
        "name": "Workday Inc",
        "shares": 4000,
        "entry_price": 135.86,
        "entry_date": "2026-03-20",
        "thesis": "Software meltdown — AI panic overdone for HCM/finance. 18% implied WACC on Baa1 credit.",
        "project": "P-11 (Software Meltdown)",
        "falsification": "NRR < 110%, NRR decel > 5pts QoQ, or F500 full platform replacement",
    },
]


def build_ira_positions_sheet(ws, today):
    """Build sheet tracking IRA equity positions (conviction picks)."""
    s = get_styles()

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1, value="IRA Equity Positions -- Conviction Picks")
    cell.font = Font(name="Calibri", size=14, bold=True, color="1A3C6E")
    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1,
            value=f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = \
        Font(name="Calibri", size=10, italic=True)

    # Headers
    row = 4
    headers = ["Ticker", "Name", "Shares", "Entry Price", "Entry Date",
               "Current Price", "Market Value", "P&L ($)", "P&L (%)", "Thesis"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    total_cost = 0
    total_value = 0

    for pos in IRA_POSITIONS:
        row += 1
        current_price = get_price(pos["ticker"], pos["entry_price"])
        shares = pos["shares"]
        entry_price = pos["entry_price"]
        market_value = shares * current_price
        cost = shares * entry_price
        pnl = market_value - cost
        pnl_pct = pnl / cost if cost > 0 else 0

        total_cost += cost
        total_value += market_value

        ws.cell(row=row, column=1, value=pos["ticker"]).font = Font(bold=True)
        ws.cell(row=row, column=2, value=pos["name"])
        ws.cell(row=row, column=3, value=shares).number_format = "#,##0"
        ws.cell(row=row, column=4, value=entry_price).number_format = "$#,##0.00"
        ws.cell(row=row, column=5, value=pos["entry_date"])
        ws.cell(row=row, column=6, value=current_price).number_format = "$#,##0.00"
        ws.cell(row=row, column=7, value=market_value).number_format = "$#,##0.00"

        pnl_cell = ws.cell(row=row, column=8, value=pnl)
        pnl_cell.number_format = "$#,##0.00"
        pnl_cell.font = Font(color="1B7A2B" if pnl >= 0 else "C62828")

        pct_cell = ws.cell(row=row, column=9, value=pnl_pct)
        pct_cell.number_format = "0.0%"
        pct_cell.font = Font(color="1B7A2B" if pnl_pct >= 0 else "C62828")

        ws.cell(row=row, column=10, value=pos["thesis"][:60])

    # Totals
    row += 2
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_value).number_format = "$#,##0.00"
    ws.cell(row=row, column=7).font = Font(bold=True)
    total_pnl = total_value - total_cost
    ws.cell(row=row, column=8, value=total_pnl).number_format = "$#,##0.00"
    ws.cell(row=row, column=8).font = Font(bold=True,
                                            color="1B7A2B" if total_pnl >= 0 else "C62828")

    # Falsification criteria section
    row += 3
    ws.cell(row=row, column=1, value="Falsification Criteria").font = \
        Font(name="Calibri", size=12, bold=True, color="1A3C6E")
    for pos in IRA_POSITIONS:
        row += 1
        ws.cell(row=row, column=1, value=pos["ticker"]).font = Font(bold=True)
        ws.cell(row=row, column=2, value=pos["falsification"])

    # Column widths
    widths = [8, 20, 10, 12, 12, 12, 14, 14, 10, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    n_positions = len(IRA_POSITIONS)
    total_shares = sum(p["shares"] for p in IRA_POSITIONS)
    return n_positions, total_shares, total_cost, total_value


# ============================================================================
# MAIN: BUILD WORKBOOK WITH ALL SHEETS
# ============================================================================

def build_spreadsheet(output_path):
    today = date.today()
    spy_price = get_price("SPY", 600.0)
    qqq_price = get_price("QQQ", 500.0)
    price_map = {"SPY": spy_price, "QQQ": qqq_price}

    wb = openpyxl.Workbook()

    # Compute SMAs for 80-delta sheet header
    delta_sma_map = {}
    for sym in ["SPY", "QQQ"]:
        sma = compute_sma200(sym)
        if sma:
            delta_sma_map[sym] = sma

    # Sheet 1: 80-Delta Calls
    ws1 = wb.active
    ws1.title = "80-Delta Calls"
    delta_results = build_80delta_sheet(ws1, price_map, delta_sma_map, today)

    # Sheet 2: PCS Paper Trades
    ws2 = wb.create_sheet("PCS Paper Trades")
    pcs_results = build_pcs_sheet(ws2, spy_price, today,
                                  db_path=PCS_DB_PATH)

    # Sheet 3: PCS Live Trades
    ws3 = wb.create_sheet("PCS Live Trades")
    pcs_live_results = build_pcs_sheet(ws3, spy_price, today,
                                       db_path=PCS_LIVE_DB_PATH)

    # Sheet 4: TSLA Bear Put Spread
    ws4 = wb.create_sheet("TSLA Bear Put Spread")
    tsla_results = build_tsla_spread_sheet(ws4, today)

    # Sheet 5: UPRO DD25/Cool40
    ws5 = wb.create_sheet("UPRO DD25-Cool40")
    upro_results = build_upro_sheet(ws5, today)

    # Sheet 6: IRA Equity Positions (Software Thesis + Other)
    ws6 = wb.create_sheet("IRA Equity Positions")
    ira_results = build_ira_positions_sheet(ws6, today)

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    return (delta_results, pcs_results, pcs_live_results,
            tsla_results, upro_results, ira_results)


if __name__ == "__main__":
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "position_tracker.xlsx")
    (delta_results, pcs_results, pcs_live_results,
     tsla_results, upro_results, ira_results) = build_spreadsheet(out_path)

    cost, value, pnl, delta_exp, contracts = delta_results
    print(f"\n--- 80-Delta Calls ---")
    print(f"Positions: {len(TRADES)}  |  Contracts: {contracts}")
    print(f"Cost: ${cost:,.0f}  |  Value: ${value:,.0f}  |  P&L: ${pnl:+,.0f}")
    print(f"Delta: {delta_exp:.0f} (~{int(delta_exp)} SPY shares)")

    n_open, pcs_contracts, pcs_credit, pcs_pnl = pcs_results
    print(f"\n--- PCS Paper Trades ---")
    print(f"Open: {n_open}  |  Contracts: {pcs_contracts}")
    print(f"Credit: ${pcs_credit:,.0f}  |  P&L: ${pcs_pnl:+,.0f}")

    n_live, live_contracts, live_credit, live_pnl = pcs_live_results
    print(f"\n--- PCS Live Trades ---")
    print(f"Open: {n_live}  |  Contracts: {live_contracts}")
    print(f"Credit: ${live_credit:,.0f}  |  P&L: ${live_pnl:+,.0f}")

    tsla_cost, tsla_val, tsla_pnl, tsla_pct = tsla_results
    print(f"\n--- TSLA Bear Put Spread ---")
    print(f"Cost: ${tsla_cost:,.0f}  |  Value: ${tsla_val:,.0f}  |  P&L: ${tsla_pnl:+,.0f} ({tsla_pct:+.1%})")

    upro_shares, upro_cost, upro_val, upro_pnl, upro_status = upro_results
    print(f"\n--- UPRO DD25/Cool40 ---")
    print(f"Shares: {upro_shares:,}  |  Cost: ${upro_cost:,.0f}  |  Value: ${upro_val:,.0f}  |  P&L: ${upro_pnl:+,.0f}")
    print(f"Status: {upro_status}")

    n_ira, ira_shares, ira_cost, ira_value = ira_results
    print(f"\n--- IRA Equity Positions ---")
    print(f"Positions: {n_ira}  |  Shares: {ira_shares:,}")
    print(f"Cost: ${ira_cost:,.0f}  |  Value: ${ira_value:,.0f}  |  P&L: ${ira_value-ira_cost:+,.0f}")
